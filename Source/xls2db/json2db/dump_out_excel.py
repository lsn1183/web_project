# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.styles import Border,Alignment, Font, Side
from openpyxl.styles.colors import BLUE
import json

def make_cover_sheet(spec_json, ws):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=36, bold=True)
    FontS2 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=16, bold=True)
    AlignS1 =  Alignment(horizontal='center',vertical='center')
    ws.title = "表紙".decode("utf8")
    ws.sheet_view.showGridLines = False
    for col_letter in  get_column_interval(1,10000):
        ws.column_dimensions[col_letter].width=2.7

    start_index = 16
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index+2, end_column=31)
    cell = ws.cell(column=2, row=start_index)
    cell.value = spec_json["chapterInfo"]
    cell.font = FontS1
    cell.alignment = AlignS1
    start_index += 4
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index + 2, end_column=31)
    cell = ws.cell(column=2, row=start_index)
    cell.value = spec_json["titleInfo"]
    cell.font = FontS1
    cell.alignment = AlignS1
    start_index += 4
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index + 2, end_column=31)
    cell = ws.cell(column=2, row=start_index)
    cell.value = spec_json["verInfo"]
    cell.font = FontS2
    cell.alignment = AlignS1

def make_cat_sheet(spec_json, wb):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False)
    FontS2 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=True)
    FontS3 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False, color=BLUE)
    Borderthin = Side(border_style="thin", color="000000")
    ws = wb.create_sheet("目次".decode("utf8"),1)
    ws.sheet_view.showGridLines = False
    for col_letter in get_column_interval(1, 10000):
        ws.column_dimensions[col_letter].width = 2.7

    start_index = 7
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index, end_column=5)
    ws.merge_cells(start_row=start_index, start_column=6, end_row=start_index, end_column=11)
    ws.merge_cells(start_row=start_index, start_column=12, end_row=start_index, end_column=17)
    ws.merge_cells(start_row=start_index, start_column=18, end_row=start_index, end_column=29)
    for iborder in range(2,30):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=2, row=start_index)
    cell.value = "項番".decode("utf8")
    cell.font = FontS2
    cell = ws.cell(column=6, row=start_index)
    cell.value = "大項目".decode("utf8")
    cell.font = FontS2
    cell = ws.cell(column=12, row=start_index)
    cell.value = "中項目".decode("utf8")
    cell.font = FontS2
    cell = ws.cell(column=18, row=start_index)
    cell.value = "小項目".decode("utf8")
    cell.font = FontS2

    start_index+=1
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index, end_column=5)
    ws.merge_cells(start_row=start_index, start_column=6, end_row=start_index, end_column=29)
    for iborder in range(2,30):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=2, row=start_index)
    cell.value = "2.3."
    cell.font = FontS1
    cell = ws.cell(column=6, row=start_index)
    cell.value = "概要".decode("utf8")
    cell.font = FontS3
    cell.hyperlink = "#2.3.概要!B2".decode("utf8")

    start_index += 1
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index, end_column=5)
    ws.merge_cells(start_row=start_index, start_column=6, end_row=start_index, end_column=29)
    for iborder in range(2, 30):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=2, row=start_index)
    cell.value = "2.3.0."
    cell.font = FontS1
    cell = ws.cell(column=6, row=start_index)
    cell.value = "用語定義".decode("utf8")
    cell.font = FontS3
    cell.hyperlink = "#2.3.0.用語定義!B2".decode("utf8")

    start_index += 1
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index, end_column=5)
    ws.merge_cells(start_row=start_index, start_column=6, end_row=start_index, end_column=29)
    for iborder in range(2, 30):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=2, row=start_index)
    cell.value = "2.3.1."
    cell.font = FontS1
    cell = ws.cell(column=6, row=start_index)
    cell.value = "案内使用言語".decode("utf8")
    cell.font = FontS3
    cell.hyperlink = "#2.3.1.案内使用言語!B2".decode("utf8")
    for iborder in range(2, 12):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(left=Borderthin, right=Borderthin,top=Borderthin)

    start_index += 1
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index, end_column=5)
    ws.merge_cells(start_row=start_index, start_column=6, end_row=start_index, end_column=11)
    ws.merge_cells(start_row=start_index, start_column=12, end_row=start_index, end_column=29)
    for iborder in range(2, 30):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=2, row=start_index)
    cell.value = "2.3.1.1."
    cell.font = FontS1
    cell = ws.cell(column=12, row=start_index)
    cell.value = "案内音声言語".decode("utf8")
    cell.font = FontS3
    cell.hyperlink = "#2.3.1.1.案内音声言語!B2".decode("utf8")
    for iborder in range(2, 12):
        cell = ws.cell(column=iborder, row=start_index)
        cell.border = Border(left=Borderthin, right=Borderthin,bottom=Borderthin)



def make_def_sheet(spec_json, wb):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False)
    FontS2 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=True)
    Borderthin = Side(border_style="thin", color="000000")
    AlignS1 = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws = wb.create_sheet(spec_json["chapterInfo"][1:-1]+".0."+"用語定義".decode("utf8"))
    ws.sheet_view.showGridLines = False
    for col_letter in get_column_interval(1, 10000):
        ws.column_dimensions[col_letter].width = 2.7

    start_index = 2
    ws.merge_cells(start_row=start_index, start_column=2, end_row=start_index, end_column=5)
    cell = ws.cell(column=2, row=start_index)
    cell.value = spec_json["chapterInfo"][1:-1]+".0."
    cell.font = FontS2
    for iborer in range(2,6):
        cell = ws.cell(column=iborer, row=start_index)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)

    cell = ws.cell(column=6, row=start_index)
    cell.value = "用語定義".decode("utf8")
    cell.font = FontS2

    start_index+=2
    for def_info in spec_json["nameDefList"]:
        nameDefList = def_info["nameDef"]
        nameInfo = def_info["nameInfo"]
        ws.merge_cells(start_row=start_index, start_column=4,
                       end_row=start_index+len(nameDefList)-1, end_column=8)
        for iRow in range(0,len(nameDefList)):
            ws.merge_cells(start_row=start_index+ iRow, start_column=9,
                           end_row=start_index + iRow, end_column=29)
            if iRow == 0:
                cell = ws.cell(column=4, row=start_index+iRow)
                cell.value = nameInfo["dataValue"][0]
                cell.alignment = AlignS1
                cell = ws.cell(column=9, row=start_index + iRow)
                cell.value = nameDefList[iRow]["dataValue"][0]
                cell.alignment = AlignS1
            else:
                cell = ws.cell(column=9, row=start_index + iRow)
                cell.value = nameDefList[iRow]["dataValue"][0]
                cell.alignment = AlignS1

        for iborer in range(4, 30):
            iRow = 0
            cell = ws.cell(column=iborer, row=start_index+iRow)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin)
            cell.font = FontS1
            iRow+=1
            while (iRow<len(nameDefList)):
                cell = ws.cell(column=iborer, row=start_index + iRow)
                cell.border = Border(left=Borderthin, right=Borderthin)
                cell.font = FontS1
                iRow += 1

            cell = ws.cell(column=iborer, row=start_index + iRow - 1)
            cell.font = FontS1
            if iRow > 1:
                cell.border = Border(bottom=Borderthin, left=Borderthin, right=Borderthin)
            else:
                cell.border = Border(top=Borderthin, bottom=Borderthin, left=Borderthin, right=Borderthin)

        start_index += len(nameDefList)

def make_func_sheet(spec_json, wb):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False)
    FontS2 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=True)
    Borderthin = Side(border_style="thin", color="000000")

    for fun_info in spec_json["chapterInfoList"]:
        title_info = fun_info["chapterInfo"]+fun_info["titleInfo"]
        ws = wb.create_sheet(title_info)
        ws.sheet_view.showGridLines = False
        for col_letter in get_column_interval(1, 10000):
            ws.column_dimensions[col_letter].width = 2.7

        row_index = 2
        col_index = 2
        ws.merge_cells(start_row=row_index, start_column=col_index,
                       end_row=row_index, end_column=col_index+3)
        cell = ws.cell(column=col_index, row=row_index)
        cell.value = fun_info["chapterInfo"]
        cell.font = FontS2
        for iborer in range(col_index, col_index+4):
            cell = ws.cell(column=iborer, row=row_index)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
        cell = ws.cell(column=col_index+4, row=row_index)
        cell.value = fun_info["titleInfo"]
        cell.font = FontS2

        row_index+=2
        col_index+=1
        ws.merge_cells(start_row=row_index, start_column=col_index,
                       end_row=row_index, end_column=col_index + 1)
        cell = ws.cell(column=col_index, row=row_index)
        cell.value = "概要".decode("utf8")
        cell.font = FontS1
        for iborer in range(col_index, col_index+2):
            cell = ws.cell(column=iborer, row=row_index)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
        for iRow in range(0,len(fun_info["summaryInfo"])):
            cell = ws.cell(column=col_index+2, row=row_index)
            cell.value = fun_info["summaryInfo"][iRow]["dataValue"][0]
            cell.font = FontS1
            row_index+=1

        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=col_index,
                       end_row=row_index, end_column=col_index + 1)
        cell = ws.cell(column=col_index, row=row_index)
        cell.value = "前提".decode("utf8")
        cell.font = FontS1
        for iborer in range(col_index, col_index + 2):
            cell = ws.cell(column=iborer, row=row_index)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
        for iRow in range(0, len(fun_info["preCondition"])):
            cell = ws.cell(column=col_index + 2, row=row_index)
            cell.value = fun_info["preCondition"][iRow]["dataValue"][0]
            cell.font = FontS1
            row_index += 1

        row_index += 1
        ws.merge_cells(start_row=row_index, start_column=col_index,
                       end_row=row_index, end_column=col_index + 1)
        cell = ws.cell(column=col_index, row=row_index)
        cell.value = "呼出".decode("utf8")
        cell.font = FontS1
        for iborer in range(col_index, col_index + 2):
            cell = ws.cell(column=iborer, row=row_index)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
        for iRow in range(0, len(fun_info["callInfo"])):
            cell = ws.cell(column=col_index + 2, row=row_index)
            cell.value = fun_info["callInfo"][iRow]["dataValue"][0]
            cell.font = FontS1
            row_index += 1

        row_index += 1
        make_func_list(ws, fun_info["appRange"], row_index,col_index)


def make_func_list(ws, func_dict_list, start_row, start_col):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False)
    Borderthin = Side(border_style="thin", color="000000")
    w_row = start_row
    w_col = start_col

    for func_dict in func_dict_list:
        cell = ws.cell(column=w_col, row=w_row)
        cell.font = FontS1

        w_col_e = 2
        if func_dict["funcType"] == "APPRANGE":
            cell.value = "適合範囲".decode("utf8")
            w_col_e = 3
        elif func_dict["funcType"] == "FUNCTION":
            cell.value = "機能".decode("utf8")
        elif func_dict["funcType"] == "SUPPLY":
            cell.value = "補足".decode("utf8")
        elif func_dict["funcType"] == "EXCEPT":
            cell.value = "例外".decode("utf8")
        elif func_dict["funcType"] == "DESTINATION":
            cell.value = "仕向".decode("utf8")
        elif func_dict["funcType"] == "TABLE":
            w_row = make_func_table(ws, func_dict["funContent"],w_row, w_col)
            if len(func_dict["subFuncList"]) > 0:
                w_row = make_func_list(ws, func_dict["subFuncList"], w_row + 1, w_col + 1)
            w_row += 1
            continue

        ws.merge_cells(start_row=w_row, start_column=w_col,
                       end_row=w_row, end_column=w_col + w_col_e-1)
        for iborer in range(w_col, w_col + w_col_e):
            cell = ws.cell(column=iborer, row=w_row)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
        for iRow in range(0, len(func_dict["funContent"])):
            cell = ws.cell(column=w_col + w_col_e, row=w_row)
            cell.value = func_dict["funContent"][iRow]["dataValue"][0]
            cell.font = FontS1
            w_row += 1

        if len(func_dict["funContent"]) == 0:
            w_row += 1

        if len(func_dict["subFuncList"]) > 0:
            w_row = make_func_list(ws, func_dict["subFuncList"], w_row+1, w_col+1)

        w_row += 1

    return w_row

def make_func_table(ws, table_content_list, start_row, start_col):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False)
    Borderthin = Side(border_style="thin", color="000000")

    w_row = start_row
    w_col = start_col
    ws.merge_cells(start_row=w_row, start_column=w_col,
                   end_row=w_row, end_column=w_col+1)
    cell = ws.cell(column=w_col, row=w_row)
    cell.font = FontS1
    cell.value = "表".decode("utf8")
    for iborer in range(w_col, w_col+2):
        cell = ws.cell(column=iborer, row=w_row)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=w_col + 2, row=w_row)
    cell.value = table_content_list[0]["dataValue"][0]
    cell.font = FontS1

    w_row+=2
    max_str_list = [0]*len(table_content_list[2]["dataValue"])
    row_line_list = [1]*(len(table_content_list)-2)
    for ir in range(2, len(table_content_list)):
        for ic in range(0,len(table_content_list[ir]["dataValue"])):
            tmpContent = table_content_list[ir]["dataValue"][ic]
            if len(tmpContent.split("\\n")) > row_line_list[ir-2]:
                row_line_list[ir-2] = len(tmpContent.split("\\n"))
            for tmpData in tmpContent.split("\\n"):
                if len(tmpData) > max_str_list[ic]:
                    max_str_list[ic] = len(tmpData)
    col_merge_list = []
    for d in max_str_list:
        dd = int(d/1.5+1.5)
        if dd > 20:
            dd=20
        col_merge_list.append(dd)

    tmp_row = w_row
    for ir in range(0, sum(row_line_list)):
        col_merge_offset = 0
        for imerge in range(0, len(col_merge_list)):
            ws.merge_cells(start_row=tmp_row, start_column=w_col + col_merge_offset,
                       end_row=tmp_row, end_column=w_col + col_merge_offset + col_merge_list[imerge]-1)
            col_merge_offset+=col_merge_list[imerge]
        tmp_row+=1

    tmp_row = w_row
    for ir_number in row_line_list:
        if ir_number == 1:
            for iborder in range(0, sum(col_merge_list)):
                cell = ws.cell(column=w_col+iborder, row=tmp_row)
                cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
            tmp_row+=1
        else:
            for iborder in range(0, sum(col_merge_list)):
                cell = ws.cell(column=w_col+iborder, row=tmp_row)
                cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin)
            tmp_row+=1
            for ii in range(1,ir_number-1):
                for iborder in range(0, sum(col_merge_list)):
                    cell = ws.cell(column=w_col + iborder, row=tmp_row)
                    cell.border = Border(left=Borderthin, right=Borderthin)
                tmp_row += 1
            for iborder in range(0, sum(col_merge_list)):
                cell = ws.cell(column=w_col + iborder, row=tmp_row)
                cell.border = Border(left=Borderthin, right=Borderthin, bottom=Borderthin)
            tmp_row += 1

    tmp_row = w_row
    for ir in range(2, len(table_content_list)):
        col_merge_offset = 0
        for ic in range(0,len(col_merge_list)):
            tmpContent = table_content_list[ir]["dataValue"][ic]
            for ii,tmpData in enumerate(tmpContent.split("\\n"),0):
                cell = ws.cell(column=w_col + col_merge_offset, row=tmp_row+ii)
                cell.value = tmpData
                cell.font = FontS1
            col_merge_offset += col_merge_list[ic]
        tmp_row += row_line_list[ir-2]

    w_row+=sum(row_line_list)

    return w_row

def make_history_sheet(spec_json, wb):
    FontS1 = Font(name='ＭＳ ゴシック'.decode("utf8"), size=10.5, bold=False)
    Borderthin = Side(border_style="thin", color="000000")

    if not "history" in spec_json.keys():
        return
    ws = wb.create_sheet("更新履歴".decode("utf8"))
    ws.sheet_view.showGridLines = False
    for col_letter in get_column_interval(1, 10000):
        ws.column_dimensions[col_letter].width = 2.7

    s_row=2
    ws.merge_cells(start_row=s_row, start_column=2,
                   end_row=s_row, end_column=3)
    ws.merge_cells(start_row=s_row, start_column=4,
                   end_row=s_row, end_column=10)
    ws.merge_cells(start_row=s_row, start_column=11,
                   end_row=s_row, end_column=23)
    ws.merge_cells(start_row=s_row, start_column=24,
                   end_row=s_row, end_column=36)
    ws.merge_cells(start_row=s_row, start_column=37,
                   end_row=s_row, end_column=41)
    for iBorder in range(2,42):
        cell = ws.cell(column=iBorder, row=s_row)
        cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
    cell = ws.cell(column=2, row=s_row)
    cell.value = "No."
    cell.font = FontS1
    cell = ws.cell(column=4, row=s_row)
    cell.value = "項番".decode("utf8")
    cell.font = FontS1
    cell = ws.cell(column=11, row=s_row)
    cell.value = "更新前内容".decode("utf8")
    cell.font = FontS1
    cell = ws.cell(column=24, row=s_row)
    cell.value = "更新後内容".decode("utf8")
    cell.font = FontS1
    cell = ws.cell(column=37, row=s_row)
    cell.value = "更新日".decode("utf8")
    cell.font = FontS1

    s_row +=1
    for i_index,history_data in enumerate(spec_json["history"],1):
        ws.merge_cells(start_row=s_row, start_column=2,
                       end_row=s_row, end_column=3)
        ws.merge_cells(start_row=s_row, start_column=4,
                       end_row=s_row, end_column=10)
        ws.merge_cells(start_row=s_row, start_column=11,
                       end_row=s_row, end_column=23)
        ws.merge_cells(start_row=s_row, start_column=24,
                       end_row=s_row, end_column=36)
        ws.merge_cells(start_row=s_row, start_column=37,
                       end_row=s_row, end_column=41)
        for iBorder in range(2, 42):
            cell = ws.cell(column=iBorder, row=s_row)
            cell.border = Border(top=Borderthin, left=Borderthin, right=Borderthin, bottom=Borderthin)
        cell = ws.cell(column=2, row=s_row)
        cell.value = "%d"%i_index
        cell.font = FontS1

        cell = ws.cell(column=4, row=s_row)
        cell.value=history_data["chapterInfo"]
        cell.font = FontS1

        cell = ws.cell(column=11, row=s_row)
        old_str = ""
        for o_k, o_v in history_data["old_content"].iteritems():
            old_str += o_v[0]["dataValue"][0] + "\n"
        cell.value = old_str
        cell.font = FontS1

        cell = ws.cell(column=24, row=s_row)
        new_str = ""
        for n_k, n_v in history_data["new_content"].iteritems():
            new_str += n_v[0]["dataValue"][0] + "\n"
        cell.value = new_str
        cell.font = FontS1

        cell = ws.cell(column=37, row=s_row)
        cell.value = history_data["modify_date"]
        cell.font = FontS1
        s_row += 1


if __name__ == '__main__':
    with open("spec.json", 'rb') as f:
        spec_json = json.load(f)

    wb = Workbook()
    ws = wb.active
    #make cover sheet
    make_cover_sheet(spec_json, ws)
    make_def_sheet(spec_json, wb)
    make_func_sheet(spec_json, wb)
    make_cat_sheet(spec_json, wb)
    make_history_sheet(spec_json, wb)


    wb.save("./o_"+spec_json["fileName"].replace("xlsm", "xlsx"))





