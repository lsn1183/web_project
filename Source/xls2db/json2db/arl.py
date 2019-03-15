# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import os
import re
import datetime
import json
from Source.xls2db.common import log
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from spec import SpecBase
PROJ_NAME = 'proj_name'
SPEC_NAME = 'spec_name'
SPEC_VERSION = 'version'
SPEC_FILE_NAME = 'file_name'
RECORD_START_ROW = 6  # 记录开始行号


class SpecArl(SpecBase):
    """
    """
    def __init__(self):
        SpecBase.__init__(self)
        self._path = None
        self._log = log.common_log.instance().logger('REQUIREMENT')

    def set_path(self, path):
        self._path = path

    def store(self):
        if not self._path:
            self._log.error('Does not indicate Path.')
            return
        self._pg.connect2()
        book = load_workbook(self._path, data_only=True)
        # book = open_workbook(self._path)
        print datetime.datetime.now()
        sheet = book.get_sheet_by_name('ARL')
        max_row = 50116  # sheet.max_row
        print 'max_row:', max_row
        row = RECORD_START_ROW
        arl_models = self.get_models(sheet, row=row,
                                     from_col=51, to_col=114)
        obj_model = ArlModel()
        row = 210
        while row < max_row:
            print row
            values1 = self.get_row(sheet, row, from_col=1, to_col=50)
            values2 = self.get_row(sheet, row, from_col=115, to_col=118)
            self._replace_nul_ch(values1)
            self._replace_nul_ch(values2)
            # values1[41] = self._cnt_data_type(values1[41])
            file_name = values1[44]
            if file_name:
                if type(file_name) in (str, unicode):
                    file_name = file_name.strip('\n')
                    values1[44] = file_name
            chapter_list = []
            if values1[1]:
                chapter = values1[45]
                if type(chapter) in (str, unicode):
                    chapter_list = chapter.split('\n')
                else:
                    chapter_list.append(chapter)
            # chapter_num, chapter_title = None, None
            # for chapter in chapter_list:
            #     chapter_num, chapter_title = self._get_chapter_info(chapter)
            params = values1 + values2
            arl_record_id = self._insert_record(params)
            model_vals = self.get_row(sheet, row, 50, 113)
            for model_list, val in zip(arl_models, model_vals):
                # model = '-'.join(model_list)
                model_id = obj_model.get_model_id(model_list)
                self._insert_model_rel(arl_record_id, model_id, val)
            row += 1
        self._pg.commit2()

    def get_row(self, sheet, row=1, from_col=1, to_col=1):
        values = []
        j = from_col
        while j <= to_col:
            val = sheet.cell(row=row, column=j).value
            if type(val) == unicode:
                val = val.replace(u'_x000D_', u'')
            values.append(val)
            j += 1
        return values

    def _cnt_data_type(self, val):
        if type(val) in (float,):
            return int(val)
        else:
            return val

    def _deal_with_category(self, category, level=None):
        pass

    def _replace_nul_ch(self, values):
        if values:
            for i in range(0, len(values)):
                s = values[i]
                if type(s) in (str, unicode):
                    if '\x00' in s:
                        values[i] = s.replace('\x00', '')

    def _get_chapter_info(self, chapter):
        result_num, result_title = None, None
        # r_list = [r"^\d+(\.\d+)+(\.|\s)", r"^\d+(\.\d+)+"]
        r_list = [r"^\d+(\.\d+)+((\.|\s)*|([A-Za-z_]{0}))",
                  r"^\d+(\.\d+)+(\.|\s)*",
                  r"^別紙(\d|０|１|２|３|４|５|６|７|８|９])*\.*(_|\b)*",
                  ]
        if chapter:
            if type(chapter) in (int, float):
                print chapter
                # if type(chapter) in (float,):
                #     chapter = int(chapter)
                chapter = str(chapter)
            chapter = chapter.strip('\n')
        for r in r_list:
            m = re.match(r, chapter)
            if m:
                num = m.group(0)
                title = chapter[m.end():]
                if num:
                    num = num.rstrip('_')
                    result_num = self._del_blank(num)
                    result_title = title
                    break
        return result_num, result_title

    def _del_blank(self, s):
        r = r"^\d+(\.\d+)+\.*"
        m = re.match(r, s)
        if m:
            return m.group(0)
        return None

    def _insert_record(self, params):
        sqlcmd = """
        INSERT INTO spec.arl(
            mm_new_num, charge,
            mm_item, tagl_exclude, category,
            id1, major_category, level1,
            id2, medium_catetory, level2,
            id3, small_category, level3,
            id4, detail, level4,
            func_summary_jp, func_summary_en, supply,
            subid, arl_id, req_post,
            remark, exception, status,
            trigger, action, arl_user,
            dealer, developer, supplier, 
            company_rule, law, old_bug,
            policy, hmi_spec_no, hmi_version,
            hmi_file_name, hmi_chapter, hmi_page,
            func_spec_no, func_version, func_file_name,
            func_chapter, func_page, if_spec,
            center_spec, other_spec, same_req, 
            sys_conf_id, author, future_req, 
            req_omission --, func_chapter_num, func_chapter_title
            )
        VALUES (%s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s)
        RETURNING arl_record_id;
        """
        self._pg.execute2(sqlcmd, params)
        arl_record_id = self.fetch_id()
        self._pg.commit2()
        return arl_record_id

    def get_models(self, sheet, row=6, from_col=19, to_col=110):
        col = from_col
        models = []
        for col in range(from_col, to_col + 1):
            curr_row = row
            leaf_model = sheet.cell(row=curr_row, column=col).value
            if not leaf_model:
                leaf_model = ''
                # self._log.error('row=%s, col=%s' % (row, col))
                # curr_row -= 1
                # leaf_model = sheet.cell_value(curr_row, col)
            else:
                leaf_model = leaf_model.replace('\n', '')
            parent_model = self._get_parent_model(sheet, row, curr_row, col)
            model = json.dumps(parent_model + [leaf_model],
                               ensure_ascii=False,
                               encoding='utf8')
            models.append(model)
        return models

    def _get_parent_model(self, sheet, start_row, cur_row, col):
        if cur_row <= 0:
            return []
        p_row = cur_row - 1
        if start_row - p_row > 2:
            return []
        val = self._get_cell_value(sheet, p_row, col)
        if val:
            p = self._get_parent_model(sheet, start_row, p_row, col)
            return p + [val]
        else:
            return []

    def _get_cell_value(self, sheet, row, column):
        b_merged, row_col = self._is_merged_cell(sheet, row, column)
        if b_merged:
            row_low, row_colum = row_col
            return sheet.cell(row=row_low, column=row_colum).value
        else:
            return sheet.cell(row=row, column=column).value

    def _is_merged_cell(self, sheet, row, column):
        coor = sheet.cell(row=row, column=column).coordinate
        for coor in sheet.merged_cells:
            # return True, (row, column)
            for irange in sheet.merged_cell_ranges:
                min_col, min_row, max_col, max_row = range_boundaries(irange)
                if(row in range(min_row, max_row + 1) and
                   column in range(min_col, max_col + 1)):
                    return True, (min_row, min_col)
            # row_low, row_high, column_low, column_high = cell_range
            # if(row in xrange(row_low, row_high) and
            #    column in xrange(column_low, column_high)):
            #     return True, (row_low, column_low)
        return False, (row, column)

    def _insert_model_rel(self, arl_id, model_id, val):
        # print val
        sqlcmd = """
        INSERT INTO spec.arl_model_rel(
                    arl_record_id, model_id, val)
            VALUES (%s, %s, %s);
        """
        self._pg.execute2(sqlcmd, (arl_id, model_id, val))

    def convert_time(self, long_time):
        if type(long_time) in (int, long):
            # excel中，用浮点数1表示1899年12月31日
            base = datetime.date(1899, 12, 31).toordinal()
            new_date = datetime.date.fromordinal(base + long_time - 1)
            return new_date.strftime("%Y-%m-%d %H:%M:%S")
        return long_time


class ArlModel(SpecBase):
    # __instance = None

    # @staticmethod
    # def instance():
    #     """create a instance"""
    #     if AnalysisModel.__instance is None:
    #         AnalysisModel.__instance = AnalysisModel()
    #     return AnalysisModel.__instance

    def __init__(self):
        SpecBase.__init__(self)

    def get_model_id_dict(self, modelList):
        for model_dict in modelList:
            for model in model_dict.iterkeys():
                model_id = self._attr_dict.get(model)
                if not model_id:
                    model_id = self._get_model_id(model)
                    self._attr_dict[model] = model_id
        return self._attr_dict

    def get_model_id(self, model):
        model_id = self._attr_dict.get(model)
        if not model_id:
            model_id = self._get_model_id(model)
            self._attr_dict[model] = model_id
        return model_id

    def _get_model_id(self, model):
        sqlcmd = """
        SELECT model_id
          FROM spec.arl_model_type
          WHERE model = %s;
        """
        self._pg.execute2(sqlcmd, (model,))
        row = self._pg.fetchone2()
        if row and row[0]:
            return row[0]
        sqlcmd = """
        INSERT INTO spec.arl_model_type(model)
            VALUES (%s) RETURNING model_id;
        """
        self._pg.execute2(sqlcmd, (model,))
        model_id = self.fetch_id()
        self._pg.commit2()
        return model_id


def strQ2B(ustring):
    """把字符串全角转半角"""
    rstring = ""
    for uchar in ustring:
        inside_code = ord(uchar)
        if inside_code == 0x3000:
            inside_code = 0x0020
        else:
            inside_code -= 0xfee0
        # 转完之后不是半角字符返回原来的字符
        if inside_code < 0x0020 or inside_code > 0x7e:
            rstring += uchar
        else:
            rstring += unichr(inside_code)
    return rstring
