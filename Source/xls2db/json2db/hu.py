# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import os
import re
import datetime
import json
from definition import DefinitionSpec
from Source.xls2db.common import log
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from arl import SpecArl, ArlModel
HU_RECORD_START_ROW = 9  #
import time


class SpecHU(SpecArl):
    """
    """
    def __init__(self):
        SpecArl.__init__(self)
        self._log = log.common_log.instance().logger('SpecHU')

    def set_path(self, path):
        self._path = path

    def store(self):
        if not self._path:
            self._log.error('Does not indicate Path.')
            return
        self._pg.connect2()
        try:
            book = load_workbook(self._path, data_only=True)
        except:
            return
        # print datetime.datetime.now()
        sheet = book.get_sheet_by_name(r'HU要件定義書')
        max_row = sheet.max_row
        # print 'max_row:', max_row
        row = HU_RECORD_START_ROW
        arl_models = self.get_models(sheet, row=row,
                                     from_col=31, to_col=60)
        obj_model = HuModel()
        row += 1
        count = 0
        while row <= max_row:
            # print row
            # 担当, ARLID
            values1 = self.get_row(sheet, row, from_col=1, to_col=2)
            # H/U要件定義ID ==> H/U分類ID
            values2 = self.get_row(sheet, row, from_col=13, to_col=30)
            # 備考 ==> 未要件分析
            values3 = self.get_row(sheet, row, from_col=59, to_col=71)
            # Group ==> TAGL要件提出
            values4 = self.get_row(sheet, row, from_col=78, to_col=79)
            arl_id = values1[1]
            if not arl_id:
                break
            hu_id = values2[0]
            unique_id = values2[1]
            new_date = values4[0]  # 日付
            if not new_date:
                print 'file=%s, not new date, hu_id==%s' % (self._path, hu_id)
            else:
                values4[0] = self.convert_time(new_date)
            if type(unique_id) in (str, unicode):
                unique_id = int((hu_id.split('.'))[-1])
                values2[1] = unique_id
            params = values1 + values2 + values3 + values4
            try:
                hu_record_id = self._insert_record(params)
                model_vals = self.get_row(sheet, row, 31, 60)
                for model_list, val in zip(arl_models, model_vals):
                    # model = '-'.join(model_list)
                    model_id = obj_model.get_model_id (model_list)
                    self._insert_model_rel(hu_record_id, model_id, val)
                self._pg.commit2()
                count += 1
            except:
                self._pg.conn2.rollback()
                print 'Already exists. file=%s, hu_id=%s' % (self._path, hu_id)
            row += 1
        self._pg.commit2()
        print 'file=%s, count=%s' % (self._path, count)

    def _insert_record(self, params):
        sqlcmd = """
        INSERT INTO spec.hu(
            author, arl_id, hu_id,
            unique_id, amp, dsrc,
            dcm, rse, touch_pad,
            separate_disp, system_conf, rel_requirement,
            exception, dcu_status, dcu_trigger, 
            dcu_action, meu_status, meu_trigger,
            meu_action, hu_category_id, remark,
            sys_spec_chapter, common_chapter, common_seq_spec,
            common_seq_no, common_cmd_guide, common_opc,
            inter_loc_spec, inter_chapter, other_chapter,
            other_doc, test_results, future_req,
            new_date, reason)
        VALUES (%s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s)
        RETURNING hu_record_id;
        """
        self._pg.execute2(sqlcmd, params)
        hu_record_id = self.fetch_id()
        # self._pg.commit2()
        return hu_record_id

    def _insert_model_rel(self, arl_id, model_id, val):
        # print val
        sqlcmd = """
        INSERT INTO spec.hu_model_rel(
                    hu_record_id, model_id, val)
            VALUES (%s, %s, %s);
        """
        self._pg.execute2(sqlcmd, (arl_id, model_id, val))


class HuModel(ArlModel):
    # __instance = None

    # @staticmethod
    # def instance():
    #     """create a instance"""
    #     if AnalysisModel.__instance is None:
    #         AnalysisModel.__instance = AnalysisModel()
    #     return AnalysisModel.__instance

    def __init__(self):
        ArlModel.__init__(self)

    def get_model_id_dict(self, modelList):
        for model_dict in modelList:
            for model in model_dict.iterkeys():
                model_id = self._attr_dict.get(model)
                if not model_id:
                    model_id = self._get_model_id(model)
                    self._attr_dict[model] = model_id
        return self._attr_dict

    def get_model_id(self, model):
        model_id = self._attr_dict.get(model)
        if not model_id:
            model_id = self._get_model_id(model)
            self._attr_dict[model] = model_id
        return model_id

    def _get_model_id(self, model):
        sqlcmd = """
        SELECT model_id
          FROM spec.hu_model_type
          WHERE model = %s;
        """
        self._pg.connect2()
        self._pg.execute2(sqlcmd, (model,))
        row = self._pg.fetchone2()
        if row and row[0]:
            return row[0]
        sqlcmd = """
        INSERT INTO spec.hu_model_type(model)
            VALUES (%s) RETURNING model_id;
        """
        self._pg.execute2(sqlcmd, (model,))
        model_id = self.fetch_id()
        self._pg.commit2()
        return model_id