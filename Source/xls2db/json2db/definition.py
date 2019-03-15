# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import os
import re
from Source.xls2db.common import log
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from arl import SpecArl, ArlModel
from spec import SpecBase
PROJ_NAME = 'proj_name'
SPEC_NAME = 'spec_name'
SPEC_VERSION = 'version'
SPEC_FILE_NAME = 'file_name'
RECORD_START_ROW = 5  # 记录开始行号


class DefinitionSpec(SpecArl):
    """
    """
    def __init__(self):
        SpecArl.__init__(self)
        self._path = None
        self._log = log.common_log.instance().logger('DefinitionSpec')

    def set_path(self, path):
        self._path = path

    def store(self):
        if not self._path:
            self._log.error('Does not indicate Path.')
            return
        self._pg.connect2()
        try:
            book = load_workbook(self._path, data_only=True)
        except:
            return
        sheet = book.get_sheet_by_name(r'TAGL要件定義')
        max_row = sheet.max_row
        # print 'max_row:', max_row
        row = RECORD_START_ROW
        arl_models = self.get_models(sheet, row=row, from_col=23, to_col=48)
        obj_model = DefModel()
        row += 1
        count = 0
        while row <= max_row:
            # print row
            values1 = self.get_row(sheet, row, from_col=1, to_col=4)  # 4个
            values2 = self.get_row(sheet, row, from_col=11, to_col=11)  # 1个
            values3 = self.get_row(sheet, row, from_col=19, to_col=22)  # 4个
            values4 = self.get_row(sheet, row, from_col=49, to_col=56)  # 8个
            hu_id = values1[1]
            if not hu_id:
                break
            definition_id = values1[2]
            try:
                req_def_id = self._insert_def_record(values1 + values2 + values3 + values4)
                model_vals = self.get_row(sheet, row, 24, 49)
                for model_list, val in zip(arl_models, model_vals):
                    model_id = obj_model.get_model_id(model_list)
                    self._insert_model_rel(req_def_id, model_id, val)
                self._pg.commit2()
            except:
                self._pg.conn2.rollback()
                print 'file=%s, def_id=%s' % (self._path, definition_id)
                row += 1
                continue
            count += 1
            row += 1
        self._pg.commit2()
        print 'file=%s, count=%s' % (self._path, count)

    def _split_requirement_id(self, requirement_id):
        chapter_num, id = None, None
        if requirement_id:
            p = re.compile(r'\d+\s*$')
            id_list = p.findall(requirement_id)
            if id_list:
                id = id_list[0]
                chapter_num = requirement_id[:len(requirement_id) - len(id)]
        return chapter_num, id

    def _set_temp_chapter_num(self):
        sqlcmd = """
        UPDATE spec.requirement_def_record AS a 
          SET spec_chapter_num = chapter_number
          from spec.spec_chapter as b
          where req_def_id = b.chapter_id;
        """
        self._pg.execute2(sqlcmd)
        self._pg.commit2()

    def _parser_spec_info(self, path):
        spec_info = {
                     # PROJ_NAME: 'TAGL',
                     # SPEC_NAME: 'RequirementDefinition',
                     # SPEC_VERSION: '0.09',
                     # SPEC_FILE_NAME: 'TAGL_RequirementDefinitionVer0.09.xlsx',
                     }
        base_name = os.path.basename(path)
        name, file_ext = os.path.splitext(base_name)
        r = r'.+_+'
        m = re.match(r, name)
        if m:
            proj_name = m.group(0)
            if proj_name:
                name = name[m.end():]
                spec_info[PROJ_NAME] = proj_name.rstrip('_')
        r = r'\.*[V|v]er\d+\.*\d*'
        s = re.search(r, name)
        if s:
            ver = s.group(0)
            if ver:
                name = name[:s.start()]
                ver = ver.lstrip(r'.')
                spec_info[SPEC_VERSION] = ver[3:]
        spec_info[SPEC_NAME] = name
        spec_info[SPEC_FILE_NAME] = base_name
        return spec_info

    def _store_spec_info(self, spec_info):
        sqlcmd = """
        INSERT INTO spec.requirement_spec(proj, spec_name, ver, file_name)
          VALUES(%s, %s, %s, %s) RETURNING requiremnt_spec_id;
        """
        self._pg.execute2(sqlcmd, (spec_info.get(PROJ_NAME),
                                   spec_info.get(SPEC_NAME),
                                   spec_info.get(SPEC_VERSION),
                                   spec_info.get(SPEC_FILE_NAME)
                                   )
                          )
        self.set_id(self.fetch_id())
        self._pg.commit2()

    def _insert_def_record(self, params):
        sqlcmd = """
        INSERT INTO spec.definition(
            author_name, hu_def_id, definition_id,
            unique_id, exception, dcu_meu,
            pf_status, pf_trigger, pf_action,
            remark,  notice, rel_hal_design,
            rel_flow_diagram, other_spec, implementation, 
            analysis, unrequire)
            VALUES (%s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s)
            RETURNING def_rc_id;
        """
        self._pg.execute2(sqlcmd, params)
        return self.fetch_id()

    def _insert_model_rel(self, def_rc_id, model_id, val):
        sqlcmd = """
        INSERT INTO spec.definition_model_rel(
                    def_rc_id, model_id, val)
            VALUES (%s, %s, %s);
        """
        self._pg.execute2(sqlcmd, (def_rc_id, model_id, val))


class AnalysisSpec(DefinitionSpec):
    """
    """
    def __init__(self):
        DefinitionSpec.__init__(self)
        self._log = log.common_log.instance().logger('ANALYSIS')

    def store(self):
        if not self._path:
            self._log.error('Does not indicate Path.')
            return
        self._pg.connect2()
        self._pg.connect2()
        book = load_workbook(self._path, data_only=True)
        sheet = book.get_sheet_by_name(r'TAGL要件分析')
        max_row = 145  # sheet.max_row
        print 'max_row:', max_row
        row = 5
        analysis_models = self.get_models(sheet, row=row, from_col=20, to_col=111)
        obj_model = AnalysisModel()
        row += 1
        while row <= max_row:
            print row
            values1 = self.get_row(sheet, row, from_col=1, to_col=1)  # 1个
            values2 = self.get_row(sheet, row, from_col=3, to_col=4)  # 2个
            values3 = self.get_row(sheet, row, from_col=11, to_col=11)  # 1个
            values4 = self.get_row(sheet, row, from_col=16, to_col=19)  # 5个
            values5 = self.get_row(sheet, row, from_col=112, to_col=114)  # 3个
            params = values1 + values2 + values3 + values4 + values5
            analysis_id = self._insert_analysis_record(params)
            # Get Model Values.
            model_vals = self.get_row(sheet, row, 20, 111)
            for model_list, val in zip(analysis_models, model_vals):
                # model = '-'.join(model_list)
                model_id = obj_model.get_model_id(model_list)
                self._insert_model_rel(analysis_id, model_id, val)
            row += 1
        self._pg.commit2()
        # self._set_temp_chapter_num()

    def _get_parent_model(self, sheet, start_row, cur_row, col):
        if cur_row <= 0:
            return []
        p_row = cur_row - 1
        if start_row - p_row > 3:
            return []
        val = self._get_cell_value(sheet, p_row, col)
        if val:
            p = self._get_parent_model(sheet, start_row, p_row, col)
            return p + [val]
        else:
            return []

    def _insert_analysis_record(self, params):
        sqlcmd = """
        INSERT INTO spec.analysis(
            author_name, definition_id, unique_id,
            exception, seq_diagram, application,
            kernel, systemd, supple_spec,
            uncheck, remark)
            VALUES (%s, %s, %s, 
                    %s, %s, %s, 
                    %s, %s, %s, 
                    %s, %s)
            RETURNING analysis_rc_id;
        """
        self._pg.execute2(sqlcmd, params)
        return self.fetch_id()

    def _insert_model_rel(self, def_rc_id, model_id, val):
        sqlcmd = """
        INSERT INTO spec.analysis_model_rel(
                    analysis_rc_id, model_id, val)
            VALUES (%s, %s, %s);
        """
        self._pg.execute2(sqlcmd, (def_rc_id, model_id, val))


class DefModel(ArlModel):
    # __instance = None

    # @staticmethod
    # def instance():
    #     """create a instance"""
    #     if AnalysisModel.__instance is None:
    #         AnalysisModel.__instance = AnalysisModel()
    #     return AnalysisModel.__instance

    def __init__(self):
        ArlModel.__init__(self)

    def get_model_id_dict(self, modelList):
        for model_dict in modelList:
            for model in model_dict.iterkeys():
                model_id = self._attr_dict.get(model)
                if not model_id:
                    model_id = self._get_model_id(model)
                    self._attr_dict[model] = model_id
        return self._attr_dict

    def get_model_id(self, model):
        model_id = self._attr_dict.get(model)
        if not model_id:
            model_id = self._get_model_id(model)
            self._attr_dict[model] = model_id
        return model_id

    def _get_model_id(self, model):
        sqlcmd = """
        SELECT model_id
          FROM spec.definition_model_type
          WHERE model = %s;
        """
        self._pg.connect2()
        self._pg.execute2(sqlcmd, (model,))
        row = self._pg.fetchone2()
        if row and row[0]:
            return row[0]
        sqlcmd = """
        INSERT INTO spec.definition_model_type(model)
            VALUES (%s) RETURNING model_id;
        """
        self._pg.execute2(sqlcmd, (model,))
        model_id = self.fetch_id()
        self._pg.commit2()
        return model_id


class AnalysisModel(ArlModel):

    def __init__(self):
        ArlModel.__init__(self)

    def get_model_id_dict(self, modelList):
        for model_dict in modelList:
            for model in model_dict.iterkeys():
                model_id = self._attr_dict.get(model)
                if not model_id:
                    model_id = self._get_model_id(model)
                    self._attr_dict[model] = model_id
        return self._attr_dict

    def get_model_id(self, model):
        model_id = self._attr_dict.get(model)
        if not model_id:
            model_id = self._get_model_id(model)
            self._attr_dict[model] = model_id
        return model_id

    def _get_model_id(self, model):
        sqlcmd = """
        SELECT model_id
          FROM spec.analysis_model_type
          WHERE model = %s;
        """
        self._pg.execute2(sqlcmd, (model,))
        row = self._pg.fetchone2()
        if row and row[0]:
            return row[0]
        sqlcmd = """
        INSERT INTO spec.analysis_model_type(model)
            VALUES (%s) RETURNING model_id;
        """
        self._pg.execute2(sqlcmd, (model,))
        model_id = self.fetch_id()
        self._pg.commit2()
        return model_id
