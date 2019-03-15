# -*- coding: UTF-8 -*-
import os
import json
import zipfile
from flask import Flask
from flask import send_from_directory
from openpyxl import load_workbook
from flask_restful import Resource, Api
from flask_cors import *
from flask import request
from flask_apscheduler import APScheduler
from lockfile import LockFile
from celery import Celery
from kombu import Queue, Exchange
import time
app = Flask(__name__)
api = Api(app)
CORS(app, supports_credentials=True)
app.config.from_pyfile('default_config.py')
celery = Celery(app.name, broker = app.config['CELERY_BROKER_URL'])
celery.conf.update(app.config)

# @celery.task
# def import_task(file_name, file_time):
#     import sys
#     reload(sys)
#     sys.setdefaultencoding('UTF-8')
#     from Source.spec2db_server.arl.arl_import import ArlImport
#     from Source.spec2db_server.arl.arl_group import ArlUser
#
#     user_id = ArlUser().find_user_by_name(file_name)
#     import_result = {}
#     import_result["ResultInfo"] = {}
#     import_result["Result"] = 0
#
#     for root, dirs, files in os.walk(os.path.join('./import/users/', file_name, file_time)):
#         if files:
#             for file in files:
#                 if file.lower() in ('.xlsx',):
#                     sheet_import_result = ArlImport(root, file, user_id).import_to_db()
#                     if sheet_import_result[file]['result'] > 0:
#                         import_result["Result"] = 1
#                     import_result["ResultInfo"].update(sheet_import_result)
#
#     return import_result

@celery.task
def export_task(export_type, export_path, user_id=None):
    from Source.spec2db_server.arl.arl_func import ArlFunc
    ret_info_dict = {}
    ret_info_dict['Result'] = 'NG'
    if export_type == "HU_DEF_All":
        export_file_name  = "HU_RequirementDefinition_All_%s.xlsx"%(
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_hu_to_excel(export_path, export_file_name,
                                     os.path.join('./template','HU_RequirementDefinition.Ver0.17_haspoint_template.xlsx'))
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name

    elif export_type == "HU_DEF_User":
        export_file_name = "HU_RequirementDefinition_User_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_hu_to_excel_by_userid(export_path, export_file_name,
                                           os.path.join('./template','HU_RequirementDefinition.Ver0.17_haspoint_template.xlsx'),
                                           user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "HU_DEF_Group":
        export_file_name = "HU_RequirementDefinition_Group_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_group_hu_to_excel_by_userid(export_path, export_file_name,
                                                    os.path.join('./template', 'HU_RequirementDefinition.Ver0.17_haspoint_template.xlsx'),
                                                     user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "TAGL_DEF_All":
        export_file_name = "TAGL_RequirementDefinition_All_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_definition_to_excel(export_path, export_file_name,
                                     os.path.join('./template','TAGL_RequirementDefinitionVer0.12_haspoint_template.xlsx'))
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "TAGL_DEF_User":
        export_file_name = "TAGL_RequirementDefinition_User_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_definition_to_excel_by_userid(export_path, export_file_name,
                                                       os.path.join('./template', 'TAGL_RequirementDefinitionVer0.12_haspoint_template.xlsx'),
                                                       user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "TAGL_DEF_Group":
        export_file_name = "TAGL_RequirementDefinition_Group_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_group_definition_to_excel_by_userid(export_path, export_file_name,
                                                             os.path.join('./template',
                                                                          'TAGL_RequirementDefinitionVer0.12_haspoint_template.xlsx'),
                                                             user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "TAGL_ANA_All":
        export_file_name = "TAGL_RequirementAnalysis_All_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_analysis_to_excel(export_path, export_file_name,
                                                             os.path.join('./template',
                                                                          'TAGL_RequirementAnalysis.ver1.00_haspoint_template.xlsx'))
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "TAGL_ANA_User":
        export_file_name = "TAGL_RequirementAnalysis_User_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_user_analysis_to_excel(export_path, export_file_name,
                                                             os.path.join('./template',
                                                                          'TAGL_RequirementAnalysis.ver1.00_haspoint_template.xlsx'),
                                                             user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "TAGL_ANA_Group":
        export_file_name = "TAGL_RequirementAnalysis_Group_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_group_analysis_to_excel(export_path, export_file_name,
                                                             os.path.join('./template',
                                                                          'TAGL_RequirementAnalysis.ver1.00_haspoint_template.xlsx'),
                                                             user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "ARL_All":
        export_file_name = "ARL_Schedule_All_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_arl_to_object_by_all(export_path, export_file_name)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "ARL_User":
        export_file_name = "ARL_Schedule_User_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_arl_to_object_by_userid(export_path, export_file_name, user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    elif export_type == "ARL_Group":
        export_file_name = "ARL_Schedule_Group_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_arl_to_object_by_groupid(export_path, export_file_name, user_id)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name

    elif export_type == "HU_Unfinished_ALL":
        print 'HU_Unfinished_ALL'
        export_file_name = "HU_Unfinished_ALL_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_hu_to_unfinished_by_all(export_path, export_file_name)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name

    elif export_type == "DEF_Unfinished_ALL":
        export_file_name = "DEF_Unfinished_ALL_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_def_to_unfinished_by_all(export_path, export_file_name)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name

    elif export_type == "Analysis_Unfinished_ALL":
        export_file_name = "Analysis_Unfinished_ALL_%s.xlsx" % (
            time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        )
        ret_info_dict['Result'] = ArlFunc().export_ana_to_unfinished_by_all(export_path, export_file_name)
        ret_info_dict['Result_PathInfo'] = export_path
        ret_info_dict['Result_FileInfo'] = export_file_name
    return ret_info_dict

class Login(Resource):
    def get(self, username, password):
        from Source.spec2db_server.arl.login import login2
        result = login2(username, password)
        return result


class Register(Resource):
    def get(self, group_id, username, password):
        from Source.spec2db_server.arl.login import register
        result = register(group_id, username, password)
        return result


class ARL_Category(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        cat = arl.get_category_tree()
        return cat


class ARL_UserCategory(Resource):
    def get(self, user_id):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        cat = arl.get_category_tree(user_id)
        return cat

class ARL_GroupCategory(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        cat = arl.get_group_category_tree(req_json)
        return cat



class ARL_Content(Resource):
    def get(self, arl_id):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        rc = arl.get_arl_by_id(arl_id)
        return rc


class ARL_SubHu(Resource):
    def get(self, arl_id):
        from Source.spec2db_server.arl.hu_server import HuRecord
        hu = HuRecord()
        rc = hu.get_by_arl_id(arl_id)
        return rc


class Summary(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = spec.get_by_category(req_json)
        return result


class SummaryAll(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = spec.summay_for_all(req_json)
        return result

class SummaryBasic(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = spec.basic_summay_for_all(req_json)
        return result

class SummaryUser(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = spec.summary_users(req_json)
        return result


class ReQueryStatus(Resource):
    def get(self, id, type, state):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = dict()
        result['result'] = "NG"
        result['content'] = spec.change_job_status(id, type, state)
        if result['content']:
            result['result'] = "OK"
        return result

class LookStatus(Resource):
    def get(self, id, type):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = dict()
        result['result'] = "NG"
        result['content'] = spec.look_job_status(id, type)
        if result['content']:
            result['result'] = "OK"
        return result


class ArlTreeInfo(Resource):
    def get(self, arl_id, user_id, group_id):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        spec = ArlSpec()
        result = dict()
        result['result'] = "NG"
        result['content'] = spec.get_by_id_deep(arl_id, user_id, group_id)
        if result['content']:
            result['result'] = "OK"
        return result

class BasicTreeInfo(Resource):
    def get(self, type, _id, user_id, group_id):
        from Source.spec2db_server.arl.basic_hu import BasicHuRecord
        spec = BasicHuRecord()
        result = dict()
        result['result'] = "NG"
        result['content'] = spec.get_by_basic_id_deep(type, _id, user_id, group_id)
        if result['content']:
            result['result'] = "OK"
        return result

class FullContent(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlRecord
        obj = ArlRecord()
        update_time = obj.get_current_time()
        result = obj.arl_full_content_post(req_json, update_time)
        return result


class BasicContent(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlRecord
        obj = ArlRecord()
        update_time = obj.get_current_time()
        result = obj.basic_or_hu_content_post(req_json, update_time)
        return result


class ARL_Summary(Resource):
    def get(self, category_id, page_size, page_number):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        rc = arl.get_arl_by_category(category_id, page_size, page_number)
        return rc

class ARL_Summary2(Resource):
    def get(self, user_id, category_id, page_size, page_number):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        rc = arl.new_get_arl_by_category(category_id, page_size, page_number, user_id)
        return rc

class ARL_UserSummary(Resource):
    def get(self, user_id, category_id, page_size, page_number):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        arl = ArlSpec()
        rc = arl.get_arl_by_category(category_id, page_size,
                                     page_number, user_id)
        return rc


class ARL_User(Resource):
    def get(self, user_id, page_size, page_number):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        arl = ArlRecord()
        rc = arl.get_by_user_id(user_id, page_size, page_number)
        return rc


class ARL_Group(Resource):
    def get(self, group_id, page_size, page_number):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        arl = ArlRecord()
        rc = arl.get_by_group_id(group_id, page_size, page_number)
        return rc
class ARL_PostInfo(Resource):
    def get(self, arl_id):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        arl = ArlRecord()
        data = arl.get_arl_post_info(arl_id)
        return data
class ARL_AssignUser(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_server import ArlRecord
        ret = ArlRecord().assign_user(req_json)
        return ret
class ARL_Graph(Resource):
    def get(self, arl_id):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        ret = ArlRecord().get_graph(arl_id)
        return ret
class ARL_Update_GroupName(Resource):
    def get(self, group_id, new_group_name):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        ret = ArlGroup().update_group_name(group_id, new_group_name)
        return ret
class HU_Summary(Resource):
    def get(self, category_id, page_size, page_number):
        from Source.spec2db_server.arl.hu_server import HuRecord
        hu = HuRecord()
        rc = hu.get_by_category(category_id, page_size, page_number)
        return rc


class HU_Summary2(Resource):
    def get(self, user_id, category_id, page_size, page_number):
        from Source.spec2db_server.arl.hu_server import HuRecord
        hu = HuRecord()
        rc = hu.get_by_category2(category_id, page_size, page_number, user_id)
        return rc


class HU_Option(Resource):
    def get(self):
        from Source.spec2db_server.arl.hu_server import HuRecord
        hu = HuRecord()
        rc = hu.get_option()
        return rc
class HU_Content(Resource):
    def post(self):
        """Add a new"""
        req_json = request.get_json(force=True)
        print req_json
        result = {"result": 'NG'}
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        result = obj.add_list(req_json)
        return result

    def put(self, hu_id):
        """update"""
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        obj.update(hu_id, req_json)
        return {"result": 'OK'}

    def get(self, hu_id):
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        data = obj.get_info(hu_id)
        return data

    def delete(self, hu_id):
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        obj.delete(hu_id)
        return {"result": 'OK'}
class HU_ContentAdd(Resource):
    def post(self):
        """Add a new"""
        data = request.get_json(force=True)
        print data
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        obj.add(data)
        return {"result": 'OK'}


class HU_ModelList(Resource):
    def get(self):
        """Add a new"""
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        result = obj.get_all_model_list()
        return result


class HU_ContentDelete(Resource):
    def get(self, hu_id):
        """Add a new"""
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        obj.delete(hu_id)
        return {"result": 'OK'}


class HU_id(Resource):
    def get(self, arl_id):
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        data = obj.get_new_hu_id(arl_id)
        return data


class HU_PostInfo(Resource):
    def get(self, hu_id):
        from Source.spec2db_server.arl.hu_server import HuRecord
        obj = HuRecord()
        data = obj.get_post_info(hu_id)
        return data


class HU_SubDef(Resource):
    def get(self, hu_id):
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        def_obj = Cdb_definition()
        rc = def_obj.get_record_hu_def_id(hu_id)
        return rc


class DEF_id(Resource):
    def get(self, hu_id):
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        obj = Cdb_definition()
        data = obj.get_new_id(hu_id)
        return data


class DEF_Content(Resource):
    def post(self):
        """Add a new"""
        data_list = request.get_json(force=True)
        print data_list
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        db_def = Cdb_definition()
        if db_def.insert_record_list(data_list):
            return {"result": 'OK'}
        return {"result": 'NG'}

    def put(self, def_id):
        """update"""
        data = request.get_json(force=True)
        print data
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        db_def = Cdb_definition()
        db_def.update_record(def_id, data)
        return def_id

    def get(self, def_id):
        from Source.spec2db_server.arl.def_server import DefRecord
        db_obj = DefRecord()
        data = db_obj.get_by_id(def_id)
        return data


class DEF_PostInfo(Resource):
    def get(self, def_id):
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        obj = Cdb_definition()
        data = obj.get_post_info(def_id)
        return data


class DEF_ContentDel(Resource):
    def get(self, def_id):
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        db_obj = Cdb_definition()
        return {"result": db_obj.delete(def_id)}


class DEF_Summary(Resource):
    def get(self, category_id, page_size, page_number):
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        db_obj = Cdb_definition()
        data = db_obj.get_by_category(category_id, page_size, page_number)
        return data


class DEF_Summary2(Resource):
    def get(self, user_id, category_id, page_size, page_number):
        from Source.spec2db_server.arl.def_server import DefRecord
        def_obj = DefRecord()
        data = def_obj.get_by_category2(category_id, page_size, page_number, user_id)
        return data


class DEF_ModeList(Resource):
    def get(self):
        """Add a new"""
        from Source.spec2db_server.arl.db_operate import Cdb_definition
        db_obj = Cdb_definition()
        result = db_obj.get_all_model_list()
        return result


class Analysis_Content(Resource):
    def post(self):
        """Add a new"""
        data = request.get_json(force=True)
        print data
        from Source.spec2db_server.arl.db_operate import Cdb_analysis
        db_analysis = Cdb_analysis()
        if data:
            old_id = data.get(u"ALS_id")
            if old_id:
                db_analysis.update_record(old_id, data)
            else:
                db_analysis.insert_record(data)
            return {"result": 'OK'}
        return {"result": 'NG'}

    def put(self, def_id):
        """update"""
        data = request.get_json(force=True)
        print data
        from Source.spec2db_server.arl.db_operate import Cdb_analysis
        db_obj = Cdb_analysis()
        db_obj.update_record(def_id, data)
        return def_id

    def get(self, def_id):
        from Source.spec2db_server.arl.db_operate import Cdb_analysis
        db_obj = Cdb_analysis()
        data = db_obj.get_record_definition_id(def_id)
        return data


class Analysis_ContentDel(Resource):
    def get(self, def_id):
        from Source.spec2db_server.arl.db_operate import Cdb_analysis
        db_obj = Cdb_analysis()
        return {"result": db_obj.delete(def_id)}


class Analysis_Summary(Resource):
    def get(self, category_id, page_size, page_number):
        from Source.spec2db_server.arl.db_operate import Cdb_analysis
        db_obj = Cdb_analysis()
        data = db_obj.get_by_category(category_id, page_size, page_number)
        return data


class Analysis_Summary2(Resource):
        def get(self, user_id, category_id, page_size, page_number):
            from Source.spec2db_server.arl.analysis_service import AnalysisRecord
            db_obj = AnalysisRecord()
            data = db_obj.get_by_category2(category_id, page_size, page_number, user_id)
            return data

# Group
class Group_Add(Resource):
    def get(self, group_name):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.add(group_name)
        return data


class Group_Del(Resource):
    def get(self, group_id):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.delete(group_id)
        return data


class Group_Modify(Resource):
    def get(self, group_id, new_group_name):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.delete(group_id, new_group_name)
        return data


class Group_Members(Resource):
    def get(self, group_id):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.get_members(group_id)
        return data

class Group_Get_Members(Resource):
    def get(self, group_id):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.get_group_members(group_id)
        return data

class Gp_Member_Add(Resource):
    def post(self):
        req = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.add_member(req)
        return data


class Gp_Member_Del(Resource):
    def get(self, group_id, user_id):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.del_member(group_id, user_id)
        return data


class Group_AllGroups(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.get_all_groups()
        return data


class Group_UpdateRole(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.update_group_role(req_json)
        return data


class Group_GroupDetail(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_group import ArlGroup
        obj = ArlGroup()
        data = obj.get_groups_detail()
        return data


# User
class All_User(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_group import ArlUser
        obj = ArlUser()
        data = obj.get_all_user()
        return data


class User_Content(Resource):
    def post(self):
        data = request.get_json(force=True)
        # {"user_id": "id", "user_name": "xx",
        #  "old_password": "xxx", "new_password": "xxx"}
        if data:
            from Source.spec2db_server.arl.arl_group import ArlUser
            obj = ArlUser()
            if obj.modify_user_info(data):
                return {"result": 'OK'}
        return {"result": 'NG'}

    def get(self, user_id):
        from Source.spec2db_server.arl.arl_group import ArlUser
        obj = ArlUser()
        data = obj.get_user(user_id)
        return data


class User_Del(Resource):
    def get(self, user_id):
        from Source.spec2db_server.arl.arl_group import ArlUser
        obj = ArlUser()
        data = obj.delete(user_id)
        return data

class User_MyGroup(Resource):
    def get(self, user_id):
        from Source.spec2db_server.arl.arl_group import ArlUser
        obj = ArlUser()
        data = obj.find_group_by_userId(user_id)
        return data

class Email_Alter(Resource):
    def get(self, user_id, email):
        from Source.spec2db_server.arl.arl_group import ArlUser
        obj = ArlUser()
        data = obj.alter_email(user_id, email)
        return data


class Summarize_By_User(Resource):
    def get(self, date=None):
        from Source.spec2db_server.arl.arl_summarize import ArlSummarize
        obj = ArlSummarize()
        data = obj.summarize_by_user(date)
        return data


class Summarize_By_Category(Resource):
    def get(self, date=None):
        from Source.spec2db_server.arl.arl_summarize import ArlSummarize
        obj = ArlSummarize()
        data = obj.summarize_by_category(date)
        return data


class Summarize_By_Category2(Resource):
    def get(self, date=None):
        from Source.spec2db_server.arl.arl_summarize import ArlSummarize
        obj = ArlSummarize()
        data = obj.summarize_by_category2(date)
        return data


class WhiteList(Resource):
    def get(self):
        from Source.spec2db_server.arl.whitelist_srv import WhitelistSrv
        obj = WhitelistSrv()
        result = {"result": "NG"}
        white_list = obj.get_white_list()
        if white_list:
            result["result"] = "OK"
            result["white_list"] = white_list
        return result

    def post(self):
        data = request.get_json(force=True)
        result = {"result": "NG"}
        white_list = data.get("white_list")
        user_id = int(data.get("user_id"))
        # white_list, user_id = data.get("white_list", "user_id")
        if white_list and user_id:
            from Source.spec2db_server.arl.whitelist_srv import WhitelistSrv
            obj = WhitelistSrv()
            if obj.agree(white_list, user_id):
                result["result"] = "OK"
        return result


#doc_type is [HU_DEF, TAGL_DEF,TAGL_ANA]
#range is [All, User, Group]
#user_id is user's id
#date is now date info
#api.add_resource(Export_Info, '/Export/<string:doc_type>/<string:range>/<int:user_id>/<string:date>')
#api.add_resource(Export_Process, '/ExportProcess/<string:doc_type>/<string:range>/<int:user_id>/<string:date>/<task_id>')
#api.add_resource(Export_File, '/ExportFile/<string:doc_type>/<string:range>/<int:user_id>/<string:date>')
class Export_Info(Resource):
    def get(self, doc_type, range, user_id, date):
        if not os.path.exists(os.path.join('./','export')):
            os.system('mkdir export')
        if not os.path.exists(os.path.join('./export','users')):
            os.system('mkdir export/users')
        if not os.path.exists(os.path.join('./export/users','%d'%(user_id))):
            os.system('mkdir ./export/users/%d'%(user_id))

        task = export_task.apply_async(args=(doc_type+"_"+range,
                                             './export/users/%d'%(user_id),
                                             user_id),
                                       queue='export_task')
        result = {}
        result['result'] = "WAITING"
        result['datapath'] = "/ExportProcess/%s" % (task.id)

        return result

class Export_Process(Resource):
    def get(self, task_id):
        task = export_task.AsyncResult(task_id)
        result = {}
        if task.successful():
            result['result'] = task.info['Result']
            if result['result'] == 'OK':
                result['datapath'] = "/ExportFile/%s" % (task_id)
            else:
                result['datapath'] = ""
        else:
            result['result'] = "WAITING"
            result['datapath'] = "/ExportProcess/%s" % (task_id)

        return result


class Export_File(Resource):
    def get(self, task_id):
        from flask import send_from_directory
        task = export_task.AsyncResult(task_id)
        if task.info['Result'] == 'OK':
            return send_from_directory(task.info['Result_PathInfo'],
                                       task.info['Result_FileInfo'], as_attachment=True)


class Export_Checklist(Resource):
    def get(self, start_time, end_time):
        result = {'result': 'NG'}
        from flask import send_from_directory
        from Source.spec2db_server.arl.checklist_export import ExportCkecklist
        check = ExportCkecklist()
        try:
            path, zip_Name = check.export_excel(start_time, end_time)
            file_url = os.path.join(path, zip_Name)
            result['result'] = 'OK'
            result['file_path'] = file_url
        except Exception as e:
            result['error'] = e.message
        return result


class DownloadChecklist(Resource):
    def get(self, file_path):
        directory, filename = os.path.split(file_path)
        data = send_from_directory(directory, filename, as_attachment=True)
        return data


class ImportFromWeb(Resource):
    def post(self):
        file_upload = request.files['file']
        classify = request.form['type']
        user_id = request.form['user_id']
        data_json = request.form['check_list']
        check_list = json.loads(data_json)
        fileName = file_upload.filename
        if not os.path.exists(os.path.join('./', 'import')):
            os.system('mkdir import')
        file_upload.save(os.path.join('./import/', fileName))
        file_path = os.path.join('./import/', fileName)
        file_list = []
        update_time = time.strftime("%Y-%m-%d %H:%M:%S")
        from Source.spec2db_server.spec_import.new_import import ArlImport
        import_obj = ArlImport(file_list, user_id, update_time)
        data = {}
        try:
            data = import_obj.sheet_import(file_path, classify, check_list)
        except Exception:
            data['result'] = 1
            data['error_list'] = '服务器错误，请联系管理员'
        return data


class ImportArl(Resource):
    def post(self):
        file_upload = request.files['file']
        user_id = request.form['user_id']
        file_name = file_upload.filename
        if not os.path.exists(os.path.join('./', 'import')):
            os.system('mkdir import')
        file_upload.save(os.path.join('./import/', file_name))
        file_path = os.path.join('./import/', file_name)
        file_list = []
        from Source.spec2db_server.spec_import.new_import import ArlImport
        import_obj = ArlImport(file_list, user_id, '')
        result = import_obj.import_arl(file_path, user_id)
        return result


class ImportBlock(Resource):
    def post(self):
        file_upload = request.files['file']
        classify = request.form['type']
        user_id = request.form['user_id']
        up = 'no'  # request.form['up']
        down = 'no'  # request.form['down']
        # classify = request.get("classify")
        file_name = file_upload.filename
        if not os.path.exists(os.path.join('./', 'import')):
            os.system('mkdir import')
        file_upload.save(os.path.join('./import/', file_name))
        file_path = os.path.join('./import/', file_name)
        file_list = []
        from Source.spec2db_server.spec_import.new_import import ArlImport
        import_obj = ArlImport(file_list, user_id, '')
        result = import_obj.import_block(file_path, classify, user_id, up, down)
        return result


class ImportSchedule(Resource):
    def post(self):
        file_upload = request.files['file']
        user_id = request.form['user_id']
        # classify = request.get("classify")
        file_name = file_upload.filename
        if not os.path.exists(os.path.join('./', 'import')):
            os.system('mkdir import')
        file_upload.save(os.path.join('./import/', file_name))
        file_path = os.path.join('./import/', file_name)
        file_list = []
        from Source.spec2db_server.spec_import.new_import import ArlImport
        import_obj = ArlImport(file_list, user_id, '')
        result = import_obj.import_schedule(file_path, user_id)
        return result


class ImportWhitelist(Resource):
    def post(self):
        file_upload = request.files['file']
        user_id = request.form['user_id']
        classify = request.form['type']
        file_name = file_upload.filename
        if not os.path.exists(os.path.join('./', 'import')):
            os.system('mkdir import')
        file_upload.save(os.path.join('./import/', file_name))
        file_path = os.path.join('./import/', file_name)
        file_list = []
        from Source.spec2db_server.spec_import.new_import import ArlImport
        import_obj = ArlImport(file_list, user_id, '')
        result = import_obj.import_whitelist(file_path, classify, user_id)
        return result


class ImportFile(Resource):
    def post(self):
        file_upload = request.files['file']
        user_id = int(request.form['user_id'])
        group_name = str(request.form['group_name'])
        group_name = group_name.replace(' ', '_')
        group_id = int(request.form['group_id'])
        data = {"result": "NG"}
        try:
            fileName = file_upload.filename
            from Source.spec2db_server.arl.arl_group import ArlGroup
            # group_obj = ArlGroup()
            # group_id = group_obj.get_group_by_name(group_name)
            file_path = os.path.join('./', 'astafile')
            if not os.path.exists(file_path):
                os.mkdir(file_path)
                # os.system('mkdir file')
            file_path1 = os.path.join('./astafile/', group_name)
            if not os.path.exists(file_path1):
                os.mkdir(file_path1)
            create_time = time.strftime("%Y-%m-%d %H:%M:%S")
            file_time = create_time.replace("-", "")
            file_time = file_time.replace(":", "")
            file_time = file_time.replace(" ", "")  # 去掉空格
            file_path2 = os.path.join(file_path1+'/', file_time)
            os.mkdir(file_path2)
            file_upload.save(os.path.join(file_path2, fileName))
            file_url = os.path.join(file_path2+'/', fileName)
            from Source.spec2db_server.astafile_server.astafile_service import AstaFileRecord
            base_obj = AstaFileRecord()
            base_obj.import_file(user_id, group_id, fileName, file_url, create_time)
        except Exception as e:
            return data
        data["result"] = "OK"
        return data


class GetFileList(Resource):
    def get(self, group_id):
        from Source.spec2db_server.astafile_server.astafile_service import AstaFileRecord
        base_obj = AstaFileRecord()
        result = base_obj.get_file_list(group_id)
        return result


# class FileDownload(Resource):
#     def post(self):
#         data = request.get_json(force=True)
#         filePath = str(data[0].get("file_url"))
#         file_name = str(data[0].get("file_name"))
#         from flask import send_from_directory
#         file_url = filePath.replace('/' + file_name, '')
#         try:
#             data = send_from_directory(file_url, file_name, as_attachment=True)
#         except Exception as e:
#             print e.message
#         return data


class DownLoad(Resource):
    def get(self, rc_id):
        from Source.spec2db_server.astafile_server.astafile_service import AstaFileRecord
        base_obj = AstaFileRecord()
        file_name, file_url = base_obj.get_file_by_id(rc_id)
        file_url = file_url.replace('/' + file_name, '')

        data = send_from_directory(file_url, file_name, as_attachment=True)
        return data


class DeleteFile(Resource):
    def get(self, rc_id):
        from Source.spec2db_server.astafile_server.astafile_service import AstaFileRecord
        base_obj = AstaFileRecord()
        data = base_obj.delete_file(rc_id)
        return data


class DowAllGroups(Resource):
    def get(self):
        from Source.spec2db_server.astafile_server.astafile_service import AstaFileRecord
        base_obj = AstaFileRecord()
        filepath, filename = base_obj.get_zipfile()
        return send_from_directory(filepath, filename, as_attachment=True)


# class Import_Info(Resource):
#     def post(self):
#         file_upload = request.files['file']
#         fileName = file_upload.filename
#         file_name = str(fileName.split('.', 1)[0])
#         if not os.path.exists(os.path.join('./', 'import')):
#             os.system('mkdir import')
#         if not os.path.exists(os.path.join('./import', 'users')):
#             os.system('mkdir -p import/users')
#         if not os.path.exists(os.path.join('./export/users/', file_name)):
#             os.system('mkdir -p import/users/%s' % (file_name))
#         file_upload.save(os.path.join('./import/', fileName))
#         zip_file = zipfile.ZipFile(os.path.join('./import/', fileName))
#         file_time = time.strftime("%Y%m%d%H%M%S")
#         zip_file.extractall(os.path.join('./import/', 'users/', file_name, file_time))
#         zip_file.close()
#
#         # task = import_task.apply_async(args=(file_name, file_time),
#         #                                queue='import_task')
#         #
#         # result = {}
#         # result['Result'] = 2  # waiting
#         # result['Location'] = "/ImportResult/%s" % (task.id)
#         # return result
#         data = self.import_task(file_name, file_time)
#         return data

    def import_task(self, file_name, file_time):
        import sys
        reload(sys)
        sys.setdefaultencoding('UTF-8')
        from Source.spec2db_server.spec_import.arl_import import ArlImport
        from Source.spec2db_server.arl.arl_group import ArlUser

        user_id = ArlUser().find_user_by_name(file_name)
        # import_result = {}
        # import_result["ResultInfo"] = {}
        # import_result["Result"] = 0

        for root, dirs, files in os.walk(os.path.join('./import/users/', file_name, file_time)):
            # print root
            # print dirs
            # print files
            file_list = []
            if files:
                for file in files:
                    base_name, ext_name = os.path.splitext(os.path.join(root, file))
                    if ext_name.lower() in ('.xlsx',):
                        xil_filename = file
                        file_list.append(os.path.join(root, xil_filename))
        update_time = time.strftime("%Y-%m-%d %H:%M:%S")
        import_obj = ArlImport(file_list, user_id, update_time )
        sheet_import_result = import_obj.sheet_classify()
        # if sheet_import_result[import_obj.decode_filename]['result'] > 0:
        #     import_result["Result"] = 1
        # import_result["ResultInfo"].update(sheet_import_result)

        return sheet_import_result

# class Import_Result(Resource):
#     def get(self, task_id):
#         task = import_task.AsyncResult(task_id)
#         result = {}
#         if task.successful():
#             result['Result'] = task.info['Result']
#             result['Location'] = "/ImportResult/%s" % (task_id)
#             result['ResultInfo'] = task.info['ResultInfo']
#         else:
#             result['Result'] = 2     # waiting
#             result['Location'] = "/ImportResult/%s" % (task_id)
#
#         return result

class Update_Style_Tool(Resource):
    def post(self):
        file_upload = request.files['file']
        file_name = file_upload.filename
        file_path = '%d'%int(time.time())

        if not os.path.exists(os.path.join('./', 'temp')):
            os.system('mkdir temp')
        if not os.path.exists(os.path.join('./temp', file_path)):
            os.system('mkdir -p %s'%(os.path.join('./temp', file_path, 'input')))
            os.system('mkdir -p %s' % (os.path.join('./temp', file_path, 'result')))

        file_upload.save(os.path.join('./temp/', file_path, 'input', file_name))

        zip_file = zipfile.ZipFile(os.path.join('./temp/', file_path, 'input', file_name))
        zip_file.extractall(os.path.join('./temp/', file_path, 'input'))
        zip_file.close()

        #os.system('rm %s'%(os.path.join('./temp/', file_path, 'input', file_name)))
        # hu_data_list = []
        # tagl_def_data_list = []
        # tagl_ana_data_list = []

        for root, dirs, files in os.walk(os.path.join('./temp/', file_path, 'input')):
            if files:
                for file in files:
                    if file.split('.')[-1] != 'xlsx':
                        continue

                    os.system('cp %s %s'%(os.path.join(root, file.replace(' ', r'\ ')), \
                                        os.path.join('./temp', file_path, 'tmpworkexcel.xlsx')))

                    try:
                        dest_file_name = file.decode("utf8").encode("utf8")
                    except:
                        dest_file_name = file.decode("gbk").encode("utf8")

                    style_tool_cmd = '''java -Xms2g -Xmx8g -jar tools/export/jar/UpdateStyle.jar %s %s %s %s %s''' % \
                                     (os.path.join('./temp', file_path, 'tmpworkexcel.xlsx'),
                                      os.path.join('./template', 'HU_RequirementDefinition.Ver0.17_template.xlsx'),
                                      os.path.join('./template', 'TAGL_RequirementDefinitionVer0.12_template.xlsx'),
                                      os.path.join('./template', 'TAGL_RequirementAnalysis.ver1.00_template.xlsx'),
                                      os.path.join(os.path.join('./temp', file_path, 'result', dest_file_name.replace(' ', r'\ '))))

                    print style_tool_cmd
                    os.system(style_tool_cmd)
                    os.system('rm %s'%(os.path.join('./temp', file_path, 'tmpworkexcel.xlsx')))

        #             wb = load_workbook(filename=os.path.join(root, file))
        #             if ('HU要件定義書' in wb.get_sheet_names()):
        #                 ws = wb.get_sheet_by_name('HU要件定義書'.decode("UTF8"))
        #                 work_list = hu_data_list
        #                 iStartRow=10
        #                 iEndCol = 80
        #             elif ('TAGL要件定義' in wb.get_sheet_names()):
        #                 ws = wb.get_sheet_by_name('TAGL要件定義'.decode("UTF8"))
        #                 work_list = tagl_def_data_list
        #                 iStartRow = 6
        #                 iEndCol = 61
        #             elif ('TAGL要件分析' in wb.get_sheet_names()):
        #                 ws = wb.get_sheet_by_name('TAGL要件分析'.decode("UTF8"))
        #                 work_list = tagl_ana_data_list
        #                 iStartRow = 6
        #                 iEndCol = 136
        #             else:
        #                 ws = None
        #
        #             if ws:
        #                 for irow in range(iStartRow, ws.max_row+1):
        #                     row_list = []
        #                     for icol in range(1, iEndCol):
        #                         row_list.append(ws.cell(row=irow, column=icol).value)
        #                     if not row_list[1]:
        #                         break
        #                     work_list.append(row_list)
        #                 continue
        #
        # wb = load_workbook(filename=os.path.join('./template','HU_RequirementDefinition.ver0.17_noformat.xlsx'))
        # ws = wb.get_sheet_by_name('HU要件定義書'.decode("UTF8"))
        # ws._current_row=9
        # for row in hu_data_list:
        #     ws.append(list(row))
        # wb.save(os.path.join(os.path.join('./temp', file_path, 'result', 'all_hu_noformat.xlsx')))
        #
        # wb = load_workbook(filename=os.path.join('./template', 'TAGL_RequirementDefinition.ver0.10_noformat.xlsx'))
        # ws = wb.get_sheet_by_name('TAGL要件定義'.decode("UTF8"))
        # ws._current_row = 5
        # for row in tagl_def_data_list:
        #     ws.append(list(row))
        # wb.save(os.path.join(os.path.join('./temp', file_path, 'result', 'all_tagl_def_noformat.xlsx')))
        #
        # wb = load_workbook(filename=os.path.join('./template', 'TAGL_RequirementAnalysis.ver0.10_noformat.xlsx'))
        # ws = wb.get_sheet_by_name('TAGL要件分析'.decode("UTF8"))
        # ws._current_row = 5
        # for row in tagl_ana_data_list:
        #     ws.append(list(row))
        # wb.save(os.path.join(os.path.join('./temp', file_path, 'result', 'all_tagl_ana_noformat.xlsx')))

        zip_file_cmd = '''
                        cd %s;
                        zip result.zip -r result;'''%(os.path.join('./temp/', file_path))
        print zip_file_cmd
        os.system(zip_file_cmd)

        ret = {}
        ret['result'] = "OK"
        ret['data_path'] = "/StyleToolFile/%s"%file_path

        return ret

class Update_Style_ToolFile(Resource):
    def get(self, file_path):
        from flask import send_from_directory
        return send_from_directory(os.path.join('./temp/', file_path), 'result.zip', as_attachment=True)


class CommitLog(Resource):
    def get(self, group_id, user_id, start_time, end_time, size, page):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log(start_time, end_time, group_id, user_id, size, page)
        return data


class CommitLogByUser(Resource):
    def get(self, user_id, start_time, end_time):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log_by_user(user_id, start_time, end_time)
        return data


class CommitLogByGroup(Resource):
    def get(self, group_id, start_time, end_time):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log_by_group(group_id, start_time, end_time)
        return data


class DetailLog(Resource):
    def get(self, commit_id, commit_log_ref_id):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log_detail(commit_id, commit_log_ref_id)
        return data


class BriefLog(Resource):
    def get(self, commit_id, page_size, page_number):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log_brief(commit_id, page_size, page_number)
        return data


class CommitLogByClassifyAndRecord(Resource):
    def get(self, classify, r_id):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log_by_classify_and_record_id(classify, r_id)
        return data


class CommitLogByClassifyAndId(Resource):
    def get(self, classify, _id):
        from Source.spec2db_server.arl.commit_log import CommitLog
        obj = CommitLog()
        data = obj.get_commit_log_by_id(classify, _id)
        return data

class Role(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_group import ArlRole
        obj = ArlRole()
        result = obj.get_all_roles()
        return result


class ForestModelTypes(Resource):
    def get(self, _type):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        obj = ArlSpec()
        result = obj.get_forest_model_type(_type)
        return result


class AllJobStatus(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        obj = ArlSpec()
        result = obj.get_all_job_status()
        return result


class Change_Lock_Status(Resource):
    def get(self, _type, _id, lock):
        from Source.spec2db_server.arl.arl_base import ServiceBase
        obj = ServiceBase()
        result = obj.change_lock_status(lock, _type, _id)
        return result


# lock
class Lock_message(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        obj = ArlRecord()
        data = obj.get_lock_message()
        return data


# lock
class BlockWhiteList(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_server import ArlSpec
        obj = ArlSpec()
        data = obj.get_block_white_list()
        return data


class Req_Content(Resource):
    def get(self, type):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        obj = ArlRecord()
        data = obj.get_req_content(type)
        return data

class Check_Content(Resource):
    def get(self, type):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        obj = ArlRecord()
        data = obj.get_check_content(type)
        return data

class ServiceStatus(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_srv_status import ArlSrvStatus
        obj = ArlSrvStatus()
        result = {"result": "OK"}
        if obj.get_srv_status():
            result["result"] = "NG"
        return result


class Release(Resource):
    def post(self):
        result = {"result": "NG"}
        data = request.get_json(force=True)
        version = data.get("version")
        rls_date = data.get("date")
        pre_rls_date = data.get("pre_version")
        user_id = data.get("user_id")
        tmc_issue_url = data.get("tmc_issue")
        suntec_confirm_url = data.get("suntec_confirm")
        blocklist_url = data.get("blocklist")
        from Source.spec2db_server.arl.arl_release import ArlRelease, play
        obj = ArlRelease()
        play(start=True)
        rst, attach_file = obj.start(version, rls_date, pre_rls_date, user_id,
                                     tmc_issue_url, suntec_confirm_url, blocklist_url)
        play(start=False)
        if attach_file:
            print attach_file
            obj.copy_release_2_ftp(attach_file, rls_date)
        if rst == "OK":
            result["result"] = "OK"
            result["error"] = None
            result["attach_file"] = attach_file
        else:
            result["result"] = "NG"
            result["error"] = rst
            result["attach_file"] = attach_file
        return result

    def get(self):
        from Source.spec2db_server.arl.arl_release import ArlRelease
        obj = ArlRelease()
        result = {"result": "NG"}
        release_list = obj.get_releases_date()
        if release_list:
            result["result"] = "OK"
            result["release_list"] = release_list
        return result


class ReleaseVersion(Resource):
    def get(self):
        from Source.spec2db_server.arl.arl_release import ArlRelease
        obj = ArlRelease()
        result = {"result": "NG"}
        release_list = obj.get_releases()
        if release_list:
            result["result"] = "OK"
            result["release_list"] = release_list
        return result


class DownloadRelease(Resource):
    def get(self, attach_file):
        directory, filename = os.path.split(attach_file)
        data = send_from_directory(directory, filename, as_attachment=True)
        return data


# login
api.add_resource(Login, '/login/<username>/<password>')
# 注册
api.add_resource(Register, '/register/<username>/<password>/<email>')

api.add_resource(Summary, '/Summary')
api.add_resource(SummaryAll, '/SummaryAll')
api.add_resource(SummaryUser, '/SummaryUsers')
api.add_resource(SummaryBasic, '/SummaryBasic')
api.add_resource(ArlTreeInfo, '/ArlTreeInfo/<arl_id>/<int:user_id>/<int:group_id>')
api.add_resource(BasicTreeInfo, '/BasicTreeInfo/<type>/<_id>/<int:user_id>/<int:group_id>')
api.add_resource(FullContent, '/PostFullContent')
api.add_resource(BasicContent, '/PostBasicContent')
api.add_resource(ForestModelTypes, '/ForestModelTypes/<_type>')
api.add_resource(AllJobStatus, '/AllJobStatus')
api.add_resource(Change_Lock_Status, '/ChangeLockStatus/<_type>/<_id>/<int:lock>')

# ARL
api.add_resource(ARL_Category, '/ARLCategory', '/')
api.add_resource(ARL_UserCategory, '/ARLUserCategory/<int:user_id>')
api.add_resource(ARL_GroupCategory, '/ARLGroupCategory')
api.add_resource(ARL_Content, '/ARLContent/<arl_id>')
api.add_resource(ARL_SubHu, '/ARLSubHu/<arl_id>')
api.add_resource(ARL_Summary, '/ARLSummary/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(ARL_Summary2, '/ARLSummary2/<int:user_id>/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(ARL_User, '/ARLToUser', '/ARLByUser/<user_id>/<int:page_size>/<int:page_number>')
api.add_resource(ARL_UserSummary, '/ARLUserSummary/<user_id>/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(ARL_Group, '/ARLByGroup/<group_id>/<int:page_size>/<int:page_number>')
api.add_resource(ARL_PostInfo, '/ARLPostInfo/<arl_id>')
api.add_resource(ARL_AssignUser, '/ARLAssignUser')
api.add_resource(ARL_Graph, '/ARLGraph/<arl_id>')
api.add_resource(ARL_Update_GroupName, '/ARLUpdateGroupName/<group_id>/<new_group_name>')

# HU
api.add_resource(HU_Content, '/HUContent', '/HUContent/<hu_id>')
api.add_resource(HU_ContentAdd, '/HUContentAdd', '/HUContentAdd/<hu_id>')
api.add_resource(HU_ModelList, '/HUModelList')
api.add_resource(HU_id, '/HuNewId', '/HuNewId/<arl_id>')
api.add_resource(HU_PostInfo, '/HUPostInfo', '/HUPostInfo/<hu_id>')
api.add_resource(HU_SubDef, '/HuSubDef', '/HuSubDef/<hu_id>')
api.add_resource(HU_ContentDelete, '/HuDelete', '/HuDelete/<hu_id>')
api.add_resource(HU_Summary, '/HUSummary/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(HU_Summary2, '/HUSummary2/<int:user_id>/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(HU_Option, '/HuOption')
# DEF
api.add_resource(DEF_id, '/DefNewId', '/DefNewId/<hu_id>')
api.add_resource(DEF_Content, '/DEFContent', '/DEFContent/<def_id>')
api.add_resource(DEF_ContentDel, '/DEFDelete', '/DEFDelete/<def_id>')
api.add_resource(DEF_PostInfo, '/DEFPostInfo', '/DEFPostInfo/<def_id>')
api.add_resource(DEF_Summary, '/DEFSummary/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(DEF_Summary2, '/DEFSummary2/<int:user_id>/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(DEF_ModeList, '/DEFModelList')
# Analysis
api.add_resource(Analysis_Content, '/AnalysisContent', '/AnalysisContent/<def_id>')
api.add_resource(Analysis_Summary, '/AnalysisSummary/<category_id>/<int:page_size>/<int:page_number>')
api.add_resource(Analysis_Summary2, '/AnalysisSummary2/<int:user_id>/<category_id>/<int:page_size>/<int:page_number>')

api.add_resource(Analysis_ContentDel, '/AnalysisDelete', '/AnalysisDelete/<def_id>')
# Group
api.add_resource(Group_Add, '/GroupAdd', '/GroupAdd/<group_name>')
api.add_resource(Group_Del, '/GroupDelete', '/GroupDelete/<int:group_id>')
api.add_resource(Group_Modify, '/GroupModify', '/GroupModify/<int:group_id>/<group_name>')
api.add_resource(Group_Members, '/GroupMembers', '/GroupMembers/<int:group_id>')
api.add_resource(Group_Get_Members, '/GroupGetMembers', '/GroupGetMembers/<int:group_id>')
api.add_resource(Gp_Member_Add, '/GpMemberAdd')
api.add_resource(Gp_Member_Del, '/GpMemberDel', '/GpMemberDel/<int:group_id>/<int:user_id>')
api.add_resource(Group_AllGroups, '/GroupAllGroups')
api.add_resource(Group_UpdateRole, '/GpUpdateRole')
api.add_resource(Group_GroupDetail, '/GpGroupDetail')
# User
api.add_resource(All_User, '/AllUsers')
api.add_resource(User_Content, '/UserContent', '/UserContent/<int:user_id>')
api.add_resource(User_Del,  '/UserDel/<int:user_id>')
api.add_resource(Email_Alter, '/EmailAlter/<int:user_id>/<email>')
api.add_resource(Role, '/AllRoles')
api.add_resource(User_MyGroup, '/UserMyGroup/<int:user_id>')
# WhiteList
api.add_resource(WhiteList, '/WhiteList')

# ArlSummarize
api.add_resource(Summarize_By_User,  '/SummarizeByUser', '/SummarizeByUser/<date>')
api.add_resource(Summarize_By_Category,  '/SummarizeByCategory', '/SummarizeByCategory/<date>')
api.add_resource(Summarize_By_Category2,  '/SummarizeByCategory2', '/SummarizeByCategory2/<date>')

# Export
#doc_type is [HU_DEF, TAGL_DEF,TAGL_ANA]
#range is [All, User, Group]
#user_id is user's id
#date is now date info
api.add_resource(Export_Info, '/Export/<string:doc_type>/<string:range>/<int:user_id>/<string:date>')
api.add_resource(Export_Process, '/ExportProcess/<task_id>')
api.add_resource(Export_File, '/ExportFile/<task_id>')
api.add_resource(Export_Checklist, '/ExportChecklist/<string:start_time>/<string:end_time>')
api.add_resource(DownloadChecklist, '/DownloadChecklist/<path:file_path>')

# Import
# api.add_resource(Import_Info, '/Import')
# api.add_resource(Import_Result, '/ImportResult/<task_id>')
api.add_resource(ImportFromWeb, '/ImportFromWeb')
api.add_resource(ImportArl, '/ImportArl')
api.add_resource(ImportBlock, '/ImportBlock')
api.add_resource(ImportSchedule, '/ImportSchedule')
api.add_resource(ImportWhitelist, '/ImportWhitelist')

# astafile
api.add_resource(ImportFile, '/ImportFile')
api.add_resource(GetFileList, '/GetFileList/<int:group_id>')
# api.add_resource(FileDownload, '/FileDownload')
api.add_resource(DownLoad, '/DownLoad/<int:rc_id>')
api.add_resource(DeleteFile, '/DeleteFile/<int:rc_id>')
api.add_resource(DowAllGroups, '/DowAllGroups')
# Tool
api.add_resource(Update_Style_Tool, '/StyleTool')
api.add_resource(Update_Style_ToolFile, '/StyleToolFile/<string:file_path>')

# commit log
api.add_resource(CommitLog, '/CommitLog/<int:group_id>/<int:user_id>/<start_time>'
                            '/<end_time>/<int:size>/<int:page>')
api.add_resource(CommitLogByUser, '/CommitLogByUser/<int:user_id>/<string:start_time>/<string:end_time>')
api.add_resource(CommitLogByGroup, '/CommitLogByGroup/<int:group_id>/<string:start_time>/<string:end_time>')
api.add_resource(DetailLog, '/DetailLog/<int:commit_id>/<int:commit_log_ref_id>')
api.add_resource(BriefLog, '/BriefLog/<int:commit_id>/<int:page_size>/<int:page_number>')
api.add_resource(CommitLogByClassifyAndRecord, '/CommitLogByClassifyAndRecord/<classify>/<int:record_id>')
api.add_resource(CommitLogByClassifyAndId, '/CommitLogByClassifyAndId/<classify>/<string:_id>')

# lock
api.add_resource(Lock_message, '/LockMessage')
api.add_resource(BlockWhiteList, '/BlockWhiteList')
api.add_resource(Req_Content, '/ReqContent/<string:type>')
api.add_resource(Check_Content, '/CheckContent/<string:type>')
# status
api.add_resource(ReQueryStatus, '/ReQueryStatus/<string:id>/<string:type>/<int:state>')
api.add_resource(LookStatus, '/LookStatus/<string:id>/<string:type>')

# Release
api.add_resource(ServiceStatus, '/ServiceStatus')
api.add_resource(Release, '/Release')
api.add_resource(ReleaseVersion, '/ReleaseVersion')
api.add_resource(DownloadRelease, '/DownloadRelease/<path:attach_file>')


###########################################################################################################
# # HMI
###########################################################################################################
class HmiCategory(Resource):
    def get(self):
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        cat = obj.get_category_tree()
        return cat


class SummaryHmi(Resource):
    def get(self, condition_data, page, size):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        try:
            rowcount, summary_data = obj.summary_category(condition_data, page, size)
            result['rowcount'] = rowcount
            result['data'] = summary_data
        except Exception as e:
            print e
            result['result'] = 'NG'
        return result


class HmiDevStatus(Resource):
    """开发状态
    """
    def get(self, classify):
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        from Source.spec2db_server.hmi.hmi_screen_it import HmiScreenIt
        from Source.spec2db_server.hmi.hmi_it_progress_report import HmiItProgressReport
        if classify == 'req':
            obj = HmiReq()
        elif classify == 'itpro':
            obj = HmiItProgressReport()
        elif classify == 'it_dev':
            obj = HmiItProgressReport()
        else:
            obj = HmiScreenIt()
        result = {"result": "NG"}
        status = obj.get_dev_status(classify)
        if status:
            result["result"] = "OK"
            result["content"] = status
        return result


class AddDevStatus(Resource):
    def post(self):
        result = {'result': 'OK'}
        req_json = request.get_json(force=True)
        classify = req_json.get('classify')
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        from Source.spec2db_server.hmi.hmi_screen_it import HmiScreenIt
        from Source.spec2db_server.hmi.hmi_it_progress_report import HmiItProgressReport
        if classify == 'req':
            obj = HmiReq()
        elif classify == 'itpro':
            obj = HmiItProgressReport()
        elif classify == 'it_dev':
            obj = HmiItProgressReport()
        else:
            obj = HmiScreenIt()
        try:
            obj.add_dev_status(req_json)
        except Exception as e:
            print e
            result['result'] = 'NG'
        return result


class UpdateDevStatus(Resource):
    def post(self):
        result = {'result': 'OK'}
        req_json = request.get_json(force=True)
        classify = req_json.get('classify')
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        from Source.spec2db_server.hmi.hmi_screen_it import HmiScreenIt
        from Source.spec2db_server.hmi.hmi_it_progress_report import HmiItProgressReport
        if classify == 'req':
            obj = HmiReq()
        elif classify == 'itpro':
            obj = HmiItProgressReport()
        elif classify == 'it_dev':
            obj = HmiItProgressReport()
        else:
            obj = HmiScreenIt()
        try:
            obj.update_dev_status(req_json)
        except Exception as e:
            print e
            result['result'] = 'NG'
        return result


class HmiExport(Resource):
    def post(self):
        path = os.path.join('.', 'export', 'hmi')
        if not os.path.exists(path):
            os.makedirs(path)
        req_json = request.get_json(force=True)
        export_path = 'hmi'
        result = {}
        result['result'] = "NG"
        if req_json['type'] == 'Req':
            filter_dict = {}
            for _k,_v in req_json.iteritems():
                if _k == "type":
                    continue
                if len(_v) >0:
                    filter_dict[_k] = _v

            from Source.spec2db_server.hmi.req_export import ReqExport
            export_file_name = "HMI_Detail_All_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            ReqExport().do_export(
                os.path.join('./template', 'HMI_Detail_template_v005.xlsx'),
                os.path.join(os.path.join('export', export_path), export_file_name),
                filter_dict)

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name
        elif req_json['type'] == 'ARLPLAN':
            from Source.spec2db_server.hmi.hmi_info import HmiInfo
            obj = HmiInfo()
            template = r'./template/HMI_ARL_Schedule_template_v002.xlsx'
            sheet_name = 'HMI画面详情'
            file_name = 'HMI_ARL_Schedule_%s.xlsx' % obj.get_current_date()
            category_list = obj.load_data(template, sheet_name)
            obj.arl_hmi_info(category_list, template, sheet_name,
                             os.path.join('export', export_path, file_name))
            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = file_name
        elif req_json['type'] == 'InternalQA':
            from Source.spec2db_server.hmi.internalQA_export import InternalQAExport
            export_file_name = "HMI_Internal_QA_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            InternalQAExport().do_export(
                os.path.join('./template', 'HMI_InternalQA_v001.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name
        elif req_json['type'] == 'ExternalQA':
            from Source.spec2db_server.hmi.externalQA_export import ExternalQAExport
            export_file_name = "HMI_External_QA_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            ExternalQAExport().do_export(
                os.path.join('./template', 'HMI_ExternalQA_v002.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name
        elif req_json['type'] == 'BugManage':
            from Source.spec2db_server.hmi.externalBug_export import ExternalBugExport
            export_file_name = "HMI_External_Bug_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            ExternalBugExport().do_export(
                os.path.join('./template', 'HMI_ExternalBug_v002.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name
        elif req_json['type'] == 'daily_finished_detail':
            date = req_json.get('date')
            from Source.spec2db_server.hmi.req_export import ReqExport
            obj = ReqExport()
            out_path = os.path.join('export', export_path)
            template_file = os.path.join('./template', 'HMI_Detail_template_v005.xlsx')
            export_file_name = obj.do_export_daily_finished_detail(template_file, out_path, date)
            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name
        elif req_json['type'] == 'HMIScreen':
            from Source.spec2db_server.hmi.hmi_screen_export import HmiScreenExport
            export_file_name = "HMI_Screen_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            HmiScreenExport().do_export(
                os.path.join('./template', 'HMI_Screen_v001.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name

        elif req_json['type'] == 'HMIScreenIt':
            from Source.spec2db_server.hmi.hmi_screen_it_export import HmiScreenItExport
            export_file_name = "Hmi_Screen_Docker_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            HmiScreenItExport().do_export(
                os.path.join('./template', 'Hmi_Screen_Docker.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name

        elif req_json['type'] == 'HMIItScreen':
            from Source.spec2db_server.hmi.hmi_it_screen_export import HmiItScreenExport
            export_file_name = "Hmi_It_Screen__%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            HmiItScreenExport().do_export(
                os.path.join('./template', 'HMI_It_Screen.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name

        elif req_json['type'] == 'HmiItProgressReport':
            filter_dict = {}
            for _k, _v in req_json.iteritems():
                if _k == "type":
                    continue
                if len(_v) > 0:
                    filter_dict[_k] = _v

            from Source.spec2db_server.hmi.hmi_it_progress_report_export import HmiItProgressReportExport
            export_file_name = "IT_Progress_Report_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )

            HmiItProgressReportExport().do_export(
                os.path.join('./template', 'IT_Progress_Report.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name), filter_dict)

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name

        elif req_json['type'] == 'ItProgressQA':
            from Source.spec2db_server.hmi.it_progress_qa_export import ItProgressQAExport
            export_file_name = "IT_Progress_QA_%s.xlsx" % (
                time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            )
            ItProgressQAExport().do_export(
                os.path.join('./template', 'IT_Progress_QA.xlsx'),
                os.path.join(os.path.join('export', export_path),
                             export_file_name))

            result['result'] = "OK"
            result['Result_PathInfo'] = export_path
            result['Result_FileInfo'] = export_file_name
        return result


class HmiImport(Resource):
    def post(self):
        file_upload = request.files['file']
        user_id = request.form['user_id']
        doc_type = request.form['type']
        path = os.path.join('import', 'hmi')
        if not os.path.exists(path):
            os.makedirs(path)
        file_name = file_upload.filename
        file_path = os.path.join(path, file_name)
        file_upload.save(file_path)
        ret_data = {'result': 'NG'}
        if doc_type == 'Req':
            update_cols = request.form['update_cols']
            from Source.spec2db_server.hmi.req_import import ReqImport
            obj = ReqImport()
            ret_data = obj.do_import(file_path, update_cols, user_id, obj.get_current_time())
        elif doc_type == 'InternalQA':
            from Source.spec2db_server.hmi.internalQA_import import InternalQAImport
            obj = InternalQAImport()
            ret_data = obj.do_import(file_path, user_id, obj.get_current_time())
        elif doc_type == 'ExternalQA':
            from Source.spec2db_server.hmi.externalQA_import import ExternalQAImport
            obj = ExternalQAImport()
            ret_data = obj.do_import(file_path, user_id, obj.get_current_time())
        elif doc_type == 'BugManage':
            from Source.spec2db_server.hmi.externalBug_import import ExternalBugImport
            obj = ExternalBugImport()
            ret_data = obj.do_import(file_path, user_id, obj.get_current_time())
        elif doc_type == 'Hmiscreen':
            from Source.spec2db_server.hmi.hmi_screen_import import HmiScreenImport
            obj = HmiScreenImport()
            ret_data = obj.do_import(file_path, user_id, obj.get_current_time())
        elif doc_type == 'HmiScreenIt':
            from Source.spec2db_server.hmi.hmi_screen_it_import import HmiScreenItImport
            obj = HmiScreenItImport()
            ret_data = obj.do_import(file_path, user_id, obj.get_current_time())
        elif doc_type == 'HmiItResultModeTransition':
            from Source.spec2db_server.hmi.hmi_it_screen_import import HmiItScreenImport
            obj = HmiItScreenImport()
            ret_data = obj.do_imports(file_name, file_path, user_id, obj.get_current_time())
        elif doc_type == 'HmiItResultModeTransition2':
            from Source.spec2db_server.hmi.hmi_it_screen_import import HmiItScreenImport
            obj = HmiItScreenImport()
            ret_data = obj.do_imports(file_name, file_path, user_id, obj.get_current_time())
        elif doc_type == 'HmiItProgressReport':
            update_cols = request.form['update_cols']
            from Source.spec2db_server.hmi.it_progress_report_import import HmiItProgressReportImport
            obj = HmiItProgressReportImport()
            ret_data = obj.do_import(file_path, update_cols, user_id, obj.get_current_time())
        elif doc_type == 'ItProgressQA':
            from Source.spec2db_server.hmi.it_progress_qa_import import ItProgressImport
            obj = ItProgressImport()
            ret_data = obj.do_import(file_path, user_id, obj.get_current_time())
        return ret_data


class HmiExportFile(Resource):
    def get(self, path_info, file_info):
        print path_info,file_info
        return send_from_directory(os.path.join('./export', path_info), file_info, as_attachment=True)


class HmiDetailByID(Resource):
    def get(self, hu_id):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        try:
            result_data = obj.get_detail_by_hu_id(hu_id)
            result['detail'] = result_data
        except Exception as e:
            print e
            result['result'] = 'NG'
        return result


class HmiDetailBySreenId(Resource):
    def get(self, screen_id):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        try:
            result_data = obj.get_hmi_by_screen(screen_id)
            result['detail'] = result_data
        except Exception as e:
            print e
            result['result'] = 'NG'
        return result


class HmiTableCols(Resource):
    def get(self, table_name='req'):
        result = {'result': 'OK'}
        from Source.spec2db_server.arl.arl_table_manage import ArlTableMng
        obj = ArlTableMng.instance()
        try:
            col_list = obj.get_col_list(table_name)
            result['content'] = col_list
        except:
            result['result'] = 'NG'
        return result


class HmiChargers(Resource):
    """负责人
    """
    def get(self):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        try:
            chargers = obj.get_chargers()
            result['content'] = chargers
        except:
            result['result'] = 'NG'
        return result


class HmiFullContent(Resource):
    def post(self):
        req_json = request.get_json(force=True)
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        result = obj.hmi_full_content_post(req_json)
        return result


class HmiSteps(Resource):
    def get(self):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_req import HmiReq
        obj = HmiReq()
        try:
            steps = obj.get_steps()
            result['content'] = steps
        except:
            result['result'] = 'NG'
        return result


class HmiInternalQA(Resource):
    def get(self, page, size):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_internal_qa import HmiInternalQA
        obj = HmiInternalQA()
        rowcount, qa_list = obj.get_qa_list(size, page)
        if qa_list:
            result['content'] = qa_list
            result['rowcount'] = rowcount
        else:
            result['result'] = 'NG'
        return result

    def post(self):
        from Source.spec2db_server.hmi.hmi_internal_qa import HmiInternalQA
        req_json = request.get_json(force=True)
        result = HmiInternalQA().update_datalist(req_json['datas'])
        return result


class HmiInternalQADel(Resource):
    def get(self, db_rc_id):
        from Source.spec2db_server.hmi.hmi_internal_qa import HmiInternalQA
        result = HmiInternalQA().delete_by_dbid(db_rc_id)
        return result

    def post(self):
        from Source.spec2db_server.hmi.hmi_internal_qa import HmiInternalQA
        req_json = request.get_json(force=True)
        result = HmiInternalQA().delete_by_idlist(req_json['datas'])
        return result


class HmiExternalQA(Resource):
    def get(self, page, size):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_external_qa import HmiExternalQA
        obj = HmiExternalQA()
        rowcount, qa_list = obj.get_qa_list(page, size)
        if qa_list:
            result['content'] = qa_list
            result['rowcount'] = rowcount
        else:
            result['result'] = 'NG'
        return result

    def post(self):
        from Source.spec2db_server.hmi.hmi_external_qa import HmiExternalQA
        req_json = request.get_json(force=True)
        result = HmiExternalQA().update_datalist(req_json['datas'])
        return result


class HmiExternalQADel(Resource):
    def get(self, db_rc_id):
        from Source.spec2db_server.hmi.hmi_external_qa import HmiExternalQA
        result = HmiExternalQA().delete_by_dbid(db_rc_id)
        return result

    def post(self):
        from Source.spec2db_server.hmi.hmi_external_qa import HmiExternalQA
        req_json = request.get_json(force=True)
        result = HmiExternalQA().delete_by_idlist(req_json['datas'])
        return result


class HmiExternalBug(Resource):
    def get(self, page, size):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_external_bug import HmiExternalBug
        obj = HmiExternalBug()
        rowcount, qa_list = obj.get_qa_list(page, size)
        if qa_list:
            result['content'] = qa_list
            result['rowcount'] = rowcount
        else:
            result['result'] = 'NG'
        return result

    def post(self):
        from Source.spec2db_server.hmi.hmi_external_bug import HmiExternalBug
        req_json = request.get_json(force=True)
        result = HmiExternalBug().update_datalist(req_json['datas'])
        return result


class HmiExternalBugDel(Resource):
    def get(self, db_rc_id):
        from Source.spec2db_server.hmi.hmi_external_bug import HmiExternalBug
        result = HmiExternalBug().delete_by_dbid(db_rc_id)
        return result

    def post(self):
        from Source.spec2db_server.hmi.hmi_external_bug import HmiExternalBug
        req_json = request.get_json(force=True)
        result = HmiExternalBug().delete_by_idlist(req_json['datas'])
        return result


class Screen(Resource):
    def get(self, screen_id=None):
        result = {'result': 'NG'}
        from Source.spec2db_server.hmi.hmi_screen import HmiScreen
        obj = HmiScreen()
        screen_data = obj.get_by_id(screen_id)
        if screen_data:
            result['content'] = screen_data
            result['result'] = 'OK'
        return result

    def post(self):
        req_json = request.get_json(force=True)
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_screen import HmiScreen
        obj = HmiScreen()
        screen_list = obj.summary(req_json)
        if screen_list:
            result['content'] = screen_list
        else:
            result['result'] = 'NG'
        return result


class Report(Resource):
    def get(self, classify, date, step=None):
        from Source.spec2db_server.hmi.hmi_report import HmiReport
        from Source.spec2db_server.hmi.it_progress_schedule import It_Schedule
        result = {"result": "NG"}
        if classify == 'it_progress_report':
            obj = It_Schedule()
            daily_report = obj.schedule_info_for_web(date)
        else:
            obj = HmiReport()
            daily_report = obj.get_daily_report(classify, date, step)
        if daily_report:
            result["content"] = daily_report[:-1]
            result["sum"] = daily_report[-1]
            result["result"] = "OK"
        return result


class TimeReport(Resource):
    def get(self, start_date, end_date, step, classify):
        from Source.spec2db_server.hmi.hmi_report import HmiReport
        obj = HmiReport()
        result = {"result": "NG"}
        report_list = obj.get_time_report(start_date, end_date, step, classify)
        if report_list:
            result["result"] = "OK"
            result["content"] = report_list
        return result


class ScreenItReport(Resource):
    def get(self, classify, date):
        from Source.spec2db_server.hmi.hmi_screen_it_report import HmiScreenItReport
        obj = HmiScreenItReport()
        result = {"result": "NG"}
        daily_report = obj.get_daily_report(classify, date)
        if daily_report:
            result["content"] = daily_report[:-1]
            result["sum"] = daily_report[-1]
            result["result"] = "OK"
        return result


class ReportExport(Resource):
    def get(self, date, step):
        result = dict()
        if not os.path.exists(os.path.join('./','export')):
            os.system('mkdir export')
        if not os.path.exists(os.path.join('./export','hmi')):
            os.system('mkdir export/hmi')
        export_path = 'hmi'
        from Source.spec2db_server.hmi.hmi_report import HmiReport
        obj = HmiReport()
        template_name = r'./template/report_template.xlsx'
        file_name = 'report_by_author_' + date + '.xlsx'
        file_url = os.path.join(os.path.join('export', export_path), file_name)
        obj.export_report(date, step, template_name, file_url)
        result['result'] = "OK"
        result['Result_PathInfo'] = export_path
        result['Result_FileInfo'] = file_name
        return result


class ScreenReportExport(Resource):
    def get(self, date):
        result = dict()
        if not os.path.exists(os.path.join('./','export')):
            os.system('mkdir export')
        if not os.path.exists(os.path.join('./export','hmi')):
            os.system('mkdir export/hmi')
        export_path = 'hmi'
        from Source.spec2db_server.hmi.hmi_screen_it_report import HmiScreenItReport
        obj = HmiScreenItReport()
        template_name = r'./template/screen_report_template.xlsx'
        file_name = 'hmi_screen_report' + date + '.xlsx'
        file_url = os.path.join(os.path.join('export', export_path), file_name)
        obj.export_report(date, template_name, file_url)
        result['result'] = "OK"
        result['Result_PathInfo'] = export_path
        result['Result_FileInfo'] = file_name
        return result


class ItProgressExport(Resource):
    def get(self, date):
        result = dict()
        if not os.path.exists(os.path.join('./', 'export')):
            os.system('mkdir export')
        if not os.path.exists(os.path.join('./export', 'hmi')):
            os.system('mkdir export/hmi')
        export_path = 'hmi'
        from Source.spec2db_server.hmi.it_progress_schedule import It_Schedule
        obj = It_Schedule()
        template = r'./template/TAGL_IT_Schedule_template_v001.xlsx'
        sheet_name1 = '今日统计表'
        sheet_name2 = '今日与昨日对比统计表'
        file_name = 'TAGL_IT_Schedule_%s.xlsx' % obj.get_current_date()
        obj.schedule_info(date, template,  sheet_name1.decode("utf8"), sheet_name2.decode("utf8"),
                          os.path.join('export', export_path, file_name))
        result['result'] = "OK"
        result['Result_PathInfo'] = export_path
        result['Result_FileInfo'] = file_name
        return result


def cron_bakup_hmi():
    from Source.spec2db_server.hmi.hmi_req import HmiReq
    obj = HmiReq()
    from Source.spec2db_server.hmi.hmi_it_progress_report import HmiItProgressReport
    it_obj = HmiItProgressReport()
    print 'Start Auto Backup %s' % obj.get_current_time()
    obj.backup_req()  # 备份要件
    it_obj.backup_It_Progress()
    print 'End Auto Backup %s' % obj.get_current_time()
    from Source.spec2db_server.hmi.hmi_report import HmiReport
    report = HmiReport()
    report.issue_report()  # 生成汇总报告


class ItAuthors(Resource):
    """负责人
    """
    def get(self):
        result = {'result': 'OK'}
        from Source.spec2db_server.hmi.hmi_it_progress_report import HmiItProgressReport
        obj = HmiItProgressReport()
        try:
            chargers = obj.get_authors()
            result['content'] = chargers
        except:
            result['result'] = 'NG'
        return result


class ItSaveUpdateTime(Resource):

    def post(self):
        result = {'result': 'OK'}
        req_json = request.get_json(force=True)
        from Source.spec2db_server.hmi.hmi_it_progress_report import HmiItProgressReport
        obj = HmiItProgressReport()
        try:
            obj.author_update_time(req_json)
        except Exception as e:
            print e.message
            result['result'] = "NG"
        return result


api.add_resource(HmiCategory, '/HmiCategory')
api.add_resource(SummaryHmi, '/SummaryHmi/<string:condition_data>/<int:page>/<int:size>')
api.add_resource(HmiDetailBySreenId, '/HmiDetailBySreenId/<string:screen_id>')
api.add_resource(HmiDevStatus, '/HmiDevStatus/<string:classify>')
api.add_resource(AddDevStatus, '/AddDevStatus')
api.add_resource(UpdateDevStatus, '/UpdateDevStatus')
api.add_resource(HmiImport, '/HmiImport')
api.add_resource(HmiExport, '/HmiExport')
api.add_resource(HmiExportFile, '/HmiExportFile/<string:path_info>/<string:file_info>')
api.add_resource(HmiDetailByID, '/HmiDetailByID/<string:hu_id>')
api.add_resource(HmiTableCols, '/HmiTableCols/<string:table_name>')
api.add_resource(HmiChargers, '/HmiChargers')
api.add_resource(HmiSteps, '/HmiSteps')
api.add_resource(HmiInternalQA, '/HmiInternalQA/<int:page>/<int:size>')
api.add_resource(HmiInternalQADel, '/HmiInternalQADel/<db_rc_id>', '/HmiInternalQADel')
api.add_resource(HmiExternalQA, '/HmiExternalQA/<int:page>/<int:size>')
api.add_resource(HmiExternalQADel, '/HmiExternalQADel/<db_rc_id>', '/HmiExternalQADel')
api.add_resource(HmiExternalBug, '/HmiExternalBug/<int:page>/<int:size>')
api.add_resource(HmiExternalBugDel, '/HmiExternalBugDel/<db_rc_id>', '/HmiExternalBugDel')
api.add_resource(HmiFullContent, '/HmiFullContent')
api.add_resource(Screen, '/Screen/<string:screen_id>', '/Screen')
api.add_resource(ScreenReportExport, '/ScreenReportExport/<string:date>')
api.add_resource(Report, '/Report/<string:classify>/<string:date>/<string:step>')
api.add_resource(TimeReport, '/TimeReport/<string:start_date>/<string:end_date>/<string:step>/<string:classify>')
api.add_resource(ScreenItReport, '/ScreenItReport/<string:classify>/<string:date>')
api.add_resource(ReportExport, '/ReportExport/<string:date>/<string:step>')
api.add_resource(ItAuthors, '/ItAuthors')
api.add_resource(ItProgressExport, '/ItProgressExport/<string:date>')
api.add_resource(ItSaveUpdateTime, '/ItSaveUpdateTime')

if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    lock = LockFile('APScheduler')
    if not lock.is_locked():
        try:
            lock.acquire(1)
            scheduler = APScheduler()
            scheduler.init_app(app)
            scheduler.start()
            app.run(debug=True, host='', threaded=True)
            lock.release()
        except Exception as e:
            print e
            app.run(debug=True, host='', threaded=True)
    else:
        app.run(debug=True, host='', threaded=True)

