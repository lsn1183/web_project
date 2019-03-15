# -*- coding: UTF-8 -*-
import platform
import os
import time
from Source.spec2db_server.hmi.hmi_base import HMIBase
from openpyxl import Workbook
from openpyxl import load_workbook
HMI_DEV_STATUS_FINISH = u'完成'  # 完成
HMI_DEV_STATUS_QA = u'QA'  # "需要QA"
HMI_DEV_STATUS_UNFINISH = u'未完成'  # 空格
DEV_STATUS_EXCLUDE_LIST = u'对象外'
DEV_STATUS_FINISH_LIST = [HMI_DEV_STATUS_FINISH]
DEV_STATUS_UNFINISH_LIST = [HMI_DEV_STATUS_UNFINISH]
DEV_STATUS_QA_LIST = [HMI_DEV_STATUS_QA]
DEV_STATUS_UNDELAY_LIST = DEV_STATUS_FINISH_LIST + [DEV_STATUS_EXCLUDE_LIST] + DEV_STATUS_QA_LIST
DEV_STATUS_ALL = [HMI_DEV_STATUS_FINISH, HMI_DEV_STATUS_QA, HMI_DEV_STATUS_UNFINISH, DEV_STATUS_EXCLUDE_LIST]
if platform.system() == 'Windows':
    ISSUSE_ROOT = r'Z:\user\maotianyuan\hmi_daily_report'
else:
    ISSUSE_ROOT = r'ftp/user/maotianyuan/hmi_daily_report'


class HmiReport(HMIBase):
    def __init__(self):
        HMIBase.__init__(self)

    def get_time_report(self, start_date, end_date, step, report_type):
        self._pg.connect()
        date_list = self.get_date_list(self._pg, start_date, end_date)
        report_list = []
        classify = 'user'
        for one_date in date_list:
            one_day_report = self.get_daily_report(classify, one_date, step)
            charger_one_day = self.sum_all_by_charger(one_day_report, report_type)
            charger_one_day['date'] = one_date
            report_list.append(charger_one_day)
        return report_list

    def sum_all_by_charger(self, one_day_report, report_type):
        if report_type == 'no':
            key = 'all_unfinished'
        elif report_type == 'qa':
            key = 'all_qa'
        elif report_type == 'delay':
            key = 'all_delay'
        charger_dict = dict()
        for charger_report in one_day_report[:-1]:
            charger = charger_report.get('charger')
            user_list = charger_report.get('user_list')
            sum = 0
            for user in user_list:
                sum += user[key]
            charger_dict[charger] = sum
        return charger_dict

    def sum_finish_by_charger(self, one_day_report):
        charger_dict = dict()
        for charger_report in one_day_report[:-1]:
            charger = charger_report.get('charger')
            user_list = charger_report.get('user_list')
            sum = 0
            for user in user_list:
                sum += user['all_finished']
            charger_dict[charger] = sum
        return charger_dict


    def get_daily_report(self, classify, today, step):
        new_date = self.get_current_date()
        if today < new_date:
            is_today = False
        elif today == new_date:
            is_today = True
        else:
            return []
        daily_report = None
        try:
            if classify == 'user':
                daily_report = self.get_daily_report_by_user(today, step, is_today)
            elif classify == 'category':
                daily_report = self.get_daily_report_by_category(today, step, is_today)
        except Exception as e:
            print e
        finally:
            self._pg.close()
        if daily_report:
            return daily_report
        return []

    def get_yesterday_finish(self, pg, today, condition, parms):
        sqlcmd = """
            SELECT a.represent_req
            FROM (
            SELECT *
            FROM hmi.req_history
            WHERE backup_date = %s and 
            dev_status in (
            SELECT dev_status_id
            FROM hmi.dev_status
            where dev_major_category = '完成'
            )
            ) AS A
            LEFT JOIN (
            SELECT hu_id, dev_status
            FROM hmi.req_history
            WHERE backup_date = %s and
            dev_status in (
            SELECT dev_status_id
            FROM hmi.dev_status
            where dev_major_category = '完成'
            )
            ) AS B
            on A.hu_id = B.hu_id
            where (A.dev_status <> B.dev_status or B.dev_status is null)
            {condition}
            ORDER BY a.hu_id
        """.format(condition=condition)
        yesterday = self.get_yesterday(pg, today)
        sum_rep = 0
        sum_not_rep = 0
        if yesterday:
            pg.execute(sqlcmd, (today, yesterday)+tuple(parms))
            rows = pg.fetchall()
            if rows:
                for row in rows:
                    if row[0] == '○':
                        sum_rep += 1
                    else:
                        sum_not_rep += 1
        return sum_rep, sum_not_rep

    def get_yesterday(self, pg, today):
        sqlcmd = """SELECT
                    backup_date
                    from hmi.req_history where
                    backup_date < %s
                    limit 1"""
        pg.execute(sqlcmd, (today,))
        row = pg.fetchone()
        yesterday = None
        if row:
            yesterday = row[0]
        return yesterday

    def get_daily_report_by_user(self, today, step, is_today):
        condition = ''
        if is_today:
            table_name = 'hmi.req'
            if step != 'all':
                condition = 'WHERE step = %s'
        else:
            table_name = 'hmi.req_history'
            if step != 'all':
                condition = 'WHERE step = %s AND backup_date = %s'
        sqlcmd = """
        SELECT charger,
               author, 
               count(hu_id),
               array_agg(hu_id) hu_id_list, 
               array_agg(dev_major_category) dev_major_category_list,
               array_agg(apl_schedule) apl_schedule_list, -- APL日程
               array_agg(it_schedule) it_schedule_list,     -- 结合日程
               array_agg(represent_req) represent_req_list,  -- 代表要件
               array_agg(it_progress) it_progress_list      --结合测试进度
          FROM (
                SELECT t1.hu_id, 
                       (case when charger is null or charger = '' then '未分配'
                        else charger end) charger,
                       (case when author is null or author = '' then '未分配'
                        else author end) author,
                       apl_schedule, it_schedule, it_progress,
                       t2.dev_major_category, t1.dev_status,
                       represent_req
                  FROM {table_name} as t1
                  left join hmi.dev_status as t2
                  on t1.dev_status = t2.dev_status_id
                  {condition}
          ) as t1
          --where new_status in (1, 13) -- 完成
          group by charger, author
          order by charger = '未分配', charger, author
        """.format(table_name=table_name, condition=condition)
        self._pg.connect()
        if step == 'all':
            self._pg.execute(sqlcmd)
        else:
            if is_today:
                self._pg.execute(sqlcmd, (step, ))
            else:
                self._pg.execute(sqlcmd, (step, today))
        rows = self._pg.fetchall()
        daily_report = []
        sum_data = {"all_num": 0, "all_exclude": 0, "all_finished": 0, "all_unfinished": 0,
                    "all_qa": 0, "all_delay": 0,  "all_num_represent": 0, "all_exclude_represent": 0,
                    "all_unfinished_represent": 0, "all_finished_represent": 0, "all_delay_represent": 0,
                    "all_qa_represent": 0, "all_num_not_rep": 0,  "all_exclude_not_rep": 0,
                    "all_finished_not_rep": 0, "all_unfinished_not_rep": 0, "it_progress_success": 0,
                    "all_qa_not_rep": 0, "all_delay_not_rep": 0,
                    "today_plan": 0, "today_plan_reprent": 0, "today_plan_not_rep": 0,
                    "yesterday_finished_reprent": 0, "yesterday_finished_not_rep": 0
                    }
        from req_export import ReqExport
        req_obj = ReqExport()
        req_finish_count_by_user = req_obj.count_daily_finished_detail_by_user(self._pg, today, step, represent='Yes')
        finish_count_by_user = req_obj.count_daily_finished_detail_by_user(self._pg, today, step, represent='No')
        for row in rows:
            data = dict()
            charger, data["author"] = row[0:2]  #
            dev_major_category_list, apl_schedule_list, it_schedule_list, represent_req_list, it_progress_list = row[4:]
            ####################################################################################################
            # 所有总数
            data["all_num"] = row[2]
            # 对象外
            all_exclude = self.count_status(dev_major_category_list, DEV_STATUS_EXCLUDE_LIST)
            data["all_exclude"] = all_exclude
            # 所有已完了
            all_finished = self.count_status(dev_major_category_list, DEV_STATUS_FINISH_LIST,
                                             represent_req_list=represent_req_list, represent=None)
            data["all_finished"] = all_finished
            # 所有未完了
            all_unfinished = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                               reserve=False,
                                               represent_req_list=represent_req_list, represent=None)
            data["all_unfinished"] = all_unfinished
            # 所有QA
            all_qa = self.count_status(dev_major_category_list, DEV_STATUS_QA_LIST)
            data["all_qa"] = all_qa
            # 所有delay
            # all_delay = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
            #                               schedule_list=apl_schedule_list, today=today,
            #                               cls='before', reserve=True,
            #                               represent_req_list=represent_req_list, represent=None)
            # 代表Delay
            all_delay_represent = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                    schedule_list=it_schedule_list, today=today,
                                                    cls='before', reserve=True,
                                                    represent_req_list=represent_req_list, represent='Yes',
                                                    include_today=False)
            # 非代表Delay
            all_delay_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                  schedule_list=it_schedule_list, today=today,
                                                  cls='before', reserve=True,
                                                  represent_req_list=represent_req_list, represent='No',
                                                  include_today=False)
            data["all_delay"] = all_delay_represent + all_delay_not_rep
            ####################################################################################################
            # 代表（结合日程）
            data["all_num_represent"] = self.count_status(dev_major_category_list, [],
                                                          reserve=False,
                                                          represent_req_list=represent_req_list, represent='Yes')
            # 对象外: 代表
            all_exclude_represent = self.count_status(dev_major_category_list, DEV_STATUS_EXCLUDE_LIST,
                                                      represent_req_list=represent_req_list, represent='Yes')
            data["all_exclude_represent"] = all_exclude_represent
            # 所有未完了: 代表
            all_unfinished_represent = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                         reserve=False,
                                                         represent_req_list=represent_req_list, represent='Yes')
            data["all_unfinished_represent"] = all_unfinished_represent
            # 所有已完了: 代表
            all_finished_represent = self.count_status(dev_major_category_list, DEV_STATUS_FINISH_LIST,
                                                       represent_req_list=represent_req_list, represent='Yes')
            data["all_finished_represent"] = all_finished_represent
            # 所有delay: 代表
            all_delay_represent = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                    schedule_list=it_schedule_list, today=today,
                                                    cls='before', reserve=True,
                                                    represent_req_list=represent_req_list, represent='Yes',
                                                    include_today=False)
            data["all_delay_represent"] = all_delay_represent
            # 所有QA: 代表
            all_qa_represent = self.count_status(dev_major_category_list, DEV_STATUS_QA_LIST,
                                                 represent_req_list=represent_req_list, represent='Yes')
            data["all_qa_represent"] = all_qa_represent
            ####################################################################################################
            # 非代表（结合日程）
            data["all_num_not_rep"] = self.count_status(dev_major_category_list, [],
                                                        reserve=False,
                                                        represent_req_list=represent_req_list, represent='No')
            # 对象外: 非代表
            all_exclude_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_EXCLUDE_LIST,
                                                    represent_req_list=represent_req_list, represent='No')
            data["all_exclude_not_rep"] = all_exclude_not_rep
            # 所有已完了: 非代表
            all_finished_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_FINISH_LIST,
                                                     represent_req_list=represent_req_list, represent='No')
            data["all_finished_not_rep"] = all_finished_not_rep
            # 所有未完了: 非代表
            all_unfinished_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                       reserve=False,
                                                       represent_req_list=represent_req_list, represent='No')
            data["all_unfinished_not_rep"] = all_unfinished_not_rep
            # 结合测试完成: 非代表
            it_progress_success = self.count_status(it_progress_list, DEV_STATUS_FINISH_LIST,
                                                    represent_req_list=represent_req_list, represent='No')
            data["it_progress_success"] = it_progress_success
            # 所有QA: 非代表
            all_qa_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_QA_LIST,
                                               represent_req_list=represent_req_list, represent='No')
            data["all_qa_not_rep"] = all_qa_not_rep
            # 所有delay: 非代表
            all_delay_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                  schedule_list=it_schedule_list, today=today,
                                                  cls='before', reserve=True,
                                                  represent_req_list=represent_req_list, represent='No',
                                                  include_today=False)

            data["all_delay_not_rep"] = all_delay_not_rep
            ####################################################################################################
            # 今日
            # 今日计划完成总数: 代表
            today_plan_reprent = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=it_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='Yes')
            # 今日计划完成总数: 非代表
            today_plan_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=apl_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='No')
            data["today_plan"] = today_plan_reprent + today_plan_not_rep  # 今日计划完成总数
            # 今日代表要件数
            today_plan_reprent = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=it_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='Yes')
            data["today_plan_reprent"] = today_plan_reprent
            # 今日非代表要件数
            today_plan_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=apl_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='No')
            data["today_plan_not_rep"] = today_plan_not_rep
            # 昨日代表要件完成件数
            sum_rep = req_finish_count_by_user.get((charger, data["author"]))
            if not sum_rep:
                sum_rep = 0
            sum_not_rep = finish_count_by_user.get((charger, data["author"]))
            if not sum_not_rep:
                sum_not_rep = 0
            data["yesterday_finished_reprent"] = sum_rep
            # 昨日非代表要件完成件数
            data["yesterday_finished_not_rep"] = sum_not_rep
            self.count_sum(sum_data, data, sum_data.keys())
            if not daily_report or charger != daily_report[-1].get("charger"):
                user_list = [data]
                daily_report.append({"charger": charger,
                                     "user_list": user_list})
            else:
                user_list = daily_report[-1].get("user_list")
                user_list.append(data)
        sum_data["charger"] = u'总计'
        daily_report.append(sum_data)
        return daily_report

    def count_status(self, status_list, filter_status_list,
                     schedule_list=None, today=None,
                     cls='equal', reserve=False,
                     represent_req_list=None, represent=None,
                     include_today=True):
        """
        :param status_list:
        :param filter_status_list:
        :param schedule_list:
        :param today:
        :param cls:
        :param reserve:
        :param represent_req_list: 代表要件
        :param represent: Yes--代表要件, No--非代表要件， None--所有要件
        :return:
        """
        count = 0
        if not schedule_list:
            schedule_list = [None] * len(status_list)
        if not represent_req_list:
            represent_req_list = [None] * len(status_list)
        for status, schedule, represent_req in zip(status_list, schedule_list, represent_req_list):
            if not status:  # 空认为未完成
                status = HMI_DEV_STATUS_UNFINISH
            if represent == 'Yes':  # 选择代表要件
                if not represent_req:
                    continue
            elif represent == 'No':  # 选择非代表要件
                if represent_req:
                    continue
            else:  # 所有要件
                pass
            if(not reserve and (status in filter_status_list or not filter_status_list) or
               (reserve and status not in filter_status_list)):
                if cls == 'equal':  # 当天
                    if schedule == today:
                        count += 1
                elif cls == 'before':  # 当天及之前
                    if include_today:
                        if schedule <= today:
                            count += 1
                    else:
                        if schedule < today:
                            count += 1
                else:
                    count += 1
        return count

    def get_daily_report_by_category(self, today, step, is_today):
        condition = ''
        if is_today:
            table_name = 'hmi.req'
            if step != 'all':
                condition = 'WHERE step = %s'
        else:
            table_name = 'hmi.req_history'
            if step != 'all':
                condition = 'WHERE step = %s AND backup_date = %s'
        sqlcmd = """
        SELECT major_category, medium_category, small_category,
               array_agg(hu_id) hu_id_list,
               array_agg(dev_major_category) dev_major_category_list,
               array_agg(apl_schedule) apl_schedule_list, -- APL日程
               array_agg(it_schedule) it_schedule_list,     -- 结合日程
               array_agg(represent_req) represent_req_list,  -- 代表要件
               array_agg(it_progress) it_progress_list      --结合测试进度
          FROM (
                SELECT major_category, medium_category, small_category,
                       t1.hu_id, 
                       apl_schedule, it_schedule,
                       t2.dev_major_category, t1.dev_status,
                       it_progress, represent_req
                  FROM {table_name} as t1
                  left join hmi.dev_status as t2
                  on t1.dev_status = t2.dev_status_id
                  {condition}
          ) as t1
          group by major_category, medium_category, small_category
          order by major_category, medium_category, small_category
        """.format(table_name=table_name, condition=condition)
        self._pg.connect()
        if step == 'all':
            self._pg.execute(sqlcmd)
        else:
            if is_today:
                self._pg.execute(sqlcmd, (step, ))
            else:
                self._pg.execute(sqlcmd, (step, today))
        rows = self._pg.fetchall()
        daily_report = []
        sum_data = {"all_num": 0, "all_exclude": 0, "all_finished": 0, "all_unfinished": 0,
                    "all_qa": 0, "all_delay": 0, "all_num_represent": 0, "all_exclude_represent": 0,
                    "all_unfinished_represent": 0, "all_finished_represent": 0, "all_delay_represent": 0,
                    "all_qa_represent": 0, "all_num_not_rep": 0, "all_exclude_not_rep": 0,
                    "all_finished_not_rep": 0, "all_unfinished_not_rep": 0, "it_progress_success": 0,
                    "all_qa_not_rep": 0, "all_delay_not_rep": 0,
                    "today_plan": 0, "today_plan_reprent": 0, "today_plan_not_rep": 0,
                    "yesterday_finished_reprent": 0, "yesterday_finished_not_rep": 0
                    }
        from req_export import ReqExport
        req_obj = ReqExport()
        req_finish_count_by_category = req_obj.count_daily_finished_detail_by_category(self._pg, today, step, represent='Yes')
        finish_count_by_category = req_obj.count_daily_finished_detail_by_category(self._pg, today, step, represent='No')
        for row in rows:
            data = dict()
            major_category = row[0]
            medium_category = row[1]
            data["small_category"] = row[2]
            dev_major_category_list, apl_schedule_list, it_schedule_list, represent_req_list, it_progress_list = row[4:]
            ####################################################################################################
            # 所有总数
            data["all_num"] = len(row[3])
            # 对象外
            all_exclude = self.count_status(dev_major_category_list, DEV_STATUS_EXCLUDE_LIST)
            data["all_exclude"] = all_exclude
            # 所有已完了
            all_finished = self.count_status(dev_major_category_list, DEV_STATUS_FINISH_LIST,
                                             represent_req_list=represent_req_list, represent=None)
            data["all_finished"] = all_finished
            # 所有未完了
            all_unfinished = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                               reserve=False,
                                               represent_req_list=represent_req_list, represent=None)
            data["all_unfinished"] = all_unfinished
            # 所有QA
            all_qa = self.count_status(dev_major_category_list, DEV_STATUS_QA_LIST)
            data["all_qa"] = all_qa
            # 所有delay
            all_delay_represent = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                    schedule_list=it_schedule_list, today=today,
                                                    cls='before', reserve=True,
                                                    represent_req_list=represent_req_list, represent='Yes',
                                                    include_today=False)
            # 非代表Delay
            all_delay_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                  schedule_list=it_schedule_list, today=today,
                                                  cls='before', reserve=True,
                                                  represent_req_list=represent_req_list, represent='No',
                                                  include_today=False)
            data["all_delay"] = all_delay_represent + all_delay_not_rep
            ####################################################################################################
            # 代表（结合日程）
            data["all_num_represent"] = self.count_status(dev_major_category_list, [],
                                                          reserve=False,
                                                          represent_req_list=represent_req_list, represent='Yes')
            # 对象外: 代表
            all_exclude_represent = self.count_status(dev_major_category_list, DEV_STATUS_EXCLUDE_LIST,
                                                      represent_req_list=represent_req_list, represent='Yes')
            data["all_exclude_represent"] = all_exclude_represent
            # 所有未完了: 代表
            all_unfinished_represent = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                         reserve=False,
                                                         represent_req_list=represent_req_list, represent='Yes')
            data["all_unfinished_represent"] = all_unfinished_represent
            # 所有已完了: 代表
            all_finished_represent = self.count_status(dev_major_category_list, DEV_STATUS_FINISH_LIST,
                                                       represent_req_list=represent_req_list, represent='Yes')
            data["all_finished_represent"] = all_finished_represent
            # 所有delay: 代表
            all_delay_represent = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                    schedule_list=it_schedule_list, today=today,
                                                    cls='before', reserve=True,
                                                    represent_req_list=represent_req_list, represent='Yes',
                                                    include_today=False)
            data["all_delay_represent"] = all_delay_represent
            # 所有QA: 代表
            all_qa_represent = self.count_status(dev_major_category_list, DEV_STATUS_QA_LIST,
                                                 represent_req_list=represent_req_list, represent='Yes')
            data["all_qa_represent"] = all_qa_represent
            ####################################################################################################
            # 非代表（结合日程）
            data["all_num_not_rep"] = self.count_status(dev_major_category_list, [],
                                                        reserve=False,
                                                        represent_req_list=represent_req_list, represent='No')
            # 对象外: 非代表
            all_exclude_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_EXCLUDE_LIST,
                                                    represent_req_list=represent_req_list, represent='No')
            data["all_exclude_not_rep"] = all_exclude_not_rep
            # 所有已完了: 非代表
            all_finished_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_FINISH_LIST,
                                                     represent_req_list=represent_req_list, represent='No')
            data["all_finished_not_rep"] = all_finished_not_rep
            # 所有未完了: 非代表
            all_unfinished_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                       reserve=False,
                                                       represent_req_list=represent_req_list, represent='No')
            data["all_unfinished_not_rep"] = all_unfinished_not_rep
            # 结合测试完成: 非代表
            it_progress_success = self.count_status(it_progress_list, DEV_STATUS_FINISH_LIST,
                                                    represent_req_list=represent_req_list, represent='No')
            data["it_progress_success"] = it_progress_success
            # 所有QA: 非代表
            all_qa_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_QA_LIST,
                                               represent_req_list=represent_req_list, represent='No')
            data["all_qa_not_rep"] = all_qa_not_rep
            # 所有delay: 非代表
            all_delay_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNDELAY_LIST,
                                                  schedule_list=it_schedule_list, today=today,
                                                  cls='before', reserve=True,
                                                  represent_req_list=represent_req_list, represent='No',
                                                  include_today=False)
            data["all_delay_not_rep"] = all_delay_not_rep
            ####################################################################################################
            # 今日
            # 今日计划完成总数: 代表
            today_plan_reprent = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=it_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='Yes')
            # 今日计划完成总数: 非代表
            today_plan_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=apl_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='No')
            data["today_plan"] = today_plan_reprent + today_plan_not_rep  # 今日计划完成总数
            # 今日代表要件数
            today_plan_reprent = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=it_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='Yes')
            data["today_plan_reprent"] = today_plan_reprent
            # 今日非代表要件数
            today_plan_not_rep = self.count_status(dev_major_category_list, DEV_STATUS_UNFINISH_LIST,
                                                   schedule_list=apl_schedule_list, today=today,
                                                   represent_req_list=represent_req_list, represent='No')
            data["today_plan_not_rep"] = today_plan_not_rep
            # 昨日代表要件完成件数
            sum_rep = req_finish_count_by_category.get((major_category, medium_category, data["small_category"]))
            if not sum_rep:
                sum_rep = 0
            sum_not_rep = finish_count_by_category.get((major_category, medium_category, data["small_category"]))
            if not sum_not_rep:
                sum_not_rep = 0
            # condition = "AND major_category = %s AND medium_category = %s AND small_category = %s"
            # parms = [major_category, medium_category, data["small_category"]]
            # sum_rep, sum_not_rep = self.get_yesterday_finish(self._pg, today, condition, parms)
            data["yesterday_finished_reprent"] = sum_rep
            # 昨日非代表要件完成件数
            data["yesterday_finished_not_rep"] = sum_not_rep
            self.count_sum(sum_data, data, sum_data.keys())
            if not daily_report or major_category != daily_report[-1].get("category_name"):
                small_list = [data]
                medium_list = [{"category_name": medium_category, "sub_list": small_list}]
                daily_report.append({"category_name": major_category,
                                     "sub_list": medium_list})
            else:
                medium_list = daily_report[-1].get("sub_list")
                if medium_category != medium_list[-1].get("category_name"):
                    small_list = [data]
                    medium_list.append({"category_name": medium_category, "sub_list": small_list})
                else:
                    small_list = medium_list[-1].get("sub_list")
                    small_list.append(data)
        sum_data["category_name"] = u'总计'
        daily_report.append(sum_data)
        return daily_report

    def count_sum(self, sum_data, curr_data, keys):
        for key in keys:
            sum_val = sum_data.get(key)
            if not sum_val:
                sum_val = 0
            curr_val = curr_data.get(key)
            if not curr_val:
                curr_val = 0
            sum_val = sum_val + curr_val
            sum_data[key] = sum_val

    def export_report(self, today, step, template_name, file_url):
        self._pg.connect()
        wb = load_workbook(template_name)
        ws = wb.get_sheet_by_name('Sheet')
        sum_data = ["all_num", "all_exclude", "all_finished", "all_unfinished",
                    "all_qa", "all_delay", "all_num_represent", "all_finished_represent",
                    "all_unfinished_represent", "all_qa_represent", "all_delay_represent",
                     "all_num_not_rep",
                    "all_finished_not_rep", "all_unfinished_not_rep", "it_progress_success", "all_qa_not_rep",
                    "all_delay_not_rep",
                    "today_plan", "today_plan_reprent", "today_plan_not_rep",
                    "yesterday_finished_reprent", "yesterday_finished_not_rep"
                    ]
        report_data = self.get_daily_report('user', today, step)
        row_i = 4
        for report in report_data:
            charger = report.get('charger')
            user_list = report.get('user_list')
            if not user_list:
                continue
            for user in user_list:
                user_row = []
                author = user.get('author')
                user_row.append(charger)
                user_row.append(author)
                user_row += self.get_params(user, sum_data)
                ws.append(user_row)
            rs = 'A'+str(row_i)+':'+'A'+str(row_i+len(user_list)-1)
            ws.merge_cells(range_string=rs, start_row=row_i, start_column='1', end_row=row_i+len(user_list)-1,
                           end_column='1')
            row_i = row_i+len(user_list)
        info_data = report_data[-1]
        row = ['总计', '', info_data['all_num'], info_data['all_exclude'], info_data['all_finished'],
               info_data['all_unfinished'], info_data['all_qa'], info_data['all_delay'], info_data['all_num_represent'],
               info_data['all_finished_represent'], info_data['all_unfinished_represent'],
               info_data['all_qa_represent'], info_data['all_delay_represent'],
               info_data['all_num_not_rep'], info_data['all_finished_not_rep'],
               info_data['all_unfinished_not_rep'], info_data['it_progress_success'], info_data['all_qa_not_rep'],
               info_data['all_delay_not_rep'],
               info_data['today_plan'], info_data['today_plan_reprent'],
               info_data['today_plan_not_rep'], info_data['yesterday_finished_reprent'],
               info_data['yesterday_finished_not_rep']]
        rs = 'A'+str(row_i)+':'+'B'+str(row_i)
        ws.append(row)
        ws.merge_cells(rs)
        self._pg.close()
        wb.save(file_url)

    def get_steps(self):
        sqlcmd = """
        SELECT distinct step
          FROM hmi.req
          order by step
        """
        steps = []
        self._pg.connect()
        self._pg.execute(sqlcmd)
        rows = self._pg.fetchall()
        for row in rows:
            steps.append(row[0])
        self._pg.close()
        return steps

    def issue_report(self):
        template_name = r'./template/report_template.xlsx'
        obj = HmiReport()
        date = obj.get_current_date()
        issue_path = os.path.join('issue', 'hmi')
        if os.path.exists(issue_path):
            try:
                # 清空
                cmd = 'rm -rf %s' % issue_path
                os.system(cmd)
                os.makedirs(issue_path)
            except Exception as e:
                print e
                return None
        else:
            os.makedirs(issue_path)
        for step in self.get_steps():
            file_name = 'TAGL_HMI_DailyReport_%s_%s.xlsx' % (step, date)
            file_url = os.path.join(issue_path, file_name)
            obj.export_report(date, step, template_name, file_url)
        # 导出昨日完成要件明细
        from Source.spec2db_server.hmi.req_export import ReqExport
        export = ReqExport()
        export.do_export_daily_finished_detail('./template/HMI_Detail_template_v005.xlsx', issue_path, date)
        self.copy_release_2_ftp(issue_path, date)

    def copy_release_2_ftp(self, release_path, rls_date):
        # rls_date = self.conver_date_format(rls_date)
        if platform.system() == 'Windows':
            dest = ISSUSE_ROOT
        else:
            dest = os.path.join(os.path.expanduser('~'), ISSUSE_ROOT)
        if not os.path.isdir(dest):
            os.mkdir(dest)
        file_name = self.get_release_dest_file(rls_date, dest)
        dest = os.path.join(dest, file_name)
        cmd = 'cp -rf {src} {dest}'.format(src=release_path, dest=dest)
        print cmd
        print 'Copy Result to %s' % dest
        if not os.system(cmd):
            return True
        return False

    def get_release_dest_file(self, release_path, dest):
        base_name = os.path.basename(release_path)
        file_name, ext_name = os.path.splitext(base_name)
        i = 1
        while True:
            if not os.path.exists(os.path.join(dest, base_name)):
                break
            else:
                base_name = '_'.join([file_name, str(i)])
                if ext_name:
                    base_name = '.'.join([base_name, ext_name])
            i += 1
        return base_name

    def get_date_list(self, pg, start_date, end_date):
        sqlcmd = """
            SELECT distinct backup_date
            FROM hmi.req_history where backup_date >= %s and backup_date<= %s
            order by backup_date
        """
        pg.execute(sqlcmd, (start_date, end_date))
        rows = pg.fetchall()
        date_list = []
        if rows:
            for row in rows:
                date_list.append(row[0])
        pg.close()
        return date_list





def main():
    obj = HmiReport()
    # from req_export import ReqExport
    # req_obj = ReqExport()
    # obj._pg.connect()
    # data = req_obj.get_daily_finished_detail(obj._pg, '2018-03-14')
    # print data
    # obj.issue_report()
    data1 = obj.get_daily_report('user', '2018-04-01', 'SP31')
    data2 = obj.get_daily_report('category', '2018-04-02', 'all')
    print 'finish'
    # result = obj.get_daily_report_by_user(today='2018-03-07', step='SP30', is_today=True)
    # result = obj.get_daily_report_by_category(today='2018-03-01', step='SP30')
    # print result
    # template_name = './template/report_template.xlsx'
    # file_url = './test.xlsx'
    # obj.export_report(today='2018-03-14', step='SP30', template_name=template_name, file_url=file_url)


if __name__ == '__main__':
    import sys, os
    os.chdir('../')
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    main()
