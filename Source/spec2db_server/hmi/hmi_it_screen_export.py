# -*- coding: UTF-8 -*-
from Source.spec2db_server.hmi.hmi_base import HMIBase
from openpyxl import load_workbook

class HmiItScreenExport(HMIBase):
    def __init__(self):
        HMIBase.__init__(self)

    def do_export(self, template_file, out_file):
        wb = load_workbook(template_file)

        sheet_name_list = [u"モード遷移", u"画面遷移", u"Notify", u"内部遷移", u"アプリ初期化・終了処理"]
        self._pg.connect()
        for ws in sheet_name_list:
            if u"モード遷移" == ws:
                sheet = wb.create_sheet(ws)
                MODETRANSITION_EXPORT = """
                            SELECT distinct 
                                file_name, it_no, mode_name, category, video_content_callback, 
                                video_process_callback, audio_content_callback, audio_process_callback, 
                                event_trigger, action, action_process, test_item, remark, obstacle_id
                            FROM hmi.it_result_mode_transition
                            Order by file_name,it_no;
                        """
                self._pg.execute(MODETRANSITION_EXPORT)
                arl_db_list = self._pg.fetchall()
                title_list = ["文件名", "番号", "モードネーム", "処理種類",
                              "映像コンテントコールバック(APFWからコール)","映像コールバック処理内容",
                              "音声コンテントコールバック(APFWからコール)",
                              "音声コールバック処理内容", "イベント（トリガー）", "アクション", "アクション処理内容",
                              "テスト項目", "備考欄", "障害ID"]
                sheet.append(title_list)
                for row in arl_db_list:
                    sheet.append(list(row))
            if u"画面遷移" == ws:
                sheet = wb.create_sheet(ws)
                HMISCREENITMOVE_EXPORT = """
                           SELECT distinct
                                file_name, it_no, category, move_unit_id, move_first_id, 
                                move_id, move_unit_api, move_unit_content, move_first_api, move_first_content, 
                                uac_stc, uac_process_content, sb_uac, sb_process_content, sb_uac_callback, 
                                rel_service, test_item, remark, obstacle_id
                           FROM hmi.it_screen_move
                           Order by file_name,it_no;
                       """
                self._pg.execute(HMISCREENITMOVE_EXPORT)
                arl_db_list = self._pg.fetchall()
                title_list = ["文件名", "番号", "処理種類", "遷移元画面ID", "遷移先画面ID", "遷移イベント", "遷移元Exit処理API",
                              "遷移元Exit処理内容", "遷移先Entry処理API", "遷移先Entry処理内容", "インターフェース(STCからコール)",
                              "インターフェース処理内容", "インターフェース(UACからコール)", "インターフェース処理内容",
                              "UACコールバック呼び出し", "関連Service", "テスト項目", "備考欄", "障害ID"]
                sheet.append(title_list)
                for row in arl_db_list:
                    sheet.append(list(row))

            if u"Notify" == ws:
                sheet = wb.create_sheet(ws)
                NOTIFY_EXPORT = """
                            SELECT distinct
                                file_name, it_no, category, rel_screen_id, rel_service, 
                                notify_interface, notify_content, notify_process, trigger, action, 
                                action_process, uac_if, uac_if_process, test_item, remark, obstacle_id
                            FROM hmi.it_result_notify
                            Order by file_name,it_no;
                        """
                self._pg.execute(NOTIFY_EXPORT)
                arl_db_list = self._pg.fetchall()
                title_list = ["文件名", " 番号", "処理種類", "関連画面ID", "関連Service", "Notifyインターフェース", "Notify内容",
                              "Notifyに対する処理", "イベント（トリガー）", "アクション", "アクション処理内容",
                              "インターフェース(ServiceBridge/MCからのコール)", "インターフェース処理内容", "テスト項目",
                              "備考欄", " 障害ID"]
                sheet.append(title_list)
                for row in arl_db_list:
                    sheet.append(list(row))
            if u"内部遷移" == ws:
                sheet = wb.create_sheet(ws)
                INTERIOR_EXPORT_SQL = """ 
                                SELECT
                                    file_name, it_no, category, screen_id, classify, component_id, 
                                    length, stc_apfw, stc_process_content, uac_stc, uac_process_content, 
                                    sb_uac, sb_process_content, sb_uac_callback, rel_service, test_item, 
                                    remark, obstacle_id
                                FROM hmi.it_inside_move
                                Order by file_name,it_no;    
                            """
                self._pg.execute(INTERIOR_EXPORT_SQL)
                arl_db_list = self._pg.fetchall()
                title_list = ["文件名", "番号", "処理種類", "画面ID", "種別", "部品ID", "イベント",
                              "押下イベントの処理関数(APFWからコール)", "処理内容", "インターフェース(STCからコール)",
                              "インターフェース処理内容", "インターフェース(UACからコール)", "インターフェース処理内容",
                              "UACコールバック呼び出し", "関連Service", "テスト項目", "備考欄", "障害ID"]
                sheet.append(title_list)
                for row in arl_db_list:
                    sheet.append(list(row))
            if u"アプリ初期化・終了処理" == ws:
                sheet = wb.create_sheet(ws)
                ITRESULTINITEND_EXPOPRT = """
                                SELECT 
                                    file_name, it_no, category, inteface, if_process_content, 
                                    mc_process_content, uac_interface, uac_process_content, uac_nofity_no, 
                                    sb_inteface, sb_process_content, sb_nofity_no, sb_uac_callback, 
                                    rel_service, test_item, remark, obstacle_id
                                FROM hmi.it_result_init_end
                                Order by file_name,it_no;
                            """
                self._pg.execute(ITRESULTINITEND_EXPOPRT)
                arl_db_list = self._pg.fetchall()
                title_list = ["文件名", "番号", "処理種類", "インタフェース(APFWからコール)", "インターフェース処理内容",
                              "MC処理内容", "インターフェース(MC/STCからコール)", "インターフェース処理内容",
                              "UserActionConverter（UAC）: 関連Nofity番号", "インターフェース(UACからコール)",
                              "インターフェース処理内容", "関連Nofity番号", "UACコールバック呼び出し", "関連Service",
                              "テスト項目", "備考欄", "障害ID"]
                sheet.append(title_list)
                for row in arl_db_list:
                    sheet.append(list(row))

        wb.save(out_file)
        self._pg.close()

