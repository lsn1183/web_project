# -*- coding: UTF-8 -*-
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from Source.spec2db_server.hmi.hmi_base import HMIBase
HMI_START_ROW_NUM = 3


class HmiScreenItExport(HMIBase):
    def __init__(self):
        HMIBase.__init__(self)

    def do_export(self, template_file, out_file):
        export_sql = '''
        SELECT app_name, first_layer, second_layer, third_layer, layer4, layer5, screen_id, 
               charger, ft.status, in_migration_date, fs.status, 
               followup_migration_date, test_docker, remark
        FROM hmi.hmi_screen_it as hs
        left join hmi.hmi_screen_it_followup_status as fs
        on hs.followup_migration_id = fs.dev_status_id
        left join hmi.hmi_screen_it_status as ft
        on hs.in_migration_id = ft.dev_status_id
        order by app_name, first_layer, second_layer, third_layer, layer4, layer5,screen_id
        '''
        self._pg.connect()
        self._pg.execute(export_sql)
        screen_db_list = []
        for db_data in self._pg.fetchall():
            screen_db_list.append(db_data)
        openpyxl_wb = load_workbook(template_file)
        ws = openpyxl_wb.get_sheet_by_name('画面进度'.decode('utf8'))
        xls_row = HMI_START_ROW_NUM
        self.make_xls_validation(ws, len(screen_db_list), self._pg)
        for screen_data in screen_db_list:
            ws['B%d' % xls_row] = screen_data[0]
            ws['C%d' % xls_row] = screen_data[1]
            ws['D%d' % xls_row] = screen_data[2]
            ws['E%d' % xls_row] = screen_data[3]
            ws['F%d' % xls_row] = screen_data[4]
            ws['G%d' % xls_row] = screen_data[5]
            ws['H%d' % xls_row] = screen_data[6]
            ws['I%d' % xls_row] = screen_data[7]
            ws['J%d' % xls_row] = screen_data[8]
            ws['K%d' % xls_row] = screen_data[9]
            ws['L%d' % xls_row] = screen_data[10]
            ws['M%d' % xls_row] = screen_data[11]
            ws['N%d' % xls_row] = screen_data[12]
            ws['O%d' % xls_row] = screen_data[13]
            xls_row += 1
        openpyxl_wb.save(out_file)
        self._pg.close()

    def get_it_status_list(self, pg):
        """迁入开发分类
        :param pg:
        :return:
        """
        sqlcmd = """
        SELECT status
          FROM hmi.hmi_screen_it_status
          order by dev_status_id
        """
        pg.execute(sqlcmd)
        rows = pg.fetchall()
        it_status_list = []
        for row in rows:
            it_status_list.append(row[0])
        return it_status_list

    def get_it_followup_status_list(self, pg):
        """后续画面开发分类
        """
        sqlcmd = """
        SELECT status
          FROM hmi.hmi_screen_it_followup_status
          order by dev_status_id
        """
        pg.execute(sqlcmd)
        rows = pg.fetchall()
        it_followup_status_list = []
        for row in rows:
            it_followup_status_list.append(row[0])
        return it_followup_status_list

    def make_xls_validation(self, ws, row_count, pg):
        # 迁入开发分类
        it_status_list = self.get_it_status_list(pg)
        status_str = '"%s"' % ', '.join(it_status_list)
        dv_it_result = DataValidation(type="list", formula1=status_str, allow_blank=True)
        dv_it_result.sqref = 'J%s:J%s' % (HMI_START_ROW_NUM, row_count + HMI_START_ROW_NUM)
        ws.add_data_validation(dv_it_result)
        # 后续画面开发分类
        it_followup_status_list = self.get_it_followup_status_list(pg)
        followup_status_str = '"%s"' % ', '.join(it_followup_status_list)
        dv_is_dalian = DataValidation(type="list", formula1=followup_status_str, allow_blank=True)
        dv_is_dalian.sqref = 'L%s:L%s' % (HMI_START_ROW_NUM, row_count + HMI_START_ROW_NUM)
        ws.add_data_validation(dv_is_dalian)


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    HmiScreenItExport().do_export(r'..\template\Hmi_Screen_Docker.xlsx', 'ta.xlsx')
