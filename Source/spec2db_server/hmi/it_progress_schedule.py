# -*- coding: UTF-8 -*-
"""
Created on 2018-2-26

@author:
"""
from Source.spec2db_server.hmi.hmi_base import HMIBase
from openpyxl import load_workbook
IT_INFO_START_ROW = 3
IT_INFO_START_COL = 1

class It_Schedule(HMIBase):
    def __init__(self):
        HMIBase.__init__(self)

    def schedule_info_for_web(self, info_date):
        new_date = self.get_current_date()
        self._pg.connect()
        if info_date == new_date:
            schedule_data = self.get_it_status(self._pg)
        else:
            schedule_data = self.get_old_it_status(self._pg, info_date)
        process_status_dict, dev_status_dict2 = self.get_dev_status(self._pg)
        info_data = self.today_info(process_status_dict, dev_status_dict2, schedule_data, new_date)
        web_data_list = self._info_data_for_web(self._pg, info_data, info_date)
        self._pg.close()
        return web_data_list

    def _info_data_for_web(self, pg, info_data, info_date):
        info_list = ['author', 'sum_delay', 'sum_all', 'sum_not_object', 'sum_qa', 'sum_out',
                     'sum_block', 'sum_finish', 'sum_stop', 'sum_same', 'sum_analogy', 'sum_no_finish',
                     'no_start', 'sum_dev_fw16', 'sum_dev_model', 'sum_dev_ut',
                     'sum_dev_it', 'sum_dev_itn', 'update_time']
        web_data_list = []
        author_dict = self._get_update_time(pg, info_date)
        for key in info_data:
            user_dict = dict()
            data_list = info_data.get(key)
            update_time = author_dict.get(key)
            if key not in author_dict.iterkeys():
                self._insert_update_time(pg, key, info_date)
            data_list.append(update_time)
            user_dict["user_list"] = [dict(zip(info_list, data_list))]
            web_data_list.append(user_dict)
        sum_dict = {'author': '总计', 'sum_delay': 0, 'sum_all': 0, 'sum_not_object': 0,
                    'sum_qa': 0, 'sum_out': 0, 'sum_block': 0,
                    'sum_finish': 0, 'sum_stop': 0, 'sum_same': 0, 'sum_analogy': 0, 'sum_no_finish': 0,
                    'no_start': 0, 'sum_dev_fw16': 0,
                    'sum_dev_model': 0, 'sum_dev_ut': 0, 'sum_dev_it': 0, 'sum_dev_itn': 0}
        for web_data in web_data_list:
            user_list = web_data.get('user_list')
            for info in info_list[1:-1]:
                sum_dict[info] += user_list[0].get(info)
        web_data_list.append(sum_dict)
        return web_data_list

    def get_update_time(self, pg, date_time):
        author_dict = self._get_update_time(pg, date_time)
        return author_dict

    def _get_update_time(self, pg, date_time):
        sqlcmd = """
            SELECT author, update_time
            FROM hmi.it_update_date WHERE date_time = %s
        """
        pg.execute(sqlcmd, (date_time,))
        rows = pg.fetchall()
        author_dict = dict()
        if rows:
            for row in rows:
                author_dict[row[0]] = row[1]
        return author_dict

    def _insert_update_time(self, pg, author, date_time):
        sqlcmd = """
            INSERT INTO hmi.it_update_date(
            author, date_time)
            VALUES (%s, %s)
        """
        pg.execute(sqlcmd, (author, date_time))
        pg.commit()

    def _get_closest_date(self, pg, date):
        sqlcmd = """
        SELECT distinct backup_date
          FROM hmi.it_progress_report_history
          where backup_date < %s
          order by backup_date DESC
          limit 1
        """
        pg.execute(sqlcmd, [date])
        row = pg.fetchone()
        if row:
            return row[0]
        return ''

    def get_it_status(self, pg):
        sqlcmd = """
            SELECT (case when author is null or author = '' then '未分配'
                        else author end) author, req_id, dev_fw16_status, 
            dev_model_status, dev_ut_status,  
           dev_it_status, dev_itn_status, dev_status, plan_date
           FROM hmi.it_progress_report
        """
        pg.execute(sqlcmd)
        rows = pg.fetchall()
        it_dict_by_aurhor = dict()
        for row in rows:
            author = row[0]
            if author not in it_dict_by_aurhor.iterkeys():
                it_dict_by_aurhor[author] = []
                it_dict_by_aurhor[author].append(row)
            else:
                it_dict_by_aurhor[author].append(row)
        return it_dict_by_aurhor

    def get_old_it_status(self, pg, backup_date):
        self._pg.connect()
        sqlcmd = """
            SELECT author, req_id, dev_fw16_status, 
            dev_model_status, dev_ut_status,  
           dev_it_status, dev_itn_status, dev_status, plan_date
           FROM hmi.it_progress_report_history where backup_date = %s
        """
        pg.execute(sqlcmd, (backup_date,))
        rows = pg.fetchall()
        it_dict_by_aurhor = dict()
        for row in rows:
            author = row[0]
            if author not in it_dict_by_aurhor.iterkeys():
                it_dict_by_aurhor[author] = []
                it_dict_by_aurhor[author].append(row)
            else:
                it_dict_by_aurhor[author].append(row)
        return it_dict_by_aurhor

    def get_dev_status(self, pg):
        process_status_dict = {}
        dev_status_dict2 = {}
        pg.execute('select  process_status_id, category from hmi.it_process_status')
        for db_data in self._pg.fetchall():
            process_status_dict[db_data[0]] = db_data[1]
        pg.execute('select dev_status_id, category, dev_status from hmi.it_dev_status')
        for db_data in self._pg.fetchall():
            dev_status_dict2[db_data[0]] = {"category": db_data[1], "dev_status": db_data[2]}
        return process_status_dict, dev_status_dict2

    def schedule_info(self, info_date, template_name, sheet_name1, sheet_name2, file_url):
        wb = load_workbook(template_name)
        ws1 = wb.get_sheet_by_name(sheet_name1)
        ws2 = wb.get_sheet_by_name(sheet_name2)
        new_date = self.get_current_date()
        self._pg.connect()
        if info_date == new_date:
            closest_date = self._get_closest_date(self._pg, new_date)
            schedule_data = self.get_it_status(self._pg)
            yesterday_schedule_data = self.get_old_it_status(self._pg, closest_date)
        else:
            closest_date = self._get_closest_date(self._pg, info_date)
            schedule_data = self.get_old_it_status(self._pg, info_date)
            yesterday_schedule_data = self.get_old_it_status(self._pg, closest_date)
        process_status_dict, dev_status_dict2 = self.get_dev_status(self._pg)
        today_info = self.today_info(process_status_dict, dev_status_dict2, schedule_data, new_date)
        today_author_dict = self._get_update_time(self._pg, info_date)
        yesterday_author_dict = self._get_update_time(self._pg, closest_date)
        start_row = IT_INFO_START_ROW
        for key in today_info:
            row = today_info.get(key)
            update_time = today_author_dict.get(key)
            if key not in today_author_dict.iterkeys():
                self._insert_update_time(self._pg, key, info_date)
            row.append(update_time)
            self.write_excel_cell(ws1, start_row, row)
            start_row += 1
        if yesterday_schedule_data:
            yesterday_info = self.today_info(process_status_dict, dev_status_dict2, yesterday_schedule_data,
                                             closest_date)
            contrast_info = self.contrast_yesterday_info(today_info, yesterday_info)
            start_row = IT_INFO_START_ROW
            for key in contrast_info:
                today_update_time = today_author_dict.get(key)
                yesterday_update_time = yesterday_author_dict.get(key)
                if key not in yesterday_author_dict.iterkeys():
                    self._insert_update_time(self._pg, key, closest_date)
                contrast_row = contrast_info.get(key)
                contrast_row.append(today_update_time)
                contrast_row.append(yesterday_update_time)
                self.write_excel_cell(ws2, start_row, contrast_row)
                start_row += 1
        else:
            print 'yesterday not backup!!!'

        wb.save(file_url)
        self._pg.close()

    def today_info(self, process_status_dict, dev_status_dict2, schedule_data, new_date):
        today_info = dict()
        for key in schedule_data:
            row = [key]
            data_list = schedule_data.get(key)
            sum_all = len(data_list)  # 总数
            sum_not_object = 0  # 对象外
            sum_finish = 0  # 完成数
            sum_qa = 0  # qa数
            sum_out = 0  # 转出数
            sum_block = 0  # block数
            sum_delay = 0  # Delay数
            sum_stop = 0  # 暂停数
            sum_same = 0  # 同件
            sum_analogy = 0  # 类似件
            sum_no_finish = 0
            no_finish_list = []
            for data in data_list:
                if not dev_status_dict2.get(data[-2]):
                    dev_category = None
                    dev_status = None
                else:
                    dev_category = dev_status_dict2.get(data[-2]).get('category')  # 状态分类
                    dev_status = dev_status_dict2.get(data[-2]).get('dev_status')  # 开发状态
                plan_date = data[-1]  # 完了予定日

                if dev_category == '对象外':
                    sum_not_object += 1
                elif dev_category == '完成':
                    sum_finish += 1
                elif dev_category == 'QA':
                    sum_qa += 1
                elif dev_category == '转出':
                    sum_out += 1
                elif dev_category == 'Block':
                    sum_block += 1
                elif dev_category == '暂停':
                    sum_stop += 1
                else:
                    if dev_status == '类似件':
                        sum_analogy += 1
                    if plan_date < new_date:
                        sum_delay += 1
                    if dev_status == '同件':
                        sum_same += 1
                    else:
                        sum_no_finish += 1  # 如果是同件就不记残件里
                        no_finish_dict = {'dev_fw16_status': process_status_dict.get(data[2]),
                                          'dev_model_status': process_status_dict.get(data[3]),
                                          'dev_ut_status': process_status_dict.get(data[4]),
                                          'dev_it_status': process_status_dict.get(data[5]),
                                          'dev_itn_status': process_status_dict.get(data[6]),
                                          'dev_status': dev_category}
                        no_finish_list.append(no_finish_dict)
            result_list = self.info_no_finish(no_finish_list)
            row += [sum_delay, sum_all, sum_not_object,  sum_qa, sum_out, sum_block, sum_finish, sum_stop,
                    sum_same, sum_analogy, sum_no_finish]
            row += result_list
            today_info[key] = row
        return today_info

    def contrast_yesterday_info(self, today_info, yesterday_info):
        contrast_info = dict()
        for key in today_info:
            today_user = today_info.get(key)[:-1]
            yesterday_user = yesterday_info.get(key)
            if yesterday_user:
                for i in range(1, len(today_user)):
                    if today_user[i] > yesterday_user[i]:
                        today_user[i] = str(today_user[i]) + '↑'+str(today_user[i]-yesterday_user[i])
                    elif today_user[i] < yesterday_user[i]:
                        today_user[i] = str(today_user[i]) + '↓'+str(yesterday_user[i]-today_user[i])
            contrast_info[key] = today_user
        return contrast_info

    def write_excel_cell(self, sheet, row, row_data):
        start_col = IT_INFO_START_COL
        end_col = len(row_data) + 1
        while start_col < end_col:
            sheet.cell(row=row, column=start_col).value = row_data[start_col-IT_INFO_START_COL]
            start_col += 1

    def info_no_finish(self, no_finish_list):
        sum_dev_fw16 = 0
        sum_dev_model = 0
        sum_dev_ut = 0
        sum_dev_it = 0
        sum_dev_itn = 0
        no_start = 0
        for no_finish in no_finish_list:
            if no_finish.get('dev_fw16_status') == '未完成':
                sum_dev_fw16 += 1
            elif no_finish.get('dev_fw16_status') in ('', None):
                no_start += 1
            elif no_finish.get('dev_model_status') in ('未完成', '', None):
                sum_dev_model += 1
            elif no_finish.get('dev_ut_status') in ('未完成', '', None):
                sum_dev_ut += 1
            elif no_finish.get('dev_it_status') in ('未完成', '', None):
                sum_dev_it += 1
            elif no_finish.get('dev_itn_status') in ('未完成', '', None):
                sum_dev_itn += 1
            elif no_finish.get('dev_status') == '未完成':
                sum_dev_itn += 1
        # no_start = len(no_finish_list) - sum_dev_fw16 - sum_dev_model - sum_dev_ut - sum_dev_it - sum_dev_itn
        return [no_start, sum_dev_fw16, sum_dev_model, sum_dev_ut, sum_dev_it, sum_dev_itn]


if __name__ == '__main__':
    import sys
    reload(sys)
    import os
    os.chdir('../')
    sys.setdefaultencoding('UTF-8')
    test_obj = It_Schedule()
    xls_filename = r'./template/TAGL_IT_Schedule_template_v001.xlsx'
    sheet_name1 = '今日统计表'
    sheet_name2 = '今日与昨日对比统计表'
    print test_obj.get_current_time()
    file_url = './export/hmi/TAGL_IT_Schedule_test.xlsx'
    test_obj.schedule_info('2018-04-03', xls_filename, sheet_name1.decode("utf8"), sheet_name2.decode("utf8"), file_url)
    print test_obj.get_current_time()
    # data = test_obj.schedule_info_for_web('2018-03-28')
    # print data
