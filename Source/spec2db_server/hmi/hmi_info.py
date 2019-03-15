# -*- coding: UTF-8 -*-
import os
from hmi_req import HmiReq
from openpyxl import load_workbook
from openpyxl import Workbook
HMI_INFO_START_ROW = 5
HMI_INFO_START_COL = 8
HMI_INFO_END_COL = 39
HMI_INFO_COL_NUM = HMI_INFO_END_COL - HMI_INFO_START_COL + 1


class HmiInfo(HmiReq):

    def __int__(self):
        HmiInfo.__init__(self)

    def get_row(self, sheet, row=1, from_col=1, to_col=1):
        values = []
        j = from_col
        while j <= to_col:
            val = sheet.cell(row=row, column=j).value
            # if type(val) == unicode:
            #     val = val.replace(u'_x000D_', u'')
            #     val = val.replace(u'_x0000_', u'')
            values.append(val)
            j += 1
        return values

    def summary_category(self, pg, parms, condition):
        sqlcmd = self.list_2_select_sql(self.table_name, self.attr_list, condition)
        pg.execute(sqlcmd, parms)
        rows = pg.fetchall()
        data_list = []
        for row in rows:
            data_dict = dict()
            for i in range(0, len(self.attr_list)):
                data_dict[self.attr_list[i]] = row[i]
            data_list.append(data_dict)
        return data_list

    def select_no_object(self, pg):
        sqlcmd = """
        SELECT dev_status_id
        FROM hmi.dev_status where dev_major_category like '对象外%'
        """
        pg.execute(sqlcmd)
        rows = pg.fetchall()
        status_list = []
        for row in rows:
            status_list.append(row[0])
        return status_list

    def select_finish(self, pg):
        sqlcmd = """
        SELECT dev_status_id
        FROM hmi.dev_status where dev_major_category = '完成'
        """
        pg.execute(sqlcmd)
        rows = pg.fetchall()
        status_list = []
        for row in rows:
            status_list.append(row[0])
        return status_list

    def get_dev_status_by_category(self, pg, category=u'QA'):
        sqlcmd = """
        SELECT dev_status_id
        FROM hmi.dev_status where dev_major_category = %s
        """
        pg.execute(sqlcmd, [category])
        rows = pg.fetchall()
        status_list = []
        for row in rows:
            status_list.append(row[0])
        return status_list

    def select_screen(self, pg, hu_list):
        sqlcmd = """
            SELECT distinct screen_id
            FROM hmi.req_screen_rel where hu_id = %s
        """
        screen_list = []
        for hu_id in hu_list:
            pg.execute(sqlcmd, (hu_id, ))
            rows = pg.fetchall()
            for row in rows:
                if row[0] not in screen_list:
                    screen_list.append(row[0])
        return screen_list

    def load_data(self, xls_filename, sheet_name):
        xls_wb = load_workbook(xls_filename)
        ws = xls_wb.get_sheet_by_name(sheet_name)
        max_row = 168
        start_row = HMI_INFO_START_ROW
        category_list = []
        for row in range(start_row, max_row + 1):
            values = self.get_row(ws, row, from_col=5, to_col=7)  # 大中小分类
            category_dict = dict()
            category_dict['major_category'] = values[0]
            category_dict['medium_category'] = values[1]
            category_dict['small_category'] = values[2]
            category_list.append(category_dict)
        return category_list

    def write_excel_cell(self, sheet, row, row_data):
        start_col = HMI_INFO_START_COL
        end_col = HMI_INFO_END_COL
        while start_col <= end_col:
            sheet.cell(row=row, column=start_col).value = row_data[start_col - HMI_INFO_START_COL]
            start_col += 1

    def count_file_nos(self, it_file_names, it_nos, represent_req, dev_status):
        file_name_list = it_file_names.split('&')
        file_name_list = [_d.strip() for _d in file_name_list if _d]  # 去空和空格
        file_nos_list = it_nos.split('&')
        file_nos_list = [_d.strip() for _d in file_nos_list if _d]  # 去空和空格
        it_file_list = []
        for i in range(len(file_name_list)):
            it_file_dict = dict()
            if i+1 > len(file_nos_list):
                break
            it_nos = file_nos_list[i].replace('\r', '')
            it_nos_list = it_nos.split('\n')
            it_nos_list = [_d.strip() for _d in it_nos_list if _d]  # 去空和空格
            it_file_dict['file_name'] = file_name_list[i]
            it_file_dict['it_nos_list'] = set(it_nos_list)
            it_file_dict['represent_req'] = represent_req
            it_file_dict['dev_status'] = dev_status
            it_file_list.append(it_file_dict)
        return it_file_list

    def arl_hmi_info(self, category_list, template_name, sheet_name, file_url):
        self._pg.connect()
        wb = load_workbook(template_name)
        ws = wb.get_sheet_by_name(sheet_name)
        small_finish_id_list = self.get_small_status(self._pg)
        finish_id_list = self.select_finish(self._pg)  # 完成状态
        qa_dev_id_list = self.get_dev_status_by_category(self._pg, u'QA')
        no_object_list = self.select_no_object(self._pg)  # 对象外状态
        for i, category in enumerate(category_list, HMI_INFO_START_ROW):
            row = []
            parms = []
            condition = []
            major_category = category.get('major_category')
            medium_category = category.get('medium_category')
            small_category = category.get('small_category')
            if not major_category and not medium_category and not small_category:
                continue
            if major_category == u'-' or major_category == '-':
                row = [''] * HMI_INFO_COL_NUM
                self.write_excel_cell(ws, i, row)
                continue
            if major_category != u'*' or major_category != '*':
                parms.append(major_category)
                condition.append('major_category')
            if medium_category != u'*' or medium_category != '*':
                parms.append(medium_category)
                condition.append('medium_category')
            if small_category != u'*' or small_category != '*':
                parms.append(small_category)
                condition.append('small_category')
            hmi_list = self.summary_category(self._pg, parms, condition)
            if not hmi_list:
                row.append('该分类不存在！')
                row += [''] * (HMI_INFO_COL_NUM - 1)
                self.write_excel_cell(ws, i, row)
                continue
            hu_id_list = []
            hu_sum_num = len(hmi_list)
            hu_no_object = 0
            hu_finish = 0
            hu_no_finish = 0
            sum_ng = 0
            step = set()
            sum_screen = 0
            screen_finish = 0
            sum_db_arl = 0
            db_no_object = 0
            db_finish = 0
            db_no_finish = 0
            charge_list = []
            sum_no_screen_id = 0
            sum_cs_success = 0
            small_finish_num = 0
            hu_qa_num = 0  # HU需要QA件数
            db_qa_num = 0  # 代表要件需要QA件数
            apl_schedule_set = set()  # APL日程
            it_schedule_set = set()  # 结合日程
            it_pro_list = []
            for hmi in hmi_list:
                it_dict = dict()
                hu_id = hmi.get('hu_id')
                hu_id_list.append(hu_id)
                represent_req = hmi.get('represent_req') # 代表要件
                dev_status = hmi.get('dev_status')
                it_file_name = hmi.get('it_file_name')  # 结合测试书名称
                it_nos = hmi.get('it_nos')  # 结合测试书番号
                it_progress = hmi.get('it_progress')
                if it_file_name and it_nos:
                    if it_progress in ('完成', u'完成'):
                        if dev_status in small_finish_id_list:
                            small_finish_num += 1
                        it_pro_list += self.count_file_nos(it_file_name, it_nos, represent_req, dev_status)
                if dev_status in finish_id_list:
                    hu_finish += 1
                if dev_status in no_object_list:
                    hu_no_object += 1
                if dev_status in qa_dev_id_list:  # QA
                    hu_qa_num += 1
                ng_num = hmi.get('ng_num')
                if ng_num:
                    sum_ng += ng_num
                step.add(hmi.get("step"))
                hmi_ds_screen_id = hmi.get('hmi_ds_screen_id')
                if not hmi_ds_screen_id:
                    sum_no_screen_id += 1
                if it_progress == '完成' or it_progress == u'完成':
                    sum_cs_success += 1
                charger = hmi.get('charger')
                if charger not in charge_list:
                    charge_list.append(charger)
                if represent_req == '○' or represent_req == u'○':  # 代表要件
                    sum_db_arl += 1
                    if dev_status in finish_id_list:
                        db_finish += 1
                    if dev_status in no_object_list:
                        db_no_object += 1
                    if dev_status in qa_dev_id_list:  # QA
                        db_qa_num += 1
                apl_schedule_set.add(hmi.get("apl_schedule"))
                it_schedule_set.add(hmi.get("it_schedule"))
            screen_list = self.select_screen(self._pg, hu_id_list)
            sum_screen = len(screen_list)
            for screen in screen_list:
                id_list = self.select_hu_id_by_screen(self._pg, screen)
                success = True
                for hu_id in id_list:
                    if hu_id in hu_id_list:
                        # hmi_detail = self.get_detail_by_hu_id(hu_id)
                        hmi_detail = self.get_old_data(self._pg, hu_id)
                        status = self.get_dev_status_by_id(self._pg, hmi_detail[0]['dev_status'])
                        hmi_detail[0]['dev_status'] = status
                        screen_dev_status = hmi_detail[0].get('dev_status')
                        if screen_dev_status not in finish_id_list:
                            success = False
                            break
                if success:
                    screen_finish += 1
            db_no_finish = sum_db_arl - db_finish - db_no_object - db_qa_num
            hu_no_finish = hu_sum_num - hu_finish - hu_no_object - hu_qa_num
            apl_progress = '%.2f%%' % (float(hu_finish)/float(hu_sum_num) * 100)  # 求百分比
            ce_progress = '%.2f%%' % (float(sum_cs_success)/float(hu_sum_num) * 100)
            name_list = []
            for charger in charge_list:
                if not charger:
                    charger = ''
                name_list.append(charger)
            chargeStr = '\n'.join(name_list)
            if not apl_schedule_set or apl_schedule_set.issubset(set([None, ''])):
                row.append('-')  # APL日程
            else:
                row.append(max(apl_schedule_set))  # APL日程
            if not it_schedule_set or it_schedule_set.issubset(set([None, ''])):
                row.append('-')  # 结合日程
            else:
                row.append(max(it_schedule_set))  # 结合日程
            project_list = self.it_pro_info(self._pg, it_pro_list)
            info_list = [sum_screen, screen_finish, sum_db_arl, db_no_object, db_finish, db_qa_num, db_no_finish,
                         hu_sum_num, hu_no_object, hu_finish, hu_qa_num, hu_no_finish, sum_ng] + project_list + \
                        [small_finish_num, '', '\n'.join(step), sum_no_screen_id, apl_progress, ce_progress, chargeStr]
            row += info_list
            self.write_excel_cell(ws, i, row)
        wb.save(file_url)
        self._pg.close()

    def it_pro_info(self, pg, it_pro_list):
        check_list = []
        check_req_dict = dict()
        pro_num = 0
        obj_pro_num = 0
        pro_ing_num = 0
        pro_o_num = 0
        pro_x_num = 0
        pro_b_num = 0
        pro_na_num = 0
        pro_part_na_num = 0
        pro_blank_num = 0
        obstacle_id_num = []
        for it_pro in it_pro_list:
            file_name = it_pro.get('file_name')
            it_nos_list = it_pro.get('it_nos_list')
            represent_req = it_pro.get('represent_req')
            dev_status = it_pro.get('dev_status')
            for it_no in it_nos_list:
                check = ':'.join([file_name, it_no])
                if check in check_list:
                    if represent_req in ('○', u'○'):  # 同个项目存在多个要件， 是代表要件的只记一个
                        if check_req_dict.get(check) not in ('○', u'○'):
                            obj_pro_num += 1
                            check_req_dict[check] = '○'
                    continue
                check_list.append(check)
                pro_num += 1
                pro_dict = self.get_pro(file_name, it_no, pg)
                if not pro_dict:
                    pro_blank_num += 1
                if represent_req in ('○', u'○'):
                    obj_pro_num += 1
                    check_req_dict[check] = '○'
                if pro_dict:
                    test_item = pro_dict.get('test_item')
                    obstacle_id = pro_dict.get('obstacle_id')
                    obstacle_id_list = obstacle_id.split('\n')
                    if obstacle_id_list:
                        obstacle_id_list = [_d.strip() for _d in obstacle_id_list if _d]  # 去空和空格
                        for obstacle_id in obstacle_id_list:
                            if obstacle_id not in obstacle_id_num:
                                obstacle_id_num.append(obstacle_id)
                    if test_item in ('○', 'x', '●', '×'):
                        pro_ing_num += 1
                    if test_item == '○':
                        pro_o_num += 1
                    if test_item in ('●', '×'):
                        pro_x_num += 1
                    if test_item == 'B':
                        pro_b_num += 1
                    if test_item == 'NA':
                        pro_na_num += 1
                    if test_item == 'NA(部分)':
                        pro_part_na_num += 1
                    if test_item in (None, ''):
                        pro_blank_num += 1
        obstacle_id_str = '\n'.join(obstacle_id_num)
        return [pro_num, obj_pro_num, pro_ing_num, pro_o_num, pro_x_num,
                pro_b_num, pro_na_num, pro_part_na_num, pro_blank_num,
                obstacle_id_str]

    def get_pro(self, file_name, it_no, pg):
        sqlcmd = """
            SELECT test_item, obstacle_id
            FROM hmi.it_result_init_end
            WHERE file_name = %s AND it_no = %s
            UNION
            SELECT test_item, obstacle_id
            FROM hmi.it_result_mode_transition
            WHERE file_name = %s AND it_no = %s
            UNION
            SELECT test_item, obstacle_id
            FROM hmi.it_screen_move
            WHERE file_name = %s AND it_no = %s
	        UNION
            SELECT test_item, obstacle_id
            FROM hmi.it_inside_move
            WHERE file_name = %s AND it_no = %s
            UNION
            SELECT test_item, obstacle_id
            FROM hmi.it_result_notify
            WHERE file_name = %s AND it_no = %s
        """
        pg.execute(sqlcmd, (file_name, it_no, file_name, it_no,
                            file_name, it_no, file_name, it_no, file_name, it_no))
        row = pg.fetchone()
        if row:
            return {'test_item': row[0], 'obstacle_id': row[1]}
        else:
            return None

    def get_small_status(self, pg):
        sqlcmd = """
           SELECT dev_status_id
            FROM hmi.dev_status where status = '完成' 
        """
        pg.execute(sqlcmd)
        rows = pg.fetchall()
        status_list = []
        for row in rows:
            status_list.append(row[0])
        return status_list




if __name__ == '__main__':
    import sys
    reload(sys)
    os.chdir('../')
    sys.setdefaultencoding('UTF-8')
    obj = HmiInfo()
    xls_filename = r'./template/HMI_ARL_Schedule_template_v002.xlsx'
    sheet_name = 'HMI画面详情'
    print obj.get_current_time()
    file_url = './export/hmi/HMI_ARL_Sshedule_test.xlsx'
    category_list = obj.load_data(xls_filename, sheet_name)
    obj.arl_hmi_info(category_list, xls_filename, sheet_name, file_url)
    print obj.get_current_time()


