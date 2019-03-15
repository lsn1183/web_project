# -*- coding: UTF-8 -*-
from Source.spec2db_server.hmi.hmi_base import HMIBase
from openpyxl import load_workbook

class HmiScreenExport(HMIBase):
    def __init__(self):
        HMIBase.__init__(self)

    def get_scr_process(self, db_data_list):
        i_cnt_finish = 0
        for db_data in db_data_list:
            if db_data[10] and db_data[10] in ('对象外','完成'):
                i_cnt_finish+=1

        if len(db_data_list) == 0:
            return 0

        return int(i_cnt_finish*1.0/len(db_data_list)*100)


    def do_export(self, template_file, out_file):
        export_sql = '''SELECT a.screen_id, c.hu_id, c.arl_id, c.major_category, c.medium_category, c.small_category,
                               c.seq_diagram_file, c.step, c.charger, c.author, d.dev_major_category,
                               a.screen_progress, a.external_qa, a.remark, is_enter_screen
                        FROM hmi.hmi_screen as a
                        left join hmi.req_screen_rel as b
                        on a.screen_id = b.screen_id
                        left join hmi.req as c
                        on b.hu_id = c.hu_id
                        left join hmi.dev_status as d 
                        on c.dev_status = d.dev_status_id
                        order by a.screen_id'''
        self._pg.connect()
        self._pg.execute(export_sql)
        screen_db_dict = {}
        for db_data in self._pg.fetchall():
            if db_data[0] not in screen_db_dict:
                screen_db_dict[db_data[0]] = []
            screen_db_dict[db_data[0]].append(db_data)

        openpyxl_wb = load_workbook(template_file)
        ws = openpyxl_wb.get_sheet_by_name('HMI画面进度'.decode('utf8'))

        xls_row = 3
        for db_key in sorted(screen_db_dict.keys()):
            db_key_datalist = screen_db_dict[db_key]
            for i_key_data, key_data in enumerate(db_key_datalist, 0):
                if i_key_data == 0:
                    ws['A%d' % xls_row] = key_data[0]
                    ws.merge_cells('A%d:A%d'%(xls_row, xls_row+len(db_key_datalist)-1))
                    ws['K%d' % xls_row] = self.get_scr_process(db_key_datalist)
                    ws.merge_cells('K%d:K%d' % (xls_row, xls_row + len(db_key_datalist) - 1))
                    ws['L%d' % xls_row] = key_data[11]
                    ws.merge_cells('L%d:L%d' % (xls_row, xls_row + len(db_key_datalist) - 1))
                    ws['M%d' % xls_row] = key_data[12]
                    ws.merge_cells('M%d:M%d' % (xls_row, xls_row + len(db_key_datalist) - 1))
                    ws['N%d' % xls_row] = key_data[13]
                    ws.merge_cells('N%d:N%d' % (xls_row, xls_row + len(db_key_datalist) - 1))
                    ws['O%d' % xls_row] = key_data[14]
                    ws.merge_cells('O%d:O%d' % (xls_row, xls_row + len(db_key_datalist) - 1))
                ws['B%d' % xls_row] = key_data[1]
                ws['C%d' % xls_row] = key_data[2]
                ws['D%d' % xls_row] = key_data[3]
                ws['E%d' % xls_row] = key_data[4]
                ws['F%d' % xls_row] = key_data[5]
                ws['G%d' % xls_row] = key_data[6]
                ws['H%d' % xls_row] = key_data[7]
                ws['I%d' % xls_row] = key_data[8]
                ws['J%d' % xls_row] = key_data[9]

                xls_row+=1

        openpyxl_wb.save(out_file)

        self._pg.close()

if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    HmiScreenExport().do_export('/mnt/sharedoc/hmi/HMI_Screen_v001.xlsx', '/mnt/sharedoc/hmi/ta.xlsx')
