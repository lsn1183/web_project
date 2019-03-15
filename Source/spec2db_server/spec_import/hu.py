# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import time
from openpyxl import load_workbook
from import_base import ImportBase
import re

HU_RECORD_START_ROW = 9  #
HU_TITLES = [u'担当', u'ARLID', u'大分類', u'中分類', u'小分類', u'詳細', u'基本要件',
             u'要件', u'状態', u'トリガ', u'動作', u'備考', u'H/U要件定義ID', u'ユニークID',
             u'AMP', u'DSRC', u'DCM', u'RSE', u'TouchPad', u'SeparateDisp', u'システム構成', u'関連基本要件',
             u'除外', u'状態', u'トリガー', u'動作', u'状態', u'トリガー', u'動作', None,
             u'表示', u'タッチデバイス', u'ヒーコン/オーディオパネル', u'通信用モジュ-ル', u'音認用、BT用マイク',
             u'AVC-LANアンプ', u'MOST用アンプ', u'音の出力', u'リアシ-ト用モニタ', u'2口のUSBジャック', u'USBジャック', u'ラジオアンテナ', u'衛星用アンテナ',
             u'DAB用アンテナ', u'地デジ用アンテナ', u'GPS用アンテナ', None, u'時計', None, None,
             None, None, None, None, None, None, None, None,
             None, None, None, u'章　Chapter/Section\nor ページ番号 Page No',
             u'章　Chapter/Section\nor ページ番号 Page No', u'シーケンス仕様\nSequence spec.',
             u'シーケンス番号\nSequence No.', u'コマンドガイド\nCommand guide', u'OPC',
             u'機能配置・機能仕様\nFunction location and spec.',
             u'章　Chapter/Section\nor ページ番号 Page No',
             u'ドキュメント名\n※トヨタ仕様の場合は仕様番号も記載する事',
             u'章　Chapter/Section\nor ページ番号 Page No',
             u'テスト結果', u'未要件分析', u'備考1', u'備考2',
             u'要件漏れチェック', u'最終修正担当', u'日付', u'変更理由'
             ]


class SpecHU(ImportBase):
    """
    """
    def __init__(self):
        ImportBase.__init__(self)
        self.hu_file_name = ''
        self._col_list = ["author", "arl_id", "basic_req", "hu_id",
                          "unique_id", "amp", "dsrc",
                          "dcm", "rse", "touch_pad",
                          "separate_disp", "system_conf", "rel_requirement",
                          "exception", "dcu_status", "dcu_trigger",
                          "dcu_action", "meu_status", "meu_trigger",
                          "meu_action", "hu_category_id", "remark",
                          "sys_spec_chapter", "common_chapter", "common_seq_spec",
                          "common_seq_no", "common_cmd_guide", "common_opc",
                          "inter_loc_spec", "inter_chapter", "other_doc",
                          "other_chapter",  "test_results", "future_req",
                          "remark1", "remark2", "leak_check", "last_modifier",
                          "new_date", "reason"]

        self._point_col_list = ["review_result", "pointout_no", "pointout_status", "pointout_comment",
                                "reader_check", "reader2_check", "final_check", "pointout_charger",
                                "pointout_priority", "pointout_date", "suntec_status", "fixed", "suntec_remark",
                                "arl_rel", "suntec_cannot_modify"
                                ]
        self._from = ["major_category", "medium_category", "small_category", "detail",
                      "arl_req_post", "arl_status", "arl_trigger", "arl_action", "arl_remark",
                      "group_name", 'lock_status']

        self.covert_dict = {'hu_remark': 'remark',
                            'hu_remark1': 'remark1',
                            'hu_remark2': 'remark2',
                            'hu_md5': 'md5_key'
                            }

    def set_file(self, hu_file_name):
        self.hu_file_name = hu_file_name

    def store(self, sheet, user_id, update_time):
        self._pg.connect()
        max_row = sheet.max_row
        start_row = HU_RECORD_START_ROW + 1
        hu_import_record = dict()
        hu_import_record['result'] = 0
        from import_check import HUImportCheck
        hu_check = HUImportCheck()
        error_list = []
        # error_list = hu_check.hu_check_row(self._pg, sheet, start_row, max_row)
        if error_list:
            hu_import_record['result'] = 1
            hu_import_record["error_list"] = error_list
            return hu_import_record
        else:
            exception = self._store(sheet, start_row, max_row, user_id, update_time)
            if exception:
                hu_import_record['result'] = 1
                hu_import_record["error_list"] = exception
                return hu_import_record
            return hu_import_record

    def _store(self, sheet, start_row, max_row, user_id, update_time):
        from Source.spec2db_server.arl.hu_server import HuRecord
        hu_obj = HuRecord()
        count = 0
        commit_list = []
        for row in range(start_row, max_row + 1):
            # print row
            # 担当, ARLID
            values1 = self.get_row(sheet, row, from_col=1, to_col=2)
            # 基本要件
            values5 = self.get_row(sheet, row, from_col=7, to_col=7)
            # H/U要件定義ID ==> H/U分類ID
            values2 = self.get_row(sheet, row, from_col=13, to_col=30)
            # 備考 ==> 最終修正担当
            values3 = self.get_row(sheet, row, from_col=61, to_col=77)
            # Group ==> TAGL要件提出
            values4 = self.get_row(sheet, row, from_col=78, to_col=81)
            arl_id = values1[1]
            if not arl_id:
                break
            new_date = values4[0]  # 日付
            if new_date:
                values4[0] = self.convert_time(new_date)
            hu_id, unique_id = values2[0], self.convert_2_int(values2[1])
            values2[1] = unique_id
            if not hu_id:
                hu_id = '.'.join([arl_id, str(unique_id)])
                values2[0] = hu_id
            if hu_id[0] in ('A', 'B', 'C', 'D'):
                continue
            # amp, dsrc, dcm, rse, touch_pad, separate_disp
            i = 2
            while i < 8:
                values2[i] = self.convert_2_int(values2[i])
                i += 1
            hu_category_id = self.convert_2_int(values2[-1])
            values2[-1] = hu_category_id
            # print '%s,%s\n' % (arl_id, hu_id)
            # if type(unique_id) in (str, unicode):
            #     unique_id = int((hu_id.split('.'))[-1])
            #     values2[1] = unique_id
            params = values1 + values5 + values2 + values3 + values4
            model_vals = self.get_row(sheet, row, 31, 60)
            new_data = self._convert_info(params, model_vals)
            new_data["md5_key"] = values4[-1]
            try:
                md5_check, lock_check, curr_commit_list = hu_obj.add(self._pg, new_data, self._col_list, update_time)
                if not md5_check:
                    return hu_id + " MD5 不存在！"
                elif not lock_check:
                    return hu_id + " 不能被修改！"
                elif curr_commit_list:
                    commit_list += curr_commit_list
            except Exception as e:
                print e
                return e.message
        log = {'user_id': user_id, "commit_list": commit_list}
        if commit_list:
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            flag, commit_id = commit_log.add_commit_log2(self._pg, log)
            if flag:
                print '提交'
                self._pg.commit()
            else:
                print 'log_commit errol!'
        else:
            print '不用更新'

    def _convert_info(self, params, model_vals):
        hu = dict()
        for col, param in zip(self._col_list, params):
            hu[col] = param
        hu["model_list"] = self._convert_model_list(model_vals)
        return hu

    def _convert_model_list(self, model_vals):
        model_list = []
        for model_id, val in enumerate(model_vals, 1):
            model = dict()
            model["model_id"] = model_id
            model["val"] = val
            model_list.append(model)
        return model_list

    def check_titles(self, sheet, start_row=HU_RECORD_START_ROW, from_col=1):
        length = len(HU_TITLES)
        to_col = from_col + length - 1
        tiles_list = self.get_row(sheet, start_row, from_col, to_col)
        check_result = True

        for i in range(0, len(HU_TITLES)):
            if tiles_list[i] != HU_TITLES[i]:
                check_result = False
                break

        return check_result

    def _insert_record(self, params):
        sqlcmd = """
        INSERT INTO spec.hu(
            author, arl_id, hu_id,
            unique_id, amp, dsrc,
            dcm, rse, touch_pad,
            separate_disp, system_conf, rel_requirement,
            exception, dcu_status, dcu_trigger, 
            dcu_action, meu_status, meu_trigger,
            meu_action, hu_category_id, remark,
            sys_spec_chapter, common_chapter, common_seq_spec,
            common_seq_no, common_cmd_guide, common_opc,
            inter_loc_spec, inter_chapter, other_chapter,
            other_doc, test_results, future_req,
            new_date, reason)
        VALUES (%s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s)
        RETURNING hu_record_id;
        """
        self._pg.execute(sqlcmd, params)
        hu_record_id = self.fetch_id()
        return hu_record_id
    def _update_hu(self, new_params):
        sqlcmd = '''
        UPDATE spec.hu SET author=%s, arl_id=%s, hu_id=%s,
            unique_id=%s, amp=%s, dsrc=%s,
            dcm=%s, rse=%s, touch_pad=%s,
            separate_disp=%s, system_conf=%s, rel_requirement=%s,
            exception=%s, dcu_status=%s, dcu_trigger=%s, 
            dcu_action=%s, meu_status=%s, meu_trigger=%s,
            meu_action=%s, hu_category_id=%s, remark=%s,
            sys_spec_chapter=%s, common_chapter=%s, common_seq_spec=%s,
            common_seq_no=%s, common_cmd_guide=%s, common_opc=%s,
            inter_loc_spec=%s, inter_chapter=%s, other_chapter=%s,
            other_doc=%s, test_results=%s, future_req=%s,
            new_date=%s, reason=%s WHERE hu_record_id=%s;
        '''
        self._pg.execute(sqlcmd, new_params)


    def _insert_model_rel(self, arl_id, model_id, val):
        # print val
        sqlcmd = """
        INSERT INTO spec.hu_model_rel(
                    hu_record_id, model_id, val)
            VALUES (%s, %s, %s);
        RETURNING order_no;
        """
        self._pg.execute(sqlcmd, (arl_id, model_id, val))
        order_no = self.fetch_id()
        return order_no

    def _update_model_rel(self, model_id, hu_record_id, val):
        sqlcmd = '''
        UPDATE spec.hu_model_rel SET val=%s
        WHERE model_id=%s and hu_record_id=%s
        RETURNING order_no;
        '''
        self._pg.execute(sqlcmd, (val, model_id, hu_record_id))
        order_no = self.fetch_id()
        return order_no

    # 设置导入权限
    def _import_power(self, pg, user_id, arl_id, data_list):
        result = False
        from Source.spec2db_server.arl.arl_group import ArlUser
        user_obj = ArlUser()
        from Source.spec2db_server.arl.arl_server import ArlRecord
        obj = ArlRecord()
        id = obj.get_user_by_arl_id(arl_id)[0]
        if user_id == id:
            return True
        pows_list = []
        for data in data_list:
            group_id = data.get("group_id")
            roles = data.get("roles")
            for role in roles:
                pows_dict = user_obj.get_user_powr(pg, role)
                pows_list.append(pows_dict)
            for pows in pows_list:
                if pows.get("modify_all") == 1:
                    return True
                elif pows.get("modify_group") == 1:
                    from Source.spec2db_server.arl.arl_group import ArlGroup
                    obj = ArlGroup()
                    data_list = obj.get_group_members(group_id)
                    user_list = []
                    for data in data_list:
                        user_list.append(data.get("user_id"))
                    if id in user_list:
                        return True
                    else:
                        return False
        return result

    def new_store(self, hu_data_list, user_id, update_time, role, check_list):
        hu_import_record = dict()
        hu_import_record['result'] = 0
        self._pg.connect()
        commit_list = []
        from Source.spec2db_server.arl.hu_server import HuRecord
        from Source.spec2db_server.arl.basic_hu import BasicHuRecord
        from Source.spec2db_server.arl.arl_group import ArlGroup
        from Source.spec2db_server.arl.basic_def import BasicDefRecord
        from Source.spec2db_server.spec_import.definition import DefinitionSpec
        group_obj = ArlGroup()
        hu_obj = HuRecord()
        basic_hu_obj = BasicHuRecord()
        basic_def_obj = BasicDefRecord()
        sqlcmd = group_obj.get_group_id(user_id)
        self._pg.execute(sqlcmd)
        row = self._pg.fetchone()
        group_id = None
        if row:
            group_id = row[0]
        # data_list = group_obj.get_one_member(self._pg, user_id)
        for hu_data in hu_data_list:
            hu_id = hu_data["hu_id"].get("datavalue")
            if not hu_id:
                continue
            basic_flag = False
            # if not self._import_power(self._pg, user_id, arl_id, data_list): # 导入权限验证
            #     continue
            try:
                if re.match("^[B|C|D].*", hu_id):
                    new_data = self.convert_basic_data(hu_data)
                    if new_data.get("group_name"):
                        new_data.pop("group_name")
                    new_data["group_id"] = group_id
                    new_data["user_id"] = user_id
                    arl_id = new_data.get("arl_id")
                    if not arl_id:
                        continue
                    basic_flag = True
                    md5_check, lock_check, curr_commit_list = basic_hu_obj.add(self._pg, new_data, self._col_list,
                                                                               update_time, role)
                else:
                    new_data = self.convert_data(hu_data)
                    arl_id = new_data.get("arl_id")
                    if not arl_id:
                        continue
                    md5_check, lock_check, curr_commit_list = hu_obj.add(self._pg, new_data, self._col_list,
                                                                         update_time, role)
                if not md5_check:
                    hu_import_record['result'] = 1
                    if not new_data.get("md5_key"):
                        hu_import_record["error_list"] = hu_id + " MD5 不存在！"
                    else:
                        hu_import_record["error_list"] = hu_id + " MD5 不正确！"
                    return hu_import_record
                elif not lock_check:
                    hu_import_record['result'] = 1
                    hu_import_record["error_list"] = hu_id + " 不能被修改！"
                    return hu_import_record
                if curr_commit_list:
                    if basic_flag:
                        rel_def_list = basic_def_obj._get_by_parent_id(self._pg, hu_id)
                        for def_data in rel_def_list:
                            def_data['dcu_status'] = new_data['dcu_status']
                            def_data['dcu_trigger'] = new_data['dcu_trigger']
                            def_data['dcu_action'] = new_data['dcu_action']
                            def_data['meu_status'] = new_data['meu_status']
                            def_data['meu_trigger'] = new_data['meu_trigger']
                            def_data['meu_action'] = new_data['meu_action']
                            def_data['hu_remark'] = new_data['remark']
                            sub_result = basic_def_obj.add(self._pg, def_data, DefinitionSpec()._col_list,
                                                           update_time, "Admin", excel_import=False)
                            md5_check, lock_check, sub_commit_list = sub_result
                            if sub_commit_list:
                                curr_commit_list += sub_commit_list
                    commit_list += curr_commit_list
            except Exception as e:
                hu_import_record['result'] = 1
                hu_import_record["error_list"] = e.message
                return hu_import_record
        if commit_list:
            author_log = []
            charge_log = []
            author_commit_list, charger_commit_list = self.split_commit_list(commit_list)
            if author_commit_list:
                author_log = {'user_id': user_id, "commit_list": author_commit_list, "group_id": group_id}
            if charger_commit_list:
                charge_log = {'user_id': user_id, "commit_list": charger_commit_list, "group_id": group_id}
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            if author_log:
                flag, commit_id = commit_log.add_commit_log2(self._pg, author_log)
                if flag:
                    self.import_check_list(self._pg, check_list, 'hu', commit_id, 'author')
                else:
                    print 'log_commit errol!'
            if charge_log:
                flag, commit_id = commit_log.add_commit_log2(self._pg, charge_log)
                if flag:
                    self.import_check_list(self._pg, check_list, 'hu', commit_id, 'charge')
            print '提交'
            self._pg.commit()
            self._pg.close()
        else:
            self._pg.close()
            print '不用更新'
        return hu_import_record


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    hu_obj = SpecHU()
    hu_obj.set_file(r'C:\Users\yuyin\Desktop\import\HU_RequirementDefinition_Try_Update.xlsx')
    book = load_workbook(hu_obj.hu_file_name, data_only=True)
    sheet = book.get_sheet_by_name(r'HU要件定義書')
    update_time = time.strftime("%Y-%m-%d %H:%M:%S")
    user_id = 446
    hu_obj.store(sheet, user_id, update_time)
