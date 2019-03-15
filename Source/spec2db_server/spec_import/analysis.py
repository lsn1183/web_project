# -*- coding: UTF-8 -*-
"""
Created on 2017-11-3

@author: yuyin
"""
import time
import re
from hu import SpecHU
from import_base import ImportBase
from openpyxl import load_workbook

RECORD_START_ROW = 5  # 记录开始行号

ANA_TITLES = [u'担当者', u'TAGL要件定義ID', u'ユニークID', u'大分類', u'中分類', u'小分類', u'詳細',u'基本要件', u'関連基本要件', u'除外', u'DCU/MEUどちらの要件か', u'状態', u'トリガー', u'動作',
              None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,  None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
              None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
              u'bluetooth_service', u'bluetooth_setting', u'bluetooth_phone', u'phonebook', u'bluetooth_mail', u'dcm_phone', u'DCM Service', u'Help Service', u'Cm Service'
              u'Cu Service', u'Sfm Service', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, u'ElectronicOM', u'Browser',
              None, None, None, None, None, None, None, None, None, None, u'1', u'2', u'3', u'4', u'更新日', u'更新内容', u'astaファイル名']


class AnalysisSpec(SpecHU):
    def __init__(self):
        SpecHU.__init__(self)
        self.xls_file_name = None
        self._col_list = ["author_name", "definition_id",
                          "unique_id", "basic_req", "rel_requirement", "exception", "seq_diagram",
                          # "application", "kernel", "systemd",
                          "supple_spec", "uncheck", "remark", 'ana_1', 'ana_2', 'ana_3', 'ana_4',
                          "new_date", "reason", "asta_filename",
                          "unique_id"
                          ]
        self._from = ["major_category", "medium_category", "small_category", "detail",
                      "dcu_meu", "pf_status", "pf_trigger", "pf_action", "group_id", 'lock_status'
                      ]
        self.covert_dict = {
            'ana_remark': 'remark',
            'ana_md5': 'md5_key'
        }

    def set_file(self, xls_file_name):
        self.xls_file_name = xls_file_name

    def store(self, sheet, user_id, update_time):
        self._pg.connect()
        max_row = sheet.max_row
        start_row = RECORD_START_ROW + 1
        ana_import_record = {}
        ana_import_record['result'] = 0
        ana_import_record["error_list"] = {}
        error_list = {}
        ####################################
        # ## Model list
        ####################################
        # analysis_models = self.get_models(sheet, row=start_row - 1, from_col=16, to_col=120)
        # obj_model = AnalysisModel()
        # model_list = []
        # for models in analysis_models:
        #     # model = json.dumps(models)
        #     model_id = obj_model.get_model_id(models)
        #     model_list.append(model_id)
        # return 0
        if error_list:
            ana_import_record['result'] = 1
            ana_import_record["error_list"] = error_list
            return ana_import_record
        else:
            exception = self._store(sheet, start_row, max_row, user_id, update_time)
            if exception:
                ana_import_record['result'] = 1
                ana_import_record["error_list"] = exception
                return ana_import_record
            return ana_import_record

    def _store(self, sheet, start_row, max_row, user_id, update_time):
        from Source.spec2db_server.arl.analysis_service import AnalysisRecord
        def_obj = AnalysisRecord()
        commit_list = []
        for row in range(start_row, max_row + 1):
            values1 = self.get_row(sheet, row, from_col=1, to_col=3)  # 3个
            values2 = self.get_row(sheet, row, from_col=8, to_col=10)  # 基本要件——除外
            values3 = self.get_row(sheet, row, from_col=15, to_col=15)  # 1个: 除外
            values4 = self.get_row(sheet, row, from_col=121, to_col=127)  # 補足参照仕様書==>備考==>1234
            values5 = self.get_row(sheet, row, from_col=128, to_col=130)  # 更新日==>astaファイル名
            values6 = self.get_row(sheet, row, from_col=131, to_col=132)  # 组, MD5
            # values6 = self.get_row(sheet, row, from_col=131, to_col=142)  # 指摘
            definition_id = values1[1]
            if not definition_id:
                break
            if definition_id[0] in ('A', 'B', 'C', 'D'):
                continue
            unique_id = int(values1[2])
            values1[2] = unique_id
            if type(unique_id) in (str, unicode):
                unique_id = int((definition_id.split('.'))[-1])
                values1[2] = unique_id
            new_date = values4[-2]  # 日付
            if new_date:
                values4[-2] = self.convert_time(new_date)
            # point_date = values6[9]  # 指摘日
            # if point_date:
            #     values6[9] = self.convert_time(point_date)
            # else:
            #     continue
        #     point_dict = self._convert_point_out(values6)
        #     point_dict["id"] = definition_id
        #     point_dict["classify "] = "analysis"
        #     point_dict["update_time"] = update_time
        #     from Source.spec2db_server.arl.arl_base import ServiceBase
        #     obi = ServiceBase()
        #     if obi._add_point_out(self._pg, point_dict):
        #         continue
        #     else:
        #         print "出错！"
        # self._pg.commit()
        # self._pg.close()
            params = values1 + values2 + values3 + values4 + values5
            model_vals = self.get_row(sheet, row, 16, 120)
            new_data = self._convert_info(params, model_vals)
            new_data["md5_key"] = values6[-1]
            try:
                md5_check, lock_check, curr_commit_list = def_obj.add(self._pg, new_data, self._col_list, update_time)
                if not md5_check:
                    return definition_id + " MD5 不存在！"
                elif not lock_check:
                    return definition_id + " 不能被修改！"
                elif curr_commit_list:
                    commit_list += curr_commit_list
            except Exception as e:
                print e
                return e.message
        log = {'user_id': user_id, "commit_list": commit_list}
        if commit_list:
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            flag, commit_id = commit_log.add_commit_log2(self._pg, log)
            if flag:
                print '提交'
                self._pg.commit()
            else:
                print 'log_commit errol!'
        else:
            print '不用更新'

    def _get_parent_model(self, sheet, start_row, cur_row, col):
        if cur_row <= 0:
            return []
        p_row = cur_row - 1
        if start_row - p_row > 3:
            return []
        temp_row, temp_col, val = self._get_cell_value(sheet, p_row, col)
        # if val:
        p = self._get_parent_model(sheet, start_row, temp_row, temp_col)
        return p + [val]
        # else:
        #     return []

    def check_titles(self, sheet, start_row=RECORD_START_ROW, from_col=1):
        length = len(ANA_TITLES)
        to_col = from_col + length - 1
        tiles_list = self.get_row(sheet, start_row, from_col, to_col)
        check_result = True
        for i in range(0, len(ANA_TITLES)):
            if tiles_list[i] != ANA_TITLES[i]:
                # check_result = False
                check_result = True
                break
        return check_result

    def new_store(self, ana_data_list, user_id, update_time, role, check_list):
        ana_import_record = dict()
        ana_import_record['result'] = 0
        self._pg.connect()
        commit_list = []
        from Source.spec2db_server.arl.analysis_service import AnalysisRecord
        from Source.spec2db_server.arl.basic_ana import BasicAnaRecord
        from Source.spec2db_server.arl.arl_group import ArlGroup
        basic_ana_obj = BasicAnaRecord()
        ana_obj = AnalysisRecord()
        group_obj = ArlGroup()
        sqlcmd = group_obj.get_group_id(user_id)
        self._pg.execute(sqlcmd)
        row = self._pg.fetchone()
        group_id = None
        if row:
            group_id = row[0]
        # data_list = group_obj.get_one_member(self._pg, user_id)
        for ana_data in ana_data_list:
            definition_id = ana_data["definition_id"].get("datavalue")
            if not definition_id:
                continue
            # if not self._import_power(self._pg, user_id, arl_id, data_list): # 导入权限验证
            #     continue
            try:
                if re.match("^[B|C|D].*", definition_id):
                    new_data = self.convert_basic_data(ana_data)
                    if new_data.get("definition_id") == 'B.4.2.3.1':
                        print new_data.get("unique_id")
                        pass
                    new_data["analysis_id"] = '.'.join([new_data.get("definition_id"), str(new_data.get("unique_id"))])
                    # 检查时序是否正确
                    # seq_diagram = new_data.get('seq_diagram')
                    # asta_filename = new_data.get('asta_filename')
                    # seq_diagram = seq_diagram.replace('\n', '')
                    # asta_filename = asta_filename.replace('\n', '')
                    # check_flag = basic_ana_obj.check_seq_diagram(self._pg, seq_diagram, asta_filename)
                    # if not check_flag:
                    #     ana_import_record['result'] = 1
                    #     ana_import_record["error_list"] = definition_id + " 的“シーケンス図”不在对应的Astah文件中！"
                    #     return ana_import_record
                    if new_data.get("group_name"):
                        new_data.pop("group_name")
                    new_data["group_id"] = group_id
                    new_data["user_id"] = user_id
                    md5_check, lock_check, curr_commit_list = basic_ana_obj.add(self._pg, new_data, self._col_list,
                                                                                update_time, role)
                else:
                    new_data = self.convert_data(ana_data)
                    new_data["analysis_id"] = '.'.join([new_data.get("definition_id"), str(new_data.get("unique_id"))])
                    # 检查时序是否正确
                    # seq_diagram = new_data.get('seq_diagram')
                    # asta_filename = new_data.get('asta_filename')
                    # seq_diagram = seq_diagram.replace('\n', '')
                    # asta_filename = asta_filename.replace('\n', '')
                    # check_flag = basic_ana_obj.check_seq_diagram(self._pg, seq_diagram, asta_filename)
                    # if not check_flag:
                    #     ana_import_record['result'] = 1
                    #     ana_import_record["error_list"] = definition_id + " 的“シーケンス図”或“astaファイル名”不正确！"
                    #     return ana_import_record

                    md5_check, lock_check, curr_commit_list = ana_obj.add(self._pg, new_data, self._col_list,
                                                                          update_time, role)
                if not md5_check:
                    ana_import_record['result'] = 1
                    if not new_data.get("md5_key"):
                        ana_import_record["error_list"] = definition_id + " MD5 不存在！"
                    else:
                        ana_import_record["error_list"] = definition_id + " MD5 不正确！"
                    return ana_import_record
                elif not lock_check:
                    ana_import_record['result'] = 1
                    ana_import_record["error_list"] = definition_id + " 不能被修改！"
                    return ana_import_record
                if curr_commit_list:
                    commit_list += curr_commit_list
            except Exception as e:
                ana_import_record['result'] = 1
                ana_import_record["error_list"] = e.message
                return ana_import_record
        if commit_list:
            author_commit_list, charger_commit_list = self.split_commit_list(commit_list)
            author_log = {'user_id': user_id, "commit_list": author_commit_list, "group_id": group_id}
            charge_log = {'user_id': user_id, "commit_list": charger_commit_list, "group_id": group_id}
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            if author_log:
                flag, commit_id = commit_log.add_commit_log2(self._pg, author_log)
                if flag:
                    self.import_check_list(self._pg, check_list, 'analysis', commit_id, 'author')
                else:
                    print 'log_commit errol!'
            if charge_log:
                flag, commit_id = commit_log.add_commit_log2(self._pg, charge_log)
                if flag:
                    self.import_check_list(self._pg, check_list, 'analysis', commit_id, 'charge')
            print '提交'
            self._pg.commit()
            self._pg.close()
        else:
            self._pg.close()
            print '不用更新'
        return ana_import_record


class AnalysisModel(ImportBase):
    @staticmethod
    def instance():
        """create a instance"""
        if AnalysisModel.__instance is None:
            AnalysisModel.__instance = AnalysisModel()
        return AnalysisModel.__instance

    def __init__(self):
        ImportBase.__init__(self)
        self._attr_dict = dict()
        self._pg.connect()

    def get_model_id_dict(self, modelList):
        for model_dict in modelList:
            for model in model_dict.iterkeys():
                model_id = self._attr_dict.get(model)
                if not model_id:
                    model_id = self._get_model_id(model)
                    self._attr_dict[model] = model_id
        return self._attr_dict

    def get_model_id(self, model):
        model_id = self._attr_dict.get(model)
        if not model_id:
            model_id = self._get_model_id(model)
            self._attr_dict[model] = model_id
        return model_id

    def _get_model_id(self, model):
        sqlcmd = """
        SELECT model_id
          FROM spec.analysis_model_type
          WHERE model = %s;
        """
        self._pg.execute(sqlcmd, (model,))
        row = self._pg.fetchone()
        if row and row[0]:
            return row[0]
        sqlcmd = """
        INSERT INTO spec.analysis_model_type(model)
            VALUES (%s) RETURNING model_id;
        """
        self._pg.execute(sqlcmd, (model,))
        model_id = self.fetch_id()
        self._pg.commit()
        return model_id


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    def_obj = AnalysisSpec()
    def_obj.set_file(r'C:\Users\yuyin\Desktop\import\wsl\TAGL_RequirementAnalysis_uifw-16.xlsx')
    book = load_workbook(def_obj.xls_file_name, data_only=True)
    sheet = book.get_sheet_by_name(r'TAGL要件分析')
    update_time = time.strftime("%Y-%m-%d %H:%M:%S")
    user_id = 446
    def_obj.store(sheet, user_id, update_time)



