# -*- coding: UTF-8 -*-
import time
from openpyxl import load_workbook
import re
from Source.spec2db_server.spec_import.arl import ArlSpec
from Source.spec2db_server.spec_import.hu import SpecHU
from Source.spec2db_server.spec_import.definition import DefinitionSpec
from Source.spec2db_server.spec_import.analysis import AnalysisSpec
from Source.spec2db_server.spec_import.block import Block
from Source.spec2db_server.spec_import.whitelist import WhiteList
from Source.spec2db_server.spec_import.schedule import Schedule
from lgatio import HUDocType, ReqDocType012 as ReqDocType, AnaDocType, AnaDocTypeWithLongID
from lgatio.data.dynamic import APIDynamicData
from Source.spec2db_server.arl.arl_base import ServiceBase
from Source.spec2db_server.docs.doc_xls import DocXls


class ArlImport(ServiceBase):
    def __init__(self, file_list, user_id, update_time):
        ServiceBase.__init__(self)
        self.file_list = file_list
        self.user_id = user_id
        self.update_time = update_time

    def get_role(self):
        self._pg.connect()
        role = "common"
        from Source.spec2db_server.arl.arl_group import ArlGroup
        group_obj = ArlGroup()
        data_list = group_obj.get_one_member(self._pg, self.user_id)
        for data in data_list:
            roles = data.get("roles")
            if 3 in roles:
                role = "Admin"
            elif 7 in roles:
                role = "PL"
            elif 4 in roles:
                role = "Leader"
            elif 5 in roles:
                role = "Member"
        self._pg.close()
        return role

    def point_import(self, file_path, new_date):
        docXls_obj = DocXls('POINT_OUT', '0.40')
        import_record = dict()
        try:
            point_data_list = docXls_obj.load_data(file_path)
        except Exception as e:
            import_record['result'] = 1
            import_record['error_list'] = '上传的文件不正确'
            return import_record
        from Source.spec2db_server.spec_import.pointout import SpecPointout
        obj = SpecPointout()
        import_record = obj.store(point_data_list, new_date)
        return import_record

    def first_point_import(self, file_path, classify):
        docXls_obj = DocXls(classify, '0.40')
        import_record = dict()
        from pointout import SpecPointout
        point_obj = SpecPointout()
        self.update_time = self.get_current_time()
        if classify == "HU_DEF":
            try:
                data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
        if classify == "TAGL_DEF":
            try:
                data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
        if classify == "TAGL_ANA":
            try:
                data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
        try:
            point_obj.point_store(data_list, self.update_time, classify)
            import_record['result'] = 0
        except:
            import_record['result'] = 1
        return import_record

    def import_point_out_by_user(self, file_path, classify):
        docXls_obj = DocXls(classify, '0.40')
        import_record = dict()
        from pointout import SpecPointout
        point_obj = SpecPointout()
        self.update_time = self.get_current_time()
        if classify == "HU_DEF":
            try:
                data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
        if classify == "TAGL_DEF":
            try:
                data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
        if classify == "TAGL_ANA":
            try:
                data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
        try:
            point_obj.point_store(data_list, update_time, classify)
            import_record['result'] = 0
        except:
            import_record['result'] = 1
        return import_record

    def sheet_import(self, file_path, classify, check_list):
        docXls_obj = DocXls(classify, '0.40')
        import_record = dict()
        role = self.get_role()
        import Source.spec2db_server.arl.arl_server
        spec_obj = Source.spec2db_server.arl.arl_server.ArlSpec()
        dyndata = spec_obj.get_block_white_list()
        api_dyndata = APIDynamicData(dyndata)
        if classify == "HU_DEF":
            try:
                hu_data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
            try:
                obj = HUDocType(api_dyndata)
                index, records, new_e = obj.excel2rows(file_path, r'HU要件定義書', continue_on_error=False)
            except Exception as e:
                if e.message.find('No sheet named') >= 0:
                    new_e = 'No sheet named <'+'HU要件定義書'+'>'
                else:
                    try:
                        new_e = str(e.collection[0].coord)
                    except Exception:
                        new_e = str(e)
            if new_e is None:
                print 'hu导入。。。'
                import_record = SpecHU().new_store(hu_data_list, self.user_id, self.update_time, role, check_list)
            else:
                import_record['result'] = 1
                import_record['error_list'] = new_e
        if classify == "TAGL_DEF":
            try:
                def_data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
            try:
                obj = ReqDocType(api_dyndata)
                index, records, new_e = obj.excel2rows(file_path, r'TAGL要件定義', continue_on_error=False)
            except Exception as e:
                if e.message.find('No sheet named') >= 0:
                    new_e = 'No sheet named <'+'TAGL要件定義'+'>'
                else:
                    try:
                        new_e = str(e.collection[0].coord)
                    except Exception:
                        new_e = str(e)
            if new_e:
                import_record['result'] = 1
                import_record['error_list'] = new_e
                return import_record
            new_e = self.check_basic_def_trans_info(def_data_list)
            if new_e:
                import_record['result'] = 1
                import_record['error_list'] = new_e
                return import_record
            print 'Definition导入。。。'
            import_record = DefinitionSpec().new_store(def_data_list, self.user_id, self.update_time, role, check_list)
        if classify == "TAGL_ANA":
            try:
                ana_data_list = docXls_obj.load_data(file_path)
            except Exception as e:
                print e
                import_record['result'] = 1
                import_record['error_list'] = '上传的文件不正确, %s' % str(e)
                return import_record
            try:
                obj = AnaDocTypeWithLongID(api_dyndata)
                index, records, new_e = obj.excel2rows(file_path, r'TAGL要件分析', continue_on_error=False)
            except Exception as e:
                if e.message.find('No sheet named')>=0:
                    new_e = 'No sheet named <'+'TAGL要件分析'+'>'
                else:
                    try:
                        new_e = str(e.collection[0].coord)
                    except Exception:
                        new_e = str(e)
            if new_e:
                import_record['result'] = 1
                import_record['error_list'] = new_e
                return import_record
            new_e = self.check_basic_ana_trans_info(ana_data_list)
            if new_e:
                import_record['result'] = 1
                import_record['error_list'] = new_e
                return import_record
            print 'Analysis导入。。。'
            import_record = AnalysisSpec().new_store(ana_data_list, self.user_id, self.update_time, role, check_list)
        return import_record

    def check_basic_def_trans_info(self, def_data_list):
        self._pg.connect()
        for def_data_dict in def_data_list:
            def_id = def_data_dict['definition_id']['datavalue']
            if not re.match("^B.*", def_id):
                continue
            trans_sql = """
                SELECT * FROM spec.basic_req_hu where hu_id=%s and dcu_status = %s
                    and dcu_trigger = %s and dcu_action=%s and meu_status = %s
                    and meu_trigger = %s and meu_action=%s and remark=%s
            """
            self._pg.execute(trans_sql, [def_data_dict['hu_def_id']['datavalue'],
                                         def_data_dict['dcu_status']['datavalue'],
                                         def_data_dict['dcu_trigger']['datavalue'],
                                         def_data_dict['dcu_action']['datavalue'],
                                         def_data_dict['meu_status']['datavalue'],
                                         def_data_dict['meu_trigger']['datavalue'],
                                         def_data_dict['meu_action']['datavalue'],
                                         def_data_dict['hu_remark']['datavalue']])
            data_list = self._pg.fetchall()
            if len(data_list) == 0:
                self._pg.close()
                return "Definition %s transfer content is error!"%def_id

        self._pg.close()
        return None

    def check_basic_ana_trans_info(self, ana_data_list):
        """检查要件分析转记内容"""
        self._pg.connect()
        for ana_data_dict in ana_data_list:
            def_id = ana_data_dict['definition_id']['datavalue']
            if not re.match("^[B|C].*", def_id):
                continue
            trans_sql = """
                SELECT * FROM spec.basic_req_definition where definition_id=%s and dcu_meu = %s 
                and pf_status = %s and pf_trigger = %s and pf_action=%s
            """
            self._pg.execute(trans_sql, [ana_data_dict['definition_id']['datavalue'],
                                         ana_data_dict['dcu_meu']['datavalue'],
                                         ana_data_dict['pf_status']['datavalue'],
                                         ana_data_dict['pf_trigger']['datavalue'],
                                         ana_data_dict['pf_action']['datavalue']])
            data_list = self._pg.fetchall()
            if len(data_list) == 0:
                self._pg.close()
                return "要件分析(%s) 转记内容出错(与要件定义不一致)!" % def_id

        self._pg.close()
        return None

    def import_arl(self, file_path, user_id):
        arl_obj = ArlSpec()
        arl_obj.set_file(file_path)
        book = load_workbook(arl_obj.xls_file_name, data_only=True)
        sheet = book.get_sheet_by_name(r'ARL')
        update_time = self.get_current_time()
        result = arl_obj.store(sheet, user_id, update_time)
        return result

    def import_block(self, file_path, classify, user_id, up, down):
        block = Block()
        block.set_file(file_path)
        self.update_time = self.get_current_time()
        result = block.store(classify, user_id, self.update_time, up, down)
        return result

    def import_schedule(self, file_path, user_id):
        schedule = Schedule()
        schedule.set_file(file_path)
        self.update_time = self.get_current_time()
        result = schedule.store(user_id, self.update_time)
        return result

    def import_whitelist(self, file_path, classify, user_id):
        white_list = WhiteList()
        white_list.set_file(file_path)
        self.update_time = self.get_current_time()
        result = white_list.store(classify, user_id, self.update_time)
        return result

    def sheet_classify(self):
        file_sheet_dict = {}
        print '读文件中。。。。'
        for file in self.file_list:
            work_book = load_workbook(file, data_only=True)
            sheet_dict = {}
            hu_sheet_list = []
            tagldef_sheet_list = []
            taglana_sheet_list = []
            for sheet_name in work_book.get_sheet_names():
                work_sheet = work_book.get_sheet_by_name(sheet_name)
                if SpecHU().check_titles(work_sheet):
                    hu_sheet_list.append(work_sheet)
                    continue
                elif DefinitionSpec().check_titles(work_sheet):
                    tagldef_sheet_list.append(work_sheet)
                    continue
                elif AnalysisSpec().check_titles(work_sheet):
                    taglana_sheet_list.append(work_sheet)
                    continue
            if hu_sheet_list:
                sheet_dict['hu'] = hu_sheet_list
            if tagldef_sheet_list:
                sheet_dict['definition'] = tagldef_sheet_list
            if taglana_sheet_list:
                sheet_dict['analysis'] = taglana_sheet_list
            try:
                file_sheet_dict[file.decode("gb2312")] = sheet_dict
            except Exception as e:
                file_sheet_dict[file.decode("utf8")] = sheet_dict
        if file_sheet_dict:
            ret_info = self.import_to_db(file_sheet_dict)
            return ret_info
        else:
            return

    def import_to_db(self, file_sheet_dict):
        ret_info = {}
        for file_name, sheet_dict in file_sheet_dict.items():
            # print filename
            # file_name = filename.decode("utf8")
            result_list = []
            hu_sheet_list = sheet_dict.get("hu")
            if hu_sheet_list:
                for hu_sheet in hu_sheet_list:
                    print 'hu导入。。。：', file_name
                    file_result_dict = SpecHU().store(hu_sheet, self.user_id, self.update_time)
                    result_list.append(file_result_dict)
                ret_info[file_name] = result_list
        for file_name, sheet_dict in file_sheet_dict.items():

            result_list = []
            def_sheet_list = sheet_dict.get("definition")
            if def_sheet_list:
                for def_sheet in def_sheet_list:
                    print 'definition导入。。。：', file_name
                    file_result_dict = DefinitionSpec().store(def_sheet, self.user_id, self.update_time)
                    result_list.append(file_result_dict)
                ret_info[file_name] = result_list
        for file_name, sheet_dict in file_sheet_dict.items():

            result_list = []
            ana_sheet_list = sheet_dict.get("analysis")
            if ana_sheet_list:
                for ana_sheet in ana_sheet_list:
                    print 'analysis导入。。。：', file_name
                    file_result_dict = AnalysisSpec().store(ana_sheet, self.user_id, self.update_time)
                    result_list.append(file_result_dict)
                ret_info[file_name] = result_list
            # for key, sheet_list in sheet_dict.items():
            #     if key == 'hu':
            #         for hu_sheet in sheet_list:
            #             print 'hu导入。。。：', filename
            #             file_result_dict = SpecHU().store(hu_sheet, self.user_id)
            #             if file_result_dict.get('result') > 0:
            #                 break
            #     if key == 'definition':
            #         for def_sheet in sheet_list:
            #             print 'definition导入。。。：', filename
            #             file_result_dict = DefinitionSpec().store(def_sheet, self.user_id)
            #             if file_result_dict.get('result') > 0:
            #                 break
            #     if key == 'analysis':
            #         for ana_sheet in sheet_list:
            #             print 'analysis导入。。。：', filename
            #             file_result_dict = AnalysisSpec().store(ana_sheet, self.user_id)
            #             if file_result_dict.get('result') > 0:
            #                 break


        return ret_info


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    update_time = time.strftime("%Y-%m-%d %H:%M:%S")
    obj = ArlImport([], 346, update_time)
    file_path = r"C:\Users\yuyin\Desktop\TAGL_RequirementAnalysisVer1.10_20180103.xlsx"
    data = obj.sheet_import(file_path, 'TAGL_ANA')
    print data
    # hu_list = [r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_01_PF-LAN_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_10_HMIVehicle Coop_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_20_Navigation_20180103.xlsx",
    #         r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_30_Media_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_40_BT_WIFI_DCM_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_50_Connect_PF_20180103.xlsx",
    #         r"C:\Users\yuyin\Desktop\20180103\hu\HU_RequirementDefinition.Ver1.10_20180103.xlsx"]
    # def_list =[ r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_01_PF-LAN_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_10_HMIVehicle Coop_20180103.xlsx",
    #             r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_20_Navigation_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_30_Media_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_40_BT_WIFI_DCM_20180103.xlsx",
    #             r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_50_Connect_PF_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\def\TAGL_RequirementDefinitionVer1.10_20180103.xlsx"]
    # ana_list =[r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_01_PF-LAN_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_10_HMIVehicle Coop_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_20_Navigation_20180103.xlsx",
    #             r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_30_Media_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_40_BT_WIFI_DCM_20180103.xlsx", r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_50_Connect_PF_20180103.xlsx",
    #             r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_20180103.xlsx"]
    # print len(hu_list)
    # print len(def_list)
    # print len(ana_list)
    # for file_path1 in hu_list:
    # # file_path = r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_20180103.xlsx"
    #     print file_path1
    #     data = obj.first_point_import(file_path1, 'HU_DEF')
    #     print data
    # for file_path2 in def_list:
    # # file_path = r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_20180103.xlsx"
    #     print file_path2
    #     data = obj.first_point_import(file_path2, 'TAGL_DEF')
    #     print data
    # for file_path3 in ana_list:
    # # file_path = r"C:\Users\yuyin\Desktop\20180103\ana\TAGL_RequirementAnalysisVer1.10_20180103.xlsx"
    #     print file_path3
    #     data = obj.first_point_import(file_path3, 'TAGL_ANA')
    #     print data



