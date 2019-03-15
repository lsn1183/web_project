# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import time
from openpyxl import load_workbook
from hu import SpecHU
from Source.spec2db_server.arl.arl_schedule import ArlSchedule


class Schedule(SpecHU):
    """
    """

    def __init__(self):
        SpecHU.__init__(self)
        self.file_name = ''
        self._col_list = ["arl_id", "major_category", "medium_catetory",
                          "small_category", "detail", "req_post",
                          "change_type", "group", "author",
                          "charger", "hu_date", "hu_remark",
                          "def_date", "def_remark", "analysis_date",
                          "analysis_remark"]

    def set_file(self, file_name):
        self.file_name = file_name

    def store(self, user_id, update_time):
        result = {'result': 0}
        try:
            book = load_workbook(self.file_name, data_only=True)
            sheet = book.get_sheet_by_name(r'Sheet')
        except Exception as e:
            result['result'] = 1
            result["error_list"] = [str(e)]
            return result
        start_row = 2
        max_row = sheet.max_row
        error = None
        self._pg.connect()
        update_count = 0
        try:
            update_count = self._store(sheet, start_row, max_row, user_id, update_time)
            if update_count:
                print 'commit'
            self._pg.commit()
        except Exception as e:
            error = str(e)
            self._pg.conn.rollback()
        finally:
            self._pg.close()
        if error:
            result['result'] = 1
            result["error_list"] = [error]
            result["update_count"] = 0
        else:
            result["update_count"] = update_count
        return result

    def _store(self, sheet, start_row, max_row, user_id, update_time):
        update_count = 0
        sch_obj = ArlSchedule()
        for row in range(start_row, max_row + 1):
            values = self.get_row(sheet, row, from_col=1, to_col=16)
            _id = values[0]
            if not _id:
                break
            values[10] = self.convert_time(values[10])  # HU DATE
            values[12] = self.convert_time(values[12])  # DEF DATE
            values[14] = self.convert_time(values[14])  # ANA DATE
            data = self._convert_info(values, model_vals=[])
            if "model_list" in data:
                data.pop("model_list")
            data["update_user_id"] = user_id
            curr_commit_list = sch_obj.add(self._pg, data, None, update_time)
            if curr_commit_list:
                update_count += 1
        if not update_count:
            print 'no change.'
        else:
            sqlcmd = """
            -------------------------------------------------------
            update spec.arl a set user_id = tt1.user_id
              from (
                SELECT arl_record_id, t1.arl_id, t3.user_id, user_name
                  FROM spec.arl as t1
                  inner join spec.arl_schedule as t2
                  on t1.arl_id = t2.arl_id
                  left join spec.arl_user as t3
                  on t2.author = t3.user_name
                  where (t1.user_id is null and t3.user_id is not null) or (t1.user_id <> t3.user_id)
              ) as tt1
              where a.arl_record_id = tt1.arl_record_id;
              
            -------------------------------------------------------
            update spec.arl a set group_id = tt1.group_id
              from (
                SELECT arl_record_id, t1.arl_id, t3.group_id, group_name
                  FROM spec.arl as t1
                  inner join spec.arl_schedule as t2
                  on t1.arl_id = t2.arl_id
                  left join spec.arl_group as t3
                  on t2."group" = t3.group_name
                  where (t1.group_id is null and t3.group_id is not null) or (t1.group_id <> t3.group_id)
              ) as tt1
              where a.arl_record_id = tt1.arl_record_id;
            """
            self._pg.execute(sqlcmd)
        return update_count

    def convert_time(self, long_time):
        date = SpecHU.convert_time(self, long_time)
        if date.find('-') > 0:
            date = time.strptime(date, "%Y-%m-%d")
            date = time.strftime('%Y/%m/%d', date)
        return date


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    files = [
        r'C:\Users\pset\Downloads\ARL_Schedule_All_20171214163030.xlsx',
    ]
    for file in files:
        print file
        obj = Schedule()
        user_id = 448
        update_time = obj.get_current_time()
        obj.set_file(file)
        result = obj.store(user_id, update_time)
        print result


