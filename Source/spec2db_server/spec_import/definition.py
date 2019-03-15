# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import os
import re
from openpyxl import load_workbook
import time
from hu import SpecHU
from import_base import ImportBase

PROJ_NAME = 'proj_name'
SPEC_NAME = 'spec_name'
SPEC_VERSION = 'version'
SPEC_FILE_NAME = 'file_name'
RECORD_START_ROW = 5  # 记录开始行号
DEF_TITLES = [u'担当', u'H/U要件定義ID', u'TAGL要件定義ID', u'ユニークID', u'大分類', u'中分類', u'小分類', u'詳細',u'基本要件', u'関連基本要件', u'除外', u'状態', u'トリガー', u'動作', u'状態',
             u'トリガー', u'動作', u'備考(HU要件定義)', u'DCU/MEUどちらの想定か', u'状態', u'トリガー', u'動作', u'電源管理デバイス',
             u'AVCLAN制御デバイス', u'MOST制御デバイス', u'CAN制御デバイス', u'ジカ線入力デバイス', u'不揮発メモリデバイス',
             u'音声出力デバイス', u'音声入力セレクタデバイス', u'BTデバイス', u'WIFIデバイス', u'タッチパネル制御デバイス', u'タッチパッド制御デバイス',
             u'ロータリースイッチ、キースイッチ検出デバイス', u'GPSデバイス', u'Radioチューナーデバイス（XM、DAB含む）', u'地デジチューナーデバイス', u'セキュリティチップデバイス',
             u'マイク用ADCデバイス', u'DVD映像デコーダデバイス', u'映像信号入力セレクタデバイス', u'映像信号出力デバイス(DCU→各ディスプレイ/ MEU→DCU)',
             u'CD/DVD/BD制御デバイス', u'ECNCデバイス', u'USB2.0-Host/Functionデバイス', u'APPLE認証デバイス', u'RTCデバイス', u'USBロールスイッチ', u'温度検知デバイス', None, None,
             u'参考ウォークスルー図', u'参考HAL設計書', u'その他仕様（リファハード仕様等）', u'リファレンスハード上での実現可否', u'詳細分析可否',
             u'未要件分析', u'日付', u'変更理由']


class DefinitionSpec(SpecHU):
    """
    """
    def __init__(self):
        SpecHU.__init__(self)
        self.xls_file_name = None
        self._col_list = ["author_name", "hu_def_id", "definition_id",
                          "unique_id", "basic_req", "rel_requirement", "exception", "dcu_meu",
                          "pf_status", "pf_trigger", "pf_action",
                          "remark", "notice", "rel_flow_diagram",
                          "rel_hal_design", "other_spec", "implementation",
                          "analysis", "unrequire", "new_date", "reason",
                          ]
        self._from = ["major_category", "medium_category", "small_category", "detail",
                      "dcu_status", "dcu_trigger", "dcu_action", "meu_status", "meu_trigger", "meu_action",
                      "hu_remark", "group_id", 'lock_status'
                      ]
        self.covert_dict = {'def_remark': 'remark',
                            'def_md5': 'md5_key'
                            }

    def set_file(self, xls_file_name):
        self.xls_file_name = xls_file_name

    def store(self, sheet, user_id, update_time):
        self._pg.connect()
        max_row = sheet.max_row
        start_row = RECORD_START_ROW + 1
        def_import_record = {}
        def_import_record['result'] = 0
        from Source.spec2db_server.spec_import.import_check import DEFImportCheck
        def_check = DEFImportCheck()
        # error_list = def_check.def_check_row(self._pg, sheet, start_row, max_row)
        # print error_list
        ################################################
        # ## Model List
        ################################################
        # def_models = self.get_models(sheet, row=start_row - 1, from_col=23, to_col=51)
        # obj_model = DefModel()
        # model_list = []
        # for models in def_models:
        #     # model = json.dumps(models)
        #     model_id = obj_model.get_model_id(models)
        #     model_list.append(model_id)
        # return 0
        error_list = []
        if error_list:
            def_import_record['result'] = 1
            def_import_record["error_list"] = error_list
            return def_import_record
        else:
            exception = self._store(sheet, start_row, max_row, user_id, update_time)
            if exception:
                def_import_record['result'] = 1
                def_import_record["error_list"] = exception
                return def_import_record
            return def_import_record

    def _store(self, sheet, start_row, max_row, user_id, update_time):
        from Source.spec2db_server.arl.def_server import DefRecord
        def_obj = DefRecord()
        commit_list = []
        for row in range(start_row, max_row + 1):
            # print row
            values1 = self.get_row(sheet, row, from_col=1, to_col=4)  # 4个
            values5 = self.get_row(sheet, row, from_col=9, to_col=10)  # 基本要件——关连基本要件
            values2 = self.get_row(sheet, row, from_col=11, to_col=11)  # 1个
            values3 = self.get_row(sheet, row, from_col=19, to_col=22)  # 4个
            values4 = self.get_row(sheet, row, from_col=52, to_col=63)  # 12个, 63:Md5
            hu_id = values1[1]
            if not hu_id:
                break
            definition_id, unique_id = values1[2],  values1[3]
            unique_id = int(unique_id)
            values1[3] = unique_id
            if not definition_id:
                definition_id = '.'.join([hu_id, str(unique_id)])
                values1[2] = definition_id
            if definition_id[0] in ('A', 'B', 'C', 'D'):
                continue
            # if type(unique_id) in (str, unicode):
            #     unique_id = int((definition_id.split('.'))[-1])
            #     values1[3] = unique_id
            new_date = values4[-2]  # 日付
            if new_date:
                values4[-2] = self.convert_time(new_date)
            params = values1 + values5 + values2 + values3 + values4
            model_vals = self.get_row(sheet, row, 23, 51)
            new_data = self._convert_info(params, model_vals)
            new_data["md5_key"] = values4[-1]
            try:
                md5_check, lock_check, curr_commit_list = def_obj.add(self._pg, new_data, self._col_list, update_time)
                if not md5_check:
                    return definition_id + " MD5 不存在！"
                elif not lock_check:
                    return definition_id + " 不能被修改！"
                elif curr_commit_list:
                    commit_list += curr_commit_list
            except Exception as e:
                print e
                return e.message

        log = {'user_id': user_id, "commit_list": commit_list}
        if commit_list:
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            flag, commit_id = commit_log.add_commit_log2(self._pg, log)
            if flag:
                print '提交'
                self._pg.commit()
            else:
                print 'log_commit errol!'
        else:
            print '不用更新'

    def check_titles(self, sheet, start_row=RECORD_START_ROW, from_col=1):
        length = len(DEF_TITLES)
        to_col = from_col + length - 1
        tiles_list = self.get_row(sheet, start_row, from_col, to_col)
        check_result = True
        for i in range(0, len(DEF_TITLES)):
            if tiles_list[i] != DEF_TITLES[i]:
                check_result = False
                break

        return check_result

    def _parser_spec_info(self, path):
        spec_info = {
                     # PROJ_NAME: 'TAGL',
                     # SPEC_NAME: 'RequirementDefinition',
                     # SPEC_VERSION: '0.09',
                     # SPEC_FILE_NAME: 'TAGL_RequirementDefinitionVer0.09.xlsx',
                     }
        base_name = os.path.basename(path)
        name, file_ext = os.path.splitext(base_name)
        r = r'.+_+'
        m = re.match(r, name)
        if m:
            proj_name = m.group(0)
            if proj_name:
                name = name[m.end():]
                spec_info[PROJ_NAME] = proj_name.rstrip('_')
        r = r'\.*[V|v]er\d+\.*\d*'
        s = re.search(r, name)
        if s:
            ver = s.group(0)
            if ver:
                name = name[:s.start()]
                ver = ver.lstrip(r'.')
                spec_info[SPEC_VERSION] = ver[3:]
        spec_info[SPEC_NAME] = name
        spec_info[SPEC_FILE_NAME] = base_name
        return spec_info

    def _store_spec_info(self, spec_info):
        sqlcmd = """
        INSERT INTO spec.requirement_spec(proj, spec_name, ver, file_name)
          VALUES(%s, %s, %s, %s) RETURNING requiremnt_spec_id;
        """
        self._pg.execute(sqlcmd, (spec_info.get(PROJ_NAME),
                                   spec_info.get(SPEC_NAME),
                                   spec_info.get(SPEC_VERSION),
                                   spec_info.get(SPEC_FILE_NAME)
                                   )
                          )
        self.set_id(self.fetch_id())
        self._pg.commit()

    def _insert_def_record(self, params):
        sqlcmd = """
        INSERT INTO spec.definition(
            author_name, hu_def_id, definition_id,
            unique_id, exception, dcu_meu,
            pf_status, pf_trigger, pf_action,
            remark,  notice, rel_hal_design,
            rel_flow_diagram, other_spec, implementation, 
            analysis, unrequire, new_date,
            reason)
            VALUES (%s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s)
            RETURNING def_rc_id;
        """
        self._pg.execute(sqlcmd, params)
        return self.fetch_id()

    def _insert_model_rel(self, def_rc_id, model_id, val):
        sqlcmd = """
        INSERT INTO spec.definition_model_rel(
                    def_rc_id, model_id, val)
            VALUES (%s, %s, %s);
        """
        self._pg.execute(sqlcmd, (def_rc_id, model_id, val))

    def get_row(self, sheet, row=1, from_col=1, to_col=1):
        values = []
        j = from_col
        while j <= to_col:
            val = sheet.cell(row=row, column=j).value
            if type(val) == unicode:
                val = val.replace(u'_x000D_', u'')
                val = val.replace(u'_x0000_', u'')
            values.append(val)
            j += 1
        return values

    def new_store(self, def_data_list, user_id, update_time, role, check_list):
        def_import_record = dict()
        def_import_record['result'] = 0
        self._pg.connect()
        commit_list = []
        from Source.spec2db_server.arl.def_server import DefRecord
        from Source.spec2db_server.arl.basic_def import BasicDefRecord
        from Source.spec2db_server.arl.arl_group import ArlGroup
        from Source.spec2db_server.arl.basic_ana import BasicAnaRecord
        from Source.spec2db_server.spec_import.analysis import AnalysisSpec
        def_obj = DefRecord()
        basic_def_obj = BasicDefRecord()
        basic_ana_obj = BasicAnaRecord()
        group_obj = ArlGroup()
        sqlcmd = group_obj.get_group_id(user_id)
        self._pg.execute(sqlcmd)
        row = self._pg.fetchone()
        group_id = None
        if row:
            group_id = row[0]
        # data_list = group_obj.get_one_member(self._pg, user_id)
        for def_data in def_data_list:
            definition_id = def_data["definition_id"].get("datavalue")
            if not definition_id:
                continue
            basic_flag = False
            # if not self._import_power(self._pg, user_id, arl_id, data_list): # 导入权限验证
            #     continue
            try:
                if re.match("^[B|C|D].*", definition_id):
                    new_data = self.convert_basic_data(def_data)
                    if new_data.get("group_name"):
                        new_data.pop("group_name")
                    new_data["group_id"] = group_id
                    new_data["user_id"] = user_id
                    hu_id = new_data.get("hu_def_id")
                    if not hu_id:
                        continue
                    basic_flag = True
                    md5_check, lock_check, curr_commit_list = basic_def_obj.add(self._pg, new_data, self._col_list,
                                                                                update_time, role)
                else:
                    new_data = self.convert_data(def_data)
                    hu_id = new_data.get("hu_def_id")
                    if not hu_id:
                        continue
                    md5_check, lock_check, curr_commit_list = def_obj.add(self._pg, new_data, self._col_list,
                                                                          update_time, role)
                if not md5_check:
                    def_import_record['result'] = 1
                    if not new_data.get("md5_key"):
                        def_import_record["error_list"] = definition_id + " MD5 不存在！"
                    else:
                        def_import_record["error_list"] = definition_id + " MD5 不正确！"
                    return def_import_record
                elif not lock_check:
                    def_import_record['result'] = 1
                    def_import_record["error_list"] = definition_id + " 不能被修改！"
                    return def_import_record
                if curr_commit_list:
                    if basic_flag:
                        rel_ana_list = basic_ana_obj._get_by_parent_id(self._pg, definition_id)
                        for ana_data in rel_ana_list:
                            ana_data['dcu_meu'] = new_data['dcu_meu']
                            ana_data['pf_status'] = new_data['pf_status']
                            ana_data['pf_trigger'] = new_data['pf_trigger']
                            ana_data['pf_action'] = new_data['pf_action']
                            sub_result = basic_ana_obj.add(self._pg, ana_data, AnalysisSpec()._col_list,
                                                           update_time, "Admin", excel_import=False)
                            md5_check, lock_check, sub_commit_list = sub_result
                            if sub_commit_list:
                                curr_commit_list += sub_commit_list
                    commit_list += curr_commit_list
            except Exception as e:
                def_import_record['result'] = 1
                def_import_record["error_list"] = e.message
                return def_import_record
        if commit_list:
            author_commit_list, charger_commit_list = self.split_commit_list(commit_list)
            author_log = {'user_id': user_id, "commit_list": author_commit_list, "group_id": group_id}
            charge_log = {'user_id': user_id, "commit_list": charger_commit_list, "group_id": group_id}
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            if author_log:
                flag, commit_id = commit_log.add_commit_log2(self._pg, author_log)
                if flag:
                    self.import_check_list(self._pg, check_list, 'definition', commit_id, 'author')
                else:
                    print 'log_commit errol!'
            if charge_log:
                flag, commit_id = commit_log.add_commit_log2(self._pg, charge_log)
                if flag:
                    self.import_check_list(self._pg, check_list, 'definition', commit_id, 'charge')
            print '提交'
            self._pg.commit()
            self._pg.close()
        else:
            self._pg.close()
            print '不用更新'
        return def_import_record

class DefModel(ImportBase):
    # __instance = None

    # @staticmethod
    # def instance():
    #     """create a instance"""
    #     if AnalysisModel.__instance is None:
    #         AnalysisModel.__instance = AnalysisModel()
    #     return AnalysisModel.__instance

    def __init__(self):
        ImportBase.__init__(self)
        self._attr_dict = dict()

    def get_model_id_dict(self, modelList):
        for model_dict in modelList:
            for model in model_dict.iterkeys():
                model_id = self._attr_dict.get(model)
                if not model_id:
                    model_id = self._get_model_id(model)
                    self._attr_dict[model] = model_id
        return self._attr_dict

    def get_model_id(self, model):
        model_id = self._attr_dict.get(model)
        if not model_id:
            model_id = self._get_model_id(model)
            self._attr_dict[model] = model_id
        return model_id

    def _get_model_id(self, model):
        sqlcmd = """
        SELECT model_id
          FROM spec.definition_model_type
          WHERE model = %s;
        """
        self._pg.connect()
        self._pg.execute(sqlcmd, (model,))
        row = self._pg.fetchone()
        if row and row[0]:
            return row[0]
        sqlcmd = """
        INSERT INTO spec.definition_model_type(model)
            VALUES (%s) RETURNING model_id;
        """
        self._pg.execute(sqlcmd, (model,))
        model_id = self.fetch_id()
        self._pg.commit()
        return model_id


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    def_obj = DefinitionSpec()
    def_obj.set_file(r'C:\Users\yuyin\Desktop\import\wsl\TAGL_RequirementDefinition_uifw-16.xlsx')
    book = load_workbook(def_obj.xls_file_name, data_only=True)
    sheet = book.get_sheet_by_name(r'TAGL要件定義')
    update_time = time.strftime("%Y-%m-%d %H:%M:%S")
    user_id = 446
    def_obj.store(sheet, user_id, update_time)
    # sqlcmd = "select hu_id from spec.hu"
    # hu_obj._pg.query_fromcur(sqlcmd)
    # row = hu_obj._pg.fetchone()
    # print row
