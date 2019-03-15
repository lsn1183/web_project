# -*- coding: UTF-8 -*-
"""
Created on 2017-11-06

@author: yuyin
"""
from hu import SpecHU
import time
import os
from openpyxl import load_workbook

RECORD_START_ROW = 5  # 记录开始行号

ARl_TITLES = [u'MMマスターカタスペ\n新連番',
              u'担当グループ', u'MMマスターカタスペ記載項目', u'TAGL対象外', u'カテゴリ', u'ID', u'大分類',u'Level1', u'ID', u'中分類', u'Level2', u'ID', u'小分類', u'Level3',
              u'ID', u'詳細', u'Level4', u'機能・開発項目の概要説明', u'Summary of Function and Development item', u'補足', u'SubID', u'要件ID', u'転記してきた要件',
              u'備考、不明点', u'除外', u'状態', u'トリガー', u'動作', u'エンドユーザー', u'ディーラー', u'開発者', u'サプライヤー', u'社内規定',u'法規（自工会等を含む）', u'過去不具合対応',
              u'目的、背景、ポリシー', u'仕様書番号', u'バージョン', u'ファイル名', u'章', u'Page', u'画面ID', u'仕様書番号', u'バージョン', u'ファイル名',
              u'章', u'Page', None, None, None, u'同件管理', u'T_EMV (ディスプレイ一体)', u'T_EMV (Toyota Touch Separate Display)', u'T1\nMEUなし', u'T1\n＋MEU Low', u'T1\n＋MEU High', u'L2', u'T0', u'T1', u'T2'
              u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L2', u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L1.5', u'L2', u'T1', u'T2'
              u'L1', u'L1.5', u'L2', u'T1', u'L1', u'L2', u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L2', u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)',
              u'L1', u'L2', u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L2', u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L2', u'T1', u'T2', u'L1', u'L2',
              u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L2', u'T1', u'T2', u'T_EMV(Toyota Touch Separate Display)', u'L1', u'L2', u'T1', u'T2',
              u'T_EMV(Toyota Touch Separate Display)', u'L2', None, u'担当', u'未要件分析', u'要件漏れ', u'ARL指摘No', u'变更类别']


class ArlSpec(SpecHU):

    def __init__(self):
        SpecHU.__init__(self)
        self.xls_file_name = None
        self._col_list = ["mm_new_num", "charge",
                          "mm_item", "tagl_exclude", "category", "id1", "major_category", "level1",
                          "id2", "medium_catetory", "level2", "id3", "small_category", "level3",
                          "id4", "detail", "level4", "func_summary_jp", "func_summary_en", 'supply', 'subid', 'arl_id',
                          'req_post', 'remark', 'exception', 'status', 'trigger', 'action', 'arl_user', 'dealer', 'developer', 'supplier', 'company_rule', 'law',
                          "old_bug", "policy", "hmi_spec_no", "hmi_version", "hmi_file_name", "hmi_chapter", "hmi_page", "screen_id", "func_spec_no", "func_version", "func_file_name",
                          "func_chapter", "func_page", "if_spec", "center_spec", "other_spec", "same_req", "sys_conf_id", "author", "future_req", "req_omission", "censure", "job_status",
                          "major_ver", "md5_key", "small_ver"]

    def set_file(self, xls_file_name):
        self.xls_file_name = xls_file_name

    def store(self, sheet, user_id, update_time):
        self._pg.connect()
        max_row = sheet.max_row
        start_row = RECORD_START_ROW + 1
        import_record = {"result": 0, "error_list": []}
        error = self._store(sheet, start_row, max_row, user_id, update_time)
        if error:
            import_record['result'] = 1
            import_record["error_list"] = [error]
        return import_record

    def _store(self, sheet, start_row, max_row, user_id, update_time):
        from Source.spec2db_server.arl.arl_server import ArlRecord
        def_obj = ArlRecord()
        commit_list = []
        for row in range(start_row, max_row + 1):
            values1 = self.get_row(sheet, row, from_col=1, to_col=51)  # 51个
            values2 = self.get_row(sheet, row, from_col=116, to_col=120)  # 5个
            arl_id = values1[21]
            if not values1[6]:
                break
            if not arl_id:
                continue
            # if self.check_unkown_code(values1 + values2):
            #     continue
            # print arl_id
            params = values1 + values2
            model_vals = self.get_row(sheet, row, 52, 115)
            new_data = self._convert_info(params, model_vals)
            if new_data.get("mm_item") or new_data.get("exception"):
                new_data["exclude_flag"] = True
            else:
                new_data["exclude_flag"] = False
            curr_commit_list = def_obj.add(self._pg, new_data, self._col_list, update_time)
            if curr_commit_list:
                commit_list += curr_commit_list
        log = {'user_id': user_id, "commit_list": commit_list}
        if commit_list:
            from Source.spec2db_server.arl.commit_log import CommitLog
            commit_log = CommitLog()
            flag, commit_id = commit_log.add_commit_log2(self._pg, log)
            if flag:
                self.update_cat_id(self._pg)
                print '提交'
                self._pg.commit()
            else:
                return 'log_commit errol!'
        else:
            print '不用更新'
        return None

    def check_titles(self, sheet, start_row=RECORD_START_ROW, from_col=1):
        length = len(ARl_TITLES)
        to_col = from_col + length - 1
        tiles_list = self.get_row(sheet, start_row, from_col, to_col)
        check_result = True

        for i in range(0, len(ARl_TITLES)):
            if tiles_list[i] != ARl_TITLES[i]:
                check_result = False
                break

        return check_result

    def excute_sql(self, file_name):
        """执行指定目录下的sql文件"""
        for each in os.listdir("."):

            count = 0  #读取行数
            sql = ""
            if file_name in each:
                with open(each, "r", encoding="utf-8") as f:
                    for each_line in f.readline():
                        # 过滤数据
                        if not each_line or each_line == "\n":
                            continue
                        # 读取1000行数据，拼接成sql
                        elif count < 1000:
                            sql += each_line
                            count += 1
                        # 读取达到1000行数据，进行提交
                        else:
                            self._pg.connect()
                            self._pg.execute(sql)
                            self._pg.commit()
                            self._pg.close()
                            return
                    # 当读取文件完毕，不到1000行时，也需对拼接的sql执行，提交
                    if sql:
                        self._pg.connect()
                        self._pg.execute(sql)
                        self._pg.commit()
                        self._pg.close()
                        return

    def check_unkown_code(self, vals):
        for val in vals:
            if type(val) in (unicode, str):
                if val.find(u'_x000D_') >= 0 or val.find(u'_x0000_') >= 0:
                    return False
        return True

    def update_cat_id(self, pg):
        sqlcmd = """
        INSERT INTO spec.arl_category(category, level_no)
        (
        SELECT distinct A.category, 0 AS level_no
          FROM spec.arl AS A
          LEFT JOIN spec.arl_category AS B
          ON (A.category = B.category or (A.category is null and B.category is null)) and level_no = 0
          where B.cat_id is null
        );
        
        INSERT INTO spec.arl_category(category, level_no)
        (
        SELECT distinct major_category, 1 AS level_no
          FROM spec.arl AS A
          LEFT JOIN spec.arl_category AS B
          ON (A.major_category = B.category or (A.major_category is null and B.category is null)) and level_no = 1
          where B.cat_id is null
        );
        
        INSERT INTO spec.arl_category(category, level_no)
        (
        SELECT distinct medium_catetory, 2 AS level_no
          FROM spec.arl AS A
          LEFT JOIN spec.arl_category AS B
          ON (A.medium_catetory = B.category or (A.medium_catetory is null and B.category is null)) and level_no = 2
          where B.cat_id is null
        );
        
        INSERT INTO spec.arl_category(category, level_no)
        (
        SELECT distinct small_category, 3 AS level_no
          FROM spec.arl AS A
          LEFT JOIN spec.arl_category AS B
          ON (A.small_category = B.category or (A.small_category is null and B.category is null)) and level_no = 3
          where B.cat_id is null
        );
        
        INSERT INTO spec.arl_category(category, level_no)
        (
        SELECT distinct detail, 4 AS level_no
          FROM spec.arl AS A
          LEFT JOIN spec.arl_category AS B
          ON (A.detail = B.category or (A.detail is null and B.category is null)) and level_no = 4
          where B.cat_id is null
          ORDER BY detail
        );
    
        update spec.arl as a set cat_id0 = cat_id
          FROM spec.arl_category as b
          where cat_id0 is null
                and (a.category =b.category or (a.category is null and b.category is null))
                and level_no = 0;
        
        update spec.arl as a set cat_id1 = cat_id
          FROM spec.arl_category as b
          where cat_id1 is null
                and (a.major_category =b.category or (a.major_category is null and b.category is null))
                and level_no = 1;
        
        update spec.arl as a set cat_id2 = cat_id
          FROM spec.arl_category as b
          where cat_id2 is null and
                (a.medium_catetory =b.category or (a.medium_catetory is null and b.category is null))
                and level_no = 2;
        
        update spec.arl as a set cat_id3 = cat_id
          FROM spec.arl_category as b
          where cat_id3 is null
                and (a.small_category =b.category or (a.small_category is null and b.category is null))
                and level_no = 3;
        
        update spec.arl as a set cat_id4 = cat_id
          FROM spec.arl_category as b
          where cat_id4 is null
                and (a.detail =b.category or (a.detail is null and b.category is null))
                and level_no = 4;
        """
        pg.execute(sqlcmd)


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    files = [
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_01_PF-LAN.xlsx',
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_10_HMIVehicle Coop.xlsx',
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_20_Navigation.xlsx',
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_30_Media.xlsx',
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_40_BT_WIFI_DCM.xlsx',
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_50_Connect_PF.xlsx',
        r'C:\SpiderInput\20171219_dup_id_fix\ARL.Ver0.40_other.xlsx',
    ]
    for file in files:
        arl_obj = ArlSpec()
        arl_obj.set_file(file)
        update_time = arl_obj.get_current_time()
        print update_time
        print arl_obj.xls_file_name
        book = load_workbook(arl_obj.xls_file_name, data_only=True)
        sheet = book.get_sheet_by_name(r'ARL')
        user_id = 446
        arl_obj.store(sheet, user_id, update_time)
        print arl_obj.get_current_time()
