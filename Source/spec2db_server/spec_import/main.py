# -*- coding: UTF-8 -*-
'''
Created on 2017-7-18

@author: hcz
'''
import datetime
import time
import sys
import os
from openpyxl import load_workbook
from Source.spec2db_server.spec_import.hu import SpecHU
from Source.spec2db_server.spec_import.definition import DefinitionSpec
from Source.spec2db_server.spec_import.analysis import AnalysisSpec

def load_hu(file_dir):
    if not os.path.isdir(file_dir):
        pass
    for f in sorted(os.listdir(file_dir)):
        path = os.path.join(file_dir, f)
        print path
        hu_obj = SpecHU()
        # continue
        if os.path.isfile(path):
            base_name, ext_name = os.path.splitext(path)
            if ext_name.lower() in ('.xlsx', '.xls'):

                hu_obj.set_file(path)
                book = load_workbook(hu_obj.hu_file_name, data_only=True)
                sheet = book.get_sheet_by_name(r'HU要件定義書')
                update_time = time.strftime("%Y-%m-%d %H:%M:%S")
                hu_obj.store(sheet, user_id=6, update_time=update_time)


def load_def(file_dir):
    """TAGL定义"""
    if not os.path.isdir(file_dir):
        pass
    def_obj = DefinitionSpec()
    for f in sorted(os.listdir(file_dir)):
        path = os.path.join(file_dir, f)
        print path
        # continue
        if os.path.isfile(path):
            base_name, ext_name = os.path.splitext(path)
            if ext_name.lower() in ('.xlsx', '.xls'):
                def_obj.set_file(path)
                book = load_workbook(def_obj.xls_file_name, data_only=True)
                sheet = book.get_sheet_by_name(r'TAGL要件定義')
                update_time = time.strftime("%Y-%m-%d %H:%M:%S")
                def_obj.store(sheet, user_id=6, update_time=update_time)


def load_analysis(file_dir):
    """TAGL定义"""
    if not os.path.isdir(file_dir):
        pass
    analysis_obj = AnalysisSpec()
    for f in sorted(os.listdir(file_dir)):
        path = os.path.join(file_dir, f)
        print path
        if os.path.isfile(path):
            base_name, ext_name = os.path.splitext(path)
            if ext_name.lower() in ('.xlsx', '.xls'):
                analysis_obj.set_file(path)
                book = load_workbook(analysis_obj.xls_file_name, data_only=True)
                sheet = book.get_sheet_by_name(r'TAGL要件分析')
                update_time = time.strftime("%Y-%m-%d %H:%M:%S")
                analysis_obj.store(sheet, user_id=6, update_time=update_time)


def main():
    # ###### HU ###########################
    hu_pathes = [
        r'C:\SpiderInput\2017-11-13\BlushFormat\hu'
    ]
    for hu_dir in hu_pathes:
        load_hu(hu_dir)
    # ###### DEF ###########################
    def_pathes = [
        r'C:\SpiderInput\2017-11-13\BlushFormat\def',
    ]
    for def_dir in def_pathes:
        load_def(def_dir)
    # # ###### Analysis ###########################
    def_pathes = [
        r'C:\SpiderInput\2017-11-13\BlushFormat\analysis',
    ]
    try:
        for def_dir in def_pathes:
            load_analysis(def_dir)
    except:
        pass
#     rdb_log.end()
    print datetime.datetime.now()
    return 0


if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    main()

