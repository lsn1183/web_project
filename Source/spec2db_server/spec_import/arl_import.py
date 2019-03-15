# -*- coding: UTF-8 -*-
import time
from openpyxl import load_workbook
from Source.spec2db_server.spec_import.hu import SpecHU
from Source.spec2db_server.spec_import.definition import DefinitionSpec
from Source.spec2db_server.spec_import.analysis import AnalysisSpec
from lgatio import HUDocType, ReqDocType012 as ReqDocType, AnaDocType


class ArlImport(object):
    def __init__(self, file_list, user_id, update_time):
        self.file_list = file_list
        self.user_id = user_id
        self.update_time = update_time
    #return dict format
    #{
    #  filename:
    #    {
    #        result：0/1/2
    #        error_list:
    #        {
    #           LineNo1:{ColNo:error_code, },
    #        }
    #    }
    #}

    def sheet_import(self, file_path, classify):
        import_record = dict()
        try:
            work_book = load_workbook(file_path, data_only=True)
        except Exception as e:
            import_record['result'] = 1
            import_record['error_list'] = '上传的文件不正确'
            return import_record
        if classify == "HU_DEF":
            obj = HUDocType()
            try:
                index, records, new_e = obj.excel2rows(file_path, r'HU要件定義書', continue_on_error=False)
            except Exception as e:
                if e.message.find('No sheet named') >= 0:
                    new_e = 'No sheet named <'+'HU要件定義書'+'>'
                else:
                    new_e = str(e)
            if new_e is None:
                work_sheet = work_book.get_sheet_by_name(r'HU要件定義書')
                print 'hu导入。。。'
                import_record = SpecHU().store(work_sheet, self.user_id, self.update_time)
                if import_record['result'] == 0:
                    from basic_hu import BasicHU
                    try:
                        obj = BasicHU()
                        obj.import_to_db(file_path, self.user_id, self.update_time)
                    except Exception as e:
                        import_record['result'] = 1
                        import_record['error_list'] = e.message
            else:
                import_record['result'] = 1
                import_record['error_list'] = new_e
        if classify == "TAGL_DEF":
            obj = ReqDocType()
            try:
                index, records, new_e = obj.excel2rows(file_path, r'TAGL要件定義', continue_on_error=False)
            except Exception as e:
                if e.message.find('No sheet named')>=0:
                    new_e = 'No sheet named <'+'TAGL要件定義'+'>'
                else:
                    new_e = str(e)
            if new_e is None:
                print 'Definition导入。。。'
                work_sheet = work_book.get_sheet_by_name(r'TAGL要件定義')
                import_record = DefinitionSpec().store(work_sheet, self.user_id, self.update_time)
                if import_record['result'] == 0:
                    from basic_def import BasicDef
                    try:
                        obj = BasicDef()
                        obj.import_to_db(file_path, self.user_id, self.update_time)
                    except Exception as e:
                        import_record['result'] = 1
                        import_record['error_list'] = e.message
            else:
                import_record['result'] = 1
                import_record['error_list'] = new_e
        if classify == "TAGL_ANA":
            obj = AnaDocType()
            try:
                index, records, new_e = obj.excel2rows(file_path, r'TAGL要件分析', continue_on_error=False)
            except Exception as e:
                if e.message.find('No sheet named')>=0:
                    new_e = 'No sheet named <'+'TAGL要件分析'+'>'
                else:
                    new_e = str(e)
            if new_e is None:
                print 'Analysis导入。。。'
                work_sheet = work_book.get_sheet_by_name(r'TAGL要件分析')
                import_record = AnalysisSpec().store(work_sheet, self.user_id, self.update_time)
                if import_record['result'] == 0:
                    from basic_ana import BasicAna
                    try:
                        obj = BasicAna()
                        obj.import_to_db(file_path, self.user_id, self.update_time)
                    except Exception as e:
                        import_record['result'] = 1
                        import_record['error_list'] = e.message
            else:
                import_record['result'] = 1
                import_record['error_list'] = new_e
        return import_record

    def sheet_classify(self):
        file_sheet_dict = {}
        print '读文件中。。。。'
        for file in self.file_list:
            work_book = load_workbook(file, data_only=True)
            sheet_dict = {}
            hu_sheet_list = []
            tagldef_sheet_list = []
            taglana_sheet_list = []
            for sheet_name in work_book.get_sheet_names():
                work_sheet = work_book.get_sheet_by_name(sheet_name)
                if SpecHU().check_titles(work_sheet):
                    hu_sheet_list.append(work_sheet)
                    continue
                elif DefinitionSpec().check_titles(work_sheet):
                    tagldef_sheet_list.append(work_sheet)
                    continue
                elif AnalysisSpec().check_titles(work_sheet):
                    taglana_sheet_list.append(work_sheet)
                    continue
            if hu_sheet_list:
                sheet_dict['hu'] = hu_sheet_list
            if tagldef_sheet_list:
                sheet_dict['definition'] = tagldef_sheet_list
            if taglana_sheet_list:
                sheet_dict['analysis'] = taglana_sheet_list
            try:
                file_sheet_dict[file.decode("gb2312")] = sheet_dict
            except Exception as e:
                file_sheet_dict[file.decode("utf8")] = sheet_dict
        if file_sheet_dict:
            ret_info = self.import_to_db(file_sheet_dict)
            return ret_info
        else:
            return

    def import_to_db(self, file_sheet_dict):
        ret_info = {}
        for file_name, sheet_dict in file_sheet_dict.items():
            # print filename
            # file_name = filename.decode("utf8")
            result_list = []
            hu_sheet_list = sheet_dict.get("hu")
            if hu_sheet_list:
                for hu_sheet in hu_sheet_list:
                    print 'hu导入。。。：', file_name
                    file_result_dict = SpecHU().store(hu_sheet, self.user_id, self.update_time)
                    result_list.append(file_result_dict)
                ret_info[file_name] = result_list
        for file_name, sheet_dict in file_sheet_dict.items():

            result_list = []
            def_sheet_list = sheet_dict.get("definition")
            if def_sheet_list:
                for def_sheet in def_sheet_list:
                    print 'definition导入。。。：', file_name
                    file_result_dict = DefinitionSpec().store(def_sheet, self.user_id, self.update_time)
                    result_list.append(file_result_dict)
                ret_info[file_name] = result_list
        for file_name, sheet_dict in file_sheet_dict.items():

            result_list = []
            ana_sheet_list = sheet_dict.get("analysis")
            if ana_sheet_list:
                for ana_sheet in ana_sheet_list:
                    print 'analysis导入。。。：', file_name
                    file_result_dict = AnalysisSpec().store(ana_sheet, self.user_id, self.update_time)
                    result_list.append(file_result_dict)
                ret_info[file_name] = result_list
            # for key, sheet_list in sheet_dict.items():
            #     if key == 'hu':
            #         for hu_sheet in sheet_list:
            #             print 'hu导入。。。：', filename
            #             file_result_dict = SpecHU().store(hu_sheet, self.user_id)
            #             if file_result_dict.get('result') > 0:
            #                 break
            #     if key == 'definition':
            #         for def_sheet in sheet_list:
            #             print 'definition导入。。。：', filename
            #             file_result_dict = DefinitionSpec().store(def_sheet, self.user_id)
            #             if file_result_dict.get('result') > 0:
            #                 break
            #     if key == 'analysis':
            #         for ana_sheet in sheet_list:
            #             print 'analysis导入。。。：', filename
            #             file_result_dict = AnalysisSpec().store(ana_sheet, self.user_id)
            #             if file_result_dict.get('result') > 0:
            #                 break


        return ret_info


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    update_time = time.strftime("%Y-%m-%d %H:%M:%S")
    obj = ArlImport([], 427, update_time)
    file_path = r"C:\Users\yuyin\Desktop\import\point-test\HU_RequirementDefinition.xlsx"
    data = obj.sheet_import(file_path, 'HU_DEF')
    print data

