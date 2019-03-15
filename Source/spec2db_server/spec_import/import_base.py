# -*- coding: UTF-8 -*-
"""
Created on 2017-9-14

@author: hcz
"""
import re
import json
import datetime
from openpyxl.utils import range_boundaries
from Source.spec2db_server.common.db_access_base import DBAccessBase


class ImportBase(DBAccessBase):
    """
    """
    def __init__(self):
        DBAccessBase.__init__(self)
        self._col_list = []
        self._point_col_list = []
        self._from = []
        self.covert_dict = {}

    def convert_time(self, long_time):
        if type(long_time) in (int, long):
            # excel中，用浮点数1表示1899年12月31日
            base = datetime.date(1899, 12, 31).toordinal()
            new_date = datetime.date.fromordinal(base + long_time - 1)
            return new_date.strftime("%Y-%m-%d")
        elif type(long_time) == datetime.datetime:
            return long_time.strftime("%Y-%m-%d")
        return long_time

    def convert_2_int(self, val):
        if val != '-':
            try:
                int_val = int(val)
                return int_val
            except:
                return val
        return val

    def get_row(self, sheet, row=1, from_col=1, to_col=1):
        values = []
        j = from_col
        while j <= to_col:
            val = sheet.cell(row=row, column=j).value
            # if type(val) == unicode:
            #     val = val.replace(u'_x000D_', u'')
            #     val = val.replace(u'_x0000_', u'')
            values.append(val)
            j += 1
        return values

    def _get_parent_model(self, sheet, start_row, cur_row, col):
        if cur_row <= 0:
            return []
        p_row = cur_row - 1
        if start_row - p_row > 2:
            return []
        temp_col, row, val = self._get_cell_value(sheet, p_row, col)
        if val:
            p = self._get_parent_model(sheet, start_row, row, col)
            return p + [val]
        else:
            return []

    def _get_cell_value(self, sheet, row, column):
        b_merged, row_col = self._is_merged_cell(sheet, row, column)
        if b_merged:
            row_low, col_row = row_col
            return row_low, col_row, sheet.cell(row=row_low, column=col_row).value
        else:
            return row, column, sheet.cell(row=row, column=column).value

    def _is_merged_cell(self, sheet, row, column):
        coor = sheet.cell(row=row, column=column).coordinate
        for coor in sheet.merged_cells:
            # return True, (row, column)
            for irange in sheet.merged_cell_ranges:
                min_col, min_row, max_col, max_row = range_boundaries(irange)
                if(row in range(min_row, max_row + 1) and
                   column in range(min_col, max_col + 1)):
                    return True, (min_row, min_col)
            # row_low, row_high, column_low, column_high = cell_range
            # if(row in xrange(row_low, row_high) and
            #    column in xrange(column_low, column_high)):
            #     return True, (row_low, column_low)
        return False, (row, column)

    def get_models(self, sheet, row=6, from_col=19, to_col=110):
        col = from_col
        models = []
        for col in range(from_col, to_col + 1):
            curr_row = row
            leaf_model = sheet.cell(row=curr_row, column=col).value
            if not leaf_model:
                leaf_model = ''
                # self._log.error('row=%s, col=%s' % (row, col))
                # curr_row -= 1
                # leaf_model = sheet.cell_value(curr_row, col)
            else:
                leaf_model = leaf_model.replace('\n', '')
            parent_model = self._get_parent_model(sheet, row, curr_row, col)
            model = json.dumps(parent_model + [leaf_model],
                               ensure_ascii=False,
                               encoding='utf8')
            models.append(model)
        return models

    def convert_data(self, _data):
        _data_dict = dict()
        _point_dict = dict()
        _model_list = []
        for key in _data:
            _model_dict = dict()
            db_col_name = key
            datavalue = _data[key].get("datavalue")
            model_id = _data[key].get("model_id")
            if model_id:
                if datavalue != '-':
                    _model_dict["model_id"] = model_id
                    _model_dict["val"] = datavalue
                    _model_list.append(_model_dict)
            elif db_col_name in self._from:
                continue
            elif db_col_name in self.covert_dict.keys():
                _data_dict[self.covert_dict.get(db_col_name)] = datavalue
            elif db_col_name in self._point_col_list:
                # continue
                _point_dict[db_col_name] = datavalue
            else:
                _data_dict[db_col_name] = datavalue
        _data_dict["model_list"] = _model_list
        for key in _point_dict.keys():
            if _point_dict.get(key):
                _data_dict["point_dict"] = _point_dict
                break
        return _data_dict

    def convert_basic_data(self, _data):
        _data_dict = dict()
        _point_dict = dict()
        _model_list = []
        for key in _data:
            _model_dict = dict()
            db_col_name = key
            datavalue = _data[key].get("datavalue")
            model_id = _data[key].get("model_id")
            if model_id:
                if datavalue != '-':
                    _model_dict["model_id"] = model_id
                    _model_dict["val"] = datavalue
                    _model_list.append(_model_dict)
            elif db_col_name in self.covert_dict.keys():
                _data_dict[self.covert_dict.get(db_col_name)] = datavalue
            elif db_col_name in self._point_col_list:
                # continue
                _point_dict[db_col_name] = datavalue
            else:
                _data_dict[db_col_name] = datavalue
        _data_dict["model_list"] = _model_list
        for key in _point_dict.keys():
            if _point_dict.get(key):
                _data_dict["point_dict"] = _point_dict
                break
        return _data_dict

    def is_basic_id(self, _id):
        if _id:
            if _id[0] in ('A', 'B', 'C', 'D'):
                return True
        return False

    def import_check_list(self, pg, check_list, classify, commit_id, check_role):
        from Source.spec2db_server.arl.arl_base import ServiceBase
        obj = ServiceBase()
        new_check_list = []
        if check_role == 'author':
            for check in check_list:
                new_check_dict = dict()
                id = check.get('id')
                result = check.get('check_result')
                new_check_dict['cl_item_id'] = id
                new_check_dict['author_check'] = result
                new_check_list.append(new_check_dict)
        else:
            for check in check_list:
                new_check_dict = dict()
                id = check.get('id')
                result = check.get('check_result')
                new_check_dict['cl_item_id'] = id
                new_check_dict['charger_check'] = result
                new_check_list.append(new_check_dict)
        obj.insert_check_list(pg, commit_id, classify, new_check_list)
        print commit_id

    def split_commit_list(self, commit_list):
        author_commit_list = []
        charger_commit_list = []
        for commit in commit_list:
            if commit.get('action') == 'add':
                author_commit_list.append(commit)
            else:
                col_change_list = commit.get('col_change_list')
                model_rel = commit.get('model_rel')
                point_dict = commit.get('point_dict')
                if col_change_list != ['job_status', 'update_time', 'small_ver', 'md5_key']:
                    author_commit_list.append(commit)
                else:
                    if point_dict:
                        author_commit_list.append(commit)
                        continue
                    author = False
                    for model in model_rel:
                        if model.get('action') != 'same':
                            author = True
                            break
                    if author:
                        author_commit_list.append(commit)
                        continue
                    charger_commit_list.append(commit)
        return author_commit_list, charger_commit_list










