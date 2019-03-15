# -*- coding: UTF-8 -*-
"""
Created on 2017-8-23

@author: hcz
"""
import time
from openpyxl import load_workbook
from import_base import ImportBase
from Source.spec2db_server.arl.hu_server import HuRecord
from Source.spec2db_server.arl.def_server import DefRecord
from Source.spec2db_server.arl.analysis_service import AnalysisRecord
from Source.spec2db_server.arl.basic_hu import BasicHuRecord
from Source.spec2db_server.arl.basic_def import BasicDefRecord
from Source.spec2db_server.arl.basic_ana import BasicAnaRecord
from Source.spec2db_server.arl.commit_log import CommitLog


class Block(ImportBase):
    """
    """
    def __init__(self):
        ImportBase.__init__(self)
        self.file_name = ''

    def set_file(self, file_name):
        self.file_name = file_name

    def store(self, classify, user_id, update_time, up, down):
        result = {'result': 0}
        try:
            book = load_workbook(self.file_name, data_only=True)
            sheet = book.get_sheet_by_name(r'block')
        except Exception as e:
            result['result'] = 1
            result["error_list"] = [str(e)]
            return result
        start_row = 2
        max_row = sheet.max_row
        error = None
        self._pg.connect()
        update_count = 0
        try:
            error, update_count = self._store(sheet, classify, start_row, max_row, user_id, update_time, up, down)
            if update_count:
                print 'commit'
            self._pg.commit()
        except Exception as e:
            print e
            error = str(e)
            self._pg.conn.rollback()
        finally:
            self._pg.close()
        if error:
            result['result'] = 1
            result["error_list"] = [error]
            result["update_count"] = 0
        else:
            result["update_count"] = update_count
        return result

    def _store(self, sheet, classify, start_row, max_row, user_id, update_time, up, down):
        error, update_count = None, 0
        if classify == 'HU_DEF':
            obj = HuRecord()
            basic_obj = BasicHuRecord()
        elif classify == 'TAGL_DEF':
            obj = DefRecord()
            basic_obj = BasicDefRecord()
        elif classify == 'TAGL_ANA':
            obj = AnalysisRecord()
            basic_obj = BasicAnaRecord()
        else:
            error = 'Unkown classify [%s]' % classify
            return error, update_count
        commit_list = []
        for row in range(start_row, max_row + 1):
            values = self.get_row(sheet, row, from_col=1, to_col=2)
            _id, block = values[0], values[1]
            if not _id:
                break
            lock_status = 0  # unlock
            if block == u'〇':
                lock_status = 1  # lock
            if self.is_basic_id(_id):
                curr_commit_list = basic_obj.upate_lock_status(self._pg, _id, lock_status, update_time, up, down)
            else:
                curr_commit_list = obj.upate_lock_status(self._pg, _id, lock_status, update_time, up, down)
            if curr_commit_list:
                update_count += 1
                commit_list += curr_commit_list
        log = {'user_id': user_id, "commit_list": commit_list}
        if commit_list:
            commit_log = CommitLog()
            flag, commit_id = commit_log.add_commit_log2(self._pg, log)
            if flag:
                obj.update_post_lock_status(self._pg)
                obj.update_parent_lock_status(self._pg)
            else:
                self._pg.conn.rollback()
                error = 'log_commit errol!'
        else:
            print 'no change.'
        return error, update_count


if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    files = [
        r'c:\share\user\hcz\block_sample_ana.xlsx',
    ]
    for file in files:
        print file
        obj = Block()
        user_id = 448
        classify = 'TAGL_ANA'  # HU_DEF/TAGL_DEF/TAGL_ANA
        update_time = obj.get_current_time()
        obj.set_file(file)
        obj.store(classify, user_id, update_time)
