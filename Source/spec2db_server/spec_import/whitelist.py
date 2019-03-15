# -*- coding: UTF-8 -*-
"""
Created on 2017-12-25

@author: hcz
"""
from openpyxl import load_workbook
from import_base import ImportBase


class WhiteList(ImportBase):
    """
    """

    def __init__(self):
        ImportBase.__init__(self)
        self.file_name = ''

    def set_file(self, file_name):
        self.file_name = file_name

    def store(self, classify, user_id, update_time):
        result = {'result': 0}
        try:
            book = load_workbook(self.file_name, data_only=True)
            sheet = book.get_sheet_by_name(r'whitelist')
        except Exception as e:
            result['result'] = 1
            result["error_list"] = [str(e)]
            return result
        start_row = 2
        max_row = sheet.max_row
        error = None
        self._pg.connect()
        update_count = 0
        try:
            error, update_count, repeat_count = self._store(sheet, classify, start_row, max_row, user_id, update_time)
            if update_count:
                print 'commit'
                self._pg.commit()
        except Exception as e:
            print e
            error = str(e)
            self._pg.conn.rollback()
        finally:
            self._pg.close()
        if error:
            result['result'] = 1
            result["error_list"] = [error]
            result["update_count"] = 0
            result["repeat_count"] = 0
        else:
            result["update_count"] = update_count
            result["repeat_count"] = repeat_count
        return result

    def _store(self, sheet, classify, start_row, max_row, user_id, update_time):
        error, insert_count, repeat_count = '', 0, 0
        del_count = 0
        if classify == 'HU_DEF':
            cls = "HU"
        elif classify == 'TAGL_DEF':
            cls = "DEF"
        elif classify == 'TAGL_ANA':
            cls = "ANA"
        else:
            error = 'Unkown classify [%s]' % classify
            return error
        for row in range(start_row, max_row + 1):
            values = self.get_row(sheet, row, from_col=1, to_col=2)
            _id, add_del_flag = values[0], values[1]
            if not _id:
                continue
            if add_del_flag == u'〇':
                if not self.exist_white_list(self._pg, cls, _id):
                    params = [cls, _id, user_id, update_time]
                    self._insert_white_list(self._pg, params)
                    insert_count += 1
                else:
                    repeat_count += 1
            elif add_del_flag == u'×':
                params = [cls, _id]
                if self._del_white_list(self._pg, params):
                    del_count += 1
        return error, insert_count + del_count, repeat_count

    def _insert_white_list(self, pg, params):
        sqlcmd = """
            INSERT INTO spec.white_list(classify, id, import_user_id, import_time)
              VALUES(%s, %s, %s, %s);
            """
        pg.execute(sqlcmd, params)

    def _del_white_list(self, pg, params):
        sqlcmd = """
        DELETE FROM spec.white_list WHERE classify = %s and id=%s;
        """
        pg.execute(sqlcmd, params)
        if pg.pgcur.rowcount > 0:
            return True
        return False

    def exist_white_list(self, pg, classify, _id):
        sqlcmd = """
            SELECT id
              FROM spec.white_list
              WHERE classify= %s and id = %s
            """
        pg.execute(sqlcmd, [classify, _id])
        row = pg.fetchone()
        if row:
            return True
        return False


if __name__ == '__main__':
    import sys, os
    reload(sys)
    sys.setdefaultencoding('UTF-8')
    os.chdir('../')
    files = [
        r'C:\share\user\hcz\white_list\whitelist_sample_ana.xlsx',
    ]
    for file in files:
        print file
        obj = WhiteList()
        user_id = 446
        classify = 'TAGL_ANA'  # HU_DEF/TAGL_DEF/TAGL_ANA
        update_time = obj.get_current_time()
        obj.set_file(file)
        obj.store(classify, user_id, update_time)
