# -*- coding: UTF-8 -*-
from arl_server import ServiceBase
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import os
import json
from lgatio import HUDocType, ReqDocType012 as ReqDocType, AnaDocType, AnaDocTypeWithLongID
from lgatio.data.dynamic import APIDynamicData
CHECK_DOC_TYPE_HU = 'HU'
CHECK_DOC_TYPE_DEFINITION = 'definition'
CHECK_DOC_TYPE_ANALYSIS = 'analysis'
HU_MODEL_LIST_NUM = 30
DEF_MODEL_LIST_NUM = 31
ANA_MODEL_LIST_NUM = 105
HU_OPTION_START_IDX = 14
HU_OPTION_END_IDX = 19


class ArlFunc(ServiceBase):
    def __init__(self):
        ServiceBase.__init__(self)

    def export_arl_to_object_by_userid(self, export_path, export_file_name, user_id):
        print 'Exporting User Arl'
        ARL_EXPORT_SQL = """
                       (
                       SELECT distinct
                            arl.arl_id, arl.major_category, arl.medium_catetory as medium_category, arl.small_category,
                            arl.detail, arl.req_post, schedule.change_type, groups.group_name,
                            users.user_name, gp_ch.charger,schedule.hu_date, schedule.hu_remark,
                            schedule.def_date, schedule.def_remark,  schedule.analysis_date, schedule.analysis_remark
                       from spec.arl as arl
                       INNER join (
                           SELECT distinct group_id, (array_agg(user_name))[1] as charger
                             FROM spec.arl_group_member as t1
                             left join spec.arl_user as t2
                             on t1.user_id = t2.user_id
                             where role_id = 4
                             group by group_id
                        ) as gp_ch
                       on arl.group_id = gp_ch.group_id
                       left join spec.arl_user as users
                       on arl.user_id = users.user_id
                       left join spec.arl_group as groups
                       on arl.group_id = groups.group_id
                       left join spec.arl_schedule as schedule
                       on arl.arl_id = schedule.arl_id 
                       where arl.exclude_flag=False and arl.user_id=%s
                       )
                     union
                       (
                        SELECT distinct
                            arl.arl_id, arl.major_category, arl.medium_category, arl.small_category,
                            arl.detail, arl.arl_req_post as req_post, schedule.change_type, groups.group_name,
                            users.user_name, gp_ch.charger,schedule.hu_date, schedule.hu_remark,
                            schedule.def_date, schedule.def_remark,  schedule.analysis_date, schedule.analysis_remark
                       from spec.basic_req_hu as arl
                       INNER join (
                           SELECT distinct group_id, (array_agg(user_name))[1] as charger
                                 FROM spec.arl_group_member as t1
                                 left join spec.arl_user as t2
                                 on t1.user_id = t2.user_id
                                 where role_id = 4
                                 group by group_id
                            ) as gp_ch
                       on arl.group_id = gp_ch.group_id
                       left join spec.arl_user as users
                       on arl.user_id = users.user_id
                       left join spec.arl_group as groups
                       on arl.group_id = groups.group_id
                       left join spec.arl_schedule as schedule
                       on arl.arl_id = schedule.arl_id 
                       where arl.user_id=%s
                       )
                    order by arl_id
           """
        self._pg.connect()
        self._pg.execute(ARL_EXPORT_SQL, (user_id,user_id))
        arl_db_list = self._pg.fetchall()

        if len(arl_db_list) == 0:
            return 'NODATA'

        wb = Workbook()
        ws = wb.active
        title_list = ["ARL ID", "大分類","中分類" , "小分類" , "詳細", "転記してきた要件", "变更类别", "担当组", "担当", "负责", "HU",
                      "备考", "TAGL定义", "备考", "TAGL分析" , "备考"]
        ws.append(title_list)
        for row in arl_db_list:
            ws.append(list(row))
        wb.save(os.path.join(export_path, export_file_name))
        self._pg.close()

        return "OK"

    def export_arl_to_object_by_groupid(self, export_path, export_file_name, user_id):
        print 'Exporting Group Arl'
        ARL_EXPORT_SQL = """
                       (
                       SELECT distinct
                            arl.arl_id, arl.major_category, arl.medium_catetory as medium_category, arl.small_category,
                            arl.detail, arl.req_post, schedule.change_type, groups.group_name,
                            users.user_name, gp_ch.charger,schedule.hu_date, schedule.hu_remark,
                            schedule.def_date, schedule.def_remark,  schedule.analysis_date, schedule.analysis_remark
                       from spec.arl as arl
                       INNER join (
                           SELECT distinct group_id, (array_agg(user_name))[1] as charger
                             FROM spec.arl_group_member as t1
                             left join spec.arl_user as t2
                             on t1.user_id = t2.user_id
                             where role_id = 4 and t1.user_id =%s
                             group by group_id
                        ) as gp_ch
                       on arl.group_id = gp_ch.group_id
                       left join spec.arl_user as users
                       on arl.user_id = users.user_id
                       left join spec.arl_group as groups
                       on arl.group_id = groups.group_id
                       left join spec.arl_schedule as schedule
                       on arl.arl_id = schedule.arl_id 
                       where arl.exclude_flag=False
                       )
                     union
                       (
                        SELECT distinct
                            arl.arl_id, arl.major_category, arl.medium_category, arl.small_category,
                            arl.detail, arl.arl_req_post as req_post, schedule.change_type, groups.group_name,
                            users.user_name, gp_ch.charger,schedule.hu_date, schedule.hu_remark,
                            schedule.def_date, schedule.def_remark,  schedule.analysis_date, schedule.analysis_remark
                       from spec.basic_req_hu as arl
                       INNER join (
                           SELECT distinct group_id, user_name as charger
                             FROM spec.arl_group_member as t1
                             left join spec.arl_user as t2
                             on t1.user_id = t2.user_id
                             where role_id = 4 and t1.user_id =%s
                        ) as gp_ch
                       on arl.group_id = gp_ch.group_id
                       left join spec.arl_user as users
                       on arl.user_id = users.user_id
                       left join spec.arl_group as groups
                       on arl.group_id = groups.group_id
                       left join spec.arl_schedule as schedule
                       on arl.arl_id = schedule.arl_id 
                       )
                      order by arl_id
                   """
        self._pg.connect()
        self._pg.execute(ARL_EXPORT_SQL, (user_id, user_id))
        arl_db_list = self._pg.fetchall()

        if len(arl_db_list) == 0:
            return 'NODATA'

        wb = Workbook()
        ws = wb.active
        title_list = ["ARL ID", "大分類","中分類" , "小分類" , "詳細", "転記してきた要件", "变更类别", "担当组", "担当", "负责", "HU",
                      "备考", "TAGL定义", "备考", "TAGL分析" , "备考"]
        ws.append(title_list)
        for row in arl_db_list:
            ws.append(list(row))
        wb.save(os.path.join(export_path, export_file_name))
        self._pg.close()

        return "OK"

    def export_arl_to_object_by_all(self, export_path, export_file_name):
        print 'Exporting All Arl'
        ARL_EXPORT_SQL = """
                       (
                       SELECT distinct
                            arl.arl_id, arl.major_category, arl.medium_catetory as medium_category, arl.small_category,
                            arl.detail, arl.req_post, schedule.change_type, groups.group_name,
                            users.user_name, gp_ch.charger,schedule.hu_date, schedule.hu_remark,
                            schedule.def_date, schedule.def_remark,  schedule.analysis_date, schedule.analysis_remark
                       from spec.arl as arl
                       INNER join (
                           SELECT distinct group_id, (array_agg(user_name))[1] as charger
                             FROM spec.arl_group_member as t1
                             left join spec.arl_user as t2
                             on t1.user_id = t2.user_id
                             where role_id = 4
                             group by group_id
                        ) as gp_ch
                       on arl.group_id = gp_ch.group_id
                       left join spec.arl_user as users
                       on arl.user_id = users.user_id
                       left join spec.arl_group as groups
                       on arl.group_id = groups.group_id
                       left join spec.arl_schedule as schedule
                       on arl.arl_id = schedule.arl_id 
                       where arl.exclude_flag=False
                       )
                     union
                       (
                        SELECT distinct
                        arl_id, major_category, medium_catetory, small_category,
                        detail, req_post, change_type, "group",
                        author, charger, hu_date, hu_remark,
                        def_date, def_remark,  analysis_date, analysis_remark
                        from spec.arl_schedule where arl_id like 'B%' or arl_id like 'C%' or arl_id like 'D%'
                       )
                    order by arl_id
                   """
        self._pg.connect()
        self._pg.execute(ARL_EXPORT_SQL)
        arl_db_list = self._pg.fetchall()

        if len(arl_db_list) == 0:
            return 'NODATA'

        wb = Workbook()
        ws = wb.active
        title_list = ["ARL ID", "大分類","中分類" , "小分類" , "詳細", "転記してきた要件", "变更类别", "担当组", "担当", "负责", "HU",
                      "备考", "TAGL定义", "备考", "TAGL分析" , "备考"]
        ws.append(title_list)
        for row in arl_db_list:
            ws.append(list(row))

        wb.save(os.path.join(export_path, export_file_name))
        self._pg.close()

        return "OK"

    def export_hu_to_unfinished_by_all(self, export_path, export_file_name):
        print 'Exporting All Unfinished HU'
        HU_EXPORT_SQL = """
                SELECT arl_id, hu_date
                    FROM (
                    SELECT distinct t1.arl_id, hu_date, length(t1.arl_id) as len
                    FROM spec.arl as t1
                    left join spec.hu as t2
                    on t1.arl_id = t2.arl_id
                    left join spec.arl_schedule as t3
                    on t1.arl_id = t3.arl_id
                    where exclude_flag = False and hu_date <> '-' and t2.arl_id is null
                    ) AS tt1
                order by len, arl_id;
                   """
        self._pg.connect()
        self._pg.execute(HU_EXPORT_SQL)
        hu_db_list = self._pg.fetchall()

        if len(hu_db_list) == 0:
            return 'NODATA'

        wb = Workbook()
        ws = wb.active
        title_list = ["HU ID", "预定日"]
        ws.append(title_list)
        for row in hu_db_list:
            ws.append(list(row))

        wb.save(os.path.join(export_path, export_file_name))
        self._pg.close()
        return "OK"


    def export_def_to_unfinished_by_all(self, export_path, export_file_name):
        print 'Exporting All Unfinished DEF'
        DEF_EXPORT_SQL = """
                SELECT arl_id, def_date
                    FROM (
                    SELECT distinct t1.arl_id, def_date, length(t1.arl_id) as len
                    FROM spec.arl as t1
                    left join spec.hu as t2
                    on t1.arl_id = t2.arl_id
                    left join spec.arl_schedule as t3
                    on t1.arl_id = t3.arl_id
                    left JOIN spec.definition as t4
                    on t2.hu_id = t4.hu_def_id
                    where exclude_flag = False and def_date <> '-'
                    and (t2.arl_id is null or t4.hu_def_id is null)
                    ) AS tt1
                order by len, arl_id;
                   """
        self._pg.connect()
        self._pg.execute(DEF_EXPORT_SQL)
        def_db_list = self._pg.fetchall()

        if len(def_db_list) == 0:
            return 'NODATA'

        wb = Workbook()
        ws = wb.active
        title_list = ["DEF ID", "预定日"]
        ws.append(title_list)
        for row in def_db_list:
            ws.append(list(row))

        wb.save(os.path.join(export_path, export_file_name))
        self._pg.close()
        return "OK"

    def export_ana_to_unfinished_by_all(self, export_path, export_file_name):
        print 'Exporting All Unfinished Ana'
        Ana_EXPORT_SQL = """
                SELECT arl_id, analysis_date
                    FROM (
                    SELECT distinct t1.arl_id, analysis_date, length(t1.arl_id) as len
                    FROM spec.arl as t1
                    left join spec.hu as t2
                    on t1.arl_id = t2.arl_id
                    left join spec.arl_schedule as t3
                    on t1.arl_id = t3.arl_id
                    left JOIN spec.definition as t4
                    on t2.hu_id = t4.hu_def_id
                    left join spec.analysis as t5
                    on t4.definition_id = t5.definition_id
                    where exclude_flag = False and analysis_date <> '-'
                    and (t2.arl_id is null or t4.hu_def_id is null
                    or t5.definition_id is null)
                    ) AS tt1
                order by len, arl_id;
                   """
        self._pg.connect()
        self._pg.execute(Ana_EXPORT_SQL)
        ana_db_list = self._pg.fetchall()

        if len(ana_db_list) == 0:
          return 'NODATA'

        wb = Workbook()
        ws = wb.active
        title_list = ["Analysis ID", "预定日"]
        ws.append(title_list)
        for row in ana_db_list:
            ws.append(list(row))

        wb.save(os.path.join(export_path, export_file_name))
        self._pg.close()
        return "OK"

    def export_hu_to_excel_by_userid(self, export_path, export_file_name, templateName, user_id):
        print 'Exporting User HU...'
        HU_EXPORT_SQL = """
         SELECT  distinct
                hu.author, hu.arl_id, arl.major_category, arl.medium_catetory,
                arl.small_category, arl.detail, hu.basic_req, arl.req_post,
                arl.status, arl.trigger, arl.action, arl.remark,
                hu.hu_id, hu.unique_id, hu.amp, hu.dsrc, 
                hu.dcm, hu.rse, hu.touch_pad, hu.separate_disp,
                hu.system_conf, hu.rel_requirement, hu.exception, hu.dcu_status,
                hu.dcu_trigger, hu.dcu_action, hu.meu_status, hu.meu_trigger, 
                hu.meu_action, hu.hu_category_id,
                model_ids, vals,
                hu.remark, hu.sys_spec_chapter, hu.common_chapter, hu.common_seq_spec,
                hu.common_seq_no, hu.common_cmd_guide, hu.common_opc, hu.inter_loc_spec,
                hu.inter_chapter, hu.other_doc, hu.other_chapter, hu.test_results,
                hu.future_req, hu.remark1, hu.remark2, hu.leak_check,
                hu.last_modifier, to_char(to_date(hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), hu.reason, 
                point_out.review_result, point_out.pointout_no, point_out.pointout_status,
                point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
                point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
                point_out.pointout_date, point_out.suntec_status, point_out.fixed,
                point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify,
		        case when hu.lock_status=0 then '无' when hu.lock_status=1 then '锁' when hu.lock_status=2 then '转记锁' end as lock_status,
                grp.group_name,
                hu.md5_key, 
                case when hu.job_status=1 then '未完成' when hu.job_status=2 then '已完成，待确认' when hu.job_status=3 then '已确认' end as job_status,
                arl_user.user_name, hu.update_time,
                length(hu.arl_id) as len
        from spec.hu as hu
        left join spec.arl as arl
        on hu.arl_id = arl.arl_id
        left join spec.arl_user
        on arl_user.user_id = arl.user_id
        left join (
        SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
          FROM(
            SELECT hu_record_id, model_id, val
              FROM spec.hu_model_rel    
              order by hu_record_id, model_id
          ) AS A
          GROUP BY hu_record_id
        ) as m
        on hu.hu_record_id = m.hu_record_id
        left join spec.arl_group_member as grp_member
        on grp_member.user_id = arl.user_id and arl.group_id = grp_member.group_id
        left join spec.arl_group as grp
        on grp.group_id = arl.group_id
        left join spec.hu_latest_point_out_view as point_out
        on hu.hu_id = point_out.id
        where exclude_flag = False and arl.user_id = %s
        order by len,hu.arl_id,hu.unique_id
        """
        self._pg.connect()
        self._pg.execute(HU_EXPORT_SQL, (user_id,))
        hu_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_hu_basic_req_hu_user_for_export(), (user_id,))
        basic_req_list = self._pg.fetchall()
        ret = self.do_hu_db_export((hu_db_list + basic_req_list), export_path, export_file_name, templateName)
        self._pg.close()

        return ret

    def get_sqlcmd_hu_basic_req_hu_user_for_export(self):
        SQL_CMD = '''
        SELECT
            basic_req_hu.author, basic_req_hu.arl_id,basic_req_hu.major_category, basic_req_hu.medium_category,
            basic_req_hu.small_category, basic_req_hu.detail, basic_req_hu.basic_req, basic_req_hu.arl_req_post,
            basic_req_hu.arl_status, basic_req_hu.arl_trigger, basic_req_hu.arl_action, basic_req_hu.arl_remark,
            basic_req_hu.hu_id, basic_req_hu.unique_id, basic_req_hu.amp, basic_req_hu.dsrc, 
            basic_req_hu.dcm, basic_req_hu.rse, basic_req_hu.touch_pad, basic_req_hu.separate_disp,
            basic_req_hu.system_conf, basic_req_hu.rel_requirement, basic_req_hu.exception, basic_req_hu.dcu_status,
            basic_req_hu.dcu_trigger, basic_req_hu.dcu_action, basic_req_hu.meu_status, basic_req_hu.meu_trigger, 
            basic_req_hu.meu_action, basic_req_hu.hu_category_id,
            model_ids, vals,
            basic_req_hu.remark, basic_req_hu.sys_spec_chapter, basic_req_hu.common_chapter, basic_req_hu.common_seq_spec,
            basic_req_hu.common_seq_no, basic_req_hu.common_cmd_guide, basic_req_hu.common_opc, basic_req_hu.inter_loc_spec,
            basic_req_hu.inter_chapter, basic_req_hu.other_doc, basic_req_hu.other_chapter, basic_req_hu.test_results,
            basic_req_hu.future_req, basic_req_hu.remark1, basic_req_hu.remark2, basic_req_hu.leak_check,
            basic_req_hu.last_modifier, to_char(to_date(basic_req_hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), basic_req_hu.reason,
            point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		    case when basic_req_hu.lock_status=0 then '无' when basic_req_hu.lock_status=1 then '锁' when basic_req_hu.lock_status=2 then '转记锁' end as lock_status,
            grp.group_name, 
            basic_req_hu.md5_key, 
            case when basic_req_hu.job_status=1 then '未完成' when basic_req_hu.job_status=2 then '已完成，待确认' when basic_req_hu.job_status=3 then '已确认' end as job_status,
            arl_user.user_name, basic_req_hu.update_time,
            length(basic_req_hu.arl_id) as len
        from spec.basic_req_hu as basic_req_hu
        left join (
            SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
              FROM(
                SELECT hu_record_id, model_id, val
                  FROM spec.basic_req_hu_model_rel
                  order by hu_record_id, model_id
              ) AS A
              GROUP BY hu_record_id ) as m 
              on basic_req_hu.hu_record_id = m.hu_record_id
              left join spec.arl_group as grp
                on grp.group_id = basic_req_hu.group_id
              left join spec.hu_latest_point_out_view as point_out
                on basic_req_hu.hu_id = point_out.id
              left join spec.arl_user on arl_user.user_id = basic_req_hu.user_id
              where basic_req_hu.user_id = %s
              order by len, basic_req_hu.arl_id, basic_req_hu.unique_id
        '''
        return SQL_CMD


    def export_group_hu_to_excel_by_userid(self, export_path, export_file_name, templateName, user_id):
        print 'Exporting UserGroup HU...'
        HU_EXPORT_SQL = '''
        SELECT  distinct
                hu.author, hu.arl_id, arl.major_category, arl.medium_catetory,
                arl.small_category, arl.detail, hu.basic_req, arl.req_post,
                arl.status, arl.trigger, arl.action, arl.remark,
                hu.hu_id, hu.unique_id, hu.amp, hu.dsrc, 
                hu.dcm, hu.rse, hu.touch_pad, hu.separate_disp,
                hu.system_conf, hu.rel_requirement, hu.exception, hu.dcu_status,
                hu.dcu_trigger, hu.dcu_action, hu.meu_status, hu.meu_trigger, 
                hu.meu_action, hu.hu_category_id,
                model_ids, vals,
                hu.remark, hu.sys_spec_chapter, hu.common_chapter, hu.common_seq_spec,
                hu.common_seq_no, hu.common_cmd_guide, hu.common_opc, hu.inter_loc_spec,
                hu.inter_chapter, hu.other_doc, hu.other_chapter, hu.test_results,
                hu.future_req, hu.remark1, hu.remark2, hu.leak_check,
                hu.last_modifier, to_char(to_date(hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), hu.reason, 
                point_out.review_result, point_out.pointout_no, point_out.pointout_status,
                point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
                point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
                point_out.pointout_date, point_out.suntec_status, point_out.fixed,
                point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		        case when hu.lock_status=0 then '无' when hu.lock_status=1 then '锁' when hu.lock_status=2 then '转记锁' end as lock_status,
                grp.group_name,
                hu.md5_key, 
                case when hu.job_status=1 then '未完成' when hu.job_status=2 then '已完成，待确认' when hu.job_status=3 then '已确认' end as job_status,
                arl_user.user_name, hu.update_time,
                length(hu.arl_id) as len
        from spec.hu as hu
        left join spec.arl as arl
        on hu.arl_id = arl.arl_id
        left join spec.arl_user
        on arl_user.user_id = arl.user_id
        left join (
        SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
          FROM(
            SELECT hu_record_id, model_id, val
              FROM spec.hu_model_rel    
              order by hu_record_id, model_id
          ) AS A
          GROUP BY hu_record_id
        ) as m
        on hu.hu_record_id = m.hu_record_id
        left join spec.arl_group_member as grp_member
        on grp_member.user_id = arl.user_id and arl.group_id = grp_member.group_id
        left join spec.arl_group as grp
        on grp.group_id = grp_member.group_id
        left join spec.hu_latest_point_out_view as point_out
        on hu.hu_id = point_out.id
        where exclude_flag = False and arl.group_id in (select group_id from spec.arl_group_member where user_id=%s)
        order by len,hu.arl_id,hu.unique_id
        '''
        self._pg.connect()
        self._pg.execute(HU_EXPORT_SQL, (user_id,))
        hu_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_hu_basic_req_hu_group_for_export(), (user_id,))
        basic_req_list = self._pg.fetchall()

        ret = self.do_hu_db_export((hu_db_list + basic_req_list), export_path, export_file_name, templateName)
        self._pg.close()

        return ret


    def get_sqlcmd_hu_basic_req_hu_group_for_export(self):
        SQL_CMD = '''
        SELECT
            basic_req_hu.author, basic_req_hu.arl_id,basic_req_hu.major_category, basic_req_hu.medium_category,
            basic_req_hu.small_category, basic_req_hu.detail, basic_req_hu.basic_req, basic_req_hu.arl_req_post,
            basic_req_hu.arl_status, basic_req_hu.arl_trigger, basic_req_hu.arl_action, basic_req_hu.arl_remark,
            basic_req_hu.hu_id, basic_req_hu.unique_id, basic_req_hu.amp, basic_req_hu.dsrc, 
            basic_req_hu.dcm, basic_req_hu.rse, basic_req_hu.touch_pad, basic_req_hu.separate_disp,
            basic_req_hu.system_conf, basic_req_hu.rel_requirement, basic_req_hu.exception, basic_req_hu.dcu_status,
            basic_req_hu.dcu_trigger, basic_req_hu.dcu_action, basic_req_hu.meu_status, basic_req_hu.meu_trigger, 
            basic_req_hu.meu_action, basic_req_hu.hu_category_id,
            model_ids, vals,
            basic_req_hu.remark, basic_req_hu.sys_spec_chapter, basic_req_hu.common_chapter, basic_req_hu.common_seq_spec,
            basic_req_hu.common_seq_no, basic_req_hu.common_cmd_guide, basic_req_hu.common_opc, basic_req_hu.inter_loc_spec,
            basic_req_hu.inter_chapter, basic_req_hu.other_doc, basic_req_hu.other_chapter, basic_req_hu.test_results,
            basic_req_hu.future_req, basic_req_hu.remark1, basic_req_hu.remark2, basic_req_hu.leak_check,
            basic_req_hu.last_modifier, to_char(to_date(basic_req_hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), basic_req_hu.reason, 
            point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		    case when basic_req_hu.lock_status=0 then '无' when basic_req_hu.lock_status=1 then '锁' when basic_req_hu.lock_status=2 then '转记锁' end as lock_status,
            grp.group_name, 
            basic_req_hu.md5_key, 
            case when basic_req_hu.job_status=1 then '未完成' when basic_req_hu.job_status=2 then '已完成，待确认' when basic_req_hu.job_status=3 then '已确认' end as job_status,
            arl_user.user_name, basic_req_hu.update_time,
            length(basic_req_hu.arl_id) as len
        from spec.basic_req_hu as basic_req_hu
        left join (
            SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
              FROM(
                SELECT hu_record_id, model_id, val
                  FROM spec.basic_req_hu_model_rel
                  order by hu_record_id, model_id
              ) AS A
              GROUP BY hu_record_id ) as m 
              on basic_req_hu.hu_record_id = m.hu_record_id
              left join spec.arl_group as grp
                on grp.group_id = basic_req_hu.group_id
              left join spec.hu_latest_point_out_view as point_out
                on basic_req_hu.hu_id = point_out.id
              left join spec.arl_user
                on arl_user.user_id = basic_req_hu.user_id
              where basic_req_hu.group_id in (select group_id from spec.arl_group_member where user_id=%s)
              order by len, basic_req_hu.arl_id, basic_req_hu.unique_id
        '''
        return SQL_CMD

    def export_hu_to_excel(self, export_path, export_file_name, templateName):
        print 'Exporting All HU...'
        HU_EXPORT_SQL = '''
        SELECT  distinct
                hu.author, hu.arl_id, arl.major_category, arl.medium_catetory,
                arl.small_category, arl.detail, hu.basic_req, arl.req_post,
                arl.status, arl.trigger, arl.action, arl.remark,
                hu.hu_id, hu.unique_id, hu.amp, hu.dsrc, 
                hu.dcm, hu.rse, hu.touch_pad, hu.separate_disp,
                hu.system_conf, hu.rel_requirement, hu.exception, hu.dcu_status,
                hu.dcu_trigger, hu.dcu_action, hu.meu_status, hu.meu_trigger, 
                hu.meu_action, hu.hu_category_id,
                model_ids, vals,
                hu.remark, hu.sys_spec_chapter, hu.common_chapter, hu.common_seq_spec,
                hu.common_seq_no, hu.common_cmd_guide, hu.common_opc, hu.inter_loc_spec,
                hu.inter_chapter, hu.other_doc, hu.other_chapter, hu.test_results,
                hu.future_req, hu.remark1, hu.remark2, hu.leak_check,
                hu.last_modifier, to_char(to_date(hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), hu.reason, 
                point_out.review_result, point_out.pointout_no, point_out.pointout_status,
                point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
                point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
                point_out.pointout_date, point_out.suntec_status, point_out.fixed,
                point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		        case when hu.lock_status=0 then '无' when hu.lock_status=1 then '锁' when hu.lock_status=2 then '转记锁' end as lock_status,
                grp.group_name,
                hu.md5_key, 
                case when hu.job_status=1 then '未完成' when hu.job_status=2 then '已完成，待确认' when hu.job_status=3 then '已确认' end as job_status,
                arl_user.user_name, hu.update_time,
                length(hu.arl_id) as len
        from spec.hu as hu
        left join spec.arl as arl
        on hu.arl_id = arl.arl_id
        left join spec.arl_user
        on arl_user.user_id = arl.user_id
        left join (
        SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
          FROM(
            SELECT hu_record_id, model_id, val
              FROM spec.hu_model_rel    
              order by hu_record_id, model_id
          ) AS A
          GROUP BY hu_record_id
        ) as m
        on hu.hu_record_id = m.hu_record_id
        left join spec.arl_group_member as grp_member
        on grp_member.user_id = arl.user_id and arl.group_id = grp_member.group_id
        left join spec.arl_group as grp
        on grp.group_id = arl.group_id
        left join spec.hu_latest_point_out_view as point_out
        on hu.hu_id = point_out.id
        where exclude_flag = False
        order by len,hu.arl_id,hu.unique_id
        '''
        self._pg.connect()
        self._pg.execute(HU_EXPORT_SQL)
        hu_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_hu_list_for_export())
        basic_req_list = self._pg.fetchall()

        ret = self.do_hu_db_export((hu_db_list + basic_req_list), export_path, export_file_name, templateName)
        self._pg.close()

        return ret

    def get_sqlcmd_hu_for_export(self):
        sqlcmd = """
        SELECT  distinct
                hu.author, hu.arl_id, arl.major_category, arl.medium_catetory,
                arl.small_category, arl.detail, hu.basic_req, arl.req_post,
                arl.status, arl.trigger, arl.action, arl.remark,
                hu.hu_id, hu.unique_id, hu.amp, hu.dsrc, 
                hu.dcm, hu.rse, hu.touch_pad, hu.separate_disp,
                hu.system_conf, hu.rel_requirement, hu.exception, hu.dcu_status,
                hu.dcu_trigger, hu.dcu_action, hu.meu_status, hu.meu_trigger, 
                hu.meu_action, hu.hu_category_id,
                model_ids, vals,
                hu.remark, hu.sys_spec_chapter, hu.common_chapter, hu.common_seq_spec,
                hu.common_seq_no, hu.common_cmd_guide, hu.common_opc, hu.inter_loc_spec,
                hu.inter_chapter, hu.other_doc, hu.other_chapter, hu.test_results,
                hu.future_req, hu.remark1, hu.remark2, hu.leak_check,
                hu.last_modifier, to_char(to_date(hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), hu.reason, 
                point_out.review_result, point_out.pointout_no, point_out.pointout_status,
                point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
                point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
                point_out.pointout_date, point_out.suntec_status, point_out.fixed,
                point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
                case when hu.lock_status=0 then '无' when hu.lock_status=1 then '锁' when hu.lock_status=2 then '转记锁' end as lock_status,
                grp.group_name,
                hu.md5_key, 
                case when hu.job_status=1 then '未完成' when hu.job_status=2 then '已完成，待确认' when hu.job_status=3 then '已确认' end as job_status,
                arl_user.user_name, hu.update_time,
                length(hu.arl_id) as len
        from spec.hu as hu
        left join spec.arl as arl
        on hu.arl_id = arl.arl_id
        left join spec.arl_user
        on arl_user.user_id = arl.user_id
        left join (
        SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
          FROM(
            SELECT hu_record_id, model_id, val
              FROM spec.hu_model_rel    
              order by hu_record_id, model_id
          ) AS A
          GROUP BY hu_record_id
        ) as m
        on hu.hu_record_id = m.hu_record_id
        left join spec.arl_group_member as grp_member
        on grp_member.user_id = arl.user_id and arl.group_id = grp_member.group_id
        left join spec.arl_group as grp
        on grp.group_id = arl.group_id
        left join spec.hu_latest_point_out_view as point_out
        on hu.hu_id = point_out.id
        where exclude_flag = False
        order by len,hu.arl_id,hu.unique_id
        """
        return sqlcmd

    def get_sqlcmd_basic_req_hu_list_for_export(self):
        BASIC_REQ_HU_EXPORT_SQL = '''
        SELECT
            basic_req_hu.author, basic_req_hu.arl_id,basic_req_hu.major_category, basic_req_hu.medium_category,
            basic_req_hu.small_category, basic_req_hu.detail, basic_req_hu.basic_req, basic_req_hu.arl_req_post,
            basic_req_hu.arl_status, basic_req_hu.arl_trigger, basic_req_hu.arl_action, basic_req_hu.arl_remark,
            basic_req_hu.hu_id, basic_req_hu.unique_id, basic_req_hu.amp, basic_req_hu.dsrc, 
            basic_req_hu.dcm, basic_req_hu.rse, basic_req_hu.touch_pad, basic_req_hu.separate_disp,
            basic_req_hu.system_conf, basic_req_hu.rel_requirement, basic_req_hu.exception, basic_req_hu.dcu_status,
            basic_req_hu.dcu_trigger, basic_req_hu.dcu_action, basic_req_hu.meu_status, basic_req_hu.meu_trigger, 
            basic_req_hu.meu_action, basic_req_hu.hu_category_id,
            model_ids, vals,
            basic_req_hu.remark, basic_req_hu.sys_spec_chapter, basic_req_hu.common_chapter, basic_req_hu.common_seq_spec,
            basic_req_hu.common_seq_no, basic_req_hu.common_cmd_guide, basic_req_hu.common_opc, basic_req_hu.inter_loc_spec,
            basic_req_hu.inter_chapter, basic_req_hu.other_doc, basic_req_hu.other_chapter, basic_req_hu.test_results,
            basic_req_hu.future_req, basic_req_hu.remark1, basic_req_hu.remark2, basic_req_hu.leak_check,
            basic_req_hu.last_modifier, to_char(to_date(basic_req_hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), basic_req_hu.reason,  
            point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		    case when basic_req_hu.lock_status=0 then '无' when basic_req_hu.lock_status=1 then '锁' when basic_req_hu.lock_status=2 then '转记锁' end as lock_status,
            grp.group_name, 
            basic_req_hu.md5_key, 
            case when basic_req_hu.job_status=1 then '未完成' when basic_req_hu.job_status=2 then '已完成，待确认' when basic_req_hu.job_status=3 then '已确认' end as job_status,
            arl_user.user_name, basic_req_hu.update_time,
            length(basic_req_hu.arl_id) as len
        from spec.basic_req_hu as basic_req_hu
        left join (
            SELECT hu_record_id, array_agg(val) vals, array_agg(model_id) model_ids
              FROM(
                SELECT hu_record_id, model_id, val
                  FROM spec.basic_req_hu_model_rel
                  order by hu_record_id, model_id
              ) AS A
              GROUP BY hu_record_id ) as m 
              on basic_req_hu.hu_record_id = m.hu_record_id
              left join spec.arl_group as grp
                on grp.group_id = basic_req_hu.group_id
              left join spec.hu_latest_point_out_view as point_out
                on basic_req_hu.hu_id = point_out.id
              left join spec.arl_user on arl_user.user_id = basic_req_hu.user_id
              order by len, basic_req_hu.arl_id, basic_req_hu.unique_id
        '''
        return BASIC_REQ_HU_EXPORT_SQL

    def do_hu_db_export(self, db_rows, export_path, export_file_name, templateName):
        result_data_list = []
        print len(db_rows)
        if len(db_rows) == 0:
            return 'NODATA'
        for row in db_rows:
            hu_all_model_val = ['-'] * HU_MODEL_LIST_NUM
            if row[30]:
                for model_id, model_val in zip(row[30], row[31]):
                    hu_all_model_val[model_id - 1] = model_val
            row = self.convert_option_type(row, HU_OPTION_START_IDX, HU_OPTION_END_IDX)
            result_data_list.append(map(self.cvt_json_to_str, (list(row[:30]) + hu_all_model_val + list(row[32:-1]))))

        openpyxl_wb = load_workbook(templateName)
        ws = openpyxl_wb.get_sheet_by_name('HU要件定義書'.decode('utf8'))

        start_row = 10
        ws._current_row = start_row - 1
        for row in result_data_list:
            ws.append(row)

        for ir in range(start_row, ws.max_row+1):
            ws.cell(row=ir, column=13).value='''=B%d&"."&N%d'''%(ir, ir)
            if not ws.cell(row=ir, column=2).value:
                break
            try:
                detail_float_try = float(ws.cell(row=ir, column=6).value)
                #1.2 is the text in arl document
                if detail_float_try == 1.2:
                    continue
                ws.cell(row=ir, column=6).value = detail_float_try
            except:
                pass
            finally:
                pass

        openpyxl_wb.save(os.path.join(export_path, export_file_name))

        return "OK"

    def export_definition_to_excel_by_userid(self, export_path, export_file_name, templateName, user_id):
        print 'Exporting TAGL Definition...'
        sqlcmd = '''
          SELECT distinct def.author_name,
               hu_def_id, definition_id, def.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               def.basic_req, def.rel_requirement, 
               def.exception,
               dcu_status, dcu_trigger, dcu_action,
               meu_status, meu_trigger, meu_action, hu.remark as hu_remark,
               dcu_meu, pf_status, pf_trigger, pf_action,
               model_ids, vals, 
               def.remark as def_remark, notice, rel_flow_diagram, rel_hal_design,
               def.other_spec, implementation, analysis, unrequire,
               to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), def.reason, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		       case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
               g.group_name, 
               def.md5_key, 
               case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
               u.user_name, def.update_time,
               length(hu_def_id) as len
          FROM spec.definition as def
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT def_rc_id, model_id, val
                  FROM spec.definition_model_rel
                  ORDER BY def_rc_id, model_id
              ) as m
              group by def_rc_id
          ) as model
          on def.def_rc_id = model.def_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.def_latest_point_out_view as point_out
                on def.definition_id = point_out.id
          where exclude_flag = False and a.user_id = %s
          order by len, hu_def_id, def.unique_id
                '''
        self._pg.connect()
        self._pg.execute(sqlcmd, (user_id,))
        define_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_define_user_list_for_export(), (user_id,))
        define_req_list = self._pg.fetchall()
        ret = self.do_def_db_export((define_db_list + define_req_list), export_path, export_file_name, templateName)
        self._pg.close()
        return ret

    def get_sqlcmd_basic_req_define_user_list_for_export(self):
        sqlcmd = '''
            SELECT def.author_name,
                   def.hu_def_id, def.definition_id, def.unique_id,
                   def.major_category, def.medium_category, def.small_category, def.detail,
                   def.basic_req, def.rel_requirement, 
                   def.exception,
                   def.dcu_status, def.dcu_trigger, def.dcu_action,
                   def.meu_status, def.meu_trigger, def.meu_action, def.hu_remark,
                   def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
                   model_ids, vals, 
                   def.remark as def_remark, def.notice, def.rel_flow_diagram, def.rel_hal_design,
                   def.other_spec, def.implementation, def.analysis, def.unrequire,
                   to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
                   def.reason, 
                   point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
                   case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
                   grp.group_name, 
                   def.md5_key,  
                   case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
                   arl_user.user_name, def.update_time,
                   length(def.hu_def_id) as len
              FROM spec.basic_req_definition as def
              LEFT JOIN (
                SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
                  from (
                    SELECT def_rc_id, model_id, val
                      FROM spec.basic_req_definition_model_rel
                      ORDER BY def_rc_id, model_id
                  ) as m
                  group by def_rc_id
              ) as model
              on def.def_rc_id = model.def_rc_id
              left join spec.arl_group as grp 
                on grp.group_id = def.group_id
            left join spec.def_latest_point_out_view as point_out
                on def.definition_id = point_out.id
              left join spec.arl_user on arl_user.user_id = def.user_id 
              where def.user_id = %s
              order by len, hu_def_id, def.unique_id
        '''
        return sqlcmd

    def export_group_definition_to_excel_by_userid(self, export_path, export_file_name, templateName, user_id):
        print 'Exporting TAGL Definition...'
        sqlcmd = '''
            SELECT distinct def.author_name,
               hu_def_id, definition_id, def.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               def.basic_req, def.rel_requirement, 
               def.exception,
               dcu_status, dcu_trigger, dcu_action,
               meu_status, meu_trigger, meu_action, hu.remark as hu_remark,
               dcu_meu, pf_status, pf_trigger, pf_action,
               model_ids, vals, 
               def.remark as def_remark, notice, rel_flow_diagram, rel_hal_design,
               def.other_spec, implementation, analysis, unrequire,
               to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), def.reason, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
		       case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
               g.group_name, 
               def.md5_key, 
               case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
               u.user_name, def.update_time,
               length(hu_def_id) as len
          FROM spec.definition as def
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT def_rc_id, model_id, val
                  FROM spec.definition_model_rel
                  ORDER BY def_rc_id, model_id
              ) as m
              group by def_rc_id
          ) as model
          on def.def_rc_id = model.def_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.def_latest_point_out_view as point_out
            on def.definition_id = point_out.id
          where exclude_flag = False and a.group_id in (select group_id from spec.arl_group_member where user_id=%s)
          order by len, hu_def_id, def.unique_id
          '''
        self._pg.connect()
        self._pg.execute(sqlcmd, (user_id,))
        define_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_define_group_list_for_export(), (user_id,))
        define_req_list = self._pg.fetchall()
        ret = self.do_def_db_export( (define_db_list + define_req_list), export_path, export_file_name, templateName)
        self._pg.close()
        return ret

    def get_sqlcmd_basic_req_define_group_list_for_export(self):
        sqlcmd = '''
            SELECT def.author_name,
                   def.hu_def_id, def.definition_id, def.unique_id,
                   def.major_category, def.medium_category, def.small_category, def.detail,
                   def.basic_req, def.rel_requirement, 
                   def.exception,
                   def.dcu_status, def.dcu_trigger, def.dcu_action,
                   def.meu_status, def.meu_trigger, def.meu_action, def.hu_remark,
                   def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
                   model_ids, vals, 
                   def.remark as def_remark, def.notice, def.rel_flow_diagram, def.rel_hal_design,
                   def.other_spec, def.implementation, def.analysis, def.unrequire,
                   to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
                   def.reason, 
                   point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
                   case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
                   grp.group_name, 
                   def.md5_key, 
                   case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
                   arl_user.user_name, def.update_time,
                   length(def.hu_def_id) as len
              FROM spec.basic_req_definition as def
              LEFT JOIN (
                SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
                  from (
                    SELECT def_rc_id, model_id, val
                      FROM spec.basic_req_definition_model_rel
                      ORDER BY def_rc_id, model_id
                  ) as m
                  group by def_rc_id
              ) as model
              on def.def_rc_id = model.def_rc_id
              left join spec.arl_group as grp 
                on grp.group_id = def.group_id
            left join spec.def_latest_point_out_view as point_out
                on def.definition_id = point_out.id
              left join spec.arl_user on arl_user.user_id = def.user_id
              where def.group_id in (select group_id from spec.arl_group_member where user_id=%s)
              order by len, hu_def_id, def.unique_id
        '''
        return sqlcmd


    def export_definition_to_excel(self, export_path, export_file_name, templateName):
        print 'Exporting TAGL All Definition...'
        sqlcmd = '''
            SELECT distinct def.author_name,
                hu_def_id, definition_id, def.unique_id,
                a.major_category, a.medium_catetory, small_category, detail,
                def.basic_req, def.rel_requirement, 
                def.exception,
                dcu_status, dcu_trigger, dcu_action,
                meu_status, meu_trigger, meu_action, hu.remark as hu_remark,
                dcu_meu, pf_status, pf_trigger, pf_action,
                model_ids, vals, 
                def.remark as def_remark, notice, rel_flow_diagram, rel_hal_design,
                def.other_spec, implementation, analysis, unrequire,
                to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
                def.reason,
                point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
                case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
                g.group_name, 
                def.md5_key, 
                case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
                u.user_name, def.update_time,
                length(hu_def_id) as len
              FROM spec.definition as def
              LEFT JOIN spec.hu as hu
              ON def.hu_def_id = hu.hu_id
              LEFT JOIN spec.arl as a
              on hu.arl_id = a.arl_id
              LEFT JOIN (
                SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
                  from (
                    SELECT def_rc_id, model_id, val
                      FROM spec.definition_model_rel
                      ORDER BY def_rc_id, model_id
                  ) as m
                  group by def_rc_id
              ) as model
              on def.def_rc_id = model.def_rc_id
              inner join spec.arl_user as u
              on a.user_id = u.user_id
              inner join spec.arl_group as g
              on a.group_id = g.group_id
              left join spec.def_latest_point_out_view as point_out
                on def.definition_id = point_out.id
              where exclude_flag = False
              order by len, hu_def_id, def.unique_id
        '''
        self._pg.connect()
        self._pg.execute(sqlcmd)
        define_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_define_list_for_export())
        define_req_list = self._pg.fetchall()
        ret = self.do_def_db_export((define_db_list + define_req_list), export_path, export_file_name, templateName)
        self._pg.close()
        return ret

    def get_sqlcmd_define_for_export(self):
        sqlcmd = """
        SELECT distinct def.author_name,
            hu_def_id, definition_id, def.unique_id,
            a.major_category, a.medium_catetory, small_category, detail,
            def.basic_req, def.rel_requirement, 
            def.exception,
            dcu_status, dcu_trigger, dcu_action,
            meu_status, meu_trigger, meu_action, hu.remark as hu_remark,
            dcu_meu, pf_status, pf_trigger, pf_action,
            model_ids, vals, 
            def.remark as def_remark, notice, rel_flow_diagram, rel_hal_design,
            def.other_spec, implementation, analysis, unrequire,
            to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
            def.reason,
            point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
            case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
            g.group_name, 
            def.md5_key, 
            case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
            u.user_name, def.update_time,
            length(hu_def_id) as len
          FROM spec.definition as def
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT def_rc_id, model_id, val
                  FROM spec.definition_model_rel
                  ORDER BY def_rc_id, model_id
              ) as m
              group by def_rc_id
          ) as model
          on def.def_rc_id = model.def_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.def_latest_point_out_view as point_out
                on def.definition_id = point_out.id
          where exclude_flag = False
          order by len, hu_def_id, def.unique_id
        """
        return sqlcmd

    def get_sqlcmd_basic_req_define_list_for_export(self):
        sqlcmd = '''
            SELECT def.author_name,
                   def.hu_def_id, def.definition_id, def.unique_id,
                   def.major_category, def.medium_category, def.small_category, def.detail,
                   def.basic_req, def.rel_requirement, 
                   def.exception,
                   def.dcu_status, def.dcu_trigger, def.dcu_action,
                   def.meu_status, def.meu_trigger, def.meu_action, def.hu_remark,
                   def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
                   model_ids, vals, 
                   def.remark as def_remark, def.notice, def.rel_flow_diagram, def.rel_hal_design,
                   def.other_spec, def.implementation, def.analysis, def.unrequire,
                   to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
                   def.reason, 
                   point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
                   case when def.lock_status=0 then '无' when def.lock_status=1 then '锁' when def.lock_status=2 then '转记锁' end as lock_status,
                   grp.group_name, 
                   def.md5_key, 
                   case when def.job_status=1 then '未完成' when def.job_status=2 then '已完成，待确认' when def.job_status=3 then '已确认' end as job_status,
                   arl_user.user_name, def.update_time,
                   length(def.hu_def_id) as len
              FROM spec.basic_req_definition as def
              LEFT JOIN (
                SELECT def_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
                  from (
                    SELECT def_rc_id, model_id, val
                      FROM spec.basic_req_definition_model_rel
                      ORDER BY def_rc_id, model_id
                  ) as m
                  group by def_rc_id
              ) as model
              on def.def_rc_id = model.def_rc_id
              left join spec.arl_group as grp 
                on grp.group_id = def.group_id
            left join spec.def_latest_point_out_view as point_out
                on def.definition_id = point_out.id
              left join spec.arl_user on arl_user.user_id = def.user_id
              order by len, hu_def_id, def.unique_id
        '''
        return sqlcmd

    def do_def_db_export(self, db_rows, export_path, export_file_name, templateName):
        result_data_list = []
        print len(db_rows)
        if len(db_rows) == 0:
            return 'NODATA'
        for row in db_rows:
            def_all_model_val = ['-'] * DEF_MODEL_LIST_NUM
            if row[22]:
                for model_id, model_val in zip(row[22], row[23]):
                    def_all_model_val[model_id - 1] = model_val
            result_data_list.append(map(self.cvt_json_to_str, (list(row[:22]) + def_all_model_val + list(row[24:-1]))))

        start_row = 6
        openpyxl_wb = load_workbook(templateName)
        ws = openpyxl_wb.get_sheet_by_name('TAGL要件定義'.decode('utf8'))

        ws._current_row = start_row - 1
        for row in result_data_list:
            ws.append(row)

        for ir in range(start_row, ws.max_row + 1):
            ws.cell(row=ir, column=3).value = '''=B%d&"."&D%d''' % (ir, ir)
            if not ws.cell(row=ir, column=2).value:
                break
            try:
                detail_float_try = float(ws.cell(row=ir, column=8).value)
                # 1.2 is the text in arl document
                if detail_float_try == 1.2:
                    continue
                ws.cell(row=ir, column=8).value = detail_float_try
            except:
                pass
            finally:
                pass
        openpyxl_wb.save(os.path.join(export_path, export_file_name))

        return "OK"

    def export_analysis_to_excel(self, export_path, export_file_name, templateName):
        print 'Exporting TAGL All Analysis...'
        sqlcmd = '''
            SELECT distinct ana.author_name,
               ana.definition_id, ana.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               g.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               u.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.analysis as ana
          LEFT JOIN spec.definition as def
          ON ana.definition_id = def.definition_id
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.ana_latest_point_out_view as point_out
                on ana.definition_id = point_out.id
          where exclude_flag = False
          order by len, ana.definition_id, ana.unique_id
        '''
        self._pg.connect()
        self._pg.execute(sqlcmd)
        ana_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_ana_list_for_export())
        ana_req_list = self._pg.fetchall()
        ret = self.do_ana_db_export((ana_db_list + ana_req_list), export_path, export_file_name, templateName)
        self._pg.close()
        return ret

    def get_sqlcmd_ana_for_export(self):
        sqlcmd = """
            SELECT distinct ana.author_name,
               ana.definition_id, ana.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               g.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               u.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.analysis as ana
          LEFT JOIN spec.definition as def
          ON ana.definition_id = def.definition_id
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.ana_latest_point_out_view as point_out
                on ana.definition_id = point_out.id
          where exclude_flag = False
          order by len, ana.definition_id, ana.unique_id
        """
        return sqlcmd

    def get_sqlcmd_basic_req_ana_list_for_export(self):
        sqlcmd = '''
            SELECT ana.author_name,
               ana.definition_id, ana.unique_id,
               ana.major_category, ana.medium_category, ana.small_category, ana.detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               ana.dcu_meu, ana.pf_status, ana.pf_trigger, ana.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               grp.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               arl_user.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.basic_req_analysis as ana
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.basic_req_analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          left join spec.arl_group as grp 
            on grp.group_id = ana.group_id
          left join spec.ana_latest_point_out_view as point_out
            on ana.definition_id = point_out.id
          left join spec.arl_user on arl_user.user_id = ana.user_id
          order by len, ana.definition_id, ana.unique_id
        '''
        return sqlcmd


    def export_user_analysis_to_excel(self, export_path, export_file_name, templateName, user_id):
        print 'Exporting TAGL User Analysis...'
        sqlcmd = '''
            SELECT distinct ana.author_name,
               ana.definition_id, ana.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               g.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               u.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.analysis as ana
          LEFT JOIN spec.definition as def
          ON ana.definition_id = def.definition_id
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.ana_latest_point_out_view as point_out
            on ana.definition_id = point_out.id
          where exclude_flag = False and a.user_id = %s
          order by len, ana.definition_id, ana.unique_id
        '''
        self._pg.connect()
        self._pg.execute(sqlcmd, (user_id,))
        ana_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_ana_user_list_for_export(), (user_id,))
        ana_req_list = self._pg.fetchall()
        ret = self.do_ana_db_export((ana_db_list + ana_req_list), export_path, export_file_name, templateName)
        self._pg.close()
        return ret

    def get_sqlcmd_basic_req_ana_user_list_for_export(self):
        sqlcmd = '''
            SELECT ana.author_name,
               ana.definition_id, ana.unique_id,
               ana.major_category, ana.medium_category, ana.small_category, ana.detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               ana.dcu_meu, ana.pf_status, ana.pf_trigger, ana.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               grp.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               arl_user.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.basic_req_analysis as ana
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.basic_req_analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          left join spec.arl_group as grp 
            on grp.group_id = ana.group_id
          left join spec.ana_latest_point_out_view as point_out
                on ana.definition_id = point_out.id
          left join spec.arl_user on arl_user.user_id = ana.user_id
          where ana.user_id=%s
          order by len, ana.definition_id, ana.unique_id
        '''
        return sqlcmd


    def export_group_analysis_to_excel(self, export_path, export_file_name, templateName, user_id):
        print 'Exporting TAGL Group Analysis...'
        sqlcmd = '''
            SELECT distinct ana.author_name,
               ana.definition_id, ana.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               def.dcu_meu, def.pf_status, def.pf_trigger, def.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               g.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               u.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.analysis as ana
          LEFT JOIN spec.definition as def
          ON ana.definition_id = def.definition_id
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          left join spec.ana_latest_point_out_view as point_out
                on ana.definition_id = point_out.id
          where exclude_flag = False and a.group_id in (select group_id from spec.arl_group_member where user_id=%s)
          order by len, ana.definition_id, ana.unique_id
        '''
        self._pg.connect()
        self._pg.execute(sqlcmd, (user_id,))
        ana_db_list = self._pg.fetchall()
        self._pg.execute(self.get_sqlcmd_basic_req_ana_group_list_for_export(), (user_id,))
        ana_req_list = self._pg.fetchall()
        ret = self.do_ana_db_export((ana_db_list + ana_req_list), export_path, export_file_name, templateName)
        self._pg.close()
        return ret

    def get_sqlcmd_basic_req_ana_group_list_for_export(self):
        sqlcmd = '''
            SELECT ana.author_name,
               ana.definition_id, ana.unique_id,
               ana.major_category, ana.medium_category, ana.small_category, ana.detail,
               ana.basic_req, ana.rel_requirement, 
               ana.exception,
               ana.dcu_meu, ana.pf_status, ana.pf_trigger, ana.pf_action,
               ana.seq_diagram,
               model_ids, vals, 
               ana.supple_spec, ana.uncheck, ana.remark,
               ana.ana_1 as n1, ana.ana_2 as n2, ana.ana_3 as n3, ana.ana_4 as n4,
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, ana.asta_filename, 
               point_out.review_result, point_out.pointout_no, point_out.pointout_status,
            point_out.pointout_comment, point_out.reader_check, point_out.reader2_check,
            point_out.final_check, point_out.pointout_charger, point_out.pointout_priority,
            point_out.pointout_date, point_out.suntec_status, point_out.fixed,
            point_out.suntec_remark, point_out.arl_rel, point_out.suntec_cannot_modify, 
               case when ana.lock_status=0 then '无' when ana.lock_status=1 then '锁' when ana.lock_status=2 then '转记锁' end as lock_status,
               grp.group_name, ana.md5_key, 
               case when ana.job_status=1 then '未完成' when ana.job_status=2 then '已完成，待确认' when ana.job_status=3 then '已确认' end as job_status,
               arl_user.user_name, ana.update_time,
               length(ana.definition_id) as len
          FROM spec.basic_req_analysis as ana
          LEFT JOIN (
            SELECT analysis_rc_id, array_agg(val) vals, array_agg(model_id) model_ids
              from (
                SELECT analysis_rc_id, model_id, val
                  FROM spec.basic_req_analysis_model_rel
                  ORDER BY analysis_rc_id, model_id
              ) as m
              group by analysis_rc_id
          ) as model
          on ana.analysis_rc_id = model.analysis_rc_id
          left join spec.arl_group as grp 
            on grp.group_id = ana.group_id
          left join spec.ana_latest_point_out_view as point_out
                on ana.definition_id = point_out.id
          left join spec.arl_user on arl_user.user_id = ana.user_id
          where ana.group_id in (select group_id from spec.arl_group_member where user_id=%s)
          order by len, ana.definition_id, ana.unique_id
        '''
        return sqlcmd

    def do_ana_db_export(self, db_rows, export_path, export_file_name, templateName):
        result_data_list = []
        print len(db_rows)
        if len(db_rows) == 0:
            return 'NODATA'
        for row in db_rows:
            ana_all_model_val = ['-'] * ANA_MODEL_LIST_NUM
            if row[15]:
                for model_id, model_val in zip(row[15], row[16]):
                    ana_all_model_val[model_id - 1] = model_val
            result_data_list.append(map(self.cvt_json_to_str, (list(row[:15]) + ana_all_model_val + list(row[17:-1]))))

        start_row = 6
        openpyxl_wb = load_workbook(templateName)
        ws = openpyxl_wb.get_sheet_by_name('TAGL要件分析'.decode('utf8'))

        ws._current_row = start_row - 1
        for row in result_data_list:
            ws.append(row)

        for ir in range(start_row, ws.max_row + 1):
            if not ws.cell(row=ir, column=2).value:
                break
            try:
                detail_float_try = float(ws.cell(row=ir, column=7).value)
                # 1.2 is the text in arl document
                if detail_float_try == 1.2:
                    continue
                ws.cell(row=ir, column=7).value = detail_float_try
            except:
                pass
            finally:
                pass
        
        openpyxl_wb.save(os.path.join(export_path, export_file_name))

        return "OK"

    def cvt_json_to_str(self, str_info):
        try:
            json_obj = json.loads(str_info)
            if type(json_obj["strVal"]) == unicode:
                return json_obj["strVal"].encode("utf8")
            return json_obj["strVal"]
        except:
            return str_info

    def check_data_list(self, pg, data_list, user_id):
        hu_record_id_list, def_rc_id_list, analysis_rc_id_list = self._get_record_id_list(data_list)
        import Source.spec2db_server.arl.arl_server
        spec_obj = Source.spec2db_server.arl.arl_server.ArlSpec()
        dyndata = spec_obj.get_block_white_list()
        api_dyndata = APIDynamicData(dyndata)
        if hu_record_id_list:
            hu_data_list = self._get_hu_by_ids(pg, hu_record_id_list)
            flag, e = self._check_data_list(hu_data_list, CHECK_DOC_TYPE_HU, api_dyndata)
            if not flag:
                return flag, e
        if def_rc_id_list:
            def_data_list = self._get_definition_by_ids(pg, def_rc_id_list)
            flag, e = self._check_data_list(def_data_list, CHECK_DOC_TYPE_DEFINITION, api_dyndata)
            if not flag:
                return flag, e
        if analysis_rc_id_list:
            analysis_data_list = self._get_analysis_by_ids(pg, analysis_rc_id_list)
            flag, e = self._check_data_list(analysis_data_list, CHECK_DOC_TYPE_ANALYSIS, api_dyndata)
            if not flag:
                return flag, e
        return True, None

    def _check_data_list(self, data_list, check_type, dyndata):
        if check_type == CHECK_DOC_TYPE_HU:
            doc_obj = HUDocType(dyndata)
        elif check_type == CHECK_DOC_TYPE_DEFINITION:
            doc_obj = ReqDocType(dyndata)
        elif check_type == CHECK_DOC_TYPE_ANALYSIS:
            doc_obj = AnaDocTypeWithLongID(dyndata)
        else:
            return False
        col_names = doc_obj.get_col_names()
        doc_obj.formula_funcs = {}  # don't check formula
        for row in data_list:
            if len(row) < len(col_names):
                return False
            data_dict = dict()
            for i, col in enumerate(col_names, 0):
                data_dict[col] = row[i]
                if col in (u'除外', u'備考'):
                    if row[i] is None:
                        data_dict[col] = u''
            try:
                for _k, _v in data_dict.iteritems():
                    if type(_v) in [str]:
                        data_dict[_k] = _v.decode('utf8')
                if u'日付' in data_dict:
                    data_dict[u'日付'] = datetime.datetime.strptime(data_dict.get(u'日付'), '%Y-%m-%d')
                if u'更新日' in data_dict:
                    data_dict[u'更新日'] = datetime.datetime.strptime(data_dict.get(u'更新日'), '%Y-%m-%d')
                errors = doc_obj.validate_row(data_dict, continue_on_error=False, ctx=None)
            except Exception as e:
                return False, e
            if errors:
                return False, errors
        return True, None

    def _get_record_id_list(self, data_list):
        hu_record_id_list, def_rc_id_list, analysis_rc_id_list = [], [], []
        for arl in data_list:
            hu_list = arl.get("hu_list")
            if hu_list:
                for hu in hu_list:
                    hu_record_id_list.append(hu.get("hu_record_id"))
                    definition_list = hu.get("definition_list")
                    if definition_list:
                        for definition in definition_list:
                            def_rc_id_list.append(definition.get("def_rc_id"))
                            analysis_list = definition.get("analysis_list")
                            if analysis_list:
                                for analysis in analysis_list:
                                    analysis_rc_id_list.append(analysis.get("analysis_rc_id"))
        return hu_record_id_list, def_rc_id_list, analysis_rc_id_list

    def _get_hu_by_ids(self, pg, hu_id_list):
        sqlcmd = """
         SELECT distinct
                hu.author, hu.arl_id, arl.major_category, arl.medium_catetory,
                arl.small_category, arl.detail, '' as basic_req, arl.req_post,
                arl.status, arl.trigger, arl.action, arl.remark,
                hu.hu_id, hu.unique_id, hu.amp, hu.dsrc, 
                hu.dcm, hu.rse, hu.touch_pad, hu.separate_disp,
                hu.system_conf, hu.rel_requirement, hu.exception, hu.dcu_status,
                hu.dcu_trigger, hu.dcu_action, hu.meu_status, hu.meu_trigger, 
                hu.meu_action, hu.hu_category_id, ---- Model List Here
                hu.remark, hu.sys_spec_chapter, hu.common_chapter, hu.common_seq_spec,
                hu.common_seq_no, hu.common_cmd_guide, hu.common_opc, hu.inter_loc_spec,
                hu.inter_chapter, hu.other_doc, hu.other_chapter, hu.test_results,
                hu.future_req, hu.remark1, hu.remark2, hu.leak_check,
                hu.last_modifier, to_char(to_date(hu.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), hu.reason, 
                case when hu.lock_status=0 then '无' when hu.lock_status=1 then '锁' when hu.lock_status=2 then '转记锁' end as lock_status,  hu.md5_key
        from spec.hu as hu
        left join spec.arl as arl
        on hu.arl_id = arl.arl_id
        left join spec.arl_user
        on arl_user.user_id = arl.user_id
        left join spec.arl_group_member as grp_member
        on grp_member.user_id = arl.user_id and arl.group_id = grp_member.group_id
        left join spec.arl_group as grp
        on grp.group_id = grp_member.group_id
        where exclude_flag = False and hu.hu_record_id = %s
        order by hu.arl_id, hu.unique_id
        """
        result_data_list = []
        from Source.spec2db_server.arl.hu_server import HuRecord
        hu_obj = HuRecord()
        for hu_record_id in hu_id_list:
            pg.execute(sqlcmd, (hu_record_id,))
            row = pg.fetchone()
            if row:
                model_list = hu_obj.get_model_list(pg, hu_record_id)
                all_model = [u'-'] * HU_MODEL_LIST_NUM
                for model_info in model_list:
                    model_id = model_info.get("model_id")
                    all_model[model_id - 1] = model_info.get("val")
                row = self.convert_option_type(row, HU_OPTION_START_IDX, HU_OPTION_END_IDX)
                result_data_list.append(list(row[:30]) + all_model + list(row[30:-2]))
        return result_data_list

    def _get_definition_by_ids(self, pg, definition_id_list):
        sqlcmd = """
        SELECT distinct def.author_name,
               hu_def_id, definition_id, def.unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               '' as basic_req, def.rel_requirement, 
               def.exception,
               dcu_status, dcu_trigger, dcu_action,
               meu_status, meu_trigger, meu_action, hu.remark as hu_remark,
               dcu_meu, pf_status, pf_trigger, pf_action, --------- Model list here
               def.remark as def_remark, notice, rel_hal_design, rel_flow_diagram,
               def.other_spec, implementation, analysis, unrequire,
               to_char(to_date(def.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), def.reason, 
               g.group_name, def.md5_key
          FROM spec.definition as def
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          where exclude_flag = False and def.def_rc_id = %s
          order by definition_id
         """
        result_data_list = []
        for def_record_id in definition_id_list:
            pg.execute(sqlcmd, (def_record_id,))
            row = pg.fetchone()
            if row:
                from Source.spec2db_server.arl.def_server import DefRecord
                def_obj = DefRecord()
                model_list = def_obj.get_model_list(pg, def_record_id)
                all_model = [u'-'] * DEF_MODEL_LIST_NUM
                for model_info in model_list:
                    model_id = model_info.get("model_id")
                    all_model[model_id - 1] = model_info.get("val")
                result_data_list.append(list(row[:22]) + all_model + list(row[22:-2]))
        return result_data_list

    def _get_analysis_by_ids(self, pg, analysis_id_list):
        sqlcmd = """
        SELECT distinct ana.author_name,
               ana.definition_id, 0 as unique_id,
               a.major_category, a.medium_catetory, small_category, detail,
               '' as basic_req, ana.rel_requirement, 
               ana.exception,
               dcu_meu, pf_status, pf_trigger, pf_action,  -- Definition
               seq_diagram,
               --------- Model list here
               ana.supple_spec, ana.uncheck, ana.remark,
               -- 1, 2, 3, 4
               to_char(to_date(ana.new_date, 'YYYY-MM-DD HH24:MI:SS'),'YYYY-MM-DD'), 
               ana.reason, asta_filename, def.md5_key
          FROM spec.analysis as ana
          LEFT JOIN spec.definition as def
          ON ana.definition_id = def.definition_id
          LEFT JOIN spec.hu as hu
          ON def.hu_def_id = hu.hu_id
          LEFT JOIN spec.arl as a
          on hu.arl_id = a.arl_id
          inner join spec.arl_user as u
          on a.user_id = u.user_id
          inner join spec.arl_group as g
          on a.group_id = g.group_id
          where exclude_flag = False and ana.analysis_rc_id = %s
          order by definition_id
         """
        result_data_list = []
        from Source.spec2db_server.arl.analysis_service import AnalysisRecord
        analysis_obj = AnalysisRecord()
        blank = [None] * 4  # -- 1, 2, 3, 4
        for analysis_record_id in analysis_id_list:
            pg.execute(sqlcmd, (analysis_record_id,))
            row = pg.fetchone()
            if row:
                model_list = analysis_obj.get_model_list(pg, analysis_record_id)
                all_model = [u'-'] * ANA_MODEL_LIST_NUM
                for model_info in model_list:
                    model_id = model_info.get("model_id")
                    all_model[model_id - 1] = model_info.get("val")
                result_data_list.append(list(row[:15]) + all_model + list(row[15:18]) + blank + list(row[18:]))
        return result_data_list

    def export_for_release(self, pg, export_type, export_path, export_file_name, template_name):
        print 'Exporting All %s...' % export_type
        if export_type == "HU_DEF":
            # Normal
            normal_sqlcmd = self.get_sqlcmd_hu_for_export()
            # Basic
            basic_sqlcmd = self.get_sqlcmd_basic_req_hu_list_for_export()
            conver_model_list = self.conver_model_list_hu
            sheet_name = 'HU要件定義書'.decode('utf8')
            start_row = 10
            point_start, point_end = 51, 65
        elif export_type == "TAGL_DEF":
            # Normal
            normal_sqlcmd = self.get_sqlcmd_define_for_export()
            # Basic
            basic_sqlcmd = self.get_sqlcmd_basic_req_define_list_for_export()
            conver_model_list = self.conver_model_list_def
            sheet_name = 'TAGL要件定義'.decode('utf8')
            start_row = 6
            point_start, point_end = 34, 48
        else:
            # Normal
            normal_sqlcmd = self.get_sqlcmd_ana_for_export()
            # Basic
            basic_sqlcmd = self.get_sqlcmd_basic_req_ana_list_for_export()
            conver_model_list = self.conver_model_list_ana
            sheet_name = 'TAGL要件分析'.decode('utf8')
            start_row = 6
            point_start, point_end = 27, 41
        for old_table, release_table in [('spec.hu ', 'release.hu '),
                                         ('spec.hu_model_rel ', 'release.hu_model_rel '),
                                         ('spec.definition ', 'release.definition '),
                                         ('spec.definition_model_rel ', 'release.definition_model_rel '),
                                         ('spec.analysis ', 'release.analysis '),
                                         ('spec.analysis_model_rel ', 'release.analysis_model_rel '),
                                         ('spec.basic_req_hu ', 'release.basic_req_hu '),
                                         ('spec.basic_req_hu_model_rel ', 'release.basic_req_hu_model_rel '),
                                         ('spec.basic_req_definition ', 'release.basic_req_definition '),
                                         ('spec.basic_req_definition_model_rel ', 'release.basic_req_definition_model_rel '),
                                         ('spec.basic_req_analysis ', 'release.basic_req_analysis '),
                                         ('spec.basic_req_analysis_model_rel ', 'release.basic_req_analysis_model_rel ')]:
            normal_sqlcmd = normal_sqlcmd.replace(old_table, release_table)
            basic_sqlcmd = basic_sqlcmd.replace(old_table, release_table)
        pg.execute(normal_sqlcmd)
        normal_db_list = pg.fetchall()
        pg.execute(basic_sqlcmd)
        basic_req_list = pg.fetchall()
        data_list = conver_model_list(normal_db_list + basic_req_list, point_start, point_end)
        ret = self.do_db_export_no_style(data_list, export_path, export_file_name,
                                         template_name, sheet_name, start_row)
        return ret

    def conver_model_list_hu(self, data_list, point_start=51, point_end=65):
        temp_data_list = []
        for row in data_list:
            hu_all_model_val = ['-'] * HU_MODEL_LIST_NUM
            if row[30]:
                for model_id, model_val in zip(row[30], row[31]):
                    hu_all_model_val[model_id - 1] = model_val
            row = self.convert_option_type(row, HU_OPTION_START_IDX, HU_OPTION_END_IDX)
            temp_data_list.append(list(row[:30]) + hu_all_model_val +
                                  list(row[32:point_start]) + list(row[point_end+1:-1]))
        return temp_data_list

    def conver_model_list_def(self, data_list, point_start=34, point_end=48):
        temp_data_list = []
        for row in data_list:
            def_all_model_val = ['-'] * DEF_MODEL_LIST_NUM
            if row[22]:
                for model_id, model_val in zip(row[22], row[23]):
                    def_all_model_val[model_id - 1] = model_val
            temp_data = list(row[:22]) + def_all_model_val + list(row[24:point_start]) + list(row[point_end + 1:-1])
            temp_data_list.append(temp_data)
        return temp_data_list

    def conver_model_list_ana(self, data_list, point_start=27, point_end=41):
        temp_data_list = []
        for row in data_list:
            ana_all_model_val = ['-'] * ANA_MODEL_LIST_NUM
            if row[15]:
                for model_id, model_val in zip(row[15], row[16]):
                    ana_all_model_val[model_id - 1] = model_val
            temp_data = list(row[:15]) + ana_all_model_val + list(row[17:point_start]) + list(row[point_end + 1:-1])
            temp_data_list.append(temp_data)
        return temp_data_list

    def do_db_export_no_style(self, db_rows, export_path, export_file_name,
                              template_name, sheet_name, start_row=10):
        print "Number:", len(db_rows)
        if len(db_rows) == 0:
            return 'NODATA'
        wb = load_workbook(template_name)
        ws = wb.get_sheet_by_name(sheet_name)
        ws._current_row = start_row - 1
        for i, row in enumerate(db_rows, 1):
            ws.append(row)
            # print i
        for ir in range(start_row, ws.max_row + 1):
            if not ws.cell(row=ir, column=2).value:
                break
            try:
                detail_float_try = float(ws.cell(row=ir, column=6).value)
                if detail_float_try == 1.2:
                    continue
                ws.cell(row=ir, column=6).value = detail_float_try
            except Exception as e:
                # print e
                pass
        print 'SAVING file:', export_file_name
        wb.save(os.path.join(export_path, export_file_name))
        return "OK"

    def convert_option_type(self, row, start, end):
        row = list(row)
        for i in range(start, end + 1):
            row[i] = self.conver2int(row[i])
        return row

    def conver2int(self, s):
        if s:
            try:
                i_value = int(s)
                return i_value
            except:
                return s
        return s


if __name__ == '__main__':

    # ArlFunc().export_hu_to_excel("./","export_hu.xlsx", "HU_export_template_ver0.17.xlsx")

    # ArlFunc().export_hu_to_excel_by_userid("./", "export_hu.xlsx", "HU_export_template_ver0.17.xlsx", user_id=213, )
    # ArlFunc().export_group_hu_to_excel_by_userid("./", "export_hu.xlsx", "HU_export_template_ver0.17.xlsx", user_id=213)
    # ArlFunc().export_definition_to_excel("./", "export_tagl_def.xlsx", "TAGL_def_template_ver0.09.xlsx")
    # ArlFunc().export_definition_to_excel_by_userid("./", "export_tagl_def.xlsx", "TAGL_def_template_ver0.09.xlsx", user_id=213)
    # ArlFunc().export_group_definition_to_excel_by_userid("./", "export_tagl_def.xlsx", "TAGL_def_template_ver0.09.xlsx", user_id=213)
    ArlFunc().export_arl_to_object_by_all('./', 'testSchedule.xlsx')

