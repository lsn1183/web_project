# -*- coding: UTF-8 -*-
import os
from openpyxl import Workbook
from Source.spec2db_server.arl.arl_base import ServiceBase
import json
import re

class AnaSeqInfo(ServiceBase):
    def __init__(self):
        ServiceBase.__init__(self)
        self.wb = Workbook()
        self.basic_req_list = []
        self.check_result_dict = {}
        self.create_basic_dict()

    def create_basic_dict(self):
        self._pg.connect()
        sql_cmd = '''
            (select basic_req from spec.analysis where basic_req is not null
            union
            select basic_req from spec.basic_req_analysis where basic_req is not null)
            order by basic_req
        '''
        self._pg.execute(sql_cmd)
        for db_data in self._pg.fetchall():
            if not db_data[0]:
                continue

            self.basic_req_list.append(db_data[0])

        self.basic_req_list = set(self.basic_req_list)

        sqlcmd = '''select definition_id, module_name, seq_content, check_result, reason
                    from temp_seq_checkresult order by definition_id'''
        self._pg.execute(sqlcmd)
        for db_data in self._pg.fetchall():
            if db_data[0] not in self.check_result_dict:
                self.check_result_dict[db_data[0]] = []

            self.check_result_dict[db_data[0]].append(list(db_data))

        self._pg.close()

    def get_dest_ana_info(self):
        sql_cmd = '''
            select arl.major_category, arl.medium_catetory, arl.small_category, arl.detail,
                dest.definition_id, grp.group_name, usr.user_name, 
                ana_seq.seq_id, ana_seq.seq, model_info.model_id, model_info.model, 
                ana.rel_requirement, ana.lock_status
            from public.price_ana as dest
            left join analysis_split_sequence as ana_seq on dest.definition_id = ana_seq.definition_id
            left join spec.analysis_model_type as model_info on ana_seq.model_id=model_info.model_id
            left join spec.analysis as ana on dest.definition_id = ana.definition_id
            left join spec.arl as arl on arl.arl_id = ana_seq.arl_id
            left join spec.arl_group as grp on grp.group_id=arl.group_id
            left join spec.arl_user as usr on usr.user_id=arl.user_id
            where ana_seq.seq_id > 0
            order by dest.definition_id, ana_seq.seq_id
        '''
        self._pg.connect()
        self._pg.execute(sql_cmd)

        dest_ana_info_list = [list(db_data) for db_data in self._pg.fetchall()]

        self._pg.close()

        return dest_ana_info_list


    def get_dest_basicreq_info(self):
        self._pg.connect()
        sql_cmd = '''
            select basic_req_name_chain from basic_req_call
        '''
        basic_req_name_list = []
        self._pg.execute(sql_cmd)
        for db_data in self._pg.fetchall():
            basic_req_name_list.extend(db_data[0].split("->"))
        basic_req_name_list = set(basic_req_name_list)

        sql_cmd = '''
            select arl.major_category, arl.medium_catetory, arl.small_category, arl.detail,
                ana.definition_id, grp.group_name, usr.user_name, 
                ana_seq.seq_id, ana_seq.seq, model_info.model_id, model_info.model, 
                ana.rel_requirement, ana.basic_req, ana.lock_status
            from spec.analysis as ana
            left join analysis_split_sequence as ana_seq on ana.definition_id = ana_seq.definition_id
            left join spec.analysis_model_type as model_info on ana_seq.model_id=model_info.model_id
            left join spec.arl as arl on arl.arl_id = ana_seq.arl_id
            left join spec.arl_group as grp on grp.group_id=arl.group_id
            left join spec.arl_user as usr on usr.user_id=arl.user_id
            where ana_seq.seq_id > 0 and ana.basic_req in (%s)
            order by ana.basic_req, ana.definition_id, ana_seq.seq_id
        ''' % ','.join(["'" + n + "'" for n in basic_req_name_list])
        self._pg.execute(sql_cmd)
        dest_ana_info_list = [list(db_data) for db_data in self._pg.fetchall()]

        sql_cmd = '''
                    select '' as major_category, '' as medium_catetory, '基本要件' as small_category, ana.basic_req,
                        ana.definition_id, grp.group_name, usr.user_name, 
                        ana_seq.seq_id, ana_seq.seq, model_info.model_id, model_info.model, 
                        ana.rel_requirement, ana.lock_status
                    from spec.basic_req_analysis as ana
                    left join basic_seq_analysis_split_sequence as ana_seq on ana.definition_id = ana_seq.definition_id
                    left join spec.analysis_model_type as model_info on ana_seq.model_id=model_info.model_id
                    left join spec.arl_group as grp on grp.group_id=ana.group_id
                    left join spec.arl_user as usr on usr.user_id=ana.user_id
                    where ana_seq.seq_id > 0 and ana.basic_req in (%s)
                    order by ana.basic_req, ana.definition_id, ana_seq.seq_id
                ''' % ','.join(["'" + n + "'" for n in basic_req_name_list])
        self._pg.execute(sql_cmd)
        for db_data in self._pg.fetchall():
            dest_ana_info_list.append(list(db_data))

        self._pg.close()

        return dest_ana_info_list


    def get_seq_dia(self, s):
        if s:
            p = '[\w_]+_0\d+'
            m = re.search(p, s)
            if m:
                return m.group(0)

        return None

    def get_usecase_info(self, db_line_data):
        if db_line_data[9] in [1,2]:
            return '无USECASE'

        seq_info = db_line_data[8]
        if seq_info.find("_0")>=0:
            return self.get_seq_dia(seq_info)

        if db_line_data[8].find("※") >= 0 or db_line_data[8].find("参照") >= 0:
            return '基本要件'

        model_name_list = json.loads(db_line_data[10])
        model_name_list = [_data for _data in model_name_list if _data]
        b_found_cannot_fix = False
        if db_line_data[4] in self.check_result_dict:
            for db_data in self.check_result_dict[db_line_data[4]]:
                if db_data[1] == model_name_list[-1] and db_data[2].find("(%d)" % db_line_data[7]) >= 0:
                    b_found_cannot_fix = True
                    break

        if b_found_cannot_fix == True:
            return '不要修正'
        else:
            if db_line_data[-1] == 0:
                return '要修正[无锁]'
            else:
                return '要修正[锁]'

    def out_common_module_seq(self, dest_ana_list, ws):
        line_out_list = []
        for i,dest_data in enumerate(dest_ana_list):
            if i==0:
                prev_ana_def_id = dest_data[4]
                line_out_list.extend(dest_data[:7])

            if dest_data[4]!= prev_ana_def_id:
                if dest_ana_list[i-1][11]:
                    for basic_req in dest_ana_list[i-1][11].split("\n"):
                        if basic_req in self.basic_req_list:
                            line_out_list.append('[基本要件]'+basic_req)
                            line_out_list.append('')
                            line_out_list.append('')
                line_out_list.append('END')

                ws.append(line_out_list)
                line_out_list = []
                line_out_list.extend(dest_data[:7])
                prev_ana_def_id = dest_data[4]

            model_name_list = json.loads(dest_data[10])
            model_name_list = [_data for _data in model_name_list if _data]
            line_out_list.append('->'.join(model_name_list))
            line_out_list.append(self.get_usecase_info(dest_data))
            line_out_list.append('')

        if len(line_out_list) > 0:
            if dest_ana_list[-1][11]:
                for basic_req in dest_ana_list[-1][11].split("\n"):
                    if basic_req in self.basic_req_list:
                        line_out_list.append('[基本要件]' + basic_req)
                        line_out_list.append('')
                        line_out_list.append('')
            line_out_list.append('END')
            ws.append(line_out_list)

    def out_common_uni_seq_info(self, dest_ana_list, ws):
        xls_out_list = []
        for i, dest_data in enumerate(dest_ana_list,0):
            line_data = []
            model_name_list = json.loads(dest_data[10])
            model_name_list = [_data for _data in model_name_list if _data]
            line_data.append(model_name_list[0])
            line_data.append('->'.join(model_name_list[1:-1]))
            line_data.append(model_name_list[-1])
            line_data.append(dest_data[0])
            line_data.append(dest_data[1])
            line_data.append(dest_data[2])
            line_data.append(dest_data[3])
            line_data.append(dest_data[4])
            line_data.append(self.get_usecase_info(dest_data))
            line_data.append(dest_data[8])
            #line_data.append('')

            if i == len(dest_ana_list)-1:
                line_data.append('-')
                line_data.append('-')
                line_data.append('')
                xls_out_list.append(line_data)
                continue

            if dest_data[4] == dest_ana_list[i+1][4]:
                next_model_name_list = json.loads(dest_ana_list[i+1][10])
                next_model_name_list = [_data for _data in next_model_name_list if _data]
                line_data.append(next_model_name_list[-1])
                line_data.append(self.get_usecase_info(dest_ana_list[i+1]))
                line_data.append('')
            else:
                line_data.append('-')
                line_data.append('-')
                line_data.append('')

            xls_out_list.append(line_data)

        xls_out_list = sorted(xls_out_list, key=lambda x:x[2])

        for xls_data in xls_out_list:
            ws.append(xls_data)


    def out_summary_info(self, ana_list, basicreq_list, ws):
        ana_cnt_list = [0,0,0]
        ana_noif_dict = {}
        basicreq_cnt_list = [0,0,0]
        basicreq_noif_dict = {}
        use_case_list = []

        for db_data in ana_list:
            seq_info = self.get_usecase_info(db_data)

            if not seq_info:
                ana_cnt_list[0] += 1
                continue

            if seq_info.find("_0")>=0:
                ana_cnt_list[0]+=1
                use_case_list.append(seq_info)
            elif seq_info in ('不要修正',):
                ana_cnt_list[2] += 1
                if db_data[10] in ana_noif_dict:
                    ana_noif_dict[db_data[10]]+=1
                else:
                    ana_noif_dict[db_data[10]] = 1
            else:
                ana_cnt_list[1] += 1

        for db_data in basicreq_list:
            seq_info = self.get_usecase_info(db_data)
            if not seq_info:
                basicreq_cnt_list[0] += 1
                continue

            if seq_info.find("_0")>=0:
                basicreq_cnt_list[0]+=1
                use_case_list.append(seq_info)
            elif seq_info in ('不要修正',):
                basicreq_cnt_list[2] += 1
                if db_data[10] in basicreq_noif_dict:
                    basicreq_noif_dict[db_data[10]]+=1
                else:
                    basicreq_noif_dict[db_data[10]] = 1
            else:
                basicreq_cnt_list[1] += 1

        ws.append(['','非基本要件', '基本要件','汇总','去重后的Usecase数量'])
        ws.append(['有usecase数[OK]', ana_cnt_list[0], basicreq_cnt_list[0], ana_cnt_list[0]+basicreq_cnt_list[0],
                   len(set(use_case_list))])
        ws.append(['无usecase数[需要修正]', ana_cnt_list[1], basicreq_cnt_list[1], ana_cnt_list[1] + basicreq_cnt_list[1]])
        ws.append(['无usecase数[找不到外部IF]', ana_cnt_list[2], basicreq_cnt_list[2], ana_cnt_list[2] + basicreq_cnt_list[2]])
        ws.append(['Total', sum(ana_cnt_list), sum(basicreq_cnt_list), sum(ana_cnt_list) + sum(basicreq_cnt_list)])


        merge_noif_dict = {}
        ws.append([])
        ws.append(['非基本要件找不到unitcase的统计'])
        ws.append(['Layer','Component','Unit','次数'])
        for _k, _v in ana_noif_dict.iteritems():
            if _k in merge_noif_dict:
                merge_noif_dict[_k]+=_v
            else:
                merge_noif_dict[_k] = _v
            model_name_list = json.loads(_k)
            model_name_list = [_data for _data in model_name_list if _data]
            ws.append([model_name_list[0],'->'.join(model_name_list[1:-1]), model_name_list[-1], _v])

        ws.append([])
        ws.append(['基本要件找不到unitcase的统计'])
        ws.append(['Layer', 'Component', 'Unit', '次数'])
        for _k, _v in basicreq_noif_dict.iteritems():
            if _k in merge_noif_dict:
                merge_noif_dict[_k]+=_v
            else:
                merge_noif_dict[_k] = _v
            model_name_list = json.loads(_k)
            model_name_list = [_data for _data in model_name_list if _data]
            ws.append([model_name_list[0], '->'.join(model_name_list[1:-1]), model_name_list[-1], _v])

        ws.append([])
        ws.append(['Total找不到unitcase的统计'])
        ws.append(['Layer', 'Component', 'Unit', '次数'])
        for _k, _v in merge_noif_dict.iteritems():
            model_name_list = json.loads(_k)
            model_name_list = [_data for _data in model_name_list if _data]
            ws.append([model_name_list[0], '->'.join(model_name_list[1:-1]), model_name_list[-1], _v])

    def out_basic_call_info(self, ws):
        self._pg.connect()
        sql_cmd = '''
            select definition_id, basic_req_name_chain, chain_result from basic_req_call order by definition_id
        '''
        self._pg.execute(sql_cmd)
        for db_data in self._pg.fetchall():
            ws.append(list(db_data))

        self._pg.close()

    def out_basic_req_list(self, ws):
        self._pg.connect()
        sql_cmd = '''
                    select basic_req_name_chain from basic_req_call
                '''
        basic_req_name_list = []
        self._pg.execute(sql_cmd)
        for db_data in self._pg.fetchall():
            basic_req_name_list.extend(db_data[0].split("->"))
        basic_req_name_list = set(basic_req_name_list)

        sql_cmd = '''
                select ana.basic_req,ana.definition_id
                from spec.basic_req_analysis as ana
                where ana.basic_req in (%s)
                order by ana.basic_req, ana.definition_id
            ''' % ','.join(["'" + n + "'" for n in basic_req_name_list])
        self._pg.execute(sql_cmd)
        req_name_dict = {}
        for db_data in self._pg.fetchall():
            if not db_data[0] in req_name_dict:
                req_name_dict[db_data[0]] = []
            req_name_dict[db_data[0]].append(db_data[1])

        sql_cmd = '''
                    select ana.basic_req,ana.definition_id
                    from spec.analysis as ana
                    where ana.basic_req in (%s)
                    order by ana.basic_req, ana.definition_id
                ''' % ','.join(["'" + n + "'" for n in basic_req_name_list])
        self._pg.execute(sql_cmd)
        for db_data in self._pg.fetchall():
            if not db_data[0] in req_name_dict:
                req_name_dict[db_data[0]] = []
            req_name_dict[db_data[0]].append(db_data[1])

        for _k,_v in req_name_dict.iteritems():
            ws.append([_k, _v[0], len(_v)])
            for req_id in _v[1:]:
                ws.append(['',req_id])

        self._pg.close()


    def out_to_excel(self):
        dest_ana_info_list = self.get_dest_ana_info()
        basic_req_info_list = self.get_dest_basicreq_info()
        ws = self.wb.active
        ws.title = "Summary"
        self.out_summary_info(dest_ana_info_list, basic_req_info_list, ws)

        ws = self.wb.create_sheet(u"基本要件一览")
        ws.append(['基本要件', '关联id', '关联id数量'])
        self.out_basic_req_list(ws)

        ws = self.wb.create_sheet(u"关联基本调用一览")
        ws.append(['definition_id', '关联基本要件一览', '结果'])
        self.out_basic_call_info(ws)

        ws = self.wb.create_sheet(u"非基本要件_Module时序")
        ws.append(['大分類','中分類','小分類','詳細','TAGL要件定義ID','组','担当'])
        self.out_common_module_seq(dest_ana_info_list, ws)

        ws = self.wb.create_sheet(u"非基本要件_UseCase整理_Unit单位")
        ws.append(['Layer','Component','Unit','大分類','中分類','小分類','詳細','TAGL要件定義ID','结合元UseCase','结合元testCase','结合先Unit','结合先（被结合）UseCase','结合元（被结合）testCase'])
        self.out_common_uni_seq_info(dest_ana_info_list, ws)


        ws = self.wb.create_sheet(u"基本要件_Module时序")
        ws.append(['大分類','中分類','小分類','詳細','TAGL要件定義ID','组','担当'])
        self.out_common_module_seq(basic_req_info_list, ws)

        ws = self.wb.create_sheet(u"基本要件_UseCase整理_Unit单位")
        ws.append(['Layer','Component','Unit','大分類','中分類','小分類','詳細','TAGL要件定義ID','结合元UseCase','结合元testCase','结合先Unit','结合先（被结合）UseCase','结合元（被结合）testCase'])
        self.out_common_uni_seq_info(basic_req_info_list, ws)

        self.wb.save('/mnt/sharedoc/temp/out1.xlsx')



if __name__ == '__main__':
    import sys
    reload(sys)
    sys.setdefaultencoding('UTF-8')

    #os.system('psql -v ON_ERROR_STOP=1 -h 192.168.0.56 spec2db_1120 -U postgres -f ../../../Doc/报价/split_sequence.sql')
    AnaSeqInfo().out_to_excel()