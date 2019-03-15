import os
import re
import win32com.client
import os, shutil
from openpyxl import Workbook


class DesignDoc(object):
    def __init__(self):
        pass

    def clean_doc(self, file_path):
        new_file = "new_17cy_2"
        wk = Workbook()
        ws1 = wk.create_sheet("没有设计文档的路径")
        ws2 = wk.create_sheet("程序打不开文档的路径")
        # ws3 = wk.create_sheet("基本设计文档一览")
        # ws3.append(['文件名', '文件路径'])
        ws4 = wk.create_sheet("详细设计文档一览")
        ws4.append(['文件名', '文件路径'])
        for root, dirs, files in os.walk(file_path):
            new_root = root.replace("17cy_2", new_file, 1)
            if not os.path.exists(new_root):
                os.makedirs(new_root)
            for d in dirs:
                d = os.path.join(new_root, d)
                if not os.path.exists(d):
                    os.mkdir(d)
            check = True
            print(root)
            # if root == r"\\192.168.37.35\share\users\hongcz\17cy_2\06_DetailedDesign\DCU\Service\NMTslService":
            #     print(root)
            for file in files:
                file_name, ext_name = os.path.splitext(file)
                filename = os.path.basename(file)
                if ext_name.lower() in ('.xls', '.xlsx'):
                    if "設計基準" not in filename:
                        if self.is_design_file(filename):
                            check = False
                            copy_file_path = os.path.join(new_root, file_name + ".xlsx")
                            # if os.path.exists(copy_file_path):
                            #     continue
                            if ext_name.lower() == ".xlsx":
                                shutil.copyfile(os.path.join(root, file), copy_file_path)
                                # ws3.append([filename, copy_file_path])
                                ws4.append([filename, copy_file_path])
                                continue
                            xl = win32com.client.Dispatch("Excel.Application")
                            xl.DisplayAlerts = False
                            try:
                                wb = xl.Workbooks.Open(os.path.join(root, file))
                            except:
                                ws2.append([os.path.join(root, file)])
                                continue
                            wb.SaveAs(copy_file_path, FileFormat=51)
                            # ws3.append([filename.replace('.xls', '.xlsx'), copy_file_path])
                            ws4.append([filename.replace('.xls', '.xlsx'), copy_file_path])
                            wb.Close()
                            xl.Quit()
            if not dirs and check:
                if '.git' not in root:
                    ws1.append([root])
        wk.save(r'C:\Users\yuyin\Desktop\详细设计导入文档统计.xlsx')
        wk.close()

    def test(self, emf_imag_file):
        import os, sys
        from PIL import Image
        Image._initialized = 2  # added this line
        f, e = os.path.splitext(emf_imag_file)
        outfile = f + ".jpg"
        if emf_imag_file != outfile:
            Image.open(emf_imag_file).convert('RGB').save(outfile, "JPEG")  # added "JPEG"

    def my_match(self, src, sub_str_list):

        # p_str = r'*%s*', p_str
        for sub in sub_str_list:
            p = re.compile(sub, flags=re.I)
            g = p.findall(src)
            if g:
                return True
        return False

    def is_design_file(self, file_name):
        match_str_list = ['detail[_]*design', 'basic[_]*design', 'base[_]*design', "基本设计", "详细设计"]
        if self.my_match(file_name, match_str_list):
            match_str_list2 = ['整体基本设计', '整体详细设计', 'Review', 'review', 'Checklists', '議事',
                               '设计基准书', '议事', '报告书']
            if not self.my_match(file_name, match_str_list2):
                return True
        return False


if __name__ == '__main__':
    file_path = r"\\192.168.37.35\share\users\hongcz\17cy_2\06_DetailedDesign"
    DesignDoc().clean_doc(file_path)
    # emf_imag_file = r"C:/Users/yuyin/Desktop/new_17cy/05_basic_design/Common/CL/wayland/WaylandCustomize_BaseDesign/image/Activity_activity0.emf"
    # DesignDoc().test(emf_imag_file)
    # str = r"\\192.168.37.35\share\users\hongcz\17cy_2\06_DetailedDesign\MEU\Service\Navi\Path\NearLink\DetailDesign_交通障碍的相交判定_详细设计.xlsx"
    # print (DesignDoc().is_design_file(str))

