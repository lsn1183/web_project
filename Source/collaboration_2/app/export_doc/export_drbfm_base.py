import os
from openpyxl import load_workbook


class ExportDrbfmBase(object):
    def __init__(self, start_row=None, sheet_name=None, templet_file=None, export_format=".xlsx"):
        self.export_format = export_format
        self.drbfm_data = None
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.templet_file = templet_file

    def write_drbfm_data(self, out_dir, doc_name):
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
        templet_file = self.templet_file
        drbfm_data = self.drbfm_data
        openpyxl_wb = load_workbook(templet_file)
        ws = openpyxl_wb.get_sheet_by_name('DRBFMSheet')
        start_row = self.start_row
        end_row = start_row
        for drbfm in drbfm_data:
            change_content = drbfm.get("change_content")
            model_list = drbfm.get("model_list")
            for model in model_list:
                model_name = model.get("model_name")
                failuremode_list = model.get("failuremode_list")
                failuremode = ",".join(failuremode_list)
                row_data = [change_content, model_name, failuremode]
                self.write_excel_cell(ws, end_row, row_data)
                end_row += 1
            ws.merge_cells("A%s:A%s" % (start_row, end_row-1))
            start_row = end_row
        file_url = os.path.join(out_dir, doc_name+"%s%s" % ("_DRBFM", self.export_format))
        openpyxl_wb.save(file_url)
        return file_url

    def write_excel_cell(self, sheet, row, row_data):
        start_col = 1
        end_col = len(row_data) + 1
        while start_col < end_col:
            sheet.cell(row=row, column=start_col).value = row_data[start_col-1]
            start_col += 1