# -*- coding: UTF-8 -*-
from flask import request
from flask import current_app
from flask_restful import Resource
from token_manage import auth
from app.ctrl.ctrl_ds_doc import CtrlDsDoc
from app.ctrl.ctrl_ds_section import CtrlDSSection
from app.ctrl.ctrl_comment import CtrlComment
from app.ctrl.ctrl_ds_rel_spec import CtrlDSRelSpec
from app.ctrl.ctrl_ds_scene import CtrlDSScene
from app.ctrl.ctrl_ds_checklist_item import CtrlDSCheckListItem
from app.ctrl.ctrl_ds_section import CtrlDSConsider
from app.ctrl.ctrl_ds_drbfm import CtrlDsDrbfm
from app.ctrl.ctrl_attach import CtrlDSAstah
from app.ctrl.ctrl_ds_if import CtrlDsIf
from app.ctrl.ctrl_doc_input import CtrlDSInput
from app.ctrl.ctrl_import_doc import CtrlImportDoc
from app.api_1_0.api_model import del_cache_model_top


class ApiDSDoc(Resource):
    @auth.login_required
    def post(self):
        data_json = request.get_json()
        ctrl_doc = CtrlDsDoc()
        result = ctrl_doc.update_ds_doc(data_json)
        return result

    def put(self):
        data_json = request.get_json()
        ctrl_doc = CtrlDsDoc()
        result = ctrl_doc.update_ds_doc(data_json)
        return result

    def get(self, doc_id, type=None):
        """
        :param doc_id:
        :return: 获取设计文档信息
        """
        result = {"result": "OK"}
        ctrl_doc = CtrlDsDoc()
        # ctrl_section = CtrlDSSection()
        try:
            ds_doc = ctrl_doc.get_doc(doc_id, type)
            # doc_section = ctrl_section.get_sections(doc_id)
            # sec_type_list = ctrl_section.get_section_type(orient='list')
            # for type_dict in sec_type_list:
            #     sec_type = type_dict.get('sec_type')
            #     if sec_type not in doc_section:
            #         doc_section[sec_type] = []
            # ds_doc.update(doc_section)
        # doc = {"doc_type": "BASIC",
        #        "title": "文档标题",
        #        "ver": "文档版本",
        #        "model_id": "文档所属模块",
        #        "create_time": "2018-5-9",
        #        "update_time": "2018-5-9",
        #        "creator": "文档创建者",
        #        "editor": "最后更新人/编辑者",
        #        "locked": "N",
        #        "SUMMARY": "Markdown0",
        #        "USERCASE": [{"content": "Markdown1",
        #                      "tags": [{"tag_id": 106, "tag": "NativeAPP"},
        #                               {"tag_id": 107, "tag": "Html5APP"},
        #                               ]
        #                      }
        #                     ],
        #        "SEQUENCE": [{"content": "Markdown2",
        #                      "tags": [{"tag_id": 106, "tag": "NativeAPP"},
        #                               {"tag_id": 107, "tag": "Html5APP"},
        #                               ]
        #                      }
        #                     ],  # 时序图
        #        "CLASS": [],  # 类图
        #        "BLOCK": [],
        #        "RESOURCE": [],  # 资源
        #        "IF": [],  # 接口
        #        "STD": [],  # 状态图
        #        }
            result["content"] = ds_doc
        except Exception as e:
            current_app.logger.error('%s' % e)
            result["result"] = "NG"
            result["error"] = "服务异常！请联系管理员！"
        return result

    def delete(self, _id, type, commit_user):
        """
        :param id:
        :param del_type:  doc/usercase
        :return:
        """
        result = {"result": "OK"}
        try:
            if type == "doc":
                ctrl_obj = CtrlDsDoc()
            elif type == "usercase":
                ctrl_obj = CtrlDSSection()
            error = ctrl_obj.delete(_id, commit_user)
            if error:
                result["result"] = "NG"
                result["error"] = error
        except Exception as e:
            current_app.logger.error('%s' % e)
            result["result"] = "NG"
            result["error"] = "服务异常！请联系管理员！"
        return result


class ApiNewDSDoc(Resource):
    @auth.login_required
    def get(self, doc_id):
        """
        查看基本/详细设计信息
        :param doc_id:
        :return:
        """
        result = {"result": "NG", "content": ''}
        ctrl_doc = CtrlDsDoc()
        doc_data = ctrl_doc.get_ds_doc(doc_id)
        if doc_data:
            result["result"] = "OK"
            result["content"] = doc_data
        else:
            result["error"] = "该文档不存在！"
        return result

    @auth.login_required
    def post(self):
        """
        新增设计文档
        :return:
        """
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        doc_type = data_json.get("doc_type")
        proj_id = data_json.get("proj_id")
        model_id = data_json.get("model_id")
        commit_user = data_json.get("commit_user")
        doc_id, error = CtrlDsDoc().add_new_doc(doc_type, proj_id, model_id, commit_user)
        if doc_id:
            result["result"] = "OK"
            result["content"] = doc_id
            # 删除
            print('delete_memoized: get_model_top, proj_id=%s' % proj_id)
            del_cache_model_top(proj_id)
        else:
            result["error"] = error
        return result
    

class ApiUpDocVer(Resource):
    @auth.login_required
    def post(self):
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        error = CtrlDsDoc().up_doc_ver(data_json)
        if not error:
            result['result'] = 'OK'
        else:
            result['error'] = error
        return result


class ApiDsDocIf(Resource):
    @auth.login_required
    def post(self):
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        flag, error = CtrlDsIf().add_if(data_json)
        if flag:
            result['result'] = 'OK'
        else:
            result['error'] = error
        return result

    @auth.login_required
    def put(self):
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        flag, error = CtrlDsIf().update_if(data_json)
        if flag:
            result['result'] = 'OK'
        else:
            result['error'] = error
        return result

    @auth.login_required
    def get(self, doc_id):
        result = {"result": "OK"}
        try:
            content = CtrlDsIf().get_if_url(doc_id)
            result['content'] = content
        except Exception as e:
            current_app.logger.error('%s' % e)
            result["result"] = "NG"
            result["error"] = str(e)
        return result

    @auth.login_required
    def delete(self, if_id, commit_user):
        """
        :param if_id:
        :return:
        """
        result = {"result": "NG", 'error': ''}
        flag, error = CtrlDsIf().delete_by_if_id(if_id, commit_user)
        if flag:
            result['result'] = 'OK'
        else:
            result['error'] = error
        return result


class ApiDSDocType(Resource):
    @auth.login_required
    def get(self):
        result = {"result": "NG"}
        doc_types = CtrlDsDoc().get_doc_type(orient='list')
        if doc_types:
            result["content"] = doc_types
            result["result"] = "OK"
        return result


class ApiDSDocStatus(Resource):
    @auth.login_required
    def get(self):
        result = {"result": "NG"}
        status_list = CtrlDsDoc().get_doc_status()
        if status_list:
            result["content"] = status_list
            result["result"] = "OK"
        return result

    @auth.login_required
    def post(self):
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        flag, error = CtrlDsDoc().update_doc_status(data_json)
        if flag:
            result['result'] = 'OK'
        else:
            result['error'] = error
        return result


class ApiDSectionType(Resource):
    @auth.login_required
    def get(self):
        result = {"result": "NG"}
        sec_types = CtrlDSSection().get_section_type()
        if sec_types:
            result = {"result": "OK", "content": sec_types}
        return result


class ApiDSDocUsecase(Resource):
    @auth.login_required
    def get(self, doc_id):
        """
        获取文档下所有usecase的说明
        :param doc_id:
        :return:
        """
        result = {"result": "NG"}
        content = CtrlDSSection().get_usecase_by_doc_id(doc_id, 'USERCASE')
        if content:
            result = {"result": "OK", "content": content}
        return result

    @auth.login_required
    def post(self):
        """
        保存和修改文档下usecase的说明
        :return:
        """
        result = {"result": "NG", 'error': ''}
        data_json = request.get_json()
        sec_obj = CtrlDSSection()
        flag, error = sec_obj.usecase_add(data_json)
        if flag:
            result["result"] = "OK"
        else:
            result["error"] = error
        return result


class ApiNewDsSection(Resource):
    @auth.login_required
    def get(self, doc_id, sec_type):
        result = {"result": "NG", "contnet": ""}
        section_data = CtrlDSSection().get_section_detail(doc_id, sec_type)
        if section_data:
            result["result"] = "OK"
            result["content"] = section_data
        return result

    @auth.login_required
    def post(self):
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        sec_obj = CtrlDSSection()
        flag, error = sec_obj.section_commont(data_json)
        if flag:
            result["result"] = "OK"
        else:
            result["error"] = error
        return result


class ApiDsUcSub(Resource):
    @auth.login_required
    def get(self, usecase_id, sec_type):
        """
        取usecase下的sequence
        :return:
        """
        result = {"result": "NG", "content": ''}
        sec_data = CtrlDSSection().get_sub_by_usecase(usecase_id, sec_type)
        if sec_data:
            result["result"] = "OK"
            result["content"] = {'usecase_id': usecase_id, 'sec_type': sec_type, 'section_list': sec_data}
        return result

    @auth.login_required
    def post(self):
        """
        編輯usecase下的usecase
        :return:
        """
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        sec_obj = CtrlDSSection()
        flag, error = sec_obj.update_uc_sub(data_json)
        if flag:
            result["result"] = "OK"
        else:
            result["error"] = error
        return result


class ApiDSUcLevel2(Resource):
    @auth.login_required
    def get(self, usecase_id):
        result = {"result": "NG", "content": ''}
        data_json = CtrlDSSection().get_level2_usecase(usecase_id)
        if data_json:
            result["result"] = "OK"
            result["content"] = data_json
        return result

    @auth.login_required
    def post(self):
        """
        编辑usecase下子usecase
        :return:
        """
        data_json = request.get_json()
        result = {"result": "NG", 'error': ''}
        sec_obj = CtrlDSSection()
        flag, error = sec_obj.update_usecase_level2(data_json)
        if flag:
            result["result"] = "OK"
        else:
            result["error"] = error
        return result


class ApiDSection(Resource):
    @auth.login_required
    def post(self):
        data_json = request.get_json()
        commit_user = data_json.get('commit_user')
        result = {"result": "NG", 'error': ''}
        sec_type = data_json.get("sec_type")
        # if sec_type == "SPEC":
        #     sec_obj = CtrlDSRelSpec()
        # elif sec_type == "SCENES":
        #     sec_obj = CtrlDSScene()
        # elif sec_type in ("USERCASE", "SEQUENCE", "CLASS",
        #                   "BLOCK", "RESOURCE", "STD"):
        sec_obj = CtrlDSSection()
        # elif sec_type == "DRBFM":
        #     sec_obj = CtrlDsDrbfm()
        flag, error = sec_obj.add(data_json, commit_user)
        if flag:
            result["result"] = "OK"
        else:
            result["error"] = error
        return result

    @auth.login_required
    def get(self, doc_id, sec_type, condition=None):
        """
        :param sec_id: User Case ID
        :param sec_type:
        :return:
        """
        result = {"result": "NG", "content": ''}
        # if sec_type == "SPEC":
        #     sec_obj = CtrlDSRelSpec()
        # elif sec_type == "SCENES":
        #     sec_obj = CtrlDSScene()
        # elif sec_type in ("USERCASE", "SEQUENCE", "CLASS",
        #                   "BLOCK", "RESOURCE", "STD"):
        sec_obj = CtrlDSSection()
        # elif sec_type == "DRBFM":
        #     sec_obj = CtrlDsDrbfm()
        content, status = sec_obj.get(doc_id, sec_type, condition)
        if content or status:
            if status:
                result["status"] = status
            result["result"] = "OK"
            result["content"] = content
        return result


class ApiUsecaseDetail(Resource):
    @auth.login_required
    def get(self, usecase_id):
        """
        取usecase的详细信息
        :param usecase_id:
        :return:
        """
        result = {"result": "NG", "content": ''}
        obj = CtrlDSSection()
        usecae_data = obj.get_usecase_detail(usecase_id)
        if usecae_data:
            result["result"] = "OK"
            result["content"] = usecae_data
        return result

    def post(self, data_json):
        """
        编辑usecase下sequence
        :param data_json:
        :return:
        """


class ApiAllUsecase(Resource):
    @auth.login_required
    def get(self, doc_id):
        """
        取一个文档的所有usecase
        :param doc_id:
        :return:
        """
        result = {"result": "NG", "content": []}
        obj = CtrlDSSection()
        data_list = obj.get_all_usecase_tille(doc_id)
        if data_list:
            result['result'] = "OK"
            result['content'] = data_list
        return result


class ApiDSConsider(Resource):
    @auth.login_required
    def post(self):
        data_json = request.get_json()
        result = {"result": "NG", "sec_id": 0, 'error': ''}
        sec_obj = CtrlDSConsider()
        flag, error = sec_obj.update_condsiders(data_json)
        if flag:
            result["result"] = "OK"
        else:
            result["error"] = error
        result["sec_id"] = data_json.get("sec_id")
        return result

    @auth.login_required
    def get(self, sec_id, csd_range='ALL'):
        sec_obj = CtrlDSConsider()
        result = {"result": "NG", "content": []}
        content = sec_obj.get(sec_id, csd_range)
        if content:
            result["result"] = "OK"
            result["content"] = content
        return result


class ApiDSDoc2(Resource):
    @auth.login_required
    def get(self, model_id, doc_type, tag_id):
        """
        :param model_id, doc_type, tag_id
        :return: 获取Model下某tag下的所有文档
        """
        result = {"result": "NG"}
        ctrl_doc = CtrlDsDoc()
        docs = ctrl_doc.get_ds_doc_by_tag(model_id, tag_id, doc_type)
        if docs:
            result["result"] = "OK"
            result["content"] = docs
        return result


class ApiComment(Resource):
    @auth.login_required
    def post(self):
        result = {"result": "NG", "error": ''}
        data_json = request.get_json()
        comment = CtrlComment()
        try:
            rst, message = comment.add(data_json)
            if rst:
                result["result"] = "OK"
            else:
                result["error"] = message
        except Exception as e:
            result["error"] = str(e)
        finally:
            return result

    def get(self, sec_id):
        result = {"result": "NG"}
        comment = CtrlComment()
        # 测试用
        # data = {"doc_id": 1, "sec_id": 5,
        #         "content": "评论测试5-1", "commentator": "hongcz"}
        # r = comment.add(data)
        comments = comment.get(sec_id=sec_id)
        if comments:
            result = {"result": "OK", "content": comments}
        return result


class ApiDSRelSpec(Resource):
    @auth.login_required
    def get(self, sec_id):
        result = {"result": "NG", "content": []}
        spec = CtrlDSRelSpec().get(sec_id)
        if spec:
            result["result"] = "OK"
            result["content"] = spec
        return result


class ApiUcRelSpec(Resource):
    @auth.login_required
    def get(self, doc_id):
        result = {"result": "NG", "content": []}
        spec = CtrlDSRelSpec().get(doc_id)
        micro_ver = CtrlDsDoc().get_doc_ver(doc_id)
        if spec:
            result["result"] = "OK"
            result["content"] = spec
        result["micro_ver"] = micro_ver
        return result

    @auth.login_required
    def post(self):
        result = {"result": "NG", "doc_id": 0, 'error': ''}
        data_json = request.get_json()
        doc_id, error = CtrlDSRelSpec().add(data_json)
        if doc_id:
            micro_ver = CtrlDsDoc().get_doc_ver(doc_id)
            result['result'] = "OK"
            result['doc_id'] = doc_id
            result["micro_ver"] = micro_ver
        else:
            result['error'] = error
        return result


class ApiDSDocAstah(Resource):
    @auth.login_required
    def get(self, doc_id):
        result = {"result": "NG", "content": [], 'error': ''}
        try:
            astah_list = CtrlDSAstah().get_astah_by_doc_id(doc_id)
            if astah_list:
                result["result"] = "OK"
                result["content"] = astah_list
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result

    @auth.login_required
    def post(self):
        result = {"result": "NG", "content": {}, 'error': ''}
        try:
            rst, msg_or_astah = CtrlDSAstah().add(request)
            if not rst:
                result["error"] = msg_or_astah
            else:
                result["result"] = "OK"
                result["content"] = msg_or_astah
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result

    @auth.login_required
    def delete(self, doc_id, commit_user, attach_id=None):
        result = {"result": "NG", 'error': ''}
        try:
            CtrlDSAstah().delete_attach(doc_id, attach_id, commit_user)
            result = {"result": "OK"}
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result


class ApiDSDocInputGet(Resource):
    @auth.login_required
    def post(self):
        result = {"result": "NG", "content": {}, 'error': ''}
        try:
            rst, msg_or_astah = CtrlDSInput().batch_add(request)
            if not rst:
                result["error"] = msg_or_astah
            else:
                result["result"] = "OK"
                result["content"] = msg_or_astah
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result

    @auth.login_required
    def get(self, proj_id=None, _type=None, spec_id=None):
        result = {"result": "NG", "content": [], 'error': '无数据'}
        try:
            if not _type:
                if spec_id:
                    astah_list = CtrlDSInput().get_ver_list_by_spec_id(spec_id)
                else:
                    astah_list = CtrlDSInput().get_astah_by_proj_id(proj_id)
            else:
                astah_list = CtrlDSInput().get_spec_num_by_proj_id(proj_id, _type)
            if astah_list:
                result["result"] = "OK"
                result["content"] = astah_list
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result


class ApiDSDocInput(Resource):
    @auth.login_required
    def post(self):
        result = {"result": "NG", "content": {}, 'error': ''}
        try:
            rst, msg_or_astah = CtrlDSInput().add(request)
            if not rst:
                result["error"] = msg_or_astah
            else:
                result["result"] = "OK"
                result["content"] = msg_or_astah
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result

    @auth.login_required
    def get(self, proj_id, _type):
        result = {"result": "NG", "content": [], 'error': '无数据'}
        try:
            astah_list = CtrlDSInput().get_astah_by_type(proj_id, _type)
            if astah_list:
                result["result"] = "OK"
                result["content"] = astah_list
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result

    @auth.login_required
    def delete(self, commit_user, spec_id=None, ver_id=None):
        result = {"result": "NG", 'error': ''}
        try:
            error = CtrlDSInput().delete_attach(spec_id, ver_id, commit_user)
            if not error:
                result = {"result": "OK"}
            else:
                result["error"] = error
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result


class ApiDSDocImport(Resource):
    # @auth.login_required
    def post(self):
        result = {"result": "NG", "content": {}, 'error': ''}
        try:
            rst, msg_or_astah = CtrlImportDoc().batch_upload_h(request)
            if not rst:
                result["error"] = msg_or_astah
            else:
                result["result"] = "OK"
                result["content"] = msg_or_astah
        except Exception as e:
            current_app.logger.error('%s' % str(e))
            result["error"] = str(e)
        return result

    # @auth.login_required
    def delete(self, h_id):
        """
        :param h_id:头文件ID
        :return:
        """
        result = {"result": "NG", 'error': ''}
        flag, error = CtrlDsIf().delete_by_h_id(h_id)
        if flag:
            result['result'] = 'OK'
        else:
            result['error'] = error
        return result

    #
    # @auth.login_required
    # def get(self, proj_id, _type):
    #     result = {"result": "NG", "content": [], 'error': ''}
    #     try:
    #         astah_list = CtrlDSInput().get_astah_by_type(proj_id, _type)
    #         if astah_list:
    #             result["result"] = "OK"
    #             result["content"] = astah_list
    #     except Exception as e:
    #         current_app.logger.error('%s' % str(e))
    #         result["error"] = str(e)
    #     return result
    #
    # @auth.login_required
    # def delete(self, commit_user, spec_id=None, ver_id=None):
    #     result = {"result": "NG", 'error': ''}
    #     try:
    #         CtrlDSInput().delete_attach(spec_id, ver_id, commit_user)
    #         result = {"result": "OK"}
    #     except Exception as e:
    #         current_app.logger.error('%s' % str(e))
    #         result["error"] = str(e)
    #     return result


class ApiCopyBasicToDetail(Resource):
    @auth.login_required
    def get(self, proj_id, model_id):
        result = {"result": "NG", "content": []}
        doc_list = CtrlDsDoc().get_basic_doc_for_detail(proj_id, model_id)
        if doc_list:
            result['result'] = "OK"
            result['content'] = doc_list
        return result

    # def post(self):
    #     data_json = request.get_json()
    #     doc_id = data_json.get('doc_id')
    #     commit_user = data_json.get('commit_user')
    #     result = {"result": "NG", 'error': ''}
    #     doc_id, error = CtrlDsDoc().copy_basic_to_detail(doc_id, commit_user)
    #     if error:
    #         result['error'] = error
    #     else:
    #         result['result'] = 'OK'
    #         result['doc_id'] = doc_id
    #     return result


class ApiDRBFM(Resource):
    @auth.login_required
    def get(self, doc_id):
        result = {"result": "NG", "content": []}
        drbfm = CtrlDsDrbfm().get_drbfm_new(doc_id)
        if drbfm:
            result['result'] = "OK"
            result['content'] = drbfm
        return result


class ApiDrbfmExport(Resource):
    @auth.login_required
    def get(self, proj_id, doc_id):
        result = {"result": "NG", "msg": '', "file_path": ''}
        msg, file_path = CtrlDsDrbfm().export_drbfm(proj_id, doc_id)
        if file_path:
            result['result'] = "OK"
            result['file_path'] = file_path
        else:
            result["msg"] = msg
        return result


class ApiKnowledgeDoc(Resource):
    @auth.login_required
    def get(self, sec_id):
        result = {"result": "NG", "content": []}
        knowledge_doc = CtrlDSSection().get_knowledge_doc(sec_id)
        if knowledge_doc:
            result['result'] = "OK"
            result['content'] = knowledge_doc
        return result


class ApiImportDsDoc(Resource):
    """
    导入新的文档(只会增加)
    """
    def get(self):
        #15000（17cy:10004; Guide:4; Search:5; Path:3; Map:2; Location:6; Traffic:7; dataengine:496）
        import os
        import time
        # from app.import_doc.ds_doc_import import DsDocImport
        # file_url = r'C:/Users/yuyin/Desktop/new_17cy/06_DetailedDesign/Common/Service/FCService/DetailDesign_ModeControl.xlsx'
        # file_url = file_url.replace("&", "^&")
        # json_url, result = DsDocImport().excel_to_json(file_url)
        # if result:
        #     return {'result': result}
        # # json_url = r"C:\Users\yuyin\Desktop\new_17cy\05_basic_design\Common\temp\Service\Navi\Search\BasicDesign\CategorySelection\【SUNTEC-MOP-100】ReviewMinutes_Checklists_CategorySelection_BasicDesign\json\ReviewMinutes_Checklists_CategorySelection_BasicDesign.json"
        # model_id = 365
        # creator = 'Admin'
        # proj_id = 10004
        # doc_type = 'DETAIL'
        # DsDocImport().read_json(json_url, model_id, creator, doc_type, proj_id)
        ##############################################################################
        # from openpyxl import load_workbook
        # wb = load_workbook(r"C:\Users\yuyin\Desktop\导入文档目录与模块.xlsx")
        # sheet = wb.get_sheet_by_name('Sheet2')
        # start_row = 1
        # max_row = sheet.max_row
        # f_txt = open(r"C:\Users\yuyin\Desktop\详细设计文档异常信息.txt", 'w')
        # f_txt.write("开始时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        # for row in range(start_row, max_row + 1):
        #     file_url = sheet.cell(row=row, column=1).value
        #     if "&" in file_url:
        #         file_url = file_url.replace('&', '%26')
        #     model_id = sheet.cell(row=row, column=2).value
        #     creator = "Admin"
        #     proj_id = 10004
        #     doc_type = "DETAIL"
        #     if not model_id:
        #         continue
        #     print(file_url)
        #     if "&" in file_url:
        #         file_url = file_url.replace('&', '%26')
        #     json_url, result = DsDocImport().excel_to_json(file_url)
        #     if result:
        #         f_txt.write("解析出错的文件：" + file_url + "\n")
        #         continue
        #     DsDocImport().read_json(json_url, model_id, creator, doc_type, proj_id, f_txt)
        # f_txt.write("结束时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        # f_txt.close()
        ##############################################################################
        from openpyxl import load_workbook
        wb = load_workbook(r"C:\Users\yuyin\Desktop\导入文档目录与模块.xlsx")
        # sheet1 = wb.get_sheet_by_name('基本设计文档目录与模块')
        sheet2 = wb.get_sheet_by_name('详细设计文档目录与模块')
        # f_txt1 = open(r"C:\Users\yuyin\Desktop\基本设计文档异常信息.txt", 'w')
        # f_txt1.write("开始时间："+time.strftime("%Y-%m-%d %H:%M:%S")+'\n')
        # self.distir_files(sheet1, "BASIC", f_txt1)
        # f_txt1.write("结束时间：" + time.strftime("%Y-%m-%d %H:%M:%S")+'\n')
        # f_txt1.close()
        f_txt2 = open(r"C:\Users\yuyin\Desktop\详细设计文档异常信息.txt", 'w')
        f_txt2.write("开始时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        self.distir_files(sheet2, "DETAIL", f_txt2)
        f_txt2.write("结束时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        f_txt2.close()
        return {'result': '导入成功！'}

    def distir_files(self, sheet, doc_type, f_txt):
        from app.import_doc.ds_doc_import import DsDocImport
        import os
        start_row = 2
        max_row = sheet.max_row
        for row in range(start_row, max_row + 1):
            files_url = sheet.cell(row=row, column=2).value
            model_id = sheet.cell(row=row, column=4).value
            if not model_id:
                continue
            creator = "Admin"
            proj_id = 10004
            doc_type = doc_type
            file_list = []
            print("正在执行:"+files_url)
            for root, dirs, files in os.walk(files_url):
                if files:
                    for file in files:
                        base_name, ext_name = os.path.splitext(os.path.join(root, file))
                        if ext_name.lower() in ('.xlsx',):
                            xil_filename = file
                            file_list.append(os.path.join(root, xil_filename))
            for file_url in file_list:
                if ("Review" in file_url or "review" in file_url
                        or "議事" in file_url or "Checklists" in file_url
                        or "checklist" in file_url or "Reference" in file_url):
                    continue
                # filepath, tempfilename = os.path.split(file_url)
                # file_name, extension = os.path.splitext(tempfilename)
                # file_name = file_name.replace(".xlsx", "")
                # doc_id = DsDocImport().get_doc_by_name(file_name)
                # if doc_id:
                #     continue
                if "&" in file_url:
                    file_url = file_url.replace("&", "^&")
                json_url, result = DsDocImport().excel_to_json(file_url)
                if result:
                    f_txt.write("解析出错的文件："+file_url+"\n")
                    continue
                DsDocImport().read_json(json_url, model_id, creator, doc_type, proj_id, f_txt)


# class ApiContentJson(Resource):
#     def get(self):
#         from app.ctrl.update_old_section import UpdateOldSection
#         UpdateOldSection().update()
