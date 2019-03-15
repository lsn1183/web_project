from flask_restful import Resource


class ApiImportIfDoc(Resource):
    """
    导入if文档(只会增加)
    """
    def get(self):
        #15000（17cy:10004; Guide:4; Search:5; Path:3; Map:2; Location:6; Traffic:7; dataengine:496）
        import os
        import time
        from app.import_doc.if_doc_import import IfDocImport
        # file_url = r'C:\Users\yuyin\Desktop\17cy_if_doc\05_basic_design\DCU\Control\Setting\Setting_InterfaceSpec.xlsx'
        # file_url = file_url.replace("&", "^&")
        # model_id = 14
        # creator = 'Admin'
        # proj_id = 10004
        # doc_type = 'IF'
        # f_txt = open(r"C:\Users\yuyin\Desktop\if设计文档异常信息.txt", 'w')
        # f_txt.write("开始时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        # IfDocImport().if_import(file_url, model_id, creator, doc_type, proj_id, f_txt, only_doc='NO')
        # f_txt.write("结束时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        # f_txt.close()
        ##############################################################################
        ##############################################################################
        # from openpyxl import load_workbook
        # wb = load_workbook(r"C:\Users\yuyin\Desktop\IF设计导入文档统计.xlsx")
        # sheet = wb.get_sheet_by_name('if设计文档一览')
        # f_txt = open(r"C:\Users\yuyin\Desktop\IF文档异常信息.txt", 'w')
        # f_txt.write("开始时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        # self.distir_files(sheet, "IF", f_txt)
        # f_txt.write("结束时间：" + time.strftime("%Y-%m-%d %H:%M:%S") + '\n')
        # f_txt.close()
        # return {'result': '导入成功！'}
        ##############################################################################
        from app.import_doc.if_doc_import import IfDocImport
        import os
        from openpyxl import load_workbook
        wb = load_workbook(r"C:\Users\yuyin\Desktop\IF设计导入文档统计.xlsx")
        sheet = wb.get_sheet_by_name('if设计文档一览')
        IfDocImport().insert_htm_url(sheet)

    def distir_files(self, sheet, doc_type, f_txt):
        from app.import_doc.if_doc_import import IfDocImport
        import os
        start_row = 2
        max_row = sheet.max_row
        for row in range(start_row, max_row + 1):
            file_url = sheet.cell(row=row, column=2).value
            model_id = sheet.cell(row=row, column=4).value
            only_doc = sheet.cell(row=row, column=5).value
            if not model_id:
                continue
            creator = "Admin"
            proj_id = 10004
            doc_type = doc_type
            print("正在执行:"+file_url)
            if "&" in file_url:
                file_url = file_url.replace("&", "^&")
            IfDocImport().if_import(file_url, model_id, creator, doc_type, proj_id, f_txt, only_doc)